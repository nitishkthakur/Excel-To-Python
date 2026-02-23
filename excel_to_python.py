"""
Excel-to-Python Converter

Reads an Excel workbook and generates:
1. A Python script that performs the same calculations.
2. An input template Excel file for the user to fill in hardcoded values.

Usage:
    python excel_to_python.py <input_excel_file> [--config config.yaml] [--output-dir output]
"""

import argparse
import os
import re
import sys
import textwrap

import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# Regex for valid aRGB hex color strings
_ARGB_REGEX = re.compile(r'^[0-9A-Fa-f]{8}$')

# Excel default formatting constants
DEFAULT_COLUMN_WIDTH = 8.43
DEFAULT_ROW_HEIGHT = 15
DEFAULT_FONT_SIZE = 11

from formula_converter import (
    FormulaConverter,
    HELPER_FUNCTIONS_CODE,
    cell_to_var_name,
    range_to_var_name,
    table_ref_to_var_name,
    col_letter_to_index,
    index_to_col_letter,
)


def load_config(config_path):
    """Load configuration from a YAML file."""
    defaults = {
        "delete_unreferenced_hardcoded_values": False,
    }
    if config_path and os.path.exists(config_path):
        with open(config_path, "r") as f:
            user_config = yaml.safe_load(f) or {}
        defaults.update(user_config)
    return defaults


def parse_workbook(wb):
    """Parse an Excel workbook into a structured representation.

    Returns:
        sheets: dict of sheet_name -> {
            "cells": {(col_letter, row_num): {"value": ..., "formula": ..., "number_format": ..., "font": ..., "fill": ...}},
            "merged_cells": list of merge ranges,
            "col_widths": dict,
            "row_heights": dict,
        }
        tables: dict of table_name -> {
            "sheet": str, "ref": str, "columns": list,
            "header_row": int, "data_start_row": int, "data_end_row": int,
            "col_start": str, "col_end": str,
        }
    """
    sheets = {}
    tables = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {
            "cells": {},
            "merged_cells": [str(m) for m in ws.merged_cells.ranges],
            "col_widths": {},
            "row_heights": {},
        }

        # Collect column widths
        for col_dim in ws.column_dimensions.values():
            if col_dim.width and col_dim.width != DEFAULT_COLUMN_WIDTH:
                sheet_data["col_widths"][col_dim.index] = col_dim.width

        # Collect row heights
        for row_dim_key, row_dim in ws.row_dimensions.items():
            if row_dim.height and row_dim.height != DEFAULT_ROW_HEIGHT:
                sheet_data["row_heights"][row_dim_key] = row_dim.height

        # Collect cell data
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None or cell.data_type == 'f':
                    col_letter = get_column_letter(cell.column)
                    cell_info = {
                        "value": cell.value,
                        "data_type": cell.data_type,
                        "number_format": cell.number_format,
                    }
                    # Store font info
                    if cell.font:
                        font_color = None
                        if cell.font.color and cell.font.color.rgb:
                            rgb_val = str(cell.font.color.rgb)
                            if _ARGB_REGEX.match(rgb_val):
                                font_color = rgb_val
                        cell_info["font"] = {
                            "bold": cell.font.bold,
                            "italic": cell.font.italic,
                            "size": cell.font.size,
                            "color": font_color,
                        }
                    # Store fill info
                    if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                        rgb = cell.fill.start_color.rgb
                        if isinstance(rgb, str) and rgb != '00000000' and _ARGB_REGEX.match(rgb):
                            cell_info["fill_color"] = rgb
                    # Store alignment
                    if cell.alignment:
                        cell_info["alignment"] = {
                            "horizontal": cell.alignment.horizontal,
                            "vertical": cell.alignment.vertical,
                            "wrap_text": cell.alignment.wrap_text,
                        }

                    sheet_data["cells"][(col_letter, cell.row)] = cell_info

        # Collect tables
        for tbl in ws.tables.values():
            tbl_ref = tbl.ref
            tbl_name = tbl.name
            # Parse the reference range
            m = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', tbl_ref)
            if m:
                col_start, row_start, col_end, row_end = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
                # Get column names from header row
                columns = []
                cs = col_letter_to_index(col_start)
                ce = col_letter_to_index(col_end)
                for ci in range(cs, ce + 1):
                    cl = index_to_col_letter(ci)
                    header_cell = sheet_data["cells"].get((cl, row_start))
                    if header_cell:
                        columns.append(str(header_cell["value"]))
                    else:
                        columns.append(f"Column{ci - cs + 1}")

                tables[tbl_name] = {
                    "sheet": sheet_name,
                    "ref": tbl_ref,
                    "columns": columns,
                    "header_row": row_start,
                    "data_start_row": row_start + 1,
                    "data_end_row": row_end,
                    "col_start": col_start,
                    "col_end": col_end,
                }

        sheets[sheet_name] = sheet_data

    return sheets, tables


def classify_cells(sheets, tables):
    """Classify cells as formula cells or hardcoded value cells.

    Returns:
        formula_cells: list of (sheet, col, row, formula_str, cell_info)
        hardcoded_cells: list of (sheet, col, row, value, cell_info)
    """
    formula_cells = []
    hardcoded_cells = []

    for sheet_name, sheet_data in sheets.items():
        for (col, row), cell_info in sheet_data["cells"].items():
            val = cell_info["value"]
            if isinstance(val, str) and val.startswith("="):
                formula_cells.append((sheet_name, col, row, val, cell_info))
            else:
                hardcoded_cells.append((sheet_name, col, row, val, cell_info))

    return formula_cells, hardcoded_cells


def find_all_references(formula_cells, tables):
    """Find all cell/range/table references across all formulas.

    Returns:
        referenced_cells: set of (sheet, col, row)
        referenced_ranges: set of (sheet, col1, row1, col2, row2)
        referenced_tables: set of (table_name, column_name)
    """
    all_ref_cells = set()
    all_ref_ranges = set()
    all_ref_tables = set()

    for sheet_name, col, row, formula, cell_info in formula_cells:
        converter = FormulaConverter(sheet_name, tables)
        converter.convert(formula)
        all_ref_cells.update(converter.referenced_cells)
        all_ref_ranges.update(converter.referenced_ranges)
        all_ref_tables.update(converter.referenced_tables)

    return all_ref_cells, all_ref_ranges, all_ref_tables


def expand_ranges_to_cells(referenced_ranges):
    """Expand range references to individual cell references."""
    cells = set()
    for sheet, c1, r1, c2, r2 in referenced_ranges:
        ci1 = col_letter_to_index(c1)
        ci2 = col_letter_to_index(c2)
        for r in range(r1, r2 + 1):
            for c in range(ci1, ci2 + 1):
                cells.add((sheet, index_to_col_letter(c), r))
    return cells


def filter_hardcoded_cells(hardcoded_cells, referenced_cells, referenced_ranges,
                           referenced_tables, tables, delete_unreferenced):
    """Filter hardcoded cells based on whether they are referenced.

    If delete_unreferenced is True, only keep hardcoded cells that are
    referenced by some formula (directly or via range/table).
    """
    if not delete_unreferenced:
        return hardcoded_cells

    # Expand ranges to cells
    range_cells = expand_ranges_to_cells(referenced_ranges)

    # Expand table references to cells
    table_cells = set()
    for tbl_name, col_name in referenced_tables:
        if tbl_name in tables:
            tbl = tables[tbl_name]
            # Find the column index
            if col_name in tbl["columns"]:
                col_idx = tbl["columns"].index(col_name)
                col_abs = index_to_col_letter(
                    col_letter_to_index(tbl["col_start"]) + col_idx
                )
                for r in range(tbl["data_start_row"], tbl["data_end_row"] + 1):
                    table_cells.add((tbl["sheet"], col_abs, r))

    all_referenced = referenced_cells | range_cells | table_cells

    filtered = []
    for sheet, col, row, val, cell_info in hardcoded_cells:
        if (sheet, col, row) in all_referenced:
            filtered.append((sheet, col, row, val, cell_info))

    return filtered


def build_dependency_order(formula_cells, tables):
    """Build a topological order for computing formula cells.

    Returns a list of (sheet, col, row, formula, cell_info) in computation order.
    """
    # Build adjacency: each formula cell depends on the cells it references
    cell_key = lambda s, c, r: (s, c, r)
    formula_map = {}
    deps = {}

    for sheet, col, row, formula, cell_info in formula_cells:
        key = cell_key(sheet, col, row)
        formula_map[key] = (sheet, col, row, formula, cell_info)

        converter = FormulaConverter(sheet, tables)
        converter.convert(formula)

        cell_deps = set()
        for ref in converter.referenced_cells:
            cell_deps.add(ref)
        for s, c1, r1, c2, r2 in converter.referenced_ranges:
            ci1 = col_letter_to_index(c1)
            ci2 = col_letter_to_index(c2)
            for r in range(r1, r2 + 1):
                for c in range(ci1, ci2 + 1):
                    cell_deps.add((s, index_to_col_letter(c), r))
        deps[key] = cell_deps

    # Topological sort (Kahn's algorithm)
    all_keys = set(formula_map.keys())
    in_degree = {k: 0 for k in all_keys}
    reverse_deps = {k: [] for k in all_keys}

    for key, dep_set in deps.items():
        for d in dep_set:
            if d in all_keys:
                in_degree[key] = in_degree.get(key, 0) + 1
                if d not in reverse_deps:
                    reverse_deps[d] = []
                reverse_deps[d].append(key)

    queue = [k for k in all_keys if in_degree.get(k, 0) == 0]
    ordered = []

    while queue:
        node = queue.pop(0)
        ordered.append(node)
        for neighbor in reverse_deps.get(node, []):
            in_degree[neighbor] -= 1
            if in_degree[neighbor] == 0:
                queue.append(neighbor)

    # Add any remaining (circular refs) at the end
    for key in all_keys:
        if key not in ordered:
            ordered.append(key)

    return [formula_map[k] for k in ordered if k in formula_map]


def generate_python_script(sheets, tables, formula_cells, hardcoded_cells,
                           ordered_formulas, config):
    """Generate the Python calculation script as a string."""
    lines = []

    # Header
    lines.append('"""')
    lines.append("Auto-generated Python script from Excel workbook.")
    lines.append("Reads input values from an input Excel file, performs calculations,")
    lines.append("and writes results to an output Excel file.")
    lines.append("")
    lines.append("Usage:")
    lines.append("    python calculate.py <input_excel> <output_excel>")
    lines.append('"""')
    lines.append("")
    lines.append("import sys")
    lines.append("import os")
    lines.append("import copy")
    lines.append("from openpyxl import Workbook, load_workbook")
    lines.append("from openpyxl.styles import Font, PatternFill, Alignment, numbers")
    lines.append("from openpyxl.utils import get_column_letter")
    lines.append("")
    lines.append("# ---- Excel helper functions ----")
    lines.append(HELPER_FUNCTIONS_CODE)
    lines.append("")
    lines.append("")

    # Main function
    lines.append("def main(input_file, output_file):")
    lines.append('    """Read inputs, compute formulas, write output."""')
    lines.append("")
    lines.append("    # Load input workbook")
    lines.append("    inp_wb = load_workbook(input_file, data_only=True)")
    lines.append("")

    # Read hardcoded input values
    lines.append("    # ---- Read input (hardcoded) values ----")
    input_sheets = set()
    for sheet, col, row, val, cell_info in hardcoded_cells:
        input_sheets.add(sheet)

    for sheet_name in sheets.keys():
        if sheet_name not in input_sheets:
            continue
        safe_sheet = re.sub(r'[^a-zA-Z0-9]', '_', sheet_name)
        lines.append(f"    inp_ws_{safe_sheet} = inp_wb[{repr(sheet_name)}] if {repr(sheet_name)} in inp_wb.sheetnames else None")

    lines.append("")

    # Read values from input sheet
    for sheet, col, row, val, cell_info in hardcoded_cells:
        var = cell_to_var_name(sheet, col, row)
        safe_sheet = re.sub(r'[^a-zA-Z0-9]', '_', sheet)
        default = repr(val)
        lines.append(
            f"    {var} = inp_ws_{safe_sheet}[{repr(f'{col}{row}')}].value "
            f"if inp_ws_{safe_sheet} is not None and inp_ws_{safe_sheet}[{repr(f'{col}{row}')}].value is not None "
            f"else {default}"
        )

    lines.append("")

    # Build range variables from hardcoded cells
    all_ranges = set()
    all_table_refs = set()
    for sheet, col, row, formula, cell_info in formula_cells:
        converter = FormulaConverter(sheet, tables)
        converter.convert(formula)
        all_ranges.update(converter.referenced_ranges)
        all_table_refs.update(converter.referenced_tables)

    # Generate range builder code
    if all_ranges:
        lines.append("    # ---- Build range variables ----")
        for sheet, c1, r1, c2, r2 in sorted(all_ranges):
            var = range_to_var_name(sheet, c1, r1, c2, r2)
            lines.append(f"    def _build_{var}():")
            lines.append(f"        rows = []")
            ci1 = col_letter_to_index(c1)
            ci2 = col_letter_to_index(c2)
            for r in range(r1, r2 + 1):
                row_cells = []
                for c in range(ci1, ci2 + 1):
                    cl = index_to_col_letter(c)
                    cell_var = cell_to_var_name(sheet, cl, r)
                    row_cells.append(cell_var)
                lines.append(f"        rows.append([{', '.join(row_cells)}])")
            lines.append(f"        return rows")
            lines.append("")

    # Generate table reference code
    if all_table_refs:
        lines.append("    # ---- Build table column references ----")
        for tbl_name, col_name in sorted(all_table_refs):
            if tbl_name in tables:
                tbl = tables[tbl_name]
                var = table_ref_to_var_name(tbl_name, col_name)
                if col_name in tbl["columns"]:
                    col_idx = tbl["columns"].index(col_name)
                    abs_col = index_to_col_letter(
                        col_letter_to_index(tbl["col_start"]) + col_idx
                    )
                    lines.append(f"    def _build_{var}():")
                    lines.append(f"        return [")
                    for r in range(tbl["data_start_row"], tbl["data_end_row"] + 1):
                        cv = cell_to_var_name(tbl["sheet"], abs_col, r)
                        lines.append(f"            {cv},")
                    lines.append(f"        ]")
                    lines.append("")

    lines.append("")
    lines.append("    # ---- Compute formulas (in dependency order) ----")

    for sheet, col, row, formula, cell_info in ordered_formulas:
        converter = FormulaConverter(sheet, tables)
        py_expr = converter.convert(formula)
        var = cell_to_var_name(sheet, col, row)

        # Rebuild any ranges/tables needed for this formula
        range_rebuilds = []
        for ref_range in converter.referenced_ranges:
            rvar = range_to_var_name(*ref_range)
            range_rebuilds.append(f"    {rvar} = _build_{rvar}()")
        for ref_tbl in converter.referenced_tables:
            tvar = table_ref_to_var_name(*ref_tbl)
            range_rebuilds.append(f"    {tvar} = _build_{tvar}()")

        for rb in range_rebuilds:
            lines.append(rb)

        # Wrap in try/except for robustness
        lines.append(f"    try:")
        lines.append(f"        {var} = {py_expr}")
        lines.append(f"    except Exception:")
        lines.append(f"        {var} = None")

    lines.append("")

    # Write output workbook
    lines.append("    # ---- Write output workbook ----")
    lines.append("    out_wb = Workbook()")
    lines.append("    # Remove default sheet")
    lines.append("    if 'Sheet' in out_wb.sheetnames:")
    lines.append("        del out_wb['Sheet']")
    lines.append("")

    for sheet_name in sheets.keys():
        safe_sheet = re.sub(r'[^a-zA-Z0-9]', '_', sheet_name)
        lines.append(f"    ws_{safe_sheet} = out_wb.create_sheet({repr(sheet_name)})")

        sheet_data = sheets[sheet_name]

        # Write column widths
        for col_key, width in sheet_data["col_widths"].items():
            if isinstance(col_key, str):
                lines.append(f"    ws_{safe_sheet}.column_dimensions[{repr(col_key)}].width = {width}")
            elif isinstance(col_key, int):
                cl = get_column_letter(col_key)
                lines.append(f"    ws_{safe_sheet}.column_dimensions[{repr(cl)}].width = {width}")

        # Write row heights
        for row_key, height in sheet_data["row_heights"].items():
            lines.append(f"    ws_{safe_sheet}.row_dimensions[{row_key}].height = {height}")

        # Merged cells
        for merge_range in sheet_data["merged_cells"]:
            lines.append(f"    ws_{safe_sheet}.merge_cells({repr(merge_range)})")

        lines.append("")

        # Write all cell values (hardcoded + computed)
        all_cells_in_sheet = {}
        for s, c, r, v, ci in hardcoded_cells:
            if s == sheet_name:
                all_cells_in_sheet[(c, r)] = (v, ci, False)
        for s, c, r, f, ci in formula_cells:
            if s == sheet_name:
                all_cells_in_sheet[(c, r)] = (f, ci, True)

        for (col, row), (val, cell_info, is_formula) in sorted(all_cells_in_sheet.items(), key=lambda x: (x[0][1], x[0][0])):
            var = cell_to_var_name(sheet_name, col, row)
            cell_ref = f"{col}{row}"
            lines.append(f"    ws_{safe_sheet}[{repr(cell_ref)}] = {var}")

            # Apply formatting
            nf = cell_info.get("number_format")
            if nf and nf != "General":
                lines.append(f"    ws_{safe_sheet}[{repr(cell_ref)}].number_format = {repr(nf)}")

            font_info = cell_info.get("font")
            if font_info:
                font_args = []
                if font_info.get("bold"):
                    font_args.append("bold=True")
                if font_info.get("italic"):
                    font_args.append("italic=True")
                if font_info.get("size") and font_info["size"] != DEFAULT_FONT_SIZE:
                    font_args.append(f"size={font_info['size']}")
                if font_info.get("color"):
                    font_args.append(f"color={repr(font_info['color'])}")
                if font_args:
                    lines.append(f"    ws_{safe_sheet}[{repr(cell_ref)}].font = Font({', '.join(font_args)})")

            fill_color = cell_info.get("fill_color")
            if fill_color:
                lines.append(f"    ws_{safe_sheet}[{repr(cell_ref)}].fill = PatternFill(start_color={repr(fill_color)}, end_color={repr(fill_color)}, fill_type='solid')")

            align_info = cell_info.get("alignment")
            if align_info:
                align_args = []
                if align_info.get("horizontal"):
                    align_args.append(f"horizontal={repr(align_info['horizontal'])}")
                if align_info.get("vertical"):
                    align_args.append(f"vertical={repr(align_info['vertical'])}")
                if align_info.get("wrap_text"):
                    align_args.append("wrap_text=True")
                if align_args:
                    lines.append(f"    ws_{safe_sheet}[{repr(cell_ref)}].alignment = Alignment({', '.join(align_args)})")

        lines.append("")

    lines.append("    out_wb.save(output_file)")
    lines.append(f"    print(f'Output saved to {{output_file}}')")
    lines.append("")
    lines.append("")
    lines.append('if __name__ == "__main__":')
    lines.append("    if len(sys.argv) < 3:")
    lines.append('        print("Usage: python calculate.py <input_excel> <output_excel>")')
    lines.append("        sys.exit(1)")
    lines.append("    main(sys.argv[1], sys.argv[2])")
    lines.append("")

    return "\n".join(lines)


def generate_input_template(sheets, hardcoded_cells, output_path):
    """Generate an input template Excel file.

    Creates a workbook with the same sheet names where the user fills in
    input values. Only the hardcoded value cells are included.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    sheets_with_inputs = set()
    for sheet, col, row, val, cell_info in hardcoded_cells:
        sheets_with_inputs.add(sheet)

    for sheet_name in sheets.keys():
        if sheet_name not in sheets_with_inputs:
            continue
        ws = wb.create_sheet(sheet_name)

        sheet_data = sheets[sheet_name]
        # Replicate column widths
        for col_key, width in sheet_data["col_widths"].items():
            if isinstance(col_key, str):
                ws.column_dimensions[col_key].width = width
            elif isinstance(col_key, int):
                ws.column_dimensions[get_column_letter(col_key)].width = width

        # Replicate row heights
        for row_key, height in sheet_data["row_heights"].items():
            ws.row_dimensions[row_key].height = height

    # Write hardcoded values as defaults
    for sheet, col, row, val, cell_info in hardcoded_cells:
        ws = wb[sheet]
        cell_ref = f"{col}{row}"
        ws[cell_ref] = val

        # Apply formatting
        nf = cell_info.get("number_format")
        if nf and nf != "General":
            ws[cell_ref].number_format = nf

        font_info = cell_info.get("font")
        if font_info:
            kwargs = {}
            if font_info.get("bold"):
                kwargs["bold"] = True
            if font_info.get("italic"):
                kwargs["italic"] = True
            if font_info.get("size") and font_info["size"] != DEFAULT_FONT_SIZE:
                kwargs["size"] = font_info["size"]
            if font_info.get("color"):
                kwargs["color"] = font_info["color"]
            if kwargs:
                ws[cell_ref].font = Font(**kwargs)

        fill_color = cell_info.get("fill_color")
        if fill_color:
            ws[cell_ref].fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )

        align_info = cell_info.get("alignment")
        if align_info:
            kwargs = {}
            if align_info.get("horizontal"):
                kwargs["horizontal"] = align_info["horizontal"]
            if align_info.get("vertical"):
                kwargs["vertical"] = align_info["vertical"]
            if align_info.get("wrap_text"):
                kwargs["wrap_text"] = True
            if kwargs:
                ws[cell_ref].alignment = Alignment(**kwargs)

    wb.save(output_path)
    return output_path


def convert_excel_to_python(excel_path, config_path=None, output_dir=None):
    """Main conversion function.

    Args:
        excel_path: Path to the input Excel file.
        config_path: Path to the YAML configuration file.
        output_dir: Directory to write output files.

    Returns:
        (script_path, template_path): Paths to the generated files.
    """
    config = load_config(config_path)
    delete_unreferenced = config.get("delete_unreferenced_hardcoded_values", False)

    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(excel_path) or ".", "output")
    os.makedirs(output_dir, exist_ok=True)

    # Load workbook
    wb = load_workbook(excel_path)

    # Parse
    sheets, tables = parse_workbook(wb)
    formula_cells, hardcoded_cells = classify_cells(sheets, tables)

    # Find references
    ref_cells, ref_ranges, ref_tables = find_all_references(formula_cells, tables)

    # Filter hardcoded cells
    hardcoded_cells = filter_hardcoded_cells(
        hardcoded_cells, ref_cells, ref_ranges, ref_tables, tables, delete_unreferenced
    )

    # Build dependency order
    ordered_formulas = build_dependency_order(formula_cells, tables)

    # Generate Python script
    script_content = generate_python_script(
        sheets, tables, formula_cells, hardcoded_cells, ordered_formulas, config
    )
    script_path = os.path.join(output_dir, "calculate.py")
    with open(script_path, "w") as f:
        f.write(script_content)

    # Generate input template
    template_path = os.path.join(output_dir, "input_template.xlsx")
    generate_input_template(sheets, hardcoded_cells, template_path)

    wb.close()

    print(f"Generated Python script: {script_path}")
    print(f"Generated input template: {template_path}")
    print(f"\nTo run: python {script_path} {template_path} <output.xlsx>")

    return script_path, template_path


def main():
    parser = argparse.ArgumentParser(
        description="Convert Excel workbook to Python calculation script"
    )
    parser.add_argument("excel_file", help="Path to the input Excel file (.xlsx)")
    parser.add_argument(
        "--config", default="config.yaml", help="Path to config YAML file (default: config.yaml)"
    )
    parser.add_argument(
        "--output-dir", default=None, help="Output directory (default: ./output)"
    )
    args = parser.parse_args()

    if not os.path.exists(args.excel_file):
        print(f"Error: File '{args.excel_file}' not found.")
        sys.exit(1)

    convert_excel_to_python(args.excel_file, args.config, args.output_dir)


if __name__ == "__main__":
    main()
