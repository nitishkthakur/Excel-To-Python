"""
generate_mapping_report.py — Layer 1 of the Excel-to-Python pipeline.

Parses an Excel workbook and produces ``mapping_report.xlsx`` which captures
every cell's type (Input / Calculation / Output), formula, cached value,
formatting, and formula-group metadata.

Usage:
    python -m src.generate_mapping_report ExcelFiles/ACC-Ltd.xlsx output/
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import warnings
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter

from src.formula_parser import (
    extract_references,
    get_referenced_cells,
    to_r1c1_pattern,
    rowcol_to_cell,
)
from src.excel_utils import (
    ensure_xlsx,
    open_workbook_dual,
    extract_formatting,
    resolve_defined_names,
    safe_sheet_name,
)

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class CellData:
    sheet: str
    row: int
    col: int
    cell_ref: str            # "B3"
    formula: str | None      # "=A1+B1" or None
    value: Any               # cached / literal value
    formatting: dict[str, Any] = field(default_factory=dict)
    cell_type: str = ""      # Input / Calculation / Output
    group_id: int = 0
    group_direction: str = ""  # row / col / none
    group_size: int = 0
    pattern_formula: str = ""
    include_flag: bool = True
    references: list[str] = field(default_factory=list)    # cells THIS cell references
    referenced_by: list[str] = field(default_factory=list) # cells that reference THIS cell


# ---------------------------------------------------------------------------
# Step 1 — Collect all cells
# ---------------------------------------------------------------------------

def _collect_cells(wb_f, wb_v, sheet_names: list[str]) -> dict[tuple[str, int, int], CellData]:
    """Iterate every non-empty cell; store formula, value, formatting."""
    cells: dict[tuple[str, int, int], CellData] = {}
    for sname in sheet_names:
        ws_f = wb_f[sname]
        ws_v = wb_v[sname]
        for row in ws_f.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.value is None:
                    continue
                formula = None
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
                # Cached value from data-only workbook
                vcell = ws_v[cell.coordinate]
                value = vcell.value if formula else cell.value
                # Handle DataTableFormula objects (Excel What-If data tables)
                if not isinstance(value, (str, int, float, bool, type(None))):
                    value = vcell.value
                fmt = extract_formatting(cell)
                key = (sname, cell.row, cell.column)
                cells[key] = CellData(
                    sheet=sname,
                    row=cell.row,
                    col=cell.column,
                    cell_ref=cell.coordinate,
                    formula=formula,
                    value=value,
                    formatting=fmt,
                )
    return cells


# ---------------------------------------------------------------------------
# Step 2 — Build dependency graph
# ---------------------------------------------------------------------------

def _build_dependency_graph(
    cells: dict[tuple[str, int, int], CellData],
    sheet_names: list[str],
) -> None:
    """Populate .references and .referenced_by for every cell in-place.

    Also classifies each cell as Input / Calculation / Output.
    """
    # Forward map: (sheet,row,col) → set of (sheet,row,col) it depends on
    forward: dict[tuple[str, int, int], set[tuple[str, int, int]]] = defaultdict(set)
    # Reverse map: (sheet,row,col) → set of (sheet,row,col) that depend on it
    reverse: dict[tuple[str, int, int], set[tuple[str, int, int]]] = defaultdict(set)

    for key, cd in cells.items():
        if cd.formula:
            deps = get_referenced_cells(cd.formula, cd.sheet)
            for dep in deps:
                forward[key].add(dep)
                reverse[dep].add(key)

    # Classify
    for key, cd in cells.items():
        if not cd.formula:
            cd.cell_type = "Input"
        else:
            # Does any other formula reference this cell?
            if reverse.get(key):
                cd.cell_type = "Calculation"
            else:
                cd.cell_type = "Output"

        # Store human-readable reference lists
        refs = forward.get(key, set())
        cd.references = sorted(
            f"{s}!{rowcol_to_cell(r, c)}" for s, r, c in refs
        )
        rby = reverse.get(key, set())
        cd.referenced_by = sorted(
            f"{s}!{rowcol_to_cell(r, c)}" for s, r, c in rby
        )


# ---------------------------------------------------------------------------
# Step 3 — Detect formula groups (dragged / repeated formulas)
# ---------------------------------------------------------------------------

def _detect_formula_groups(
    cells: dict[tuple[str, int, int], CellData],
) -> list[list[tuple[str, int, int]]]:
    """Return groups of cells that share the same R1C1 pattern.

    Algorithm:
      1. Compute R1C1 pattern for every formula cell.
      2. Bucket by (sheet, pattern).
      3. Within each bucket, find maximal contiguous runs in one direction.
    """
    # Compute patterns
    pattern_map: dict[tuple[str, int, int], str] = {}
    for key, cd in cells.items():
        if cd.formula:
            pat = to_r1c1_pattern(cd.formula, cd.row, cd.col)
            pattern_map[key] = pat
            cd.pattern_formula = pat

    # Bucket by (sheet, pattern)
    buckets: dict[tuple[str, str], list[tuple[str, int, int]]] = defaultdict(list)
    for key, pat in pattern_map.items():
        buckets[(key[0], pat)].append(key)

    groups: list[list[tuple[str, int, int]]] = []
    for (_sheet, _pat), members in buckets.items():
        if len(members) < 2:
            continue
        # Sort by (row, col) and find contiguous runs
        members_sorted = sorted(members, key=lambda k: (k[1], k[2]))
        # Try row-direction runs (same row, consecutive cols)
        row_runs = _find_row_runs(members_sorted)
        # Try col-direction runs (same col, consecutive rows)
        col_runs = _find_col_runs(members_sorted)
        # Pick all runs of length ≥ 2
        for run in row_runs + col_runs:
            if len(run) >= 2:
                groups.append(run)
    return groups


def _find_row_runs(members: list[tuple[str, int, int]]) -> list[list[tuple[str, int, int]]]:
    """Find contiguous horizontal (same-row) runs."""
    by_row: dict[tuple[str, int], list[tuple[str, int, int]]] = defaultdict(list)
    for s, r, c in members:
        by_row[(s, r)].append((s, r, c))
    runs = []
    for _key, row_cells in by_row.items():
        row_cells.sort(key=lambda x: x[2])
        run: list[tuple[str, int, int]] = [row_cells[0]]
        for i in range(1, len(row_cells)):
            if row_cells[i][2] == run[-1][2] + 1:
                run.append(row_cells[i])
            else:
                if len(run) >= 2:
                    runs.append(run)
                run = [row_cells[i]]
        if len(run) >= 2:
            runs.append(run)
    return runs


def _find_col_runs(members: list[tuple[str, int, int]]) -> list[list[tuple[str, int, int]]]:
    """Find contiguous vertical (same-column) runs."""
    by_col: dict[tuple[str, int], list[tuple[str, int, int]]] = defaultdict(list)
    for s, r, c in members:
        by_col[(s, c)].append((s, r, c))
    runs = []
    for _key, col_cells in by_col.items():
        col_cells.sort(key=lambda x: x[1])
        run: list[tuple[str, int, int]] = [col_cells[0]]
        for i in range(1, len(col_cells)):
            if col_cells[i][1] == run[-1][1] + 1:
                run.append(col_cells[i])
            else:
                if len(run) >= 2:
                    runs.append(run)
                run = [col_cells[i]]
        if len(run) >= 2:
            runs.append(run)
    return runs


def _assign_groups(
    cells: dict[tuple[str, int, int], CellData],
    groups: list[list[tuple[str, int, int]]],
) -> None:
    """Assign GroupID, GroupDirection, GroupSize to cells."""
    # De-duplicate: a cell might appear in both a row-run and a col-run.
    # Prefer the longer group; break ties by preferring row direction.
    cell_best: dict[tuple[str, int, int], tuple[int, str, int]] = {}
    for gid, group in enumerate(groups, start=1):
        sheet_set = {k[0] for k in group}
        rows = {k[1] for k in group}
        cols = {k[2] for k in group}
        if len(rows) == 1:
            direction = "row"
        elif len(cols) == 1:
            direction = "col"
        else:
            direction = "block"
        size = len(group)
        for key in group:
            prev = cell_best.get(key)
            if prev is None or size > prev[2]:
                cell_best[key] = (gid, direction, size)

    for key, (gid, direction, size) in cell_best.items():
        cd = cells.get(key)
        if cd:
            cd.group_id = gid
            cd.group_direction = direction
            cd.group_size = size


# ---------------------------------------------------------------------------
# Step 4 — Write the mapping report
# ---------------------------------------------------------------------------

_HEADER = [
    "Cell", "Row", "Col", "Type", "Formula", "Value",
    "NumberFormat", "FontBold", "FontItalic", "FontSize",
    "FontColor", "FillColor", "HAlignment", "VAlignment",
    "WrapText", "GroupID", "GroupDirection", "GroupSize",
    "PatternFormula", "IncludeFlag", "ReferencedBy", "References",
]


def _write_report(
    cells: dict[tuple[str, int, int], CellData],
    sheet_names: list[str],
    output_path: str,
    original_meta: dict[str, dict[str, Any]],
    defined_names: dict[str, str] | None = None,
) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)                    # remove default sheet

    for sname in sheet_names:
        ws = wb.create_sheet(safe_sheet_name(sname))
        ws.append(_HEADER)
        # Bold header
        for c in range(1, len(_HEADER) + 1):
            ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)

        sheet_cells = sorted(
            (cd for cd in cells.values() if cd.sheet == sname),
            key=lambda cd: (cd.row, cd.col),
        )
        for cd in sheet_cells:
            fmt = cd.formatting
            row_data = [
                cd.cell_ref,
                cd.row,
                cd.col,
                cd.cell_type,
                cd.formula or "",
                _safe_value(cd.value),
                fmt.get("number_format", "General"),
                fmt.get("font_bold", False),
                fmt.get("font_italic", False),
                fmt.get("font_size", 11),
                fmt.get("font_color", ""),
                fmt.get("fill_color", ""),
                fmt.get("h_alignment", ""),
                fmt.get("v_alignment", ""),
                fmt.get("wrap_text", False),
                cd.group_id,
                cd.group_direction,
                cd.group_size,
                cd.pattern_formula,
                cd.include_flag,
                "; ".join(cd.referenced_by),
                "; ".join(cd.references),
            ]
            cur_row = ws.max_row + 1
            for ci, val in enumerate(row_data, start=1):
                cell = ws.cell(row=cur_row, column=ci)
                val = _safe_value(val)
                # Prevent openpyxl from interpreting "=..." as a formula
                if isinstance(val, str) and val.startswith("="):
                    cell.data_type = "s"
                    cell._value = val
                else:
                    cell.value = val

    # _Metadata sheet
    ws_meta = wb.create_sheet("_Metadata")
    ws_meta.append(["SheetName", "SheetIndex", "TotalCells", "InputCells",
                     "CalcCells", "OutputCells", "FormulaGroups"])
    for c in range(1, 8):
        ws_meta.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
    for idx, sname in enumerate(sheet_names):
        meta = original_meta.get(sname, {})
        ws_meta.append([
            sname, idx,
            meta.get("total", 0),
            meta.get("input", 0),
            meta.get("calc", 0),
            meta.get("output", 0),
            meta.get("groups", 0),
        ])

    # _DefinedNames sheet — stores workbook-level named ranges
    if defined_names:
        ws_dn = wb.create_sheet("_DefinedNames")
        ws_dn.append(["Name", "Reference"])
        for c in range(1, 3):
            ws_dn.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
        for name, ref in sorted(defined_names.items()):
            ws_dn.append([name, ref])

    wb.save(output_path)


_ILLEGAL_CHAR_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def _safe_value(v: Any) -> Any:
    """Ensure the value can be written to openpyxl without error."""
    if v is None:
        return ""
    if isinstance(v, (int, float, bool)):
        return v
    s = str(v)
    s = _ILLEGAL_CHAR_RE.sub('', s)
    return s


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def generate_mapping_report(excel_path: str, output_dir: str) -> str:
    """Full pipeline: Excel file → mapping_report.xlsx.

    Returns the path to the generated report.
    """
    os.makedirs(output_dir, exist_ok=True)
    xlsx_path = ensure_xlsx(excel_path, cache_dir=os.path.join(output_dir, "_converted"))

    print(f"  Opening workbook (formulas + values) …")
    wb_f, wb_v = open_workbook_dual(xlsx_path)
    sheet_names = wb_f.sheetnames

    print(f"  Collecting cells across {len(sheet_names)} sheets …")
    cells = _collect_cells(wb_f, wb_v, sheet_names)
    print(f"    → {len(cells)} non-empty cells")

    print(f"  Building dependency graph & classifying cells …")
    _build_dependency_graph(cells, sheet_names)

    n_inp = sum(1 for c in cells.values() if c.cell_type == "Input")
    n_calc = sum(1 for c in cells.values() if c.cell_type == "Calculation")
    n_out = sum(1 for c in cells.values() if c.cell_type == "Output")
    print(f"    → Input={n_inp}  Calculation={n_calc}  Output={n_out}")

    print(f"  Detecting formula groups …")
    groups = _detect_formula_groups(cells)
    _assign_groups(cells, groups)
    print(f"    → {len(groups)} formula groups detected")

    # Compute per-sheet metadata
    original_meta: dict[str, dict[str, Any]] = {}
    for sname in sheet_names:
        sc = [c for c in cells.values() if c.sheet == sname]
        grp_ids = {c.group_id for c in sc if c.group_id > 0}
        original_meta[sname] = {
            "total": len(sc),
            "input": sum(1 for c in sc if c.cell_type == "Input"),
            "calc": sum(1 for c in sc if c.cell_type == "Calculation"),
            "output": sum(1 for c in sc if c.cell_type == "Output"),
            "groups": len(grp_ids),
        }

    report_path = os.path.join(output_dir, "mapping_report.xlsx")
    print(f"  Writing mapping report → {report_path}")

    # Extract workbook-level defined names (named ranges)
    defined_names: dict[str, str] = {}
    for name, dn in wb_f.defined_names.items():
        ref = dn.attr_text
        # Skip names that reference errors like #REF!
        if '#REF!' in ref or ref.startswith('#'):
            continue
        # Skip external workbook references like [10]Sheet!$A$1
        if '[' in ref:
            continue
        # Skip multi-range references (commas outside of quotes)
        if ',' in ref:
            continue
        # Skip NA() or similar function-based definitions
        if '(' in ref:
            continue
        # Must reference a sheet in this workbook
        if '!' not in ref:
            continue
        # Only accept single-cell references (not ranges with ':')
        # Range refs need special handling and are less common in formula usage
        if ':' in ref:
            continue
        defined_names[name] = ref
    if defined_names:
        print(f"    → {len(defined_names)} defined names extracted")

    _write_report(cells, sheet_names, report_path, original_meta, defined_names)

    wb_f.close()
    wb_v.close()
    return report_path


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description="Generate mapping_report.xlsx from an Excel file.")
    ap.add_argument("excel_file", help="Path to the source Excel file (.xlsx or .xls)")
    ap.add_argument("output_dir", help="Directory to write mapping_report.xlsx into")
    args = ap.parse_args()
    path = generate_mapping_report(args.excel_file, args.output_dir)
    print(f"\n✓ Mapping report written to {path}")


if __name__ == "__main__":
    main()
