"""
compare_utils.py — Cell-by-cell comparison between output.xlsx and the original.

Reads both workbooks in data-only mode and compares every non-empty cell.
Reports mismatches with cell address, expected value, and actual value.
"""

from __future__ import annotations

import math
import os
import warnings
from dataclasses import dataclass
from typing import Any

import openpyxl

from src.excel_utils import ensure_xlsx

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---------------------------------------------------------------------------
# Tolerance
# ---------------------------------------------------------------------------

REL_TOL = 1e-6   # relative tolerance for numeric comparisons
ABS_TOL = 1e-9   # absolute tolerance


@dataclass
class Mismatch:
    sheet: str
    cell: str
    expected: Any
    actual: Any
    kind: str   # "missing", "extra", "value", "type"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def compare_workbooks(
    original_path: str,
    output_path: str,
    *,
    cache_dir: str | None = None,
    ignore_sheets: set[str] | None = None,
    compare_formulas: bool = False,
) -> list[Mismatch]:
    """Compare *output_path* against *original_path* cell-by-cell.

    Both workbooks are opened in data-only mode so that cached values
    (not formulas) are compared.

    Returns a list of :class:`Mismatch` objects.
    """
    orig_path = ensure_xlsx(original_path, cache_dir=cache_dir)

    wb_orig = openpyxl.load_workbook(orig_path, data_only=True)
    wb_out  = openpyxl.load_workbook(output_path, data_only=True)

    ignore = ignore_sheets or set()
    mismatches: list[Mismatch] = []

    for sname in wb_orig.sheetnames:
        if sname in ignore:
            continue
        if sname not in wb_out.sheetnames:
            mismatches.append(Mismatch(sname, "", None, None, "missing_sheet"))
            continue
        ws_o = wb_orig[sname]
        ws_n = wb_out[sname]

        # Collect all cells from original
        orig_cells: dict[str, Any] = {}
        for row in ws_o.iter_rows():
            for cell in row:
                if cell.value is not None:
                    orig_cells[cell.coordinate] = cell.value

        # Compare against output
        out_cells: dict[str, Any] = {}
        for row in ws_n.iter_rows():
            for cell in row:
                if cell.value is not None:
                    out_cells[cell.coordinate] = cell.value

        for coord, exp in orig_cells.items():
            act = out_cells.get(coord)
            if act is None:
                mismatches.append(Mismatch(sname, coord, exp, None, "missing"))
            elif not _values_equal(exp, act):
                mismatches.append(Mismatch(sname, coord, exp, act, "value"))

        for coord in out_cells:
            if coord not in orig_cells:
                mismatches.append(Mismatch(sname, coord, None, out_cells[coord], "extra"))

    wb_orig.close()
    wb_out.close()
    return mismatches


def summarise_mismatches(mismatches: list[Mismatch], max_print: int = 25) -> str:
    """Return a human-readable summary string."""
    if not mismatches:
        return "✓ All cells match."
    lines = [f"✗ {len(mismatches)} mismatches found:"]
    by_kind: dict[str, int] = {}
    for m in mismatches:
        by_kind[m.kind] = by_kind.get(m.kind, 0) + 1
    for k, n in sorted(by_kind.items()):
        lines.append(f"  {k}: {n}")
    lines.append("")
    for m in mismatches[:max_print]:
        exp_s = repr(m.expected)[:40]
        act_s = repr(m.actual)[:40]
        lines.append(f"  [{m.sheet}] {m.cell}: expected={exp_s}  actual={act_s}  ({m.kind})")
    if len(mismatches) > max_print:
        lines.append(f"  … and {len(mismatches) - max_print} more")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _values_equal(a: Any, b: Any) -> bool:
    """Compare two cell values with tolerance for floats."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    # Both numeric
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        if math.isnan(a) and math.isnan(b):
            return True
        return math.isclose(a, b, rel_tol=REL_TOL, abs_tol=ABS_TOL)
    # Both strings
    if isinstance(a, str) and isinstance(b, str):
        return a.strip() == b.strip()
    # Mixed types — try numeric cast
    try:
        fa, fb = float(a), float(b)
        return math.isclose(fa, fb, rel_tol=REL_TOL, abs_tol=ABS_TOL)
    except (ValueError, TypeError):
        pass
    return str(a).strip() == str(b).strip()
