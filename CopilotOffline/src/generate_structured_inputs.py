"""
generate_structured_inputs.py — Layer 2b of the Excel-to-Python pipeline.

Reads ``mapping_report.xlsx`` and produces a clean, tabular
``structured_input.xlsx`` organised by source sheet with:

  * **Index** — cross-reference to the mapping report
  * **Config** — all scalar / short-vector inputs
  * One or more data-sheet tabs per source sheet for larger input tables

Auto-transpose rule: if a patch's column headers look like financial dates
(years, periods), the table is transposed so rows = periods, columns = metrics.

Usage:
    python -m src.generate_structured_inputs output/ACC-Ltd/mapping_report.xlsx output/ACC-Ltd/
"""

from __future__ import annotations

import argparse
import os
import re
import warnings
from collections import defaultdict
from typing import Any

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from src.excel_utils import safe_sheet_name

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---------------------------------------------------------------------------
# Mapping-report reader (shared with 2a but duplicated for self-containment)
# ---------------------------------------------------------------------------

def _read_mapping_report(report_path: str) -> dict[str, list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(report_path, data_only=True)
    sheets: dict[str, list[dict[str, Any]]] = {}
    for sname in wb.sheetnames:
        if sname == "_Metadata":
            continue
        ws = wb[sname]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        rows: list[dict[str, Any]] = []
        for r in range(2, (ws.max_row or 1) + 1):
            row = {}
            for ci, h in enumerate(headers, start=1):
                row[h] = ws.cell(row=r, column=ci).value
            rows.append(row)
        sheets[sname] = rows
    wb.close()
    return sheets


# ---------------------------------------------------------------------------
# Date / period detection
# ---------------------------------------------------------------------------

_YEAR_RE = re.compile(r"^(19|20)\d{2}[A-Za-z ]*$")
_QUARTER_RE = re.compile(r"^Q[1-4]\s*(19|20)\d{2}", re.IGNORECASE)

def _is_period_label(v: Any) -> bool:
    """Return True if *v* looks like a financial-period header.

    Matches: 2020, 2020E, "2020 F", "Q1 2023", datetime years, integers ≥ 1900.
    """
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return 1900 <= v <= 2100
    s = str(v).strip()
    if _YEAR_RE.match(s) or _QUARTER_RE.match(s):
        return True
    try:
        n = float(s)
        return 1900 <= n <= 2100
    except (ValueError, TypeError):
        return False


# ---------------------------------------------------------------------------
# Patch detection — contiguous rectangular blocks of Input cells
# ---------------------------------------------------------------------------

def _find_patches(
    input_cells: list[dict[str, Any]],
    all_cells: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Identify rectangular patches of input cells in a sheet.

    A patch is a maximal contiguous rectangle of Input cells.
    Returns a list of patch dicts with keys:
      min_row, max_row, min_col, max_col, cells (list of cell dicts),
      row_labels, col_labels
    """
    if not input_cells:
        return []

    # Build a set of (row, col) for fast lookup
    occupied = {(int(c["Row"]), int(c["Col"])): c for c in input_cells}
    all_map = {(int(c["Row"]), int(c["Col"])): c for c in all_cells}
    visited = set()
    patches = []

    for (r, c) in sorted(occupied.keys()):
        if (r, c) in visited:
            continue
        # BFS / flood-fill to find connected component
        # (4-directional: up, down, left, right)
        component = set()
        queue = [(r, c)]
        while queue:
            cr, cc = queue.pop()
            if (cr, cc) in component:
                continue
            if (cr, cc) not in occupied:
                continue
            component.add((cr, cc))
            for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                nr, nc = cr + dr, cc + dc
                if (nr, nc) not in component and (nr, nc) in occupied:
                    queue.append((nr, nc))
        visited |= component

        if not component:
            continue

        min_r = min(r for r, _ in component)
        max_r = max(r for r, _ in component)
        min_c = min(c for _, c in component)
        max_c = max(c for _, c in component)

        cells_in_patch = [occupied[(r, c)] for r, c in component]

        # Try to resolve row labels: look at column (min_c - 1) for each row
        row_labels = {}
        for row in range(min_r, max_r + 1):
            label_cell = all_map.get((row, min_c - 1))
            if label_cell and label_cell.get("Value") is not None:
                row_labels[row] = label_cell["Value"]
            else:
                row_labels[row] = f"Line{row - min_r + 1}"

        # Try to resolve col labels: look at row (min_r - 1) for each col
        col_labels = {}
        for col in range(min_c, max_c + 1):
            label_cell = all_map.get((min_r - 1, col))
            if label_cell and label_cell.get("Value") is not None:
                col_labels[col] = label_cell["Value"]
            else:
                label_cell = all_map.get((min_r, col))
                if label_cell and label_cell.get("Type") != "Input":
                    col_labels[col] = f"Col{col - min_c + 1}"
                else:
                    col_labels[col] = f"Col{col - min_c + 1}"

        patches.append({
            "min_row": min_r, "max_row": max_r,
            "min_col": min_c, "max_col": max_c,
            "cells": cells_in_patch,
            "row_labels": row_labels,
            "col_labels": col_labels,
            "source_cells": occupied,
        })

    return patches


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

def generate_structured_inputs(report_path: str, output_dir: str) -> str:
    os.makedirs(output_dir, exist_ok=True)
    report = _read_mapping_report(report_path)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Index sheet ---
    ws_idx = wb.create_sheet("Index")
    ws_idx.append(["SourceSheet", "TargetSheet", "InputType", "Cell",
                    "Row", "Col", "Label", "Value"])
    for c in range(1, 9):
        ws_idx.cell(row=1, column=c).font = Font(bold=True)

    # --- Config sheet (scalars) ---
    ws_cfg = wb.create_sheet("Config")
    ws_cfg.append(["SourceSheet", "Cell", "Label", "Value"])
    for c in range(1, 5):
        ws_cfg.cell(row=1, column=c).font = Font(bold=True)

    tab_counter = 0

    for sname, rows in report.items():
        input_rows = [r for r in rows
                      if r.get("Type") == "Input"
                      and r.get("IncludeFlag") is not False
                      and r.get("IncludeFlag") != 0]
        if not input_rows:
            continue

        patches = _find_patches(input_rows, rows)

        for pidx, patch in enumerate(patches):
            n_rows = patch["max_row"] - patch["min_row"] + 1
            n_cols = patch["max_col"] - patch["min_col"] + 1
            total_cells = len(patch["cells"])

            # Classify: scalar (1 cell or 1-row × 1-col), small vector, or table
            is_scalar = total_cells <= 2 and min(n_rows, n_cols) <= 1

            if is_scalar:
                # → Config sheet
                for cd in patch["cells"]:
                    row_num = int(cd["Row"])
                    col_num = int(cd["Col"])
                    label = patch["row_labels"].get(row_num, f"R{row_num}")
                    ws_cfg.append([sname, cd["Cell"], str(label), cd["Value"]])
                    ws_idx.append([sname, "Config", "Scalar",
                                   cd["Cell"], row_num, col_num, str(label), cd["Value"]])
            else:
                # → Dedicated sheet (or shared per source sheet)
                tab_counter += 1
                # Check auto-transpose: are col_labels financial dates?
                col_vals = list(patch["col_labels"].values())
                should_transpose = (
                    len(col_vals) >= 2
                    and sum(_is_period_label(v) for v in col_vals) >= len(col_vals) * 0.5
                )

                tab_name = safe_sheet_name(
                    f"{sname[:20]}_{pidx + 1}" if len(patches) > 1 else sname
                )
                ws_tab = wb.create_sheet(tab_name)

                # Build a 2D numpy array of values
                data = {}
                for cd in patch["cells"]:
                    r = int(cd["Row"]) - patch["min_row"]
                    c = int(cd["Col"]) - patch["min_col"]
                    data[(r, c)] = cd["Value"]

                if should_transpose:
                    # Rows = periods (col labels), Cols = metrics (row labels)
                    # Header row: blank + row_labels (as column headers)
                    headers = ["Period"] + [
                        str(patch["row_labels"].get(r, f"Line{ri+1}"))
                        for ri, r in enumerate(range(patch["min_row"], patch["max_row"] + 1))
                    ]
                    ws_tab.append(headers)
                    for c in range(1, len(headers) + 1):
                        ws_tab.cell(row=1, column=c).font = Font(bold=True)

                    for ci, col_num in enumerate(range(patch["min_col"], patch["max_col"] + 1)):
                        period_label = str(patch["col_labels"].get(col_num, f"Col{ci+1}"))
                        row_vals = [period_label]
                        for ri in range(n_rows):
                            row_vals.append(data.get((ri, ci), ""))
                        ws_tab.append(row_vals)

                    # Index entries
                    for cd in patch["cells"]:
                        ws_idx.append([sname, tab_name, "Table(transposed)",
                                       cd["Cell"], int(cd["Row"]), int(cd["Col"]),
                                       "", cd["Value"]])
                else:
                    # Normal orientation: rows = metrics, cols = periods/headers
                    headers = ["Metric"] + [
                        str(patch["col_labels"].get(c, f"Col{ci+1}"))
                        for ci, c in enumerate(range(patch["min_col"], patch["max_col"] + 1))
                    ]
                    ws_tab.append(headers)
                    for c in range(1, len(headers) + 1):
                        ws_tab.cell(row=1, column=c).font = Font(bold=True)

                    for ri, row_num in enumerate(range(patch["min_row"], patch["max_row"] + 1)):
                        label = str(patch["row_labels"].get(row_num, f"Line{ri+1}"))
                        row_vals = [label]
                        for ci in range(n_cols):
                            row_vals.append(data.get((ri, ci), ""))
                        ws_tab.append(row_vals)

                    for cd in patch["cells"]:
                        ws_idx.append([sname, tab_name, "Table",
                                       cd["Cell"], int(cd["Row"]), int(cd["Col"]),
                                       "", cd["Value"]])

        # Remaining isolated input cells (not in any patch) → Config
        patch_cells = set()
        for p in patches:
            for cd in p["cells"]:
                patch_cells.add((int(cd["Row"]), int(cd["Col"])))

        for cd in input_rows:
            key = (int(cd["Row"]), int(cd["Col"]))
            if key not in patch_cells:
                ws_cfg.append([sname, cd["Cell"], cd["Cell"], cd["Value"]])
                ws_idx.append([sname, "Config", "Scalar",
                               cd["Cell"], key[0], key[1], cd["Cell"], cd["Value"]])

    out_path = os.path.join(output_dir, "structured_input.xlsx")
    wb.save(out_path)
    print(f"  → {out_path}  ({tab_counter} data tabs + Config + Index)")
    return out_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Generate structured_input.xlsx from mapping_report.xlsx")
    ap.add_argument("report", help="Path to mapping_report.xlsx")
    ap.add_argument("output_dir", help="Directory for output")
    args = ap.parse_args()
    generate_structured_inputs(args.report, args.output_dir)
    print("✓ Done")


if __name__ == "__main__":
    main()
