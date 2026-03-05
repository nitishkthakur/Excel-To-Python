"""
generate_unstructured_inputs.py — Layer 2a of the Excel-to-Python pipeline.

Reads ``mapping_report.xlsx`` and produces ``unstructured_inputs.xlsx``:
the same layout as the original workbook but with **all formula cells stripped**.
Only raw hardcoded values (Type == "Input", IncludeFlag == True) are retained,
in their original sheet positions and formatting.

Usage:
    python -m src.generate_unstructured_inputs output/ACC-Ltd/mapping_report.xlsx output/ACC-Ltd/
"""

from __future__ import annotations

import argparse
import os
import warnings
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from src.excel_utils import safe_sheet_name, apply_formatting

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---------------------------------------------------------------------------
# Mapping-report reader
# ---------------------------------------------------------------------------

def _read_mapping_report(report_path: str) -> dict[str, list[dict[str, Any]]]:
    """Read the mapping report into {sheet_name: [row_dicts, …]}."""
    wb = openpyxl.load_workbook(report_path, data_only=True)
    sheets: dict[str, list[dict[str, Any]]] = {}
    for sname in wb.sheetnames:
        if sname == "_Metadata":
            continue
        ws = wb[sname]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        rows: list[dict[str, Any]] = []
        for r in range(2, (ws.max_row or 1) + 1):
            row = {}
            for ci, h in enumerate(headers, start=1):
                row[h] = ws.cell(row=r, column=ci).value
            rows.append(row)
        sheets[sname] = rows
    wb.close()
    return sheets


# ---------------------------------------------------------------------------
# Unstructured input generator
# ---------------------------------------------------------------------------

def generate_unstructured_inputs(report_path: str, output_dir: str) -> str:
    """Create ``unstructured_inputs.xlsx`` from the mapping report.

    Only Input cells with IncludeFlag==True are written, in their
    original sheet / row / column positions with formatting.
    """
    os.makedirs(output_dir, exist_ok=True)
    report = _read_mapping_report(report_path)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sname, rows in report.items():
        ws = wb.create_sheet(safe_sheet_name(sname))
        input_count = 0
        for rd in rows:
            if rd.get("Type") != "Input":
                continue
            if rd.get("IncludeFlag") is False or rd.get("IncludeFlag") == 0:
                continue

            row_idx = int(rd["Row"])
            col_idx = int(rd["Col"])
            value = rd.get("Value", "")
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value

            # Apply formatting from the report
            fmt = {
                "number_format": rd.get("NumberFormat", "General"),
                "font_bold": bool(rd.get("FontBold", False)),
                "font_italic": bool(rd.get("FontItalic", False)),
                "font_size": rd.get("FontSize", 11),
                "font_color": str(rd.get("FontColor", "000000") or "000000"),
                "fill_color": str(rd.get("FillColor", "") or ""),
                "h_alignment": str(rd.get("HAlignment", "") or ""),
                "v_alignment": str(rd.get("VAlignment", "") or ""),
                "wrap_text": bool(rd.get("WrapText", False)),
            }
            apply_formatting(cell, fmt)
            input_count += 1
        print(f"  {sname}: {input_count} input cells written")

    out_path = os.path.join(output_dir, "unstructured_inputs.xlsx")
    wb.save(out_path)
    print(f"  → {out_path}")
    return out_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description="Generate unstructured_inputs.xlsx from mapping_report.xlsx")
    ap.add_argument("report", help="Path to mapping_report.xlsx")
    ap.add_argument("output_dir", help="Directory for output")
    args = ap.parse_args()
    generate_unstructured_inputs(args.report, args.output_dir)
    print("✓ Done")


if __name__ == "__main__":
    main()
