"""
generate_structured_calculate.py — Layer 3b code generator.

Reads ``mapping_report.xlsx`` and emits ``structured_calculate.py``, a
standalone Python script that:

  1. Reads ``structured_input.xlsx`` (tabular format with Config + data tabs)
  2. Evaluates every Calculation and Output formula in topological order
  3. Applies formatting
  4. Writes ``output.xlsx``

The formula evaluation engine is identical to Layer 3a
(``unstructured_calculate.py``).  The only difference is the **input
loading** stage: values are read from the structured tabular format
(Config sheet for scalars, data-sheet tabs for tables) rather than
from the original cell-by-cell layout.

Usage:
    python -m src.generate_structured_calculate \\
        output/ACC-Ltd/mapping_report.xlsx output/ACC-Ltd/
"""

from __future__ import annotations

import argparse
import os
import re
import textwrap
import warnings
from collections import defaultdict
from typing import Any

import openpyxl

# Reuse core infrastructure from the unstructured calculator
import src.generate_unstructured_calculate as _uc

from src.formula_parser import cell_to_rowcol

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# ---------------------------------------------------------------------------
# Build emission plan  (mirrors the logic in Layer 3a's _generate_code)
# ---------------------------------------------------------------------------

def _build_emission_plan(
    sheets: dict[str, list[dict[str, Any]]],
    sheet_order: list[str],
    defined_names: dict[str, str] | None = None,
) -> list[tuple[str, Any]]:
    """Build the emission plan from mapping-report data.

    This duplicates the SCC-based evaluation-order logic from
    ``generate_unstructured_calculate._generate_code`` so that the
    structured calculator can operate independently.

    Returns a list of ``(action, data)`` tuples where *action* is
    ``"group"`` or ``"cell"``.
    """
    # Populate module-level defined-names on the unstructured module so that
    # the formula translators (which live there) can resolve named ranges.
    _uc._DEFINED_NAMES = defined_names or {}

    # ------------------------------------------------------------------
    # 1. Collect formula cells & cell map
    # ------------------------------------------------------------------
    formula_cells: list[dict[str, Any]] = []
    all_cells_map: dict[tuple[str, str], dict] = {}
    for sname, rows in sheets.items():
        for rd in rows:
            rd["_sheet"] = sname
            key = (sname, rd["Cell"])
            all_cells_map[key] = rd
            if rd.get("Type") in ("Calculation", "Output"):
                if rd.get("IncludeFlag") is not False and rd.get("IncludeFlag") != 0:
                    formula_cells.append(rd)

    # Cell-level adjacency (dep → set[dependant])
    cell_keys = {(fc["_sheet"], fc["Cell"]) for fc in formula_cells}
    cell_adj: dict[tuple, set[tuple]] = defaultdict(set)
    cell_rev: dict[tuple, set[tuple]] = defaultdict(set)
    for fc in formula_cells:
        my = (fc["_sheet"], fc["Cell"])
        refs_str = fc.get("References", "") or ""
        for ref_token in refs_str.split("; "):
            ref_token = ref_token.strip()
            if not ref_token:
                continue
            if "!" in ref_token:
                parts = ref_token.split("!", 1)
                dep = (parts[0], parts[1])
            else:
                dep = (fc["_sheet"], ref_token)
            if dep in cell_keys and dep != my:
                cell_adj[dep].add(my)
                cell_rev[my].add(dep)

    # ------------------------------------------------------------------
    # 2. Build groups
    # ------------------------------------------------------------------
    groups: dict[int, list[dict]] = defaultdict(list)
    for fc in formula_cells:
        gid = fc.get("GroupID", 0) or 0
        if gid > 0:
            groups[gid].append(fc)

    # ------------------------------------------------------------------
    # 3. Build unit-level graph
    # ------------------------------------------------------------------
    cell_to_unit: dict[tuple, str] = {}
    unit_cells: dict[str, list[dict]] = {}

    for gid, members in groups.items():
        if len(members) >= 2:
            uid = f"g{gid}"
            unit_cells[uid] = members
            for m in members:
                cell_to_unit[(m["_sheet"], m["Cell"])] = uid

    for fc in formula_cells:
        key = (fc["_sheet"], fc["Cell"])
        if key not in cell_to_unit:
            uid = f"c{fc['_sheet']}!{fc['Cell']}"
            cell_to_unit[key] = uid
            unit_cells[uid] = [fc]

    unit_adj: dict[str, set[str]] = defaultdict(set)
    for fc in formula_cells:
        my_unit = cell_to_unit[(fc["_sheet"], fc["Cell"])]
        refs_str = fc.get("References", "") or ""
        for ref_token in refs_str.split("; "):
            ref_token = ref_token.strip()
            if not ref_token:
                continue
            if "!" in ref_token:
                parts = ref_token.split("!", 1)
                dep_key = (parts[0], parts[1])
            else:
                dep_key = (fc["_sheet"], ref_token)
            dep_unit = cell_to_unit.get(dep_key)
            if dep_unit and dep_unit != my_unit:
                unit_adj[dep_unit].add(my_unit)

    # ------------------------------------------------------------------
    # 4. Tarjan's SCC on the unit graph
    # ------------------------------------------------------------------
    all_unit_ids = list(unit_cells.keys())
    sccs = _uc._tarjan_scc(all_unit_ids, unit_adj)

    # Build unit → scc_id
    unit_to_scc: dict[str, int] = {}
    for i, scc in enumerate(sccs):
        for uid in scc:
            unit_to_scc[uid] = i

    # Groups inside a non-trivial SCC are broken into individual cells
    broken_groups: set[int] = set()
    for scc in sccs:
        if len(scc) <= 1:
            continue
        for uid in scc:
            if uid.startswith("g"):
                gid = int(uid[1:])
                broken_groups.add(gid)

    # ------------------------------------------------------------------
    # 5. Topo-sort the condensed DAG
    # ------------------------------------------------------------------
    scc_adj: dict[int, set[int]] = defaultdict(set)
    for src, dests in unit_adj.items():
        si = unit_to_scc.get(src)
        if si is None:
            continue
        for d in dests:
            di = unit_to_scc.get(d)
            if di is not None and di != si:
                scc_adj[si].add(di)
    scc_in = {i: 0 for i in range(len(sccs))}
    for si, dests in scc_adj.items():
        for di in dests:
            scc_in[di] += 1
    scc_queue = [i for i in range(len(sccs)) if scc_in[i] == 0]
    scc_order: list[int] = []
    while scc_queue:
        n = scc_queue.pop(0)
        scc_order.append(n)
        for nb in sorted(scc_adj.get(n, [])):
            scc_in[nb] -= 1
            if scc_in[nb] == 0:
                scc_queue.append(nb)
    seen_scc = set(scc_order)
    for i in range(len(sccs)):
        if i not in seen_scc:
            scc_order.append(i)

    # ------------------------------------------------------------------
    # 6. Within each SCC, cell-level Kahn sort → emission plan
    # ------------------------------------------------------------------
    emission_plan: list[tuple[str, Any]] = []

    for si in scc_order:
        scc_members = sccs[si]
        if len(scc_members) == 1:
            uid = scc_members[0]
            if uid.startswith("g") and int(uid[1:]) not in broken_groups:
                members = unit_cells[uid]
                if len(members) >= 2:
                    emission_plan.append(("group", members))
                else:
                    emission_plan.append(("cell", members[0]))
            else:
                for m in unit_cells[uid]:
                    emission_plan.append(("cell", m))
        else:
            # Non-trivial SCC: cell-level topo sort
            scc_cell_keys = set()
            scc_fc_map: dict[tuple, dict] = {}
            for uid in scc_members:
                for m in unit_cells[uid]:
                    k = (m["_sheet"], m["Cell"])
                    scc_cell_keys.add(k)
                    scc_fc_map[k] = m

            local_in: dict[tuple, int] = {k: 0 for k in scc_cell_keys}
            local_adj: dict[tuple, list[tuple]] = defaultdict(list)
            for k in scc_cell_keys:
                for dep in cell_rev.get(k, []):
                    if dep in scc_cell_keys:
                        local_adj[dep].append(k)
                        local_in[k] += 1
            lq = [k for k, d in local_in.items() if d == 0]
            local_order: list[tuple] = []
            while lq:
                n = lq.pop(0)
                local_order.append(n)
                for nb in local_adj.get(n, []):
                    local_in[nb] -= 1
                    if local_in[nb] == 0:
                        lq.append(nb)
            lo_set = set(local_order)
            for k in scc_cell_keys:
                if k not in lo_set:
                    local_order.append(k)

            emitted_in_scc: set[tuple] = set()
            group_members_seen: dict[int, int] = defaultdict(int)
            for k in local_order:
                if k in emitted_in_scc:
                    continue
                fc = scc_fc_map[k]
                gid = fc.get("GroupID", 0) or 0
                if gid > 0 and gid not in broken_groups and len(groups[gid]) >= 2:
                    group_members_seen[gid] += 1
                    in_scc_count = sum(
                        1 for m in groups[gid]
                        if (m["_sheet"], m["Cell"]) in scc_cell_keys
                    )
                    if group_members_seen[gid] >= in_scc_count:
                        emission_plan.append(("group", groups[gid]))
                        for m in groups[gid]:
                            emitted_in_scc.add((m["_sheet"], m["Cell"]))
                else:
                    emission_plan.append(("cell", fc))
                    emitted_in_scc.add(k)

    return emission_plan


# ---------------------------------------------------------------------------
# Runtime helpers for the generated script
# ---------------------------------------------------------------------------

# The structured calculator reuses ALL the runtime helpers from the
# unstructured version (Excel function emulations, _g, _s, _rng, etc.)
# and adds structured-input loading on top.

_STRUCTURED_INPUT_LOADER = r'''

# ======================================================================
# Structured-input loader
# ======================================================================

def _cell_to_rowcol(cell_ref):
    """Convert a cell reference like 'E2' to (row, col) tuple."""
    import re as _re
    ref = cell_ref.replace("$", "")
    m = _re.match(r"([A-Z]{1,3})(\d+)", ref)
    if not m:
        return 1, 1
    col_str, row_str = m.group(1), m.group(2)
    col = 0
    for ch in col_str:
        col = col * 26 + (ord(ch) - ord('A') + 1)
    return int(row_str), col


def _load_structured_input(si_path):
    """Read structured_input.xlsx and populate the workbook cells."""
    si_wb = openpyxl.load_workbook(si_path, data_only=True)

    # --- Read Config sheet (scalar inputs) ---
    if 'Config' in si_wb.sheetnames:
        ws_cfg = si_wb['Config']
        cfg_hdr = {ws_cfg.cell(1, c).value: c
                   for c in range(1, (ws_cfg.max_column or 0) + 1)}
        c_src  = cfg_hdr.get('SourceSheet', 1)
        c_cell = cfg_hdr.get('Cell', 2)
        c_val  = cfg_hdr.get('Value', 4)
        for r in range(2, (ws_cfg.max_row or 1) + 1):
            src_sheet = ws_cfg.cell(r, c_src).value
            cell_ref  = ws_cfg.cell(r, c_cell).value
            value     = ws_cfg.cell(r, c_val).value
            if src_sheet and cell_ref:
                try:
                    row_num, col_num = _cell_to_rowcol(cell_ref)
                    _s(src_sheet, row_num, col_num, value)
                except Exception:
                    pass

    # --- Read Index sheet to build tab → cell mappings ---
    if 'Index' in si_wb.sheetnames:
        ws_idx = si_wb['Index']
        idx_hdr = {ws_idx.cell(1, c).value: c
                   for c in range(1, (ws_idx.max_column or 0) + 1)}
        ci_src = idx_hdr.get('SourceSheet', 1)
        ci_tgt = idx_hdr.get('TargetSheet', 2)
        ci_typ = idx_hdr.get('InputType', 3)
        ci_row = idx_hdr.get('Row', 5)
        ci_col = idx_hdr.get('Col', 6)

        # Group entries by target sheet
        from collections import defaultdict as _ddict
        tab_entries = _ddict(list)
        for r in range(2, (ws_idx.max_row or 1) + 1):
            src = ws_idx.cell(r, ci_src).value
            tgt = ws_idx.cell(r, ci_tgt).value
            typ = ws_idx.cell(r, ci_typ).value
            row_num = ws_idx.cell(r, ci_row).value
            col_num = ws_idx.cell(r, ci_col).value
            if not src or not row_num or not col_num:
                continue
            if tgt and tgt != 'Config':
                tab_entries[tgt].append({
                    'src': src,
                    'typ': str(typ or ''),
                    'row': int(row_num),
                    'col': int(col_num),
                })

        # Read values from each data tab
        for tab_name, entries in tab_entries.items():
            if tab_name not in si_wb.sheetnames:
                continue
            ws_tab = si_wb[tab_name]
            is_transposed = any('transposed' in e['typ'].lower()
                                for e in entries)
            min_row = min(e['row'] for e in entries)
            min_col = min(e['col'] for e in entries)

            for e in entries:
                if is_transposed:
                    # Transposed: rows = periods (orig cols), cols = metrics (orig rows)
                    tab_r = (e['col'] - min_col) + 2
                    tab_c = (e['row'] - min_row) + 2
                else:
                    # Normal: rows = metrics (orig rows), cols = periods (orig cols)
                    tab_r = (e['row'] - min_row) + 2
                    tab_c = (e['col'] - min_col) + 2
                try:
                    val = ws_tab.cell(tab_r, tab_c).value
                    _s(e['src'], e['row'], e['col'], val)
                except Exception:
                    pass

    si_wb.close()

'''


# ---------------------------------------------------------------------------
# Code generation
# ---------------------------------------------------------------------------

def _generate_structured_code(
    sheets: dict[str, list[dict[str, Any]]],
    sheet_order: list[str],
    defined_names: dict[str, str] | None = None,
) -> str:
    """Generate the full ``structured_calculate.py`` script as a string."""

    emission_plan = _build_emission_plan(sheets, sheet_order, defined_names)

    # Start with the same runtime helpers as the unstructured version
    code_lines: list[str] = [_uc._RUNTIME_HELPERS]

    # Add structured-input loader
    code_lines.append(_STRUCTURED_INPUT_LOADER)

    # ------------------------------------------------------------------
    # Emit constant values (non-formula cells: labels, headers, inputs)
    # These provide the base workbook data; structured inputs override
    # Input-type cells afterwards.
    # ------------------------------------------------------------------
    code_lines.append("# ======================================================================")
    code_lines.append("# Constant cell values (labels, headers, default inputs)")
    code_lines.append("# ======================================================================")
    code_lines.append("")
    code_lines.append("def _load_constants():")
    code_lines.append('    """Populate non-formula cells from the mapping report."""')

    const_count = 0
    for sname in sheet_order:
        rows = sheets.get(sname, [])
        for rd in rows:
            cell_type = rd.get("Type", "")
            if cell_type in ("Calculation", "Output"):
                continue
            if rd.get("IncludeFlag") is False or rd.get("IncludeFlag") == 0:
                continue
            val = rd.get("Value")
            if val is None:
                continue
            # Skip non-serializable values (e.g., DataTableFormula objects)
            if not isinstance(val, (str, int, float, bool)):
                continue
            row_i = int(rd["Row"])
            col_i = int(rd["Col"])
            code_lines.append(f"    _s({repr(sname)}, {row_i}, {col_i}, {repr(val)})")
            const_count += 1

    if const_count == 0:
        code_lines.append("    pass")
    code_lines.append("")

    # ------------------------------------------------------------------
    # Formatting helper
    # ------------------------------------------------------------------
    code_lines.append("""
# ======================================================================
# Formatting helpers
# ======================================================================

def _apply_fmt(cell, fmt_dict):
    \"\"\"Apply formatting from the mapping report to a cell.\"\"\"
    cell.number_format = fmt_dict.get('nf', 'General')
    cell.font = Font(
        bold=fmt_dict.get('bold', False),
        italic=fmt_dict.get('italic', False),
        size=fmt_dict.get('size', 11),
    )
    ha = fmt_dict.get('ha') or None
    va = fmt_dict.get('va') or None
    cell.alignment = Alignment(horizontal=ha, vertical=va,
                               wrap_text=fmt_dict.get('wrap', False))
    fc = fmt_dict.get('fill', '')
    if fc and not str(fc).startswith('theme:') and not str(fc).startswith('indexed:'):
        try:
            cell.fill = PatternFill(start_color=str(fc), end_color=str(fc), fill_type='solid')
        except Exception:
            pass

""")

    # ------------------------------------------------------------------
    # main()
    # ------------------------------------------------------------------
    code_lines.append("def main(structured_input_path: str, output_path: str) -> None:")
    code_lines.append("    global _wb, _cache")
    code_lines.append("    _cache.clear()")
    code_lines.append("")
    code_lines.append("    # Create a fresh workbook with all original sheets")
    code_lines.append("    _wb = openpyxl.Workbook()")
    code_lines.append("    _wb.remove(_wb.active)")
    for sname in sheet_order:
        code_lines.append(f"    _wb.create_sheet({repr(sname)})")
    code_lines.append("")

    code_lines.append("    # Load constant values (labels, headers, default inputs)")
    code_lines.append("    _load_constants()")
    code_lines.append("")

    code_lines.append("    # Load structured inputs (overrides Input-type cells)")
    code_lines.append("    _load_structured_input(structured_input_path)")
    code_lines.append("")

    # --- Emit formula evaluations ---
    code_lines.append("    # === Evaluate formulas in topological order ===")
    code_lines.append("")

    for action, data in emission_plan:
        if action == "group":
            _uc._emit_group(code_lines, data, sheets)
        else:
            _uc._emit_single_cell(code_lines, data)

    # --- Emit formatting ---
    code_lines.append("")
    code_lines.append("    # === Apply formatting ===")
    for sname in sheet_order:
        rows = sheets.get(sname, [])
        for rd in rows:
            if rd.get("IncludeFlag") is False or rd.get("IncludeFlag") == 0:
                continue
            nf = rd.get("NumberFormat", "General") or "General"
            bold = bool(rd.get("FontBold", False))
            italic = bool(rd.get("FontItalic", False))
            size = rd.get("FontSize", 11) or 11
            fill = str(rd.get("FillColor", "") or "")
            ha = str(rd.get("HAlignment", "") or "")
            va = str(rd.get("VAlignment", "") or "")
            wrap = bool(rd.get("WrapText", False))
            if any([nf != "General", bold, italic, size != 11, fill, ha, va, wrap]):
                fmt_dict = {"nf": nf, "bold": bold, "italic": italic,
                            "size": size, "fill": fill, "ha": ha, "va": va,
                            "wrap": wrap}
                row_i = int(rd["Row"])
                col_i = int(rd["Col"])
                code_lines.append(
                    f"    _apply_fmt(_wb[{repr(sname)}].cell(row={row_i}, column={col_i}), "
                    f"{repr(fmt_dict)})"
                )

    # --- Save ---
    code_lines.append("")
    code_lines.append("    _wb.save(output_path)")
    code_lines.append("    print(f'✓ Output written to {output_path}')")
    code_lines.append("")
    code_lines.append("")
    code_lines.append("if __name__ == '__main__':")
    code_lines.append("    import sys")
    code_lines.append("    inp = sys.argv[1] if len(sys.argv) > 1 else 'structured_input.xlsx'")
    code_lines.append("    out = sys.argv[2] if len(sys.argv) > 2 else 'output.xlsx'")
    code_lines.append("    main(inp, out)")
    code_lines.append("")

    return "\n".join(code_lines)


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def generate_structured_calculate(report_path: str, output_dir: str) -> str:
    """Generate ``structured_calculate.py`` from the mapping report."""
    os.makedirs(output_dir, exist_ok=True)
    sheets, sheet_order, defined_names = _uc._read_report(report_path)
    code = _generate_structured_code(sheets, sheet_order, defined_names)
    out_path = os.path.join(output_dir, "structured_calculate.py")
    with open(out_path, "w") as f:
        f.write(code)
    print(f"  → {out_path}")
    return out_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Generate structured_calculate.py")
    ap.add_argument("report", help="Path to mapping_report.xlsx")
    ap.add_argument("output_dir", help="Output directory")
    args = ap.parse_args()
    generate_structured_calculate(args.report, args.output_dir)
    print("✓ Done")


if __name__ == "__main__":
    main()
