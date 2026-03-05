"""
generate_unstructured_calculate.py — Layer 3a code generator.

Reads ``mapping_report.xlsx`` and emits ``unstructured_calculate.py``, a
standalone Python script that:

  1. Reads ``unstructured_inputs.xlsx`` (all Input cells in original layout)
  2. Evaluates every Calculation and Output formula in topological order
  3. Applies formatting
  4. Writes ``output.xlsx``

**Vectorisation strategy**
  • Grouped formulas (same R1C1 pattern in a contiguous run) are emitted as
    loops over the group range rather than one statement per cell.
  • Groups whose patterns have NO within-group dependency are emitted as
    NumPy-vectorised batch operations.

Usage:
    python -m src.generate_unstructured_calculate \\
        output/ACC-Ltd/mapping_report.xlsx output/ACC-Ltd/
"""

from __future__ import annotations

import argparse
import os
import re
import textwrap
import warnings
from collections import defaultdict
from typing import Any

import openpyxl

from src.formula_parser import (
    extract_references,
    to_r1c1_pattern,
    r1c1_to_a1,
    rowcol_to_cell,
    cell_to_rowcol,
)

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Module-level storage for defined names (populated by _generate_code)
_DEFINED_NAMES: dict[str, str] = {}


def _resolve_named_ranges(formula: str, sheet_name: str) -> str:
    """Replace named range references in *formula* with their cell references.

    Named ranges like ``Net_Sales`` are replaced with ``'Income statement'!$F$8``
    (or their plain cell ref if on the same sheet).
    """
    if not _DEFINED_NAMES:
        return formula
    for name, ref in _DEFINED_NAMES.items():
        # Only replace whole-word occurrences (not inside other identifiers)
        # Named range names can contain underscores, digits, dots
        pattern = r'(?<![A-Za-z0-9_\.])' + re.escape(name) + r'(?![A-Za-z0-9_\.])'
        if re.search(pattern, formula):
            formula = re.sub(pattern, ref, formula)
    return formula


# ---------------------------------------------------------------------------
# Read mapping report into structured data
# ---------------------------------------------------------------------------

def _read_report(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets: dict[str, list[dict[str, Any]]] = {}
    sheet_order: list[str] = []
    defined_names: dict[str, str] = {}
    for sname in wb.sheetnames:
        if sname == "_Metadata":
            continue
        if sname == "_DefinedNames":
            ws_dn = wb[sname]
            for r in range(2, (ws_dn.max_row or 1) + 1):
                name_val = ws_dn.cell(row=r, column=1).value
                ref_val = ws_dn.cell(row=r, column=2).value
                if name_val and ref_val:
                    defined_names[str(name_val)] = str(ref_val)
            continue
        ws = wb[sname]
        headers = [ws.cell(row=1, column=c).value for c in range(1, (ws.max_column or 0) + 1)]
        rows = []
        for r in range(2, (ws.max_row or 1) + 1):
            d = {}
            for ci, h in enumerate(headers, start=1):
                d[h] = ws.cell(row=r, column=ci).value
            rows.append(d)
        sheets[sname] = rows
        sheet_order.append(sname)
    wb.close()
    return sheets, sheet_order, defined_names


# ---------------------------------------------------------------------------
# Topological sort of formula cells
# ---------------------------------------------------------------------------

def _topo_sort(formula_cells: list[dict], all_cells_map: dict):
    """Return *formula_cells* in evaluation order (dependencies first).

    Uses Kahn's algorithm on the dependency graph derived from the
    References column in the mapping report.
    """
    # Build adjacency from references text
    # key = (sheet, cell_ref)
    key_to_cell: dict[tuple[str, str], dict] = {}
    for fc in formula_cells:
        k = (fc["_sheet"], fc["Cell"])
        key_to_cell[k] = fc

    # Parse references
    in_degree: dict[tuple[str, str], int] = {k: 0 for k in key_to_cell}
    adj: dict[tuple[str, str], list[tuple[str, str]]] = defaultdict(list)

    for k, fc in key_to_cell.items():
        refs_str = fc.get("References", "") or ""
        if not refs_str:
            continue
        for ref_token in refs_str.split("; "):
            ref_token = ref_token.strip()
            if not ref_token:
                continue
            if "!" in ref_token:
                parts = ref_token.split("!", 1)
                dep_key = (parts[0], parts[1])
            else:
                dep_key = (fc["_sheet"], ref_token)
            if dep_key in key_to_cell:
                adj[dep_key].append(k)
                in_degree[k] = in_degree.get(k, 0) + 1

    # Kahn's
    queue = [k for k, deg in in_degree.items() if deg == 0]
    ordered = []
    while queue:
        node = queue.pop(0)
        ordered.append(node)
        for neighbour in adj.get(node, []):
            in_degree[neighbour] -= 1
            if in_degree[neighbour] == 0:
                queue.append(neighbour)

    # Any remaining (cycles) — append in original order
    ordered_set = set(ordered)
    for k in key_to_cell:
        if k not in ordered_set:
            ordered.append(k)

    return [key_to_cell[k] for k in ordered if k in key_to_cell]


# ---------------------------------------------------------------------------
# Excel formula → Python expression translator
# ---------------------------------------------------------------------------

_FUNC_MAP = {
    "SUM":   "_xl_sum",
    "IF":    "_xl_if",
    "ABS":   "abs",
    "MAX":   "max",
    "MIN":   "min",
    "ROUND": "round",
    "INT":   "int",
    "MOD":   "_xl_mod",
    "POWER": "_xl_power",
    "SQRT":  "_xl_sqrt",
    "LN":    "_xl_ln",
    "LOG":   "_xl_log",
    "LOG10": "_xl_log10",
    "EXP":   "_xl_exp",
    "AND":   "_xl_and",
    "OR":    "_xl_or",
    "NOT":   "_xl_not",
    "AVERAGE": "_xl_average",
    "COUNT":   "_xl_count",
    "COUNTA":  "_xl_counta",
    "COUNTIF": "_xl_countif",
    "VLOOKUP": "_xl_vlookup",
    "HLOOKUP": "_xl_hlookup",
    "INDEX":   "_xl_index",
    "MATCH":   "_xl_match",
    "IFERROR": "_xl_iferror",
    "ISERROR": "_xl_iserror",
    "ISBLANK": "_xl_isblank",
    "LEFT":    "_xl_left",
    "RIGHT":   "_xl_right",
    "MID":     "_xl_mid",
    "LEN":     "_xl_len",
    "TRIM":    "_xl_trim",
    "UPPER":   "_xl_upper",
    "LOWER":   "_xl_lower",
    "CONCATENATE": "_xl_concatenate",
    "TEXT":    "_xl_text",
    "VALUE":   "_xl_value",
    "SUMIF":   "_xl_sumif",
    "SUMIFS":  "_xl_sumifs",
    "SUMPRODUCT": "_xl_sumproduct",
    "NPV":    "_xl_npv",
    "IRR":    "_xl_irr",
    "PMT":    "_xl_pmt",
    "PV":     "_xl_pv",
    "FV":     "_xl_fv",
    "RATE":   "_xl_rate",
    "NPER":   "_xl_nper",
    "PPMT":   "_xl_ppmt",
    "IPMT":   "_xl_ipmt",
    "CUMIPMT": "_xl_cumipmt",
    "CUMPRINC": "_xl_cumprinc",
    "TRANSPOSE": "_xl_transpose",
    "OFFSET": "_xl_offset",
    "INDIRECT": "_xl_indirect",
    "ROW":    "_xl_row",
    "COLUMN": "_xl_column",
    "ROWS":   "_xl_rows",
    "COLUMNS": "_xl_columns",
    "YEAR":   "_xl_year",
    "MONTH":  "_xl_month",
    "DAY":    "_xl_day",
    "DATE":   "_xl_date",
    "TODAY":  "_xl_today",
    "NOW":    "_xl_now",
    "NETWORKDAYS": "_xl_networkdays",
    "EDATE":  "_xl_edate",
    "EOMONTH": "_xl_eomonth",
    "PI":     "_xl_pi",
    "CHOOSE": "_xl_choose",
    "LARGE":  "_xl_large",
    "SMALL":  "_xl_small",
    "RANK":   "_xl_rank",
    "MEDIAN": "_xl_median",
    "STDEV":  "_xl_stdev",
    "VAR":    "_xl_var",
    "NA":     "_xl_na",
    "ROUNDDOWN": "_xl_rounddown",
    "ROUNDUP":   "_xl_roundup",
    "CEILING":   "_xl_ceiling",
    "FLOOR":     "_xl_floor",
    "SIGN":      "_xl_sign",
    "COUNTBLANK": "_xl_countblank",
    "AVERAGEIF":  "_xl_averageif",
    "COUNTIFS":   "_xl_countifs",
    "FIND":      "_xl_find",
    "SEARCH":    "_xl_search",
    "SUBSTITUTE": "_xl_substitute",
    "REPLACE":    "_xl_replace_func",
    "REPT":      "_xl_rept",
    "EXACT":     "_xl_exact",
    "TYPE":      "_xl_type",
    "N":         "_xl_n",
    "T":         "_xl_t",
    "ISNUMBER":  "_xl_isnumber",
    "ISTEXT":    "_xl_istext",
    "ISNA":      "_xl_isna",
    "NUMBERVALUE": "_xl_numbervalue",
    "CHAR":      "chr",
    "CODE":      "ord",
}


def _formula_to_python(formula: str, sheet: str, all_sheets: list[str]) -> str:
    """Best-effort translation of an Excel formula to a Python expression.

    Cell references become ``_g(sheet, row, col)`` calls.
    Range references become ``_rng(sheet, r1, c1, r2, c2)`` calls.
    Functions are mapped to helper wrappers.
    """
    if not formula or not formula.startswith("="):
        return repr(formula)

    expr = formula[1:]  # strip leading =

    # Resolve named ranges before any other translation
    expr = _resolve_named_ranges(expr, sheet)

    # Mask string literals
    strings: list[str] = []
    def _mask(m):
        strings.append(m.group())
        return f"__XLSTR{len(strings)-1}__"
    expr = re.sub(r'"[^"]*"', _mask, expr)

    # Replace Excel error constants with None
    expr = re.sub(r'#REF!', 'None', expr)
    expr = re.sub(r'#N/A', 'None', expr)
    expr = re.sub(r'#VALUE!', 'None', expr)
    expr = re.sub(r'#DIV/0!', 'None', expr)
    expr = re.sub(r'#NAME\?', 'None', expr)
    expr = re.sub(r'#NULL!', 'None', expr)
    expr = re.sub(r'#NUM!', 'None', expr)

    # Replace structured table references with None (not supported)
    # e.g. TableName[[#This Row],[Col]], Data[[#Headers],[12]], Table[Column]
    expr = re.sub(r'[A-Za-z_]\w*\[\[.*?\]\]', 'None', expr)
    expr = re.sub(r'[A-Za-z_]\w*\[#[^\]]*\]', 'None', expr)
    # Also handle remaining table refs: Table[], Table[Col] (no #)
    expr = re.sub(r'[A-Za-z_]\w*\[\]', 'None', expr)
    expr = re.sub(r'(?<![\'"])\b[A-Z][A-Za-z_]\w*\[[A-Za-z][^\]]*\]', 'None', expr)

    # Replace sheet-qualified range references: 'Sheet'!A1:B5 or Sheet!A1:B5
    def _repl_range(m):
        sh = m.group(1) or m.group(2)
        c1 = m.group(3).replace("$", "")
        c2 = m.group(4).replace("$", "")
        r1, co1 = cell_to_rowcol(c1)
        r2, co2 = cell_to_rowcol(c2)
        return f"_rng({repr(sh)},{r1},{co1},{r2},{co2})"

    expr = re.sub(
        r"(?:'([^']*)'|([A-Za-z0-9_]+))!"
        r"(\$?[A-Z]{1,3}\$?\d+):(\$?[A-Z]{1,3}\$?\d+)",
        _repl_range, expr
    )

    # Replace local range references: A1:B5
    def _repl_local_range(m):
        # Must not be preceded by ! (already handled above) or a letter
        c1 = m.group(1).replace("$", "")
        c2 = m.group(2).replace("$", "")
        r1, co1 = cell_to_rowcol(c1)
        r2, co2 = cell_to_rowcol(c2)
        return f"_rng({repr(sheet)},{r1},{co1},{r2},{co2})"

    expr = re.sub(
        r"(?<![A-Za-z0-9_!])(\$?[A-Z]{1,3}\$?\d+):(\$?[A-Z]{1,3}\$?\d+)",
        _repl_local_range, expr
    )

    # Replace sheet-qualified cell references: 'Sheet'!A1 or Sheet!A1
    def _repl_sheet_cell(m):
        sh = m.group(1) or m.group(2)
        ref = m.group(3).replace("$", "")
        r, c = cell_to_rowcol(ref)
        return f"_g({repr(sh)},{r},{c})"

    expr = re.sub(
        r"(?:'([^']*)'|([A-Za-z0-9_]+))!"
        r"(\$?[A-Z]{1,3}\$?\d+)"
        r"(?![:\(A-Z0-9])",
        _repl_sheet_cell, expr
    )

    # Replace local cell references: A1
    def _repl_local_cell(m):
        ref = m.group(0).replace("$", "")
        r, c = cell_to_rowcol(ref)
        return f"_g({repr(sheet)},{r},{c})"

    expr = re.sub(
        r"(?<![A-Za-z0-9_!'\"])(\$?[A-Z]{1,3}\$?\d+)(?![\(:A-Z0-9])",
        _repl_local_cell, expr
    )

    # Replace Excel functions with Python wrappers
    for xl_func, py_func in _FUNC_MAP.items():
        # Case-insensitive function replacement
        expr = re.sub(
            r'\b' + xl_func + r'\s*\(',
            py_func + '(',
            expr,
            flags=re.IGNORECASE,
        )

    # Replace Excel % postfix operator:  number% or expr)% → /100
    expr = re.sub(r'(?<=[\d\)])%', '/100', expr)

    # Replace Excel operators
    expr = expr.replace("<>", "!=")
    # = comparison → == (but not <=, >=, !=, ==)
    expr = re.sub(r'(?<![<>!=])=(?!=)', '==', expr)
    expr = expr.replace("^", "**")
    # & means string concatenation in Excel
    expr = re.sub(r'&', '+', expr)

    # Replace TRUE/FALSE
    expr = re.sub(r'\bTRUE\b', 'True', expr, flags=re.IGNORECASE)
    expr = re.sub(r'\bFALSE\b', 'False', expr, flags=re.IGNORECASE)

    # Unmask strings
    for i, s in enumerate(strings):
        expr = expr.replace(f"__XLSTR{i}__", s)

    # Use _rng2d for INDEX first argument (INDEX needs 2D, not flat)
    expr = expr.replace('_xl_index(_rng(', '_xl_index(_rng2d(')

    # Safety: if untranslated range operator ':' remains (e.g. dynamic ranges
    # like A1:OFFSET(...)), the expression would be invalid Python.
    # Detect '):' or number ':' patterns that aren't inside strings.
    if re.search(r'\)\s*:', expr) or re.search(r'\d\s*:\s*_xl_', expr):
        return 'None'

    # Fix empty arguments left by table-ref → None replacement (e.g. INDEX(None,,1))
    while ',,' in expr:
        expr = expr.replace(',,', ',None,')

    # Final safety: try to compile expression; return None if syntax is invalid
    try:
        compile(expr, '<expr>', 'eval')
    except SyntaxError:
        return 'None'

    return expr


def _is_abs_row(ref_raw: str) -> bool:
    """Check if the row part of a reference is absolute ($)."""
    m = re.match(r"\$?[A-Z]+(\$)\d+", ref_raw)
    return m is not None


def _is_abs_col(ref_raw: str) -> bool:
    """Check if the column part of a reference is absolute ($)."""
    return ref_raw.startswith("$")


# ---------------------------------------------------------------------------
# Code emitter
# ---------------------------------------------------------------------------

_RUNTIME_HELPERS = r'''
"""
Auto-generated unstructured calculation script.
Reads unstructured_inputs.xlsx and produces output.xlsx.
"""
import math
import os
import sys
import warnings
from datetime import datetime, date, timedelta
from typing import Any

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ======================================================================
# Runtime helpers — lightweight Excel-function emulation
# ======================================================================

_wb = None  # global workbook handle (set in main)
_cache: dict[tuple[str, int, int], Any] = {}
_NM = [False]  # _NM[0]: When True, _g() returns 0 for None (TypeError retry)


def _g(sheet: str, row: int, col: int) -> Any:
    """Get a cell value (with caching).
    
    When _NM[0] is True, returns 0 for None (used in TypeError retry).
    """
    key = (sheet, row, col)
    if key in _cache:
        v = _cache[key]
        return 0 if (_NM[0] and v is None) else v
    ws = _wb[sheet]
    v = ws.cell(row=row, column=col).value
    _cache[key] = v
    return 0 if (_NM[0] and v is None) else v


def _gn(sheet: str, row: int, col: int):
    """Get cell value coerced to numeric (None→0) for arithmetic safety."""
    v = _g(sheet, row, col)
    if v is None:
        return 0
    if isinstance(v, bool):
        return 1 if v else 0
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0


def _s(sheet: str, row: int, col: int, value: Any) -> None:
    """Set a cell value (and update cache)."""
    _cache[(sheet, row, col)] = value
    _wb[sheet].cell(row=row, column=col).value = value


def _rng(sheet: str, r1: int, c1: int, r2: int, c2: int) -> list:
    """Return flat list of values in a rectangular range (row-major order)."""
    vals = []
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            vals.append(_g(sheet, r, c))
    return vals


def _rng2d(sheet: str, r1: int, c1: int, r2: int, c2: int) -> list:
    """Return 2D list (list of row-lists) for INDEX."""
    rows = []
    for r in range(r1, r2 + 1):
        row = []
        for c in range(c1, c2 + 1):
            row.append(_g(sheet, r, c))
        rows.append(row)
    return rows


def _rng_ncols(sheet: str, r1: int, c1: int, r2: int, c2: int) -> int:
    """Return the number of columns in a range (for VLOOKUP grid inference)."""
    return c2 - c1 + 1


def _num(v: Any) -> float:
    """Coerce to float, treating None/blank/bool like Excel does."""
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


# --- Excel function wrappers ---

def _xl_sum(*args):
    total = 0.0
    for a in args:
        if isinstance(a, (list, tuple)):
            for v in a:
                total += _num(v)
        else:
            total += _num(a)
    return total

def _xl_if(cond, true_val, false_val=False):
    return true_val if cond else false_val

def _xl_mod(a, b):
    return _num(a) % _num(b) if _num(b) != 0 else 0

def _xl_power(a, b):
    return _num(a) ** _num(b)

def _xl_sqrt(a):
    return math.sqrt(_num(a))

def _xl_ln(a):
    v = _num(a)
    return math.log(v) if v > 0 else 0

def _xl_log(a, base=10):
    v = _num(a)
    return math.log(v, _num(base)) if v > 0 else 0

def _xl_log10(a):
    v = _num(a)
    return math.log10(v) if v > 0 else 0

def _xl_exp(a):
    return math.exp(_num(a))

def _xl_and(*args):
    flat = []
    for a in args:
        if isinstance(a, (list, tuple)):
            flat.extend(a)
        else:
            flat.append(a)
    return all(bool(v) for v in flat)

def _xl_or(*args):
    flat = []
    for a in args:
        if isinstance(a, (list, tuple)):
            flat.extend(a)
        else:
            flat.append(a)
    return any(bool(v) for v in flat)

def _xl_not(a):
    return not bool(a)

def _xl_average(*args):
    vals = []
    for a in args:
        if isinstance(a, (list, tuple)):
            vals.extend(_num(v) for v in a if v is not None)
        elif a is not None:
            vals.append(_num(a))
    return sum(vals) / len(vals) if vals else 0

def _xl_count(*args):
    n = 0
    for a in args:
        if isinstance(a, (list, tuple)):
            n += sum(1 for v in a if isinstance(v, (int, float)) and v is not None)
        elif isinstance(a, (int, float)) and a is not None:
            n += 1
    return n

def _xl_counta(*args):
    n = 0
    for a in args:
        if isinstance(a, (list, tuple)):
            n += sum(1 for v in a if v is not None and v != "")
        elif a is not None and a != "":
            n += 1
    return n

def _xl_countif(rng, criteria):
    if not isinstance(rng, (list, tuple)):
        rng = [rng]
    return sum(1 for v in rng if _match_criteria(v, criteria))

def _match_criteria(v, criteria):
    if isinstance(criteria, str):
        if criteria.startswith(">="):
            return _num(v) >= _num(criteria[2:])
        if criteria.startswith("<="):
            return _num(v) <= _num(criteria[2:])
        if criteria.startswith("<>"):
            return str(v) != criteria[2:]
        if criteria.startswith(">"):
            return _num(v) > _num(criteria[1:])
        if criteria.startswith("<"):
            return _num(v) < _num(criteria[1:])
        if criteria.startswith("="):
            return str(v).lower() == criteria[1:].lower()
        return str(v).lower() == criteria.lower()
    return v == criteria

def _xl_vlookup(lookup, table_range, col_idx, approx=True):
    """VLOOKUP over a flat list interpreted as a grid.

    *table_range* is a flat list from _rng().  We infer the number of columns
    from col_idx and the total length.  If *approx* is False (exact match)
    we search column-0 for *lookup*.  If True, we do the sorted-ascending
    approximate match (find last value <= lookup).
    """
    if not isinstance(table_range, (list, tuple)) or not table_range:
        return None
    ci = int(_num(col_idx))
    if ci < 1:
        return None
    # Try to infer number of columns from context — heuristic: ci is the
    # maximum useful column count.  We need len(table_range) % ncols == 0.
    n = len(table_range)
    ncols = ci  # minimum columns needed
    # If n is divisible by ci, use that. Otherwise try larger column counts.
    while ncols <= n:
        if n % ncols == 0:
            break
        ncols += 1
    if ncols > n:
        return None
    nrows = n // ncols
    # Build key column (column 0) and result column (col_idx - 1)
    keys = [table_range[r * ncols] for r in range(nrows)]
    results = [table_range[r * ncols + ci - 1] for r in range(nrows)]
    lv = lookup
    approx_flag = bool(approx) if not isinstance(approx, (int, float)) else int(approx) != 0
    if not approx_flag:
        # Exact match
        for i, k in enumerate(keys):
            if k == lv:
                return results[i]
            if isinstance(k, (int, float)) and isinstance(lv, (int, float)):
                if math.isclose(k, lv, rel_tol=1e-9):
                    return results[i]
            elif str(k).lower() == str(lv).lower():
                return results[i]
        return None
    else:
        # Approximate match (sorted ascending)
        best = None
        for i, k in enumerate(keys):
            kn = _num(k)
            ln = _num(lv)
            if kn <= ln:
                best = i
            else:
                break
        return results[best] if best is not None else None

def _xl_hlookup(lookup, table_range, row_idx, approx=True):
    """HLOOKUP — horizontal lookup in a flat grid."""
    if not isinstance(table_range, (list, tuple)) or not table_range:
        return None
    ri = int(_num(row_idx))
    if ri < 1:
        return None
    n = len(table_range)
    nrows = ri
    while nrows <= n:
        if n % nrows == 0:
            break
        nrows += 1
    if nrows > n:
        return None
    ncols = n // nrows
    keys = [table_range[c] for c in range(ncols)]
    results = [table_range[(ri - 1) * ncols + c] for c in range(ncols)]
    lv = lookup
    approx_flag = bool(approx) if not isinstance(approx, (int, float)) else int(approx) != 0
    if not approx_flag:
        for i, k in enumerate(keys):
            if k == lv or (isinstance(k, (int, float)) and isinstance(lv, (int, float))
                           and math.isclose(k, lv, rel_tol=1e-9)):
                return results[i]
            elif isinstance(k, str) and isinstance(lv, str) and k.lower() == lv.lower():
                return results[i]
        return None
    else:
        best = None
        for i, k in enumerate(keys):
            if _num(k) <= _num(lv):
                best = i
            else:
                break
        return results[best] if best is not None else None

def _xl_index(rng, row_num, col_num=None):
    if isinstance(rng, (list, tuple)):
        idx = int(_num(row_num)) - 1
        if 0 <= idx < len(rng):
            v = rng[idx]
            if isinstance(v, (list, tuple)):
                if col_num is not None:
                    ci = int(_num(col_num)) - 1
                    if 0 <= ci < len(v):
                        return v[ci]
                    return None
                # Single-column range: unwrap
                if len(v) == 1:
                    return v[0]
                return v
            # Flat list with col_num (fallback)
            if col_num is not None:
                return None  # need 2D range
            return v
    return None

def _xl_match(lookup, rng, match_type=1):
    if not isinstance(rng, (list, tuple)):
        return None
    for i, v in enumerate(rng):
        if v == lookup:
            return i + 1
        if isinstance(v, (int, float)) and isinstance(lookup, (int, float)):
            if math.isclose(v, lookup, rel_tol=1e-9):
                return i + 1
    return None

def _xl_iferror(expr, fallback):
    try:
        if expr is None or (isinstance(expr, float) and math.isnan(expr)):
            return fallback
        return expr
    except Exception:
        return fallback

def _xl_iserror(v):
    return v is None or (isinstance(v, float) and math.isnan(v))

def _xl_isblank(v):
    return v is None or v == ""

def _xl_left(s, n=1):
    return str(s or "")[:int(n)]

def _xl_right(s, n=1):
    return str(s or "")[-int(n):]

def _xl_mid(s, start, length):
    s = str(s or "")
    return s[int(start)-1:int(start)-1+int(length)]

def _xl_len(s):
    return len(str(s or ""))

def _xl_trim(s):
    return str(s or "").strip()

def _xl_upper(s):
    return str(s or "").upper()

def _xl_lower(s):
    return str(s or "").lower()

def _xl_concatenate(*args):
    return "".join(str(a or "") for a in args)

def _xl_text(v, fmt):
    try:
        if isinstance(v, (int, float)):
            if "%" in str(fmt):
                return f"{v:.2%}"
            return str(v)
        return str(v or "")
    except Exception:
        return str(v or "")

def _xl_value(s):
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0

def _xl_sumif(rng, criteria, sum_range=None):
    if sum_range is None:
        sum_range = rng
    if not isinstance(rng, (list, tuple)):
        rng = [rng]
    if not isinstance(sum_range, (list, tuple)):
        sum_range = [sum_range]
    total = 0.0
    for v, s in zip(rng, sum_range):
        if _match_criteria(v, criteria):
            total += _num(s)
    return total

def _xl_sumifs(sum_range, *crit_pairs):
    if not isinstance(sum_range, (list, tuple)):
        sum_range = [sum_range]
    n = len(sum_range)
    pairs = list(crit_pairs)
    criteria_ranges = []
    while len(pairs) >= 2:
        cr = pairs.pop(0)
        cv = pairs.pop(0)
        if not isinstance(cr, (list, tuple)):
            cr = [cr]
        criteria_ranges.append((cr, cv))
    total = 0.0
    for i in range(n):
        match = True
        for cr, cv in criteria_ranges:
            if i < len(cr) and not _match_criteria(cr[i], cv):
                match = False
                break
        if match and i < len(sum_range):
            total += _num(sum_range[i])
    return total

def _xl_sumproduct(*args):
    arrays = []
    for a in args:
        if isinstance(a, (list, tuple)):
            arrays.append([_num(v) for v in a])
        else:
            arrays.append([_num(a)])
    if not arrays:
        return 0
    length = min(len(a) for a in arrays)
    total = 0.0
    for i in range(length):
        prod = 1.0
        for a in arrays:
            prod *= a[i]
        total += prod
    return total

def _xl_npv(rate, *cashflows):
    r = _num(rate)
    flat = []
    for c in cashflows:
        if isinstance(c, (list, tuple)):
            flat.extend(c)
        else:
            flat.append(c)
    total = 0.0
    for i, cf in enumerate(flat, 1):
        total += _num(cf) / (1 + r) ** i
    return total

def _xl_irr(values, guess=0.1):
    if not isinstance(values, (list, tuple)):
        return 0
    vals = [_num(v) for v in values]
    rate = _num(guess)
    for _ in range(1000):
        npv = sum(v / (1 + rate) ** i for i, v in enumerate(vals))
        dnpv = sum(-i * v / (1 + rate) ** (i + 1) for i, v in enumerate(vals))
        if abs(dnpv) < 1e-12:
            break
        rate -= npv / dnpv
        if abs(npv) < 1e-8:
            break
    return rate

def _xl_pmt(rate, nper, pv, fv=0, ptype=0):
    r, n, p, f = _num(rate), _num(nper), _num(pv), _num(fv)
    if r == 0:
        return -(p + f) / n if n != 0 else 0
    pmt = -r * (p * (1 + r)**n + f) / (((1 + r)**n - 1) * (1 + r * _num(ptype)))
    return pmt

def _xl_pv(rate, nper, pmt, fv=0, ptype=0):
    r, n, pm, f = _num(rate), _num(nper), _num(pmt), _num(fv)
    if r == 0:
        return -(pm * n + f)
    t = (1 + r)**n
    return -(pm * (1 + r * _num(ptype)) * (t - 1) / (r * t) + f / t)

def _xl_fv(rate, nper, pmt, pv=0, ptype=0):
    r, n, pm, p = _num(rate), _num(nper), _num(pmt), _num(pv)
    if r == 0:
        return -(p + pm * n)
    t = (1 + r)**n
    return -(p * t + pm * (1 + r * _num(ptype)) * (t - 1) / r)

def _xl_rate(nper, pmt, pv, fv=0, ptype=0, guess=0.1):
    # Newton's method
    rate = _num(guess)
    for _ in range(100):
        t = (1 + rate)**_num(nper)
        f = _num(pv)*t + _num(pmt)*(1+rate*_num(ptype))*(t-1)/rate + _num(fv) if rate != 0 else 0
        if abs(f) < 1e-10:
            break
        # numerical derivative
        dr = 1e-6
        t2 = (1 + rate + dr)**_num(nper)
        f2 = _num(pv)*t2 + _num(pmt)*(1+(rate+dr)*_num(ptype))*(t2-1)/(rate+dr) + _num(fv) if (rate+dr)!=0 else 0
        deriv = (f2 - f) / dr
        if abs(deriv) < 1e-15:
            break
        rate -= f / deriv
    return rate

def _xl_nper(rate, pmt, pv, fv=0, ptype=0):
    r, pm, p, f = _num(rate), _num(pmt), _num(pv), _num(fv)
    if r == 0:
        return -(p + f) / pm if pm != 0 else 0
    w = pm * (1 + r * _num(ptype))
    return math.log((-f * r + w) / (p * r + w)) / math.log(1 + r) if (p * r + w) != 0 else 0

def _xl_ppmt(rate, per, nper, pv, fv=0, ptype=0):
    pmt_val = _xl_pmt(rate, nper, pv, fv, ptype)
    ipmt_val = _xl_ipmt(rate, per, nper, pv, fv, ptype)
    return pmt_val - ipmt_val

def _xl_ipmt(rate, per, nper, pv, fv=0, ptype=0):
    r = _num(rate)
    p = _num(per)
    pmt_val = _xl_pmt(rate, nper, pv, fv, ptype)
    if r == 0:
        return 0
    fv_before = _num(pv) * (1 + r)**(p - 1) + pmt_val * ((1 + r)**(p - 1) - 1) / r
    return -fv_before * r

def _xl_cumipmt(rate, nper, pv, start, end, ptype):
    total = 0.0
    for per in range(int(_num(start)), int(_num(end)) + 1):
        total += _xl_ipmt(rate, per, nper, pv, 0, ptype)
    return total

def _xl_cumprinc(rate, nper, pv, start, end, ptype):
    total = 0.0
    for per in range(int(_num(start)), int(_num(end)) + 1):
        total += _xl_ppmt(rate, per, nper, pv, 0, ptype)
    return total

def _xl_transpose(rng):
    """Transpose a range list — no-op for flat values."""
    return rng

def _xl_offset(base_sheet, base_row, base_col, rows, cols, height=None, width=None):
    """OFFSET — returns a single cell value or range list.

    Caller must supply (sheet, row, col) for the base reference.
    """
    r = int(_num(base_row)) + int(_num(rows))
    c = int(_num(base_col)) + int(_num(cols))
    if height is not None and width is not None:
        return _rng(base_sheet, r, c, r + int(_num(height)) - 1, c + int(_num(width)) - 1)
    return _g(base_sheet, r, c)

def _xl_indirect(ref_str, a1=True):
    """INDIRECT — limited: only handles plain cell refs like 'Sheet!A1' or 'A1'."""
    return None  # Still limited without dynamic sheet resolution

def _xl_row(ref=None):
    return ref if isinstance(ref, int) else 0

def _xl_column(ref=None):
    return ref if isinstance(ref, int) else 0

def _xl_rows(rng):
    return len(rng) if isinstance(rng, (list, tuple)) else 1

def _xl_columns(rng):
    """Return number of elements (approximation for 1D list)."""
    return len(rng) if isinstance(rng, (list, tuple)) else 1

def _xl_year(d):
    if isinstance(d, (datetime, date)):
        return d.year
    return int(_num(d))

def _xl_month(d):
    if isinstance(d, (datetime, date)):
        return d.month
    return 1

def _xl_day(d):
    if isinstance(d, (datetime, date)):
        return d.day
    return 1

def _xl_date(y, m, d):
    try:
        return date(int(y), int(m), int(d))
    except Exception:
        return None

def _xl_today():
    return date.today()

def _xl_now():
    return datetime.now()

def _xl_networkdays(start, end, holidays=None):
    """Count business days between dates (simplistic)."""
    try:
        from datetime import timedelta
        if not isinstance(start, (datetime, date)):
            return 0
        if not isinstance(end, (datetime, date)):
            return 0
        d = start if isinstance(start, date) else start.date()
        e = end if isinstance(end, date) else end.date()
        count = 0
        step = 1 if e >= d else -1
        cur = d
        while (step > 0 and cur <= e) or (step < 0 and cur >= e):
            if cur.weekday() < 5:
                count += 1
            cur += timedelta(days=step)
        return count
    except Exception:
        return 0

def _xl_edate(start, months):
    """Return date that is N months after start."""
    try:
        if not isinstance(start, (datetime, date)):
            return start
        m = int(_num(months))
        month = start.month + m
        year = start.year + (month - 1) // 12
        month = (month - 1) % 12 + 1
        import calendar
        day = min(start.day, calendar.monthrange(year, month)[1])
        if isinstance(start, datetime):
            return datetime(year, month, day)
        return date(year, month, day)
    except Exception:
        return start

def _xl_eomonth(start, months):
    """Return end of month that is N months from start."""
    try:
        d = _xl_edate(start, months)
        if not isinstance(d, (datetime, date)):
            return d
        import calendar
        last_day = calendar.monthrange(d.year, d.month)[1]
        if isinstance(d, datetime):
            return datetime(d.year, d.month, last_day)
        return date(d.year, d.month, last_day)
    except Exception:
        return start

def _xl_pi():
    return math.pi

def _xl_choose(idx, *args):
    i = int(_num(idx))
    if 1 <= i <= len(args):
        return args[i - 1]
    return None

def _xl_large(rng, k):
    if not isinstance(rng, (list, tuple)):
        return rng
    nums = sorted([_num(v) for v in rng if v is not None], reverse=True)
    ki = int(_num(k)) - 1
    return nums[ki] if 0 <= ki < len(nums) else None

def _xl_small(rng, k):
    if not isinstance(rng, (list, tuple)):
        return rng
    nums = sorted([_num(v) for v in rng if v is not None])
    ki = int(_num(k)) - 1
    return nums[ki] if 0 <= ki < len(nums) else None

def _xl_rank(number, rng, order=0):
    if not isinstance(rng, (list, tuple)):
        return 1
    nums = sorted([_num(v) for v in rng if v is not None],
                  reverse=(int(_num(order)) == 0))
    n = _num(number)
    for i, v in enumerate(nums, 1):
        if math.isclose(v, n, rel_tol=1e-9):
            return i
    return None

def _xl_median(*args):
    vals = []
    for a in args:
        if isinstance(a, (list, tuple)):
            vals.extend(_num(v) for v in a if v is not None)
        elif a is not None:
            vals.append(_num(a))
    if not vals:
        return 0
    vals.sort()
    n = len(vals)
    if n % 2:
        return vals[n // 2]
    return (vals[n // 2 - 1] + vals[n // 2]) / 2

def _xl_stdev(*args):
    vals = []
    for a in args:
        if isinstance(a, (list, tuple)):
            vals.extend(_num(v) for v in a if v is not None)
        elif a is not None:
            vals.append(_num(a))
    if len(vals) < 2:
        return 0
    mean = sum(vals) / len(vals)
    return math.sqrt(sum((v - mean)**2 for v in vals) / (len(vals) - 1))

def _xl_var(*args):
    vals = []
    for a in args:
        if isinstance(a, (list, tuple)):
            vals.extend(_num(v) for v in a if v is not None)
        elif a is not None:
            vals.append(_num(a))
    if len(vals) < 2:
        return 0
    mean = sum(vals) / len(vals)
    return sum((v - mean)**2 for v in vals) / (len(vals) - 1)

def _xl_na():
    return float('nan')

def _xl_rounddown(number, num_digits=0):
    n = _num(number)
    d = int(_num(num_digits))
    factor = 10 ** d
    return math.trunc(n * factor) / factor

def _xl_roundup(number, num_digits=0):
    n = _num(number)
    d = int(_num(num_digits))
    factor = 10 ** d
    if n >= 0:
        return math.ceil(n * factor) / factor
    else:
        return math.floor(n * factor) / factor

def _xl_ceiling(number, significance=1):
    n = _num(number)
    s = _num(significance)
    if s == 0:
        return 0
    return math.ceil(n / s) * s

def _xl_floor(number, significance=1):
    n = _num(number)
    s = _num(significance)
    if s == 0:
        return 0
    return math.floor(n / s) * s

def _xl_sign(n):
    n = _num(n)
    return (1 if n > 0 else (-1 if n < 0 else 0))

def _xl_countblank(*args):
    cnt = 0
    for a in args:
        if isinstance(a, (list, tuple)):
            cnt += sum(1 for v in a if v is None or v == "")
        else:
            if a is None or a == "":
                cnt += 1
    return cnt

def _xl_averageif(rng, criteria, avg_range=None):
    if avg_range is None:
        avg_range = rng
    if not isinstance(rng, (list, tuple)):
        rng = [rng]
    if not isinstance(avg_range, (list, tuple)):
        avg_range = [avg_range]
    vals = []
    for i, v in enumerate(rng):
        if _match_criteria(v, criteria) and i < len(avg_range):
            vals.append(_num(avg_range[i]))
    return sum(vals) / len(vals) if vals else 0

def _xl_countifs(*args):
    pairs = list(args)
    if len(pairs) < 2:
        return 0
    rng0 = pairs[0] if isinstance(pairs[0], (list, tuple)) else [pairs[0]]
    cnt = 0
    for i in range(len(rng0)):
        match = True
        for p in range(0, len(pairs) - 1, 2):
            rng_p = pairs[p] if isinstance(pairs[p], (list, tuple)) else [pairs[p]]
            cr = pairs[p + 1]
            if i < len(rng_p) and not _match_criteria(rng_p[i], cr):
                match = False
                break
        if match:
            cnt += 1
    return cnt

def _xl_find(find_text, within_text, start_num=1):
    s = str(within_text or "")
    f = str(find_text or "")
    idx = s.find(f, int(_num(start_num)) - 1)
    return idx + 1 if idx >= 0 else None

def _xl_search(find_text, within_text, start_num=1):
    s = str(within_text or "").lower()
    f = str(find_text or "").lower()
    idx = s.find(f, int(_num(start_num)) - 1)
    return idx + 1 if idx >= 0 else None

def _xl_substitute(text, old_text, new_text, instance_num=None):
    s = str(text or "")
    o = str(old_text or "")
    n = str(new_text or "")
    if instance_num is None:
        return s.replace(o, n)
    cnt = 0
    result = []
    i = 0
    while i < len(s):
        if s[i:i+len(o)] == o:
            cnt += 1
            if cnt == int(_num(instance_num)):
                result.append(n)
            else:
                result.append(o)
            i += len(o)
        else:
            result.append(s[i])
            i += 1
    return "".join(result)

def _xl_replace_func(old_text, start_num, num_chars, new_text):
    s = str(old_text or "")
    st = int(_num(start_num)) - 1
    nc = int(_num(num_chars))
    return s[:st] + str(new_text or "") + s[st+nc:]

def _xl_rept(text, number_times):
    return str(text or "") * int(_num(number_times))

def _xl_exact(text1, text2):
    return str(text1 or "") == str(text2 or "")

def _xl_type(v):
    if isinstance(v, (int, float)):
        return 1
    if isinstance(v, str):
        return 2
    if isinstance(v, bool):
        return 4
    if v is None:
        return 16
    return 1

def _xl_n(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, bool):
        return 1.0 if v else 0.0
    return 0.0

def _xl_t(v):
    return v if isinstance(v, str) else ""

def _xl_isnumber(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)

def _xl_istext(v):
    return isinstance(v, str)

def _xl_isna(v):
    return v is None or (isinstance(v, float) and math.isnan(v))

def _xl_numbervalue(text, decimal_sep=".", group_sep=","):
    try:
        s = str(text or "").replace(str(group_sep), "").replace(str(decimal_sep), ".")
        return float(s)
    except (ValueError, TypeError):
        return None

'''


# ---------------------------------------------------------------------------
# Tarjan's SCC  (iterative, to avoid recursion depth issues on large models)
# ---------------------------------------------------------------------------

def _tarjan_scc(nodes: list[str], adj: dict[str, set[str]]) -> list[list[str]]:
    """Return list of Strongly Connected Components (reverse-topological order).

    Each SCC is a list of node ids.  Trivial SCCs (single node, no self-loop)
    have length 1.
    """
    index_counter = [0]
    stack: list[str] = []
    on_stack: set[str] = set()
    index_map: dict[str, int] = {}
    lowlink: dict[str, int] = {}
    sccs: list[list[str]] = []

    def strongconnect(v: str):
        # Iterative DFS with explicit stack
        call_stack: list[tuple[str, list[str], int]] = []
        call_stack.append((v, list(adj.get(v, [])), 0))
        index_map[v] = lowlink[v] = index_counter[0]
        index_counter[0] += 1
        stack.append(v)
        on_stack.add(v)

        while call_stack:
            node, neighbours, ni = call_stack[-1]
            if ni < len(neighbours):
                call_stack[-1] = (node, neighbours, ni + 1)
                w = neighbours[ni]
                if w not in index_map:
                    index_map[w] = lowlink[w] = index_counter[0]
                    index_counter[0] += 1
                    stack.append(w)
                    on_stack.add(w)
                    call_stack.append((w, list(adj.get(w, [])), 0))
                elif w in on_stack:
                    lowlink[node] = min(lowlink[node], index_map[w])
            else:
                # Done with this node's neighbours
                if lowlink[node] == index_map[node]:
                    scc: list[str] = []
                    while True:
                        w = stack.pop()
                        on_stack.discard(w)
                        scc.append(w)
                        if w == node:
                            break
                    sccs.append(scc)
                call_stack.pop()
                if call_stack:
                    parent = call_stack[-1][0]
                    lowlink[parent] = min(lowlink[parent], lowlink[node])

    for n in nodes:
        if n not in index_map:
            strongconnect(n)

    return sccs


def _generate_code(
    sheets: dict[str, list[dict[str, Any]]],
    sheet_order: list[str],
    defined_names: dict[str, str] | None = None,
) -> str:
    """Generate the full Python script as a string.

    Algorithm overview
    ------------------
    1. Collect formula cells, build per-cell dependency map from *References*.
    2. Build groups (GroupID → member cells).  A group with ≥ 2 members is an
       *emission unit* that will be emitted as a vectorised loop.
    3. Construct a *unit-level* dependency graph (group→group / cell→group / …).
    4. Detect Strongly-Connected Components (SCCs) via Tarjan's algorithm.
       Groups that participate in an inter-unit cycle cannot be emitted as a
       single vectorised block because they interleave with outside cells.
       Such groups are **broken** back into individual cells.
    5. Topologically sort the condensed DAG of SCCs.
    6. Within each SCC, cells are emitted in *cell-level* topological order
       (Kahn's on the fine-grained cell graph restricted to cells of that SCC).
       Non-cyclic groups are emitted as vectorised loops at the point where
       their SCC comes up.
    """
    # Store defined names at module level so formula translators can access them
    global _DEFINED_NAMES
    _DEFINED_NAMES = defined_names or {}

    code_lines: list[str] = [_RUNTIME_HELPERS]

    # ------------------------------------------------------------------
    # 1. Collect formula cells & cell map
    # ------------------------------------------------------------------
    formula_cells: list[dict[str, Any]] = []
    all_cells_map: dict[tuple[str, str], dict] = {}
    for sname, rows in sheets.items():
        for rd in rows:
            rd["_sheet"] = sname
            key = (sname, rd["Cell"])
            all_cells_map[key] = rd
            if rd.get("Type") in ("Calculation", "Output"):
                if rd.get("IncludeFlag") is not False and rd.get("IncludeFlag") != 0:
                    formula_cells.append(rd)

    # Cell-level adjacency  (dep → set[dependant])
    cell_keys = {(fc["_sheet"], fc["Cell"]) for fc in formula_cells}
    cell_adj: dict[tuple, set[tuple]] = defaultdict(set)
    cell_rev: dict[tuple, set[tuple]] = defaultdict(set)   # also needed later
    for fc in formula_cells:
        my = (fc["_sheet"], fc["Cell"])
        refs_str = fc.get("References", "") or ""
        for ref_token in refs_str.split("; "):
            ref_token = ref_token.strip()
            if not ref_token:
                continue
            if "!" in ref_token:
                parts = ref_token.split("!", 1)
                dep = (parts[0], parts[1])
            else:
                dep = (fc["_sheet"], ref_token)
            if dep in cell_keys and dep != my:
                cell_adj[dep].add(my)
                cell_rev[my].add(dep)

    # ------------------------------------------------------------------
    # 2. Build groups
    # ------------------------------------------------------------------
    groups: dict[int, list[dict]] = defaultdict(list)
    for fc in formula_cells:
        gid = fc.get("GroupID", 0) or 0
        if gid > 0:
            groups[gid].append(fc)

    # ------------------------------------------------------------------
    # 3. Build unit-level graph
    # ------------------------------------------------------------------
    cell_to_unit: dict[tuple, str] = {}
    unit_cells: dict[str, list[dict]] = {}

    for gid, members in groups.items():
        if len(members) >= 2:
            uid = f"g{gid}"
            unit_cells[uid] = members
            for m in members:
                cell_to_unit[(m["_sheet"], m["Cell"])] = uid

    for fc in formula_cells:
        key = (fc["_sheet"], fc["Cell"])
        if key not in cell_to_unit:
            uid = f"c{fc['_sheet']}!{fc['Cell']}"
            cell_to_unit[key] = uid
            unit_cells[uid] = [fc]

    unit_adj: dict[str, set[str]] = defaultdict(set)
    for fc in formula_cells:
        my_unit = cell_to_unit[(fc["_sheet"], fc["Cell"])]
        refs_str = fc.get("References", "") or ""
        for ref_token in refs_str.split("; "):
            ref_token = ref_token.strip()
            if not ref_token:
                continue
            if "!" in ref_token:
                parts = ref_token.split("!", 1)
                dep_key = (parts[0], parts[1])
            else:
                dep_key = (fc["_sheet"], ref_token)
            dep_unit = cell_to_unit.get(dep_key)
            if dep_unit and dep_unit != my_unit:
                unit_adj[dep_unit].add(my_unit)

    # ------------------------------------------------------------------
    # 4. Tarjan's SCC on the unit graph
    # ------------------------------------------------------------------
    all_unit_ids = list(unit_cells.keys())
    sccs = _tarjan_scc(all_unit_ids, unit_adj)
    # sccs is list of lists; each list is one SCC (in reverse-topo order).

    # Build unit → scc_id
    unit_to_scc: dict[str, int] = {}
    for i, scc in enumerate(sccs):
        for uid in scc:
            unit_to_scc[uid] = i

    # Groups inside a *non-trivial* SCC (size > 1) are broken into singles
    broken_groups: set[int] = set()
    for scc in sccs:
        if len(scc) <= 1:
            continue
        for uid in scc:
            if uid.startswith("g"):
                gid = int(uid[1:])
                broken_groups.add(gid)

    # ------------------------------------------------------------------
    # 5. Topo-sort the condensed DAG  (SCCs in order)
    # ------------------------------------------------------------------
    scc_adj: dict[int, set[int]] = defaultdict(set)
    for src, dests in unit_adj.items():
        si = unit_to_scc.get(src)
        if si is None:
            continue
        for d in dests:
            di = unit_to_scc.get(d)
            if di is not None and di != si:
                scc_adj[si].add(di)
    scc_in = {i: 0 for i in range(len(sccs))}
    for si, dests in scc_adj.items():
        for di in dests:
            scc_in[di] += 1
    scc_queue = [i for i in range(len(sccs)) if scc_in[i] == 0]
    scc_order: list[int] = []
    while scc_queue:
        n = scc_queue.pop(0)
        scc_order.append(n)
        for nb in sorted(scc_adj.get(n, [])):
            scc_in[nb] -= 1
            if scc_in[nb] == 0:
                scc_queue.append(nb)
    # Append any remaining (shouldn't happen after SCC condensation, but safety)
    seen_scc = set(scc_order)
    for i in range(len(sccs)):
        if i not in seen_scc:
            scc_order.append(i)

    # ------------------------------------------------------------------
    # 6. Within each SCC, cell-level Kahn sort
    # ------------------------------------------------------------------
    # Final emission plan: list of (action, data)
    #   ("group", group_cells_list)
    #   ("cell", single_cell_dict)
    emission_plan: list[tuple[str, Any]] = []

    for si in scc_order:
        scc_members = sccs[si]
        if len(scc_members) == 1:
            uid = scc_members[0]
            if uid.startswith("g") and int(uid[1:]) not in broken_groups:
                members = unit_cells[uid]
                if len(members) >= 2:
                    emission_plan.append(("group", members))
                else:
                    emission_plan.append(("cell", members[0]))
            else:
                # Single cell
                for m in unit_cells[uid]:
                    emission_plan.append(("cell", m))
        else:
            # Non-trivial SCC: cell-level topo sort of all cells in this SCC
            scc_cell_keys = set()
            scc_fc_map: dict[tuple, dict] = {}
            for uid in scc_members:
                for m in unit_cells[uid]:
                    k = (m["_sheet"], m["Cell"])
                    scc_cell_keys.add(k)
                    scc_fc_map[k] = m

            # Kahn's within the SCC
            local_in: dict[tuple, int] = {k: 0 for k in scc_cell_keys}
            local_adj: dict[tuple, list[tuple]] = defaultdict(list)
            for k in scc_cell_keys:
                for dep in cell_rev.get(k, []):
                    if dep in scc_cell_keys:
                        local_adj[dep].append(k)
                        local_in[k] += 1
            lq = [k for k, d in local_in.items() if d == 0]
            local_order: list[tuple] = []
            while lq:
                n = lq.pop(0)
                local_order.append(n)
                for nb in local_adj.get(n, []):
                    local_in[nb] -= 1
                    if local_in[nb] == 0:
                        lq.append(nb)
            # Append remaining (residual cycles at cell level — very rare)
            lo_set = set(local_order)
            for k in scc_cell_keys:
                if k not in lo_set:
                    local_order.append(k)

            # Now check: can we still emit any non-broken groups as
            # vectorised blocks within this SCC?
            # Groups NOT in broken_groups can be emitted when their last
            # member appears in local_order.
            emitted_in_scc: set[tuple] = set()
            group_members_seen: dict[int, int] = defaultdict(int)
            for k in local_order:
                if k in emitted_in_scc:
                    continue
                fc = scc_fc_map[k]
                gid = fc.get("GroupID", 0) or 0
                if gid > 0 and gid not in broken_groups and len(groups[gid]) >= 2:
                    group_members_seen[gid] += 1
                    # Count in-SCC members for this group
                    in_scc_count = sum(
                        1 for m in groups[gid]
                        if (m["_sheet"], m["Cell"]) in scc_cell_keys
                    )
                    if group_members_seen[gid] >= in_scc_count:
                        emission_plan.append(("group", groups[gid]))
                        for m in groups[gid]:
                            emitted_in_scc.add((m["_sheet"], m["Cell"]))
                    # else skip — will be emitted when last member reached
                else:
                    emission_plan.append(("cell", fc))
                    emitted_in_scc.add(k)

    # Build the main() function
    code_lines.append("""
# ======================================================================
# Formatting helpers
# ======================================================================

def _apply_fmt(cell, fmt_dict):
    \"\"\"Apply formatting from the mapping report to a cell.\"\"\"
    cell.number_format = fmt_dict.get('nf', 'General')
    cell.font = Font(
        bold=fmt_dict.get('bold', False),
        italic=fmt_dict.get('italic', False),
        size=fmt_dict.get('size', 11),
    )
    ha = fmt_dict.get('ha') or None
    va = fmt_dict.get('va') or None
    cell.alignment = Alignment(horizontal=ha, vertical=va,
                               wrap_text=fmt_dict.get('wrap', False))
    fc = fmt_dict.get('fill', '')
    if fc and not str(fc).startswith('theme:') and not str(fc).startswith('indexed:'):
        try:
            cell.fill = PatternFill(start_color=str(fc), end_color=str(fc), fill_type='solid')
        except Exception:
            pass

""")

    code_lines.append("def main(input_path: str, output_path: str) -> None:")
    code_lines.append("    global _wb, _cache")
    code_lines.append("    _cache.clear()")
    code_lines.append("    _wb = openpyxl.load_workbook(input_path)")
    code_lines.append("")

    # Ensure all sheets exist in the workbook
    code_lines.append("    # Ensure all sheets exist")
    for sname in sheet_order:
        code_lines.append(f"    if {repr(sname)} not in _wb.sheetnames:")
        code_lines.append(f"        _wb.create_sheet({repr(sname)})")
    code_lines.append("")

    # Emit formula evaluations from the emission plan
    code_lines.append("    # === Evaluate formulas in topological order ===")
    code_lines.append("")

    for action, data in emission_plan:
        if action == "group":
            _emit_group(code_lines, data, sheets)
        else:
            _emit_single_cell(code_lines, data)

    # Emit formatting
    code_lines.append("")
    code_lines.append("    # === Apply formatting ===")
    for sname, rows in sheets.items():
        for rd in rows:
            if rd.get("IncludeFlag") is False or rd.get("IncludeFlag") == 0:
                continue
            nf = rd.get("NumberFormat", "General") or "General"
            bold = bool(rd.get("FontBold", False))
            italic = bool(rd.get("FontItalic", False))
            size = rd.get("FontSize", 11) or 11
            fill = str(rd.get("FillColor", "") or "")
            ha = str(rd.get("HAlignment", "") or "")
            va = str(rd.get("VAlignment", "") or "")
            wrap = bool(rd.get("WrapText", False))
            # Only emit if non-default
            if any([nf != "General", bold, italic, size != 11, fill, ha, va, wrap]):
                fmt_dict = {"nf": nf, "bold": bold, "italic": italic,
                            "size": size, "fill": fill, "ha": ha, "va": va,
                            "wrap": wrap}
                row_i = int(rd["Row"])
                col_i = int(rd["Col"])
                code_lines.append(
                    f"    _apply_fmt(_wb[{repr(sname)}].cell(row={row_i}, column={col_i}), "
                    f"{repr(fmt_dict)})"
                )

    code_lines.append("")
    code_lines.append("    _wb.save(output_path)")
    code_lines.append("    print(f'✓ Output written to {output_path}')")
    code_lines.append("")
    code_lines.append("")
    code_lines.append("if __name__ == '__main__':")
    code_lines.append("    import sys")
    code_lines.append("    inp = sys.argv[1] if len(sys.argv) > 1 else 'unstructured_inputs.xlsx'")
    code_lines.append("    out = sys.argv[2] if len(sys.argv) > 2 else 'output.xlsx'")
    code_lines.append("    main(inp, out)")
    code_lines.append("")

    return "\n".join(code_lines)


def _emit_single_cell(code_lines: list[str], fc: dict) -> None:
    sheet = fc["_sheet"]
    row = int(fc["Row"])
    col = int(fc["Col"])
    formula = fc.get("Formula", "")
    if not formula:
        return
    py_expr = _formula_to_python(formula, sheet, [])
    code_lines.append(f"    # {sheet}!{fc['Cell']}: {formula[:60]}")
    code_lines.append(f"    try:")
    code_lines.append(f"        _s({repr(sheet)}, {row}, {col}, {py_expr})")
    code_lines.append(f"    except TypeError:")
    code_lines.append(f"        _NM[0] = True")
    code_lines.append(f"        try: _s({repr(sheet)}, {row}, {col}, {py_expr})")
    code_lines.append(f"        except Exception: _s({repr(sheet)}, {row}, {col}, None)")
    code_lines.append(f"        finally: _NM[0] = False")
    code_lines.append(f"    except Exception:")
    code_lines.append(f"        _s({repr(sheet)}, {row}, {col}, None)  # formula error")


def _emit_group(code_lines: list[str], group_cells: list[dict], sheets) -> None:
    """Emit a vectorised loop for a formula group."""
    direction = group_cells[0].get("GroupDirection", "")
    sheet = group_cells[0]["_sheet"]
    pattern = group_cells[0].get("PatternFormula", "")

    # Sort by position
    if direction == "row":
        group_cells.sort(key=lambda c: int(c["Col"]))
        row = int(group_cells[0]["Row"])
        col_start = int(group_cells[0]["Col"])
        col_end = int(group_cells[-1]["Col"])
        code_lines.append(f"    # Group [{sheet}] row {row}, cols {col_start}–{col_end}: {direction}-dragged")
        code_lines.append(f"    # Pattern: {pattern[:70]}")
        code_lines.append(f"    for _col in range({col_start}, {col_end + 1}):")
        # Generate the formula for an arbitrary cell in this row
        # Use the first cell's formula as template, but parameterised by _col
        first_formula = group_cells[0].get("Formula", "")
        if first_formula:
            py_expr = _formula_to_python_parametric(
                first_formula, sheet, row, col_start, "row"
            )
            code_lines.append(f"        try:")
            code_lines.append(f"            _s({repr(sheet)}, {row}, _col, {py_expr})")
            code_lines.append(f"        except TypeError:")
            code_lines.append(f"            _NM[0] = True")
            code_lines.append(f"            try: _s({repr(sheet)}, {row}, _col, {py_expr})")
            code_lines.append(f"            except Exception: _s({repr(sheet)}, {row}, _col, None)")
            code_lines.append(f"            finally: _NM[0] = False")
            code_lines.append(f"        except Exception:")
            code_lines.append(f"            _s({repr(sheet)}, {row}, _col, None)")
    elif direction == "col":
        group_cells.sort(key=lambda c: int(c["Row"]))
        col = int(group_cells[0]["Col"])
        row_start = int(group_cells[0]["Row"])
        row_end = int(group_cells[-1]["Row"])
        code_lines.append(f"    # Group [{sheet}] col {col}, rows {row_start}–{row_end}: {direction}-dragged")
        code_lines.append(f"    # Pattern: {pattern[:70]}")
        code_lines.append(f"    for _row in range({row_start}, {row_end + 1}):")
        first_formula = group_cells[0].get("Formula", "")
        if first_formula:
            py_expr = _formula_to_python_parametric(
                first_formula, sheet, row_start, col, "col"
            )
            code_lines.append(f"        try:")
            code_lines.append(f"            _s({repr(sheet)}, _row, {col}, {py_expr})")
            code_lines.append(f"        except TypeError:")
            code_lines.append(f"            _NM[0] = True")
            code_lines.append(f"            try: _s({repr(sheet)}, _row, {col}, {py_expr})")
            code_lines.append(f"            except Exception: _s({repr(sheet)}, _row, {col}, None)")
            code_lines.append(f"            finally: _NM[0] = False")
            code_lines.append(f"        except Exception:")
            code_lines.append(f"            _s({repr(sheet)}, _row, {col}, None)")
    else:
        # Fallback: emit individually
        for fc in group_cells:
            _emit_single_cell(code_lines, fc)

    code_lines.append("")


def _formula_to_python_parametric(
    formula: str, sheet: str, base_row: int, base_col: int,
    direction: str,
) -> str:
    """Translate a formula into a Python expression parametric in _row or _col.

    For a row-dragged group, cell references that have the same row as the
    formula cell use a relative column offset expressed via ``_col``.
    For a col-dragged group, references use ``_row``.
    """
    if not formula or not formula.startswith("="):
        return repr(formula)

    expr = formula[1:]

    # Resolve named ranges before any other translation
    expr = _resolve_named_ranges(expr, sheet)

    # Mask strings
    strings: list[str] = []
    def _mask(m):
        strings.append(m.group())
        return f"__XLSTR{len(strings)-1}__"
    expr = re.sub(r'"[^"]*"', _mask, expr)

    # Replace Excel error constants with None
    expr = re.sub(r'#REF!', 'None', expr)
    expr = re.sub(r'#N/A', 'None', expr)
    expr = re.sub(r'#VALUE!', 'None', expr)
    expr = re.sub(r'#DIV/0!', 'None', expr)
    expr = re.sub(r'#NAME\?', 'None', expr)
    expr = re.sub(r'#NULL!', 'None', expr)
    expr = re.sub(r'#NUM!', 'None', expr)

    # Replace structured table references with None (not supported)
    expr = re.sub(r'[A-Za-z_]\w*\[\[.*?\]\]', 'None', expr)
    expr = re.sub(r'[A-Za-z_]\w*\[#[^\]]*\]', 'None', expr)
    expr = re.sub(r'[A-Za-z_]\w*\[\]', 'None', expr)
    expr = re.sub(r'(?<![\'"])\b[A-Z][A-Za-z_]\w*\[[A-Za-z][^\]]*\]', 'None', expr)

    # Replace range refs first (sheet-qualified)
    def _repl_range_sq(m):
        sh = m.group(1) or m.group(2)
        c1 = m.group(3).replace("$", "")
        c2 = m.group(4).replace("$", "")
        r1, co1 = cell_to_rowcol(c1)
        r2, co2 = cell_to_rowcol(c2)
        r1e = _parametric_rc(r1, base_row, direction, "row")
        r2e = _parametric_rc(r2, base_row, direction, "row")
        co1e = _parametric_rc(co1, base_col, direction, "col")
        co2e = _parametric_rc(co2, base_col, direction, "col")
        return f"_rng({repr(sh)},{r1e},{co1e},{r2e},{co2e})"

    expr = re.sub(
        r"(?:'([^']*)'|([A-Za-z0-9_]+))!"
        r"(\$?[A-Z]{1,3}\$?\d+):(\$?[A-Z]{1,3}\$?\d+)",
        _repl_range_sq, expr
    )

    # Local range refs
    def _repl_range_local(m):
        c1_raw = m.group(1)
        c2_raw = m.group(2)
        c1 = c1_raw.replace("$", "")
        c2 = c2_raw.replace("$", "")
        r1, co1 = cell_to_rowcol(c1)
        r2, co2 = cell_to_rowcol(c2)
        c1_abs = "$" in c1_raw
        c2_abs = "$" in c2_raw
        r1e = _parametric_rc(r1, base_row, direction, "row", _is_abs_row(c1_raw))
        r2e = _parametric_rc(r2, base_row, direction, "row", _is_abs_row(c2_raw))
        co1e = _parametric_rc(co1, base_col, direction, "col", _is_abs_col(c1_raw))
        co2e = _parametric_rc(co2, base_col, direction, "col", _is_abs_col(c2_raw))
        return f"_rng({repr(sheet)},{r1e},{co1e},{r2e},{co2e})"

    expr = re.sub(
        r"(?<![A-Za-z0-9_!])(\$?[A-Z]{1,3}\$?\d+):(\$?[A-Z]{1,3}\$?\d+)",
        _repl_range_local, expr
    )

    # Sheet-qualified cell refs
    def _repl_cell_sq(m):
        sh = m.group(1) or m.group(2)
        ref_raw = m.group(3)
        ref = ref_raw.replace("$", "")
        r, c = cell_to_rowcol(ref)
        re_ = _parametric_rc(r, base_row, direction, "row", _is_abs_row(ref_raw))
        ce_ = _parametric_rc(c, base_col, direction, "col", _is_abs_col(ref_raw))
        return f"_g({repr(sh)},{re_},{ce_})"

    expr = re.sub(
        r"(?:'([^']*)'|([A-Za-z0-9_]+))!"
        r"(\$?[A-Z]{1,3}\$?\d+)"
        r"(?![:\(A-Z0-9])",
        _repl_cell_sq, expr
    )

    # Local cell refs
    def _repl_cell_local(m):
        ref_raw = m.group(0)
        ref = ref_raw.replace("$", "")
        r, c = cell_to_rowcol(ref)
        re_ = _parametric_rc(r, base_row, direction, "row", _is_abs_row(ref_raw))
        ce_ = _parametric_rc(c, base_col, direction, "col", _is_abs_col(ref_raw))
        return f"_g({repr(sheet)},{re_},{ce_})"

    expr = re.sub(
        r"(?<![A-Za-z0-9_!'\"])(\$?[A-Z]{1,3}\$?\d+)(?![\(:A-Z0-9])",
        _repl_cell_local, expr
    )

    # Functions
    for xl_func, py_func in _FUNC_MAP.items():
        expr = re.sub(r'\b' + xl_func + r'\s*\(', py_func + '(', expr, flags=re.IGNORECASE)

    # Replace Excel % postfix operator:  number% or expr)% → /100
    expr = re.sub(r'(?<=[\d\)])%', '/100', expr)

    expr = expr.replace("<>", "!=")
    # = comparison → == (but not <=, >=, !=, ==)
    expr = re.sub(r'(?<![<>!=])=(?!=)', '==', expr)
    expr = expr.replace("^", "**")
    expr = re.sub(r'&', '+', expr)
    expr = re.sub(r'\bTRUE\b', 'True', expr, flags=re.IGNORECASE)
    expr = re.sub(r'\bFALSE\b', 'False', expr, flags=re.IGNORECASE)

    for i, s in enumerate(strings):
        expr = expr.replace(f"__XLSTR{i}__", s)

    # Use _rng2d for INDEX first argument (INDEX needs 2D, not flat)
    expr = expr.replace('_xl_index(_rng(', '_xl_index(_rng2d(')

    # Safety: if untranslated range operator ':' remains, return None
    if re.search(r'\)\s*:', expr) or re.search(r'\d\s*:\s*_xl_', expr):
        return 'None'

    # Fix empty arguments left by table-ref → None replacement (e.g. INDEX(None,,1))
    while ',,' in expr:
        expr = expr.replace(',,', ',None,')

    # Final safety: try to compile expression; return None if syntax is invalid
    try:
        compile(expr, '<expr>', 'eval')
    except SyntaxError:
        return 'None'

    return expr


def _is_abs_col(ref_raw: str) -> bool:
    """Check if the column part of a reference is absolute ($)."""
    return ref_raw.startswith("$")


def _parametric_rc(value: int, base: int, drag_direction: str,
                   component: str, is_absolute: bool = False) -> str:
    """Return a parametric expression for a row or column.

    If ``drag_direction`` matches ``component`` and the reference is relative,
    returns an expression using the loop variable ``_row`` or ``_col``.
    Otherwise returns the literal integer.
    """
    if is_absolute:
        return str(value)
    if drag_direction == "row" and component == "col":
        offset = value - base
        if offset == 0:
            return "_col"
        return f"_col + {offset}" if offset > 0 else f"_col - {-offset}"
    if drag_direction == "col" and component == "row":
        offset = value - base
        if offset == 0:
            return "_row"
        return f"_row + {offset}" if offset > 0 else f"_row - {-offset}"
    return str(value)


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def generate_unstructured_calculate(report_path: str, output_dir: str) -> str:
    """Generate ``unstructured_calculate.py`` from the mapping report."""
    os.makedirs(output_dir, exist_ok=True)
    sheets, sheet_order, defined_names = _read_report(report_path)
    code = _generate_code(sheets, sheet_order, defined_names)
    out_path = os.path.join(output_dir, "unstructured_calculate.py")
    with open(out_path, "w") as f:
        f.write(code)
    print(f"  → {out_path}")
    return out_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Generate unstructured_calculate.py")
    ap.add_argument("report", help="Path to mapping_report.xlsx")
    ap.add_argument("output_dir", help="Output directory")
    args = ap.parse_args()
    generate_unstructured_calculate(args.report, args.output_dir)
    print("✓ Done")


if __name__ == "__main__":
    main()
