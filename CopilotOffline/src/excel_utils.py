"""
excel_utils.py — Shared Excel I/O utilities.

Responsibilities:
  • Convert .xls → .xlsx via LibreOffice
  • Open workbooks (formula view + data-only view)
  • Extract / apply cell formatting
  • Resolve defined names
"""

from __future__ import annotations

import copy
import os
import re
import subprocess
import tempfile
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string


# ---------------------------------------------------------------------------
# .xls → .xlsx conversion
# ---------------------------------------------------------------------------

def ensure_xlsx(path: str, cache_dir: str | None = None) -> str:
    """Return *path* unchanged if it is .xlsx, otherwise convert via LibreOffice.

    Converted files are placed in *cache_dir* (default: a ``_converted``
    subfolder next to the original).  If the converted file already exists
    and is newer than the source, conversion is skipped.
    """
    p = Path(path)
    if p.suffix.lower() == ".xlsx":
        return str(p)

    if cache_dir is None:
        cache_dir = str(p.parent / "_converted")
    os.makedirs(cache_dir, exist_ok=True)

    out_path = Path(cache_dir) / (p.stem + ".xlsx")
    if out_path.exists() and out_path.stat().st_mtime >= p.stat().st_mtime:
        return str(out_path)

    subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "xlsx",
         "--outdir", cache_dir, str(p)],
        capture_output=True, check=True, timeout=120,
    )
    if not out_path.exists():
        raise FileNotFoundError(f"LibreOffice conversion failed for {path}")
    return str(out_path)


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def open_workbook_dual(path: str):
    """Open a workbook twice: once for formulas, once for cached values.

    Returns (wb_formulas, wb_values).
    """
    wb_f = openpyxl.load_workbook(path, data_only=False)
    wb_v = openpyxl.load_workbook(path, data_only=True)
    return wb_f, wb_v


# ---------------------------------------------------------------------------
# Formatting extraction
# ---------------------------------------------------------------------------

def extract_formatting(cell: Cell) -> dict[str, Any]:
    """Return a flat dict of formatting attributes for one cell."""
    if isinstance(cell, MergedCell):
        return _empty_format()
    font = cell.font
    fill = cell.fill
    align = cell.alignment
    return {
        "number_format": cell.number_format or "General",
        "font_bold": bool(font.bold) if font.bold is not None else False,
        "font_italic": bool(font.italic) if font.italic is not None else False,
        "font_size": font.size if font.size else 11,
        "font_color": _color_to_hex(font.color),
        "fill_color": _fill_to_hex(fill),
        "h_alignment": align.horizontal or "",
        "v_alignment": align.vertical or "",
        "wrap_text": bool(align.wrap_text) if align.wrap_text is not None else False,
    }


def _empty_format() -> dict[str, Any]:
    return {
        "number_format": "General",
        "font_bold": False, "font_italic": False,
        "font_size": 11, "font_color": "000000",
        "fill_color": "", "h_alignment": "", "v_alignment": "",
        "wrap_text": False,
    }


def _color_to_hex(color) -> str:
    if color is None:
        return "000000"
    if color.type == "rgb" and color.rgb:
        rgb = str(color.rgb)
        if len(rgb) == 8:
            return rgb[2:]          # strip alpha
        return rgb
    if color.type == "theme":
        return f"theme:{color.theme}"
    if color.type == "indexed" and color.indexed is not None:
        return f"indexed:{color.indexed}"
    return "000000"


def _fill_to_hex(fill) -> str:
    if fill is None or fill.patternType is None:
        return ""
    fg = fill.fgColor
    if fg and fg.type == "rgb" and fg.rgb and str(fg.rgb) != "00000000":
        rgb = str(fg.rgb)
        return rgb[2:] if len(rgb) == 8 else rgb
    if fg and fg.type == "theme":
        return f"theme:{fg.theme}"
    if fg and fg.type == "indexed" and fg.indexed is not None:
        return f"indexed:{fg.indexed}"
    return ""


def apply_formatting(cell: Cell, fmt: dict[str, Any]) -> None:
    """Apply a formatting dict (as returned by extract_formatting) to *cell*."""
    cell.number_format = fmt.get("number_format", "General")
    cell.font = Font(
        bold=fmt.get("font_bold", False),
        italic=fmt.get("font_italic", False),
        size=fmt.get("font_size", 11),
        color=_hex_to_color(fmt.get("font_color", "000000")),
    )
    fill_hex = fmt.get("fill_color", "")
    if fill_hex and not fill_hex.startswith("theme:") and not fill_hex.startswith("indexed:"):
        cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    cell.alignment = Alignment(
        horizontal=fmt.get("h_alignment") or None,
        vertical=fmt.get("v_alignment") or None,
        wrap_text=fmt.get("wrap_text", False),
    )


def _hex_to_color(hex_str: str):
    if not hex_str or hex_str.startswith("theme:") or hex_str.startswith("indexed:"):
        return None
    return hex_str


# ---------------------------------------------------------------------------
# Defined-name resolution
# ---------------------------------------------------------------------------

def resolve_defined_names(wb) -> dict[str, list[tuple[str, int, int]]]:
    """Parse workbook defined names into {name: [(sheet, row, col), …]}."""
    result: dict[str, list[tuple[str, int, int]]] = {}
    if not hasattr(wb, "defined_names"):
        return result
    for dn in wb.defined_names.definedName:
        name = dn.name
        dests = list(dn.destinations)
        cells: list[tuple[str, int, int]] = []
        for sheet, coord in dests:
            # coord may be a range "A1:B5" or single "A1"
            if ":" in coord:
                parts = coord.replace("$", "").split(":")
                r1, c1 = _parse_coord(parts[0])
                r2, c2 = _parse_coord(parts[1])
                for r in range(r1, r2 + 1):
                    for c in range(c1, c2 + 1):
                        cells.append((sheet, r, c))
            else:
                r, c = _parse_coord(coord.replace("$", ""))
                cells.append((sheet, r, c))
        result[name] = cells
    return result


def _parse_coord(coord: str) -> tuple[int, int]:
    m = re.match(r"([A-Z]{1,3})(\d+)", coord)
    if not m:
        return 1, 1
    return int(m.group(2)), column_index_from_string(m.group(1))


# ---------------------------------------------------------------------------
# Misc
# ---------------------------------------------------------------------------

def col_range_str(start_col: int, end_col: int) -> str:
    """Human-readable column range: 'B–F'."""
    return f"{get_column_letter(start_col)}–{get_column_letter(end_col)}"


def safe_sheet_name(name: str) -> str:
    """Sanitise a string for use as an Excel sheet name (max 31 chars)."""
    name = re.sub(r"[\\/*?\[\]:]", "_", name)
    return name[:31]
