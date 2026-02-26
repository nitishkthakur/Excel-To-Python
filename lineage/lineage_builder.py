"""
Lineage Builder — extracts formula lineage from an Excel workbook.

Produces two levels of lineage:

* **Simple lineage** — shows inputs (hardcoded cells), outputs (cells
  that are not referenced by other formulas), and calculations (formula
  cells) at the *sheet/column* level.  Dragged formulas are collapsed
  into a single representative using the smart-formula-sampler
  normalisation so that 10 000 identical rows appear as one pattern.

* **Complex lineage** — preserves every unique formula pattern per
  column, cross-sheet references, external-file references, and the
  row ranges where each pattern applies.

Both lineages are persisted to Excel files via ``write_simple_lineage``
and ``write_complex_lineage``.  The companion module ``lineage_graph``
reads those Excel files and renders them as readable graphs.
"""

import os
import re
import sys
from collections import defaultdict
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Make the repository root importable
# ---------------------------------------------------------------------------
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from mcp_server.smart_formula_sampler import (
    normalise_formula,
    deduplicate_workbook_formulas,
)

# ---------------------------------------------------------------------------
# Reference regex (lightweight — just enough for lineage)
# ---------------------------------------------------------------------------

_CROSS_SHEET_RE = re.compile(
    r"(?:'([^'\[\]]+)'|([A-Za-z_]\w*))!"
    r"\$?([A-Z]{1,3})\$?(\d+)"
)
_EXTERNAL_RE = re.compile(
    r"'?\[([^\]]+)\]([^'!]+)'?!"
    r"\$?([A-Z]{1,3})\$?(\d+)"
)
_LOCAL_CELL_RE = re.compile(
    r"(?<![A-Za-z_])\$?([A-Z]{1,3})\$?(\d+)"
)


def _extract_ref_targets(formula: str, current_sheet: str):
    """Yield ``(sheet, col_letter, row)`` tuples referenced by *formula*."""
    raw = formula[1:] if formula.startswith("=") else formula
    seen = set()

    for m in _EXTERNAL_RE.finditer(raw):
        key = (m.group(1) + "|" + m.group(2), m.group(3), int(m.group(4)))
        if key not in seen:
            seen.add(key)
            yield key

    occupied = set()
    for m in _EXTERNAL_RE.finditer(raw):
        occupied.update(range(m.start(), m.end()))

    for m in _CROSS_SHEET_RE.finditer(raw):
        if any(p in occupied for p in range(m.start(), m.end())):
            continue
        sheet = m.group(1) or m.group(2)
        key = (sheet, m.group(3), int(m.group(4)))
        if key not in seen:
            seen.add(key)
            yield key
        occupied.update(range(m.start(), m.end()))

    for m in _LOCAL_CELL_RE.finditer(raw):
        if any(p in occupied for p in range(m.start(), m.end())):
            continue
        if m.start() > 0 and (raw[m.start() - 1].isalpha() or raw[m.start() - 1] == "_"):
            continue
        key = (current_sheet, m.group(1), int(m.group(2)))
        if key not in seen:
            seen.add(key)
            yield key
        occupied.update(range(m.start(), m.end()))


# ---------------------------------------------------------------------------
# Full cell scan
# ---------------------------------------------------------------------------

def _scan_workbook(path: str):
    """Return (formula_cells, value_cells) dicts keyed by (sheet, col, row)."""
    wb = load_workbook(path, data_only=False)
    formula_cells = {}   # (sheet, col, row) -> formula string
    value_cells = {}     # (sheet, col, row) -> value

    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row is None or ws.max_column is None:
            continue
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                col = get_column_letter(c)
                key = (name, col, r)
                if isinstance(v, str) and v.startswith("="):
                    formula_cells[key] = v
                elif v is not None:
                    value_cells[key] = v
    wb.close()
    return formula_cells, value_cells


# ---------------------------------------------------------------------------
# Simple lineage
# ---------------------------------------------------------------------------

def build_simple_lineage(path: str) -> dict[str, Any]:
    """Build a sheet-level simple lineage.

    Returns a dict with:
      * ``sheets`` — list of per-sheet dicts:
          - ``sheet_name``
          - ``inputs``  — list of column-level input descriptions
          - ``outputs`` — list of column-level output descriptions
          - ``calculations`` — list of unique formula pattern descriptions
      * ``cross_sheet_edges`` — list of (source_sheet, target_sheet) pairs
    """
    formula_cells, value_cells = _scan_workbook(path)
    dedup = deduplicate_workbook_formulas(path)

    # Determine which cells are referenced by at least one formula
    referenced: set[tuple] = set()
    for key, formula in formula_cells.items():
        sheet = key[0]
        for ref in _extract_ref_targets(formula, sheet):
            referenced.add(ref)

    # Build per-sheet summaries
    sheets_out = []
    cross_sheet_edges: set[tuple[str, str]] = set()

    for sheet_info in dedup["sheets"]:
        sname = sheet_info["sheet_name"]

        # Inputs: value cells in this sheet that are referenced by formulas
        input_cols: dict[str, list[int]] = defaultdict(list)
        for (s, col, row), _v in value_cells.items():
            if s == sname and (s, col, row) in referenced:
                input_cols[col].append(row)

        inputs = []
        for col in sorted(input_cols, key=lambda x: (len(x), x)):
            rows = sorted(input_cols[col])
            header = sheet_info["headers"].get(col, "")
            label = f"{sname}!{col}"
            if header:
                label += f" ({header})"
            inputs.append({
                "column": col,
                "header": header,
                "label": label,
                "row_range": [rows[0], rows[-1]] if rows else [],
                "count": len(rows),
            })

        # Calculations: deduplicated formula patterns
        calculations = []
        for col_info in sheet_info["columns"]:
            for pat in col_info["patterns"]:
                calculations.append({
                    "column": col_info["column"],
                    "header": col_info["header"],
                    "pattern": pat["pattern"],
                    "example_formula": pat["example_formula"],
                    "row_range": pat["row_range"],
                    "count": pat["count"],
                })

        # Outputs: formula cells that are NOT referenced by any other formula
        output_cols: dict[str, list[int]] = defaultdict(list)
        for (s, col, row), _f in formula_cells.items():
            if s == sname and (s, col, row) not in referenced:
                output_cols[col].append(row)

        outputs = []
        for col in sorted(output_cols, key=lambda x: (len(x), x)):
            rows = sorted(output_cols[col])
            header = sheet_info["headers"].get(col, "")
            label = f"{sname}!{col}"
            if header:
                label += f" ({header})"
            outputs.append({
                "column": col,
                "header": header,
                "label": label,
                "row_range": [rows[0], rows[-1]] if rows else [],
                "count": len(rows),
            })

        sheets_out.append({
            "sheet_name": sname,
            "inputs": inputs,
            "outputs": outputs,
            "calculations": calculations,
        })

        # Cross-sheet edges
        for (s, col, row), formula in formula_cells.items():
            if s != sname:
                continue
            for ref_s, _rc, _rr in _extract_ref_targets(formula, sname):
                if "|" not in ref_s and ref_s != sname:
                    cross_sheet_edges.add((ref_s, sname))

    return {
        "file": path,
        "sheets": sheets_out,
        "cross_sheet_edges": sorted(cross_sheet_edges),
    }


# ---------------------------------------------------------------------------
# Complex lineage
# ---------------------------------------------------------------------------

def build_complex_lineage(path: str) -> dict[str, Any]:
    """Build a detailed cell-level lineage with full dependency tracking.

    Returns a dict with:
      * ``sheets`` — per-sheet dicts with:
          - ``patterns`` — every unique formula pattern (from smart sampler)
          - ``dependencies`` — for each pattern, the set of referenced
            columns/sheets
      * ``cross_sheet_refs`` — detailed cross-sheet reference list
      * ``external_refs``    — detailed external-file reference list
      * ``dependency_edges`` — list of (source_node, target_node) edges
        at column granularity
    """
    formula_cells, value_cells = _scan_workbook(path)
    dedup = deduplicate_workbook_formulas(path)

    cross_sheet_refs = []
    external_refs = []
    dependency_edges: list[tuple[str, str]] = []

    sheets_out = []
    for sheet_info in dedup["sheets"]:
        sname = sheet_info["sheet_name"]
        patterns_detail = []

        for col_info in sheet_info["columns"]:
            col = col_info["column"]
            header = col_info["header"]

            for pat in col_info["patterns"]:
                # Resolve dependencies from the example formula
                example = pat["example_formula"]
                deps_cols: set[str] = set()  # "Sheet!Col" strings
                deps_detail = []

                for ref_s, ref_c, ref_r in _extract_ref_targets(example, sname):
                    dep_node = f"{ref_s}!{ref_c}"
                    deps_cols.add(dep_node)
                    deps_detail.append({
                        "sheet": ref_s,
                        "column": ref_c,
                        "row": ref_r,
                    })

                    # Track cross-sheet and external
                    if "|" in ref_s:
                        parts = ref_s.split("|", 1)
                        external_refs.append({
                            "source_sheet": sname,
                            "source_column": col,
                            "formula_pattern": pat["pattern"],
                            "external_file": parts[0],
                            "external_sheet": parts[1],
                            "external_column": ref_c,
                        })
                    elif ref_s != sname:
                        cross_sheet_refs.append({
                            "source_sheet": sname,
                            "source_column": col,
                            "formula_pattern": pat["pattern"],
                            "target_sheet": ref_s,
                            "target_column": ref_c,
                        })

                    # Edge for dependency graph
                    source_node = f"{sname}!{col}"
                    if dep_node != source_node:
                        edge = (dep_node, source_node)
                        if edge not in dependency_edges:
                            dependency_edges.append(edge)

                patterns_detail.append({
                    "column": col,
                    "header": header,
                    "pattern": pat["pattern"],
                    "example_cell": pat["example_cell"],
                    "example_formula": pat["example_formula"],
                    "row_range": pat["row_range"],
                    "count": pat["count"],
                    "dependencies": sorted(deps_cols),
                    "dependency_detail": deps_detail,
                })

        sheets_out.append({
            "sheet_name": sname,
            "total_formula_cells": sheet_info["total_formula_cells"],
            "unique_patterns": sheet_info["unique_patterns"],
            "patterns": patterns_detail,
        })

    return {
        "file": path,
        "sheets": sheets_out,
        "cross_sheet_refs": cross_sheet_refs,
        "external_refs": external_refs,
        "dependency_edges": dependency_edges,
    }


# ---------------------------------------------------------------------------
# Excel output — simple lineage
# ---------------------------------------------------------------------------

def write_simple_lineage(lineage: dict[str, Any], output_path: str):
    """Write the simple lineage to an Excel file.

    Creates sheets:
      * **Overview** — one row per sheet with input/output/calc counts
      * **Inputs** — all input columns across all sheets
      * **Calculations** — all unique formula patterns
      * **Outputs** — all output columns
      * **Cross-Sheet Edges** — sheet-to-sheet data flow
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # --- Overview ---
    ws = wb.create_sheet("Overview")
    headers = ["Sheet", "Inputs", "Calculations", "Outputs"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    for ri, s in enumerate(lineage["sheets"], 2):
        ws.cell(row=ri, column=1, value=s["sheet_name"])
        ws.cell(row=ri, column=2, value=len(s["inputs"]))
        ws.cell(row=ri, column=3, value=len(s["calculations"]))
        ws.cell(row=ri, column=4, value=len(s["outputs"]))
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 25

    # --- Inputs ---
    ws = wb.create_sheet("Inputs")
    headers = ["Sheet", "Column", "Header", "Row Range", "Count"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    ri = 2
    for s in lineage["sheets"]:
        for inp in s["inputs"]:
            ws.cell(row=ri, column=1, value=s["sheet_name"])
            ws.cell(row=ri, column=2, value=inp["column"])
            ws.cell(row=ri, column=3, value=inp["header"])
            ws.cell(row=ri, column=4, value=str(inp["row_range"]))
            ws.cell(row=ri, column=5, value=inp["count"])
            ri += 1
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 22

    # --- Calculations ---
    ws = wb.create_sheet("Calculations")
    headers = ["Sheet", "Column", "Header", "Pattern", "Example", "Row Range", "Count"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    ri = 2
    for s in lineage["sheets"]:
        for calc in s["calculations"]:
            ws.cell(row=ri, column=1, value=s["sheet_name"])
            ws.cell(row=ri, column=2, value=calc["column"])
            ws.cell(row=ri, column=3, value=calc["header"])
            ws.cell(row=ri, column=4, value=calc["pattern"])
            ws.cell(row=ri, column=5, value=calc["example_formula"])
            ws.cell(row=ri, column=6, value=str(calc["row_range"]))
            ws.cell(row=ri, column=7, value=calc["count"])
            ri += 1
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 28

    # --- Outputs ---
    ws = wb.create_sheet("Outputs")
    headers = ["Sheet", "Column", "Header", "Row Range", "Count"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    ri = 2
    for s in lineage["sheets"]:
        for out in s["outputs"]:
            ws.cell(row=ri, column=1, value=s["sheet_name"])
            ws.cell(row=ri, column=2, value=out["column"])
            ws.cell(row=ri, column=3, value=out["header"])
            ws.cell(row=ri, column=4, value=str(out["row_range"]))
            ws.cell(row=ri, column=5, value=out["count"])
            ri += 1
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 22

    # --- Cross-Sheet Edges ---
    ws = wb.create_sheet("Cross-Sheet Edges")
    headers = ["Source Sheet", "Target Sheet"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    for ri, (src, tgt) in enumerate(lineage["cross_sheet_edges"], 2):
        ws.cell(row=ri, column=1, value=src)
        ws.cell(row=ri, column=2, value=tgt)
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 30

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Excel output — complex lineage
# ---------------------------------------------------------------------------

def write_complex_lineage(lineage: dict[str, Any], output_path: str):
    """Write the complex lineage to an Excel file.

    Creates sheets:
      * **Summary** — per-sheet formula / pattern counts
      * **All Patterns** — every unique pattern with dependencies
      * **Dependency Edges** — column-to-column edges
      * **Cross-Sheet Refs** — detailed cross-sheet references
      * **External Refs** — detailed external-file references
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # --- Summary ---
    ws = wb.create_sheet("Summary")
    headers = ["Sheet", "Total Formulas", "Unique Patterns"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    for ri, s in enumerate(lineage["sheets"], 2):
        ws.cell(row=ri, column=1, value=s["sheet_name"])
        ws.cell(row=ri, column=2, value=s["total_formula_cells"])
        ws.cell(row=ri, column=3, value=s["unique_patterns"])
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 25

    # --- All Patterns ---
    ws = wb.create_sheet("All Patterns")
    headers = ["Sheet", "Column", "Header", "Pattern", "Example Cell",
               "Example Formula", "Row Range", "Count", "Dependencies"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    ri = 2
    for s in lineage["sheets"]:
        for pat in s["patterns"]:
            ws.cell(row=ri, column=1, value=s["sheet_name"])
            ws.cell(row=ri, column=2, value=pat["column"])
            ws.cell(row=ri, column=3, value=pat["header"])
            ws.cell(row=ri, column=4, value=pat["pattern"])
            ws.cell(row=ri, column=5, value=pat["example_cell"])
            ws.cell(row=ri, column=6, value=pat["example_formula"])
            ws.cell(row=ri, column=7, value=str(pat["row_range"]))
            ws.cell(row=ri, column=8, value=pat["count"])
            ws.cell(row=ri, column=9, value=", ".join(pat["dependencies"]))
            ri += 1
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 30

    # --- Dependency Edges ---
    ws = wb.create_sheet("Dependency Edges")
    headers = ["Source (Sheet!Col)", "Target (Sheet!Col)"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    for ri, (src, tgt) in enumerate(lineage["dependency_edges"], 2):
        ws.cell(row=ri, column=1, value=src)
        ws.cell(row=ri, column=2, value=tgt)
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 35

    # --- Cross-Sheet Refs ---
    ws = wb.create_sheet("Cross-Sheet Refs")
    headers = ["Source Sheet", "Source Column", "Pattern",
               "Target Sheet", "Target Column"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    for ri, ref in enumerate(lineage["cross_sheet_refs"], 2):
        ws.cell(row=ri, column=1, value=ref["source_sheet"])
        ws.cell(row=ri, column=2, value=ref["source_column"])
        ws.cell(row=ri, column=3, value=ref["formula_pattern"])
        ws.cell(row=ri, column=4, value=ref["target_sheet"])
        ws.cell(row=ri, column=5, value=ref["target_column"])
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 28

    # --- External Refs ---
    ws = wb.create_sheet("External Refs")
    headers = ["Source Sheet", "Source Column", "Pattern",
               "External File", "External Sheet", "External Column"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
    for ri, ref in enumerate(lineage["external_refs"], 2):
        ws.cell(row=ri, column=1, value=ref["source_sheet"])
        ws.cell(row=ri, column=2, value=ref["source_column"])
        ws.cell(row=ri, column=3, value=ref["formula_pattern"])
        ws.cell(row=ri, column=4, value=ref["external_file"])
        ws.cell(row=ri, column=5, value=ref["external_sheet"])
        ws.cell(row=ri, column=6, value=ref["external_column"])
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 28

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    """Build simple and complex lineages for an Excel file.

    Usage::

        python -m lineage.lineage_builder <workbook.xlsx> [output_dir]
    """
    if len(sys.argv) < 2:
        print("Usage: python -m lineage.lineage_builder <workbook.xlsx> [output_dir]",
              file=sys.stderr)
        sys.exit(1)

    input_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) >= 3 else "."
    os.makedirs(output_dir, exist_ok=True)

    base = os.path.splitext(os.path.basename(input_path))[0]

    print(f"Building simple lineage for {input_path} ...")
    simple = build_simple_lineage(input_path)
    simple_path = os.path.join(output_dir, f"{base}_simple_lineage.xlsx")
    write_simple_lineage(simple, simple_path)
    print(f"  Written to {simple_path}")

    print(f"Building complex lineage for {input_path} ...")
    cmplx = build_complex_lineage(input_path)
    complex_path = os.path.join(output_dir, f"{base}_complex_lineage.xlsx")
    write_complex_lineage(cmplx, complex_path)
    print(f"  Written to {complex_path}")


if __name__ == "__main__":
    main()
