"""
Lineage Graph — reads lineage Excel files and produces readable graphs.

Uses ``networkx`` for layout computation and ``matplotlib`` for rendering.
Produces PNG images with:

* Colour-coded nodes (inputs=green, calculations=orange, outputs=red).
* Readable text that fits inside each node.
* Edge labels for cross-sheet flows.

Two entry points:

* ``render_simple_graph(excel_path, output_png)`` — reads the simple-
  lineage Excel file produced by ``lineage_builder.write_simple_lineage``.
* ``render_complex_graph(excel_path, output_png)`` — reads the complex-
  lineage Excel file produced by ``lineage_builder.write_complex_lineage``.
"""

import os
import sys
import textwrap
from typing import Any

import matplotlib

matplotlib.use("Agg")  # non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import networkx as nx
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_MAX_LABEL_WIDTH = 22  # characters per line inside a node


def _wrap(text: str, width: int = _MAX_LABEL_WIDTH) -> str:
    """Word-wrap *text* for node labels."""
    return "\n".join(textwrap.wrap(text, width=width))


def _read_sheet_rows(wb, sheet_name: str) -> list[dict[str, Any]]:
    """Read all rows from *sheet_name* as a list of dicts keyed by header."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = []
    for c in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(row=1, column=c).value
        headers.append(str(v) if v is not None else f"col{c}")
    rows = []
    for r in range(2, (ws.max_row or 1) + 1):
        row = {}
        for ci, h in enumerate(headers, 1):
            row[h] = ws.cell(row=r, column=ci).value
        rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# Simple lineage graph
# ---------------------------------------------------------------------------

def render_simple_graph(excel_path: str, output_png: str):
    """Read a simple-lineage Excel file and render a **sheet-level** graph.

    Each sheet becomes a single node whose label shows the counts of
    inputs, calculations, and outputs.  Cross-sheet data-flow edges are
    drawn between sheets.  This keeps the graph compact and readable even
    for workbooks with thousands of formula cells.
    """
    wb = load_workbook(excel_path, data_only=True)

    overview = _read_sheet_rows(wb, "Overview")
    cross_edges = _read_sheet_rows(wb, "Cross-Sheet Edges")

    wb.close()

    G = nx.DiGraph()
    input_nodes: set[str] = set()
    calc_nodes: set[str] = set()
    output_nodes: set[str] = set()

    # Build sheet-level nodes
    target_sheets: set[str] = set()
    source_sheets: set[str] = set()
    for row in cross_edges:
        src = row.get("Source Sheet", "")
        tgt = row.get("Target Sheet", "")
        if src:
            source_sheets.add(src)
        if tgt:
            target_sheets.add(tgt)

    for row in overview:
        sheet = row.get("Sheet", "")
        n_in = row.get("Inputs", 0) or 0
        n_calc = row.get("Calculations", 0) or 0
        n_out = row.get("Outputs", 0) or 0

        if n_in == 0 and n_calc == 0 and n_out == 0:
            continue

        label = f"{sheet}\nIn:{n_in}  Calc:{n_calc}  Out:{n_out}"

        # Classify: pure source → input, pure sink → output, else calc
        is_source = sheet in source_sheets
        is_target = sheet in target_sheets
        if is_source and not is_target:
            kind = "input"
            input_nodes.add(sheet)
        elif is_target and not is_source:
            kind = "output"
            output_nodes.add(sheet)
        else:
            kind = "calc"
            calc_nodes.add(sheet)

        G.add_node(sheet, label=_wrap(label, 28), kind=kind, sheet=sheet)

    # Cross-sheet edges
    for row in cross_edges:
        src = row.get("Source Sheet", "")
        tgt = row.get("Target Sheet", "")
        if src in G and tgt in G:
            G.add_edge(src, tgt)

    _render_graph(G, input_nodes, calc_nodes, output_nodes, output_png,
                  title="Simple Lineage (Sheet Level)")


# ---------------------------------------------------------------------------
# Complex lineage graph
# ---------------------------------------------------------------------------

def render_complex_graph(excel_path: str, output_png: str,
                         max_nodes: int = 80):
    """Read a complex-lineage Excel file and render a graph to *output_png*.

    Nodes represent ``Sheet!Column`` endpoints.  Edges represent formula
    dependencies between columns (possibly across sheets).

    When the graph would exceed *max_nodes*, only cross-sheet edges and
    the most-connected intra-sheet nodes are kept so the graph remains
    readable.
    """
    wb = load_workbook(excel_path, data_only=True)

    edges = _read_sheet_rows(wb, "Dependency Edges")
    patterns = _read_sheet_rows(wb, "All Patterns")

    wb.close()

    G = nx.DiGraph()

    # Collect all unique column nodes from patterns
    node_meta: dict[str, dict] = {}
    for row in patterns:
        sheet = row.get("Sheet", "")
        col = row.get("Column", "")
        header = row.get("Header", "")
        node_id = f"{sheet}!{col}"
        if node_id not in node_meta:
            node_meta[node_id] = {"sheet": sheet, "col": col, "header": header}

    # Determine if a node is purely an input (never a target in edges)
    target_set = set()
    source_set = set()
    edge_tuples = []
    for row in edges:
        src = row.get("Source (Sheet!Col)", "")
        tgt = row.get("Target (Sheet!Col)", "")
        if src and tgt:
            source_set.add(src)
            target_set.add(tgt)
            edge_tuples.append((src, tgt))

    # Prioritise cross-sheet edges if we need to prune
    cross_sheet_edges = []
    intra_sheet_edges = []
    for src, tgt in edge_tuples:
        s_sheet = src.split("!")[0] if "!" in src else ""
        t_sheet = tgt.split("!")[0] if "!" in tgt else ""
        if s_sheet != t_sheet:
            cross_sheet_edges.append((src, tgt))
        else:
            intra_sheet_edges.append((src, tgt))

    # If too many nodes, keep cross-sheet edges and top intra-sheet edges
    all_nodes_in_edges = set()
    for s, t in cross_sheet_edges:
        all_nodes_in_edges.add(s)
        all_nodes_in_edges.add(t)

    if len(all_nodes_in_edges) < max_nodes:
        budget = max_nodes - len(all_nodes_in_edges)
        # Add intra-sheet edge nodes by frequency
        from collections import Counter
        freq = Counter()
        for s, t in intra_sheet_edges:
            freq[s] += 1
            freq[t] += 1
        for node, _cnt in freq.most_common():
            if node not in all_nodes_in_edges:
                all_nodes_in_edges.add(node)
                budget -= 1
                if budget <= 0:
                    break

    # Filter edges to only include kept nodes
    kept_edges = [(s, t) for s, t in edge_tuples
                  if s in all_nodes_in_edges and t in all_nodes_in_edges]

    input_nodes = set()
    calc_nodes = set()
    output_nodes = set()

    kept_targets = {t for _, t in kept_edges}
    kept_sources = {s for s, _ in kept_edges}

    for node_id in all_nodes_in_edges:
        meta = node_meta.get(node_id, {"sheet": "", "col": "", "header": ""})
        label = node_id
        if meta.get("header"):
            label += f"\n({meta['header']})"

        is_target = node_id in kept_targets
        is_source = node_id in kept_sources

        if not is_target and is_source:
            kind = "input"
            input_nodes.add(node_id)
        elif is_target and not is_source:
            kind = "output"
            output_nodes.add(node_id)
        else:
            kind = "calc"
            calc_nodes.add(node_id)

        G.add_node(node_id, label=_wrap(label), kind=kind,
                   sheet=meta.get("sheet", ""))

    for src, tgt in kept_edges:
        G.add_edge(src, tgt)

    _render_graph(G, input_nodes, calc_nodes, output_nodes, output_png,
                  title="Complex Lineage (Column Level)")


# ---------------------------------------------------------------------------
# Common rendering
# ---------------------------------------------------------------------------

def _render_graph(G: nx.DiGraph, input_nodes, calc_nodes, output_nodes,
                  output_png: str, title: str = "Lineage"):
    """Render the graph to *output_png* using matplotlib."""
    if not G.nodes:
        # Empty graph — create a minimal placeholder image
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.text(0.5, 0.5, "No lineage data", ha="center", va="center",
                fontsize=14)
        ax.set_axis_off()
        fig.savefig(output_png, dpi=150, bbox_inches="tight")
        plt.close(fig)
        return

    # --- Colour map ---
    colour_map = []
    for n in G.nodes:
        if n in input_nodes:
            colour_map.append("#90EE90")   # light green
        elif n in output_nodes:
            colour_map.append("#FFA07A")   # salmon
        else:
            colour_map.append("#FFD580")   # light orange

    # --- Layout ---
    # Use multipartite layout if we have clear layers, else spring
    try:
        for n in G.nodes:
            if n in input_nodes:
                G.nodes[n]["subset"] = 0
            elif n in output_nodes:
                G.nodes[n]["subset"] = 2
            else:
                G.nodes[n]["subset"] = 1
        pos = nx.multipartite_layout(G, subset_key="subset", align="horizontal")
    except Exception:
        pos = nx.spring_layout(G, k=2.5, iterations=80, seed=42)

    # --- Figure sizing ---
    n_nodes = len(G.nodes)
    fig_w = max(16, n_nodes * 0.8)
    fig_h = max(10, n_nodes * 0.5)
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.set_title(title, fontsize=16, fontweight="bold", pad=20)

    # --- Draw ---
    labels = {n: G.nodes[n].get("label", n) for n in G.nodes}
    node_size = max(2000, 6000 - n_nodes * 30)
    font_size = max(5, 9 - n_nodes // 40)

    nx.draw_networkx_nodes(G, pos, ax=ax, node_color=colour_map,
                           node_size=node_size, edgecolors="#555555",
                           linewidths=1.2)
    nx.draw_networkx_labels(G, pos, labels=labels, ax=ax,
                            font_size=font_size, font_weight="normal")
    nx.draw_networkx_edges(G, pos, ax=ax, edge_color="#888888",
                           arrows=True, arrowsize=15, width=1.0,
                           connectionstyle="arc3,rad=0.1",
                           min_source_margin=20, min_target_margin=20)

    # Legend
    legend_handles = [
        mpatches.Patch(color="#90EE90", label="Input"),
        mpatches.Patch(color="#FFD580", label="Calculation"),
        mpatches.Patch(color="#FFA07A", label="Output"),
    ]
    ax.legend(handles=legend_handles, loc="upper left", fontsize=10)
    ax.set_axis_off()
    fig.tight_layout()
    fig.savefig(output_png, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"Graph saved to {output_png}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    """Render lineage graphs from Excel files.

    Usage::

        python -m lineage.lineage_graph simple <simple_lineage.xlsx> <output.png>
        python -m lineage.lineage_graph complex <complex_lineage.xlsx> <output.png>
    """
    if len(sys.argv) < 4:
        print("Usage: python -m lineage.lineage_graph <simple|complex> "
              "<lineage.xlsx> <output.png>",
              file=sys.stderr)
        sys.exit(1)

    mode = sys.argv[1]
    excel_path = sys.argv[2]
    output_png = sys.argv[3]

    if mode == "simple":
        render_simple_graph(excel_path, output_png)
    elif mode == "complex":
        render_complex_graph(excel_path, output_png)
    else:
        print(f"Unknown mode: {mode}. Use 'simple' or 'complex'.",
              file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
