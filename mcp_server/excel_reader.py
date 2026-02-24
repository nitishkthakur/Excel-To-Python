"""
Core Excel reading module with formula extraction, data region detection,
unstructured sheet handling, and smart sampling for large files.
"""

import re
from typing import Any
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Workbook / Sheet loading
# ---------------------------------------------------------------------------

def open_workbook(path: str):
    """Open an Excel workbook preserving formulas (data_only=False)."""
    return load_workbook(path, data_only=False)


def open_workbook_values(path: str):
    """Open an Excel workbook reading cached values (data_only=True)."""
    return load_workbook(path, data_only=True)


def sheet_names(path: str) -> list[str]:
    """Return the list of sheet names in a workbook."""
    wb = open_workbook(path)
    names = wb.sheetnames
    wb.close()
    return names


# ---------------------------------------------------------------------------
# Data‑region detection  (handles unstructured / patchy sheets)
# ---------------------------------------------------------------------------

class DataRegion:
    """A rectangular region of contiguous data inside a worksheet."""

    def __init__(self, min_row: int, max_row: int, min_col: int, max_col: int,
                 header_row: int | None = None):
        self.min_row = min_row
        self.max_row = max_row
        self.min_col = min_col
        self.max_col = max_col
        self.header_row = header_row  # row index that holds headers (if detected)

    def row_count(self) -> int:
        return self.max_row - self.min_row + 1

    def col_count(self) -> int:
        return self.max_col - self.min_col + 1

    def __repr__(self):
        return (f"DataRegion(rows={self.min_row}-{self.max_row}, "
                f"cols={self.min_col}-{self.max_col}, header={self.header_row})")


def _row_is_blank(ws, row: int, min_col: int, max_col: int) -> bool:
    for col in range(min_col, max_col + 1):
        if ws.cell(row=row, column=col).value is not None:
            return False
    return True


def detect_regions(ws) -> list[DataRegion]:
    """
    Detect rectangular patches of data in a worksheet.

    Strategy: scan rows top‑to‑bottom.  A contiguous vertical run of
    non‑blank rows (within the used column range) forms a region.  Within
    each region the first row that is *all text* (no numbers / formulas)
    is flagged as the header row.
    """
    if ws.max_row is None or ws.max_column is None:
        return []

    max_row = ws.max_row
    max_col = ws.max_column
    regions: list[DataRegion] = []
    region_start = None

    for r in range(1, max_row + 1):
        blank = _row_is_blank(ws, r, 1, max_col)
        if not blank and region_start is None:
            region_start = r
        elif blank and region_start is not None:
            # Determine the actual column bounds for this region
            cmin, cmax = _col_bounds(ws, region_start, r - 1)
            regions.append(DataRegion(region_start, r - 1, cmin, cmax))
            region_start = None

    # Close last region
    if region_start is not None:
        cmin, cmax = _col_bounds(ws, region_start, max_row)
        regions.append(DataRegion(region_start, max_row, cmin, cmax))

    # Detect header rows inside each region
    for reg in regions:
        _detect_header(ws, reg)

    return regions


def _col_bounds(ws, start_row: int, end_row: int) -> tuple[int, int]:
    """Return (min_col, max_col) with actual data in the row range."""
    cmin = ws.max_column
    cmax = 1
    for r in range(start_row, end_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value is not None:
                cmin = min(cmin, c)
                cmax = max(cmax, c)
    return cmin, cmax


def _detect_header(ws, region: DataRegion):
    """Mark the first all‑text row in a region as its header."""
    for r in range(region.min_row, min(region.min_row + 5, region.max_row + 1)):
        all_text = True
        has_value = False
        for c in range(region.min_col, region.max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                has_value = True
                if isinstance(v, (int, float)):
                    all_text = False
                    break
                if isinstance(v, str) and v.startswith("="):
                    all_text = False
                    break
        if all_text and has_value:
            region.header_row = r
            return


# ---------------------------------------------------------------------------
# Cell extraction helpers
# ---------------------------------------------------------------------------

def _cell_addr(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def _cell_info(ws_formula, ws_value, row: int, col: int) -> dict[str, Any]:
    """Return a dict with address, formula (if any) and cached value."""
    f_cell = ws_formula.cell(row=row, column=col)
    v_cell = ws_value.cell(row=row, column=col)
    info: dict[str, Any] = {"address": _cell_addr(row, col)}

    raw = f_cell.value
    if isinstance(raw, str) and raw.startswith("="):
        info["formula"] = raw
    info["value"] = v_cell.value
    return info


# ---------------------------------------------------------------------------
# Smart sampling  (for very large sheets)
# ---------------------------------------------------------------------------

DEFAULT_SAMPLE_ROWS = 100  # max rows per region in a sample


def sample_row_indices(region: DataRegion, ws_formula,
                       max_rows: int = DEFAULT_SAMPLE_ROWS) -> list[int]:
    """
    Pick representative row indices from a region.

    Priority order:
      1. Header row (always included)
      2. Rows that contain formulas (always included, up to a budget)
      3. First N rows, last M rows, evenly spaced middle rows
    """
    total = region.row_count()
    if total <= max_rows:
        return list(range(region.min_row, region.max_row + 1))

    selected: set[int] = set()

    # 1. Header
    if region.header_row is not None:
        selected.add(region.header_row)

    # 2. Formula rows (scan up to 500 rows to stay fast)
    scan_limit = min(region.max_row, region.min_row + 500)
    for r in range(region.min_row, scan_limit + 1):
        for c in range(region.min_col, region.max_col + 1):
            v = ws_formula.cell(row=r, column=c).value
            if isinstance(v, str) and v.startswith("="):
                selected.add(r)
                break
        if len(selected) >= max_rows // 2:
            break

    # 3. Head / tail / middle
    budget = max_rows - len(selected)
    head_n = budget // 3
    tail_n = budget // 3
    mid_n = budget - head_n - tail_n

    data_start = (region.header_row or region.min_row) + 1
    data_end = region.max_row

    for r in range(data_start, min(data_start + head_n, data_end + 1)):
        selected.add(r)
    for r in range(max(data_end - tail_n + 1, data_start), data_end + 1):
        selected.add(r)
    if mid_n > 0 and data_end > data_start:
        step = max(1, (data_end - data_start) // (mid_n + 1))
        r = data_start + step
        while r < data_end and len(selected) < max_rows:
            selected.add(r)
            r += step

    return sorted(selected)


# ---------------------------------------------------------------------------
# High‑level sheet data extraction
# ---------------------------------------------------------------------------

def extract_sheet_data(path: str, sheet_name: str,
                       max_sample_rows: int = DEFAULT_SAMPLE_ROWS,
                       full: bool = False) -> dict[str, Any]:
    """
    Extract structured data from a single sheet.

    Returns a dict with:
      - sheet_name
      - regions: list of region dicts, each with headers, rows, formulas
      - sampled: whether sampling was applied
      - total_rows / sampled_rows counts
    """
    wb_f = open_workbook(path)
    wb_v = open_workbook_values(path)
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]

    regions = detect_regions(ws_f)
    result_regions: list[dict[str, Any]] = []
    total_rows = 0
    sampled_rows = 0
    was_sampled = False

    for reg in regions:
        total_rows += reg.row_count()

        if full:
            rows_to_read = list(range(reg.min_row, reg.max_row + 1))
        else:
            rows_to_read = sample_row_indices(reg, ws_f, max_sample_rows)

        if len(rows_to_read) < reg.row_count():
            was_sampled = True
        sampled_rows += len(rows_to_read)

        # Extract headers
        headers: list[str] = []
        if reg.header_row is not None:
            for c in range(reg.min_col, reg.max_col + 1):
                v = ws_f.cell(row=reg.header_row, column=c).value
                headers.append(str(v) if v is not None else "")

        # Extract rows
        row_data: list[dict[str, Any]] = []
        formulas: list[dict[str, str]] = []

        for r in rows_to_read:
            if r == reg.header_row:
                continue
            cells: list[Any] = []
            for c in range(reg.min_col, reg.max_col + 1):
                info = _cell_info(ws_f, ws_v, r, c)
                cells.append(info["value"])
                if "formula" in info:
                    formulas.append({
                        "address": info["address"],
                        "formula": info["formula"],
                        "cached_value": info["value"],
                    })
            row_data.append({"row_number": r, "values": cells})

        result_regions.append({
            "region": str(reg),
            "min_row": reg.min_row,
            "max_row": reg.max_row,
            "min_col": reg.min_col,
            "max_col": reg.max_col,
            "headers": headers,
            "rows": row_data,
            "formulas": formulas,
        })

    wb_f.close()
    wb_v.close()

    return {
        "sheet_name": sheet_name,
        "regions": result_regions,
        "sampled": was_sampled,
        "total_rows": total_rows,
        "sampled_rows": sampled_rows,
    }


def extract_formulas(path: str, sheet_name: str) -> list[dict[str, Any]]:
    """Return every formula cell in the sheet with its address and cached value."""
    wb_f = open_workbook(path)
    wb_v = open_workbook_values(path)
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]

    formulas: list[dict[str, Any]] = []
    if ws_f.max_row is None or ws_f.max_column is None:
        wb_f.close()
        wb_v.close()
        return formulas

    for r in range(1, ws_f.max_row + 1):
        for c in range(1, ws_f.max_column + 1):
            v = ws_f.cell(row=r, column=c).value
            if isinstance(v, str) and v.startswith("="):
                formulas.append({
                    "address": _cell_addr(r, c),
                    "formula": v,
                    "cached_value": ws_v.cell(row=r, column=c).value,
                })

    wb_f.close()
    wb_v.close()
    return formulas


def workbook_summary(path: str) -> dict[str, Any]:
    """
    Return a lightweight structural summary of the entire workbook:
    sheet names, row/column counts, detected regions, formula counts.
    """
    wb_f = open_workbook(path)
    wb_v = open_workbook_values(path)
    sheets: list[dict[str, Any]] = []

    for name in wb_f.sheetnames:
        ws_f = wb_f[name]
        ws_v = wb_v[name]
        regions = detect_regions(ws_f)
        formula_count = 0
        for reg in regions:
            for r in range(reg.min_row, reg.max_row + 1):
                for c in range(reg.min_col, reg.max_col + 1):
                    v = ws_f.cell(row=r, column=c).value
                    if isinstance(v, str) and v.startswith("="):
                        formula_count += 1
        sheets.append({
            "name": name,
            "max_row": ws_f.max_row,
            "max_column": ws_f.max_column,
            "regions": [str(r) for r in regions],
            "region_count": len(regions),
            "formula_count": formula_count,
        })

    wb_f.close()
    wb_v.close()
    return {"file": path, "sheets": sheets}
