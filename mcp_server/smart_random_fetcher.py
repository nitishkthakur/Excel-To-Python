"""
Vectorized Smart Random Fetcher — fast sampling using pandas & numpy.

This module reimplements the smart-random sampling idea from
``excel_reader_smart_sampler.py`` using **pandas** and **numpy** vectorized
operations so that region detection, header identification, formula
scanning, and row selection are all performed on in-memory arrays instead
of cell-by-cell iteration through openpyxl.

Key capabilities:
  * Detect contiguous data patches (regions) via vectorized null-mask ops.
  * Identify header rows using dtype / string-heuristic checks.
  * Sample a few data rows right after each header, plus tail rows and
    evenly-spaced middle rows.
  * Capture isolated non-empty rows (single-row regions) as contextual
    snippets.
  * Works for **both values and formulas** — two DataFrames are built
    (one from ``data_only=True``, one from ``data_only=False``) and the
    formula frame is used for formula detection while the value frame
    supplies cached results.

Standalone usage
----------------
::

    python smart_random_fetcher.py path/to/workbook.xlsx [sheet_name]

The script writes a *highlighted* copy of the workbook
(``<name>_highlighted.xlsx``) where sampled cells are filled with a
colour and non-sampled cells are left unhighlighted.
"""

import os
import sys
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

DEFAULT_SAMPLE_ROWS = 100
HEAD_ROWS = 5          # rows right after a header to always include
TAIL_ROWS = 5          # rows at the end of a region to always include
ISOLATED_MAX_ROWS = 3  # regions with <= this many rows are kept whole

HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00",
                             fill_type="solid")


# ---------------------------------------------------------------------------
# DataFrame builders  (vectorized Excel → pandas)
# ---------------------------------------------------------------------------

def _load_sheet_dataframe(path: str, sheet_name: str,
                          data_only: bool) -> pd.DataFrame:
    """Load a worksheet into a pandas DataFrame preserving cell positions.

    Row / column indices are 1-based to match openpyxl conventions.
    """
    wb = load_workbook(path, data_only=data_only)
    ws = wb[sheet_name]
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    if max_row == 0 or max_col == 0:
        wb.close()
        return pd.DataFrame()

    # Build a 2-D Python list, then hand it to pandas in one shot
    data = []
    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            row_vals.append(ws.cell(row=r, column=c).value)
        data.append(row_vals)

    wb.close()

    cols = list(range(1, max_col + 1))
    idx = list(range(1, max_row + 1))
    return pd.DataFrame(data, index=idx, columns=cols)


def load_sheet_frames(path: str, sheet_name: str):
    """Return ``(df_values, df_formulas)`` for the given sheet.

    *df_values*   — cached values  (``data_only=True``)
    *df_formulas* — raw cell content including formula strings
    """
    df_v = _load_sheet_dataframe(path, sheet_name, data_only=True)
    df_f = _load_sheet_dataframe(path, sheet_name, data_only=False)
    return df_v, df_f


# ---------------------------------------------------------------------------
# Vectorized region detection
# ---------------------------------------------------------------------------

def _non_null_mask(df: pd.DataFrame) -> np.ndarray:
    """Return a boolean 2-D numpy array — True where the cell is not None/NaN."""
    # pd.notna handles None, np.nan, and NaT
    return pd.notna(df).values


def detect_regions(df: pd.DataFrame):
    """Detect contiguous rectangular data patches.

    Returns a list of dicts, each with keys:
        min_row, max_row, min_col, max_col, header_row
    (all 1-based, consistent with openpyxl).
    """
    if df.empty:
        return []

    mask = _non_null_mask(df)           # shape (nrows, ncols)
    row_has_data = mask.any(axis=1)     # 1-D bool array, one per row

    regions = []
    in_region = False
    start = 0

    for i, has in enumerate(row_has_data):
        if has and not in_region:
            start = i
            in_region = True
        elif not has and in_region:
            regions.append((start, i - 1))
            in_region = False

    if in_region:
        regions.append((start, len(row_has_data) - 1))

    result = []
    for r_start, r_end in regions:
        # Actual column bounds inside this row-range
        sub_mask = mask[r_start:r_end + 1, :]
        col_has_data = sub_mask.any(axis=0)
        col_indices = np.where(col_has_data)[0]
        if len(col_indices) == 0:
            continue
        min_col = int(col_indices[0]) + 1   # back to 1-based
        max_col = int(col_indices[-1]) + 1

        min_row = int(df.index[r_start])
        max_row = int(df.index[r_end])

        header_row = _detect_header_vectorized(df, min_row, max_row,
                                               min_col, max_col)
        result.append({
            "min_row": min_row,
            "max_row": max_row,
            "min_col": min_col,
            "max_col": max_col,
            "header_row": header_row,
        })

    return result


def _detect_header_vectorized(df: pd.DataFrame, min_row: int, max_row: int,
                              min_col: int, max_col: int) -> int | None:
    """Return the 1-based row index of the first all-text row (header).

    Scans up to 5 rows from the top of the region.  A row qualifies as a
    header if every non-null cell is a non-numeric, non-formula string.
    """
    scan_end = min(min_row + 4, max_row)
    col_slice = list(range(min_col, max_col + 1))

    for r in range(min_row, scan_end + 1):
        row_data = df.loc[r, col_slice]
        non_null = row_data.dropna()
        if len(non_null) == 0:
            continue

        all_text = True
        for v in non_null:
            if isinstance(v, (int, float, np.integer, np.floating)):
                all_text = False
                break
            if isinstance(v, str) and v.startswith("="):
                all_text = False
                break
        if all_text:
            return r

    return None


# ---------------------------------------------------------------------------
# Vectorized formula detection
# ---------------------------------------------------------------------------

def _formula_row_mask(df_formulas: pd.DataFrame, min_row: int, max_row: int,
                      min_col: int, max_col: int) -> np.ndarray:
    """Return a 1-D boolean array (indexed by row position) — True if the
    row contains at least one formula cell.
    """
    rows = list(range(min_row, max_row + 1))
    cols = list(range(min_col, max_col + 1))
    sub = df_formulas.loc[rows, cols]

    def _is_formula(v):
        return isinstance(v, str) and v.startswith("=")

    # Vectorized via applymap / map (works across pandas versions)
    bool_df = sub.map(_is_formula)
    return bool_df.any(axis=1).values  # 1-D array, one per row


# ---------------------------------------------------------------------------
# Smart row selection  (vectorized where possible)
# ---------------------------------------------------------------------------

def sample_row_indices(region: dict, df_formulas: pd.DataFrame,
                       max_rows: int = DEFAULT_SAMPLE_ROWS) -> list[int]:
    """Pick representative 1-based row indices from *region*.

    Priority:
      1. Header row (always).
      2. Rows containing formulas (up to half the budget).
      3. First few data rows after the header (HEAD_ROWS).
      4. Last few data rows (TAIL_ROWS).
      5. Evenly-spaced middle rows to fill remaining budget.
    """
    min_row = region["min_row"]
    max_row = region["max_row"]
    total = max_row - min_row + 1

    if total <= max_rows:
        return list(range(min_row, max_row + 1))

    selected: set[int] = set()

    # 1. Header
    header = region["header_row"]
    if header is not None:
        selected.add(header)

    # 2. Formula rows — vectorized detection
    fm = _formula_row_mask(df_formulas, min_row, max_row,
                           region["min_col"], region["max_col"])
    formula_rows = np.array(range(min_row, max_row + 1))[fm]
    budget_formula = max_rows // 2
    for r in formula_rows[:budget_formula]:
        selected.add(int(r))

    # 3. Head rows (right after header)
    data_start = (header or min_row) + 1
    for r in range(data_start, min(data_start + HEAD_ROWS, max_row + 1)):
        selected.add(r)

    # 4. Tail rows
    tail_start = max(max_row - TAIL_ROWS + 1, data_start)
    for r in range(tail_start, max_row + 1):
        selected.add(r)

    # 5. Evenly-spaced middle
    remaining = max_rows - len(selected)
    if remaining > 0 and max_row > data_start:
        mid_indices = np.linspace(data_start, max_row, remaining + 2,
                                  dtype=int)[1:-1]
        for r in mid_indices:
            selected.add(int(r))

    return sorted(selected)[:max_rows]


# ---------------------------------------------------------------------------
# High-level extraction
# ---------------------------------------------------------------------------

def extract_sheet_data(path: str, sheet_name: str,
                       max_sample_rows: int = DEFAULT_SAMPLE_ROWS) -> dict[str, Any]:
    """Extract structured data from a single sheet using vectorized sampling.

    Returns a dict with the same schema as
    ``excel_reader_smart_sampler.extract_sheet_data`` so it can be consumed
    by the existing formatter functions (to_markdown, to_json, to_xml).
    """
    df_v, df_f = load_sheet_frames(path, sheet_name)
    if df_f.empty:
        return {
            "sheet_name": sheet_name,
            "regions": [],
            "sampled": False,
            "total_rows": 0,
            "sampled_rows": 0,
        }

    regions = detect_regions(df_f)
    result_regions: list[dict[str, Any]] = []
    total_rows = 0
    sampled_rows = 0
    was_sampled = False

    for reg in regions:
        rmin, rmax = reg["min_row"], reg["max_row"]
        cmin, cmax = reg["min_col"], reg["max_col"]
        reg_total = rmax - rmin + 1
        total_rows += reg_total

        # For tiny / isolated regions, keep everything
        if reg_total <= ISOLATED_MAX_ROWS:
            rows_to_read = list(range(rmin, rmax + 1))
        else:
            rows_to_read = sample_row_indices(reg, df_f, max_sample_rows)

        if len(rows_to_read) < reg_total:
            was_sampled = True
        sampled_rows += len(rows_to_read)

        # Headers (vectorized slice)
        headers: list[str] = []
        if reg["header_row"] is not None:
            hdr = reg["header_row"]
            cols = list(range(cmin, cmax + 1))
            hdr_vals = df_f.loc[hdr, cols]
            headers = [str(v) if pd.notna(v) else "" for v in hdr_vals]

        # Row data + formulas
        row_data: list[dict[str, Any]] = []
        formulas: list[dict[str, str]] = []
        cols = list(range(cmin, cmax + 1))

        for r in rows_to_read:
            if r == reg["header_row"]:
                continue
            val_row = df_v.loc[r, cols] if r in df_v.index else pd.Series(
                [None] * len(cols), index=cols)
            frm_row = df_f.loc[r, cols]

            cells: list[Any] = []
            for c in cols:
                v_val = val_row[c] if c in val_row.index else None
                f_val = frm_row[c] if c in frm_row.index else None

                # Normalise numpy / pandas types to plain Python
                if isinstance(v_val, (np.integer,)):
                    v_val = int(v_val)
                elif isinstance(v_val, (np.floating,)):
                    v_val = float(v_val)
                if pd.isna(v_val) if not isinstance(v_val, str) else False:
                    v_val = None

                cells.append(v_val)

                if isinstance(f_val, str) and f_val.startswith("="):
                    formulas.append({
                        "address": f"{get_column_letter(c)}{r}",
                        "formula": f_val,
                        "cached_value": v_val,
                    })

            row_data.append({"row_number": r, "values": cells})

        result_regions.append({
            "region": (f"DataRegion(rows={rmin}-{rmax}, "
                       f"cols={cmin}-{cmax}, header={reg['header_row']})"),
            "min_row": rmin,
            "max_row": rmax,
            "min_col": cmin,
            "max_col": cmax,
            "headers": headers,
            "rows": row_data,
            "formulas": formulas,
        })

    return {
        "sheet_name": sheet_name,
        "regions": result_regions,
        "sampled": was_sampled,
        "total_rows": total_rows,
        "sampled_rows": sampled_rows,
    }


# ---------------------------------------------------------------------------
# Collect all sampled cell coordinates  (for highlighting)
# ---------------------------------------------------------------------------

def sampled_cells(path: str, sheet_name: str,
                  max_sample_rows: int = DEFAULT_SAMPLE_ROWS
                  ) -> set[tuple[int, int]]:
    """Return the set of ``(row, col)`` pairs that the sampler would read.

    Useful for highlighting: every coordinate **in** this set is part of
    the sample; everything else is not.
    """
    df_v, df_f = load_sheet_frames(path, sheet_name)
    if df_f.empty:
        return set()

    regions = detect_regions(df_f)
    cells: set[tuple[int, int]] = set()

    for reg in regions:
        rmin, rmax = reg["min_row"], reg["max_row"]
        cmin, cmax = reg["min_col"], reg["max_col"]
        reg_total = rmax - rmin + 1

        if reg_total <= ISOLATED_MAX_ROWS:
            rows = list(range(rmin, rmax + 1))
        else:
            rows = sample_row_indices(reg, df_f, max_sample_rows)

        for r in rows:
            for c in range(cmin, cmax + 1):
                cells.add((r, c))

    return cells


# ---------------------------------------------------------------------------
# Highlight & save
# ---------------------------------------------------------------------------

def highlight_workbook(path: str, sheet_name: str | None = None,
                       max_sample_rows: int = DEFAULT_SAMPLE_ROWS,
                       output_path: str | None = None) -> str:
    """Create a copy of the workbook with sampled cells highlighted.

    Non-sampled cells are left with their original formatting.  Returns
    the path of the highlighted file.
    """
    wb = load_workbook(path)
    sheets = [sheet_name] if sheet_name else wb.sheetnames

    for sn in sheets:
        ws = wb[sn]
        cells_to_highlight = sampled_cells(path, sn, max_sample_rows)
        for (r, c) in cells_to_highlight:
            ws.cell(row=r, column=c).fill = HIGHLIGHT_FILL

    if output_path is None:
        base, ext = os.path.splitext(path)
        output_path = f"{base}_highlighted{ext}"

    wb.save(output_path)
    wb.close()
    return output_path


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main():
    """Read an Excel file, sample it, and write a highlighted copy.

    Usage::

        python smart_random_fetcher.py <workbook.xlsx> [sheet_name]

    If *sheet_name* is omitted every sheet is processed.  The highlighted
    workbook is saved alongside the original with a ``_highlighted`` suffix.
    """
    if len(sys.argv) < 2:
        print("Usage: python smart_random_fetcher.py <workbook.xlsx> [sheet_name]",
              file=sys.stderr)
        sys.exit(1)

    input_path = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) >= 3 else None

    # Print extraction summary
    wb = load_workbook(input_path, data_only=False)
    target_sheets = [sheet_name] if sheet_name else wb.sheetnames
    wb.close()

    for sn in target_sheets:
        data = extract_sheet_data(input_path, sn)
        print(f"\n{'='*60}")
        print(f"Sheet: {sn}")
        print(f"  Regions detected : {len(data['regions'])}")
        print(f"  Total rows       : {data['total_rows']}")
        print(f"  Sampled rows     : {data['sampled_rows']}")
        print(f"  Sampled          : {data['sampled']}")
        for i, reg in enumerate(data["regions"], 1):
            print(f"  Region {i}: rows {reg['min_row']}-{reg['max_row']}, "
                  f"cols {get_column_letter(reg['min_col'])}-"
                  f"{get_column_letter(reg['max_col'])}")
            if reg["headers"]:
                print(f"    Headers: {reg['headers']}")
            print(f"    Data rows: {len(reg['rows'])}")
            print(f"    Formulas : {len(reg['formulas'])}")

    # Write highlighted copy
    out = highlight_workbook(input_path, sheet_name)
    print(f"\nHighlighted workbook saved to: {out}")


if __name__ == "__main__":
    main()
