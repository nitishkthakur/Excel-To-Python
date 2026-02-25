"""
Smart Formula Sampler — de-duplicates formulas from an Excel workbook.

Given a sheet with 40 000 rows where every row uses the same column
formula (e.g. ``=B{n}*C{n}``), loading all 40 000 formula strings into
an LLM is wasteful.  This module normalises row references inside each
formula, groups identical patterns together, and emits a compact
representation that still conveys:

  * The column headers (so you know what each column represents).
  * Every *unique* formula pattern per column.
  * The row range where each pattern is applied.
  * A representative example with the original (un-normalised) formula.

The output is sufficient to understand all calculations and the entities
they represent without redundancy.

Standalone usage
----------------
::

    python smart_formula_sampler.py path/to/workbook.xlsx [output.json]

The result is written to ``output.json`` (default:
``<workbook>_formulas.json``) and also printed to stdout.
"""

import json
import os
import re
import sys
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def _open_formula_wb(path: str):
    return load_workbook(path, data_only=False)


def _open_value_wb(path: str):
    return load_workbook(path, data_only=True)


# ---------------------------------------------------------------------------
# Formula normalisation
# ---------------------------------------------------------------------------

_ROW_NUM_RE = re.compile(
    r"""
    (?<=[A-Z])          # preceded by a column letter
    (\d+)               # the row number we want to replace
    """,
    re.VERBOSE,
)


def normalise_formula(formula: str, row: int) -> str:
    """Replace the current row number in a formula with ``{R}``.

    Only row numbers that match *row* are replaced so that absolute
    references to other rows (e.g. a header row) are preserved.

    >>> normalise_formula("=B5*C5", 5)
    '=B{R}*C{R}'
    >>> normalise_formula("=B5*$C$1", 5)
    '=B{R}*$C$1'
    """
    row_str = str(row)

    def _replace(m: re.Match) -> str:
        return "{R}" if m.group(1) == row_str else m.group(1)

    return _ROW_NUM_RE.sub(_replace, formula)


# ---------------------------------------------------------------------------
# Per-sheet extraction
# ---------------------------------------------------------------------------

def _cell_addr(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def deduplicate_sheet_formulas(ws_f, ws_v) -> dict[str, Any]:
    """Return a compact, de-duplicated summary of all formulas in *ws_f*.

    Returns a dict with:
      * ``headers`` — column-letter → header text mapping
      * ``columns`` — list of per-column formula summaries, each with:
          - ``column``   — column letter
          - ``header``   — header text (if any)
          - ``patterns`` — list of unique normalised patterns, each with:
              * ``pattern``       — normalised formula, e.g. ``=B{R}*C{R}``
              * ``example_cell``  — one concrete cell address, e.g. ``D5``
              * ``example_formula`` — the original formula at that cell
              * ``row_range``     — ``[first_row, last_row]`` where this
                                     pattern is used
              * ``count``         — how many rows use this pattern
      * ``total_formula_cells``   — total formulas found
      * ``unique_patterns``       — total unique patterns after dedup
    """
    max_row = ws_f.max_row or 0
    max_col = ws_f.max_column or 0

    # Detect a likely header row (first all-text row in the first 5 rows)
    header_row: int | None = None
    for r in range(1, min(6, max_row + 1)):
        all_text = True
        has_value = False
        for c in range(1, max_col + 1):
            v = ws_f.cell(row=r, column=c).value
            if v is not None:
                has_value = True
                if isinstance(v, (int, float)):
                    all_text = False
                    break
                if isinstance(v, str) and v.startswith("="):
                    all_text = False
                    break
        if all_text and has_value:
            header_row = r
            break

    # Build header map
    headers: dict[str, str] = {}
    if header_row is not None:
        for c in range(1, max_col + 1):
            v = ws_f.cell(row=header_row, column=c).value
            if v is not None:
                headers[get_column_letter(c)] = str(v)

    # Scan every cell for formulas, grouping by (column, normalised_pattern)
    # Key: (col_letter, normalised_pattern)
    # Value: {example_cell, example_formula, first_row, last_row, count}
    pattern_map: dict[tuple[str, str], dict[str, Any]] = {}
    total_formula_cells = 0

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            raw = ws_f.cell(row=r, column=c).value
            if not (isinstance(raw, str) and raw.startswith("=")):
                continue
            total_formula_cells += 1
            col_letter = get_column_letter(c)
            norm = normalise_formula(raw, r)
            key = (col_letter, norm)
            if key not in pattern_map:
                pattern_map[key] = {
                    "example_cell": _cell_addr(r, c),
                    "example_formula": raw,
                    "first_row": r,
                    "last_row": r,
                    "count": 1,
                }
            else:
                entry = pattern_map[key]
                entry["last_row"] = max(entry["last_row"], r)
                entry["first_row"] = min(entry["first_row"], r)
                entry["count"] += 1

    # Organise by column
    col_groups: dict[str, list[dict[str, Any]]] = {}
    for (col_letter, norm), info in pattern_map.items():
        col_groups.setdefault(col_letter, []).append({
            "pattern": norm,
            "example_cell": info["example_cell"],
            "example_formula": info["example_formula"],
            "row_range": [info["first_row"], info["last_row"]],
            "count": info["count"],
        })

    # Sort patterns within each column by first_row
    for patterns in col_groups.values():
        patterns.sort(key=lambda p: p["row_range"][0])

    columns: list[dict[str, Any]] = []
    for col_letter in sorted(col_groups, key=lambda x: (len(x), x)):
        columns.append({
            "column": col_letter,
            "header": headers.get(col_letter, ""),
            "patterns": col_groups[col_letter],
        })

    return {
        "headers": headers,
        "columns": columns,
        "total_formula_cells": total_formula_cells,
        "unique_patterns": len(pattern_map),
    }


# ---------------------------------------------------------------------------
# Workbook-level entry point
# ---------------------------------------------------------------------------

def deduplicate_workbook_formulas(path: str) -> dict[str, Any]:
    """Return de-duplicated formula summaries for every sheet in *path*.

    Returns::

        {
            "file": "<path>",
            "sheets": [
                {
                    "sheet_name": "...",
                    "headers": {...},
                    "columns": [...],
                    "total_formula_cells": N,
                    "unique_patterns": M,
                },
                ...
            ]
        }
    """
    wb_f = _open_formula_wb(path)
    wb_v = _open_value_wb(path)
    sheets: list[dict[str, Any]] = []

    for name in wb_f.sheetnames:
        ws_f = wb_f[name]
        ws_v = wb_v[name]
        result = deduplicate_sheet_formulas(ws_f, ws_v)
        result["sheet_name"] = name
        sheets.append(result)

    wb_f.close()
    wb_v.close()
    return {"file": path, "sheets": sheets}


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    """Read an Excel file and write deduplicated formulas to disk.

    Usage::

        python smart_formula_sampler.py <workbook.xlsx> [output.json]
    """
    if len(sys.argv) < 2:
        print("Usage: python smart_formula_sampler.py <workbook.xlsx> [output.json]",
              file=sys.stderr)
        sys.exit(1)

    input_path = sys.argv[1]
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        base = os.path.splitext(os.path.basename(input_path))[0]
        output_path = f"{base}_formulas.json"

    result = deduplicate_workbook_formulas(input_path)
    text = json.dumps(result, indent=2, default=str)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(text)

    print(text)
    print(f"\nOutput written to: {output_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
