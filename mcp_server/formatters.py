"""
Output formatters — convert extracted sheet data to Markdown, JSON, or XML.
"""

import json
import xml.etree.ElementTree as ET
from typing import Any
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Markdown
# ---------------------------------------------------------------------------

def _md_table(headers: list[str], rows: list[dict[str, Any]]) -> str:
    """Render a markdown table from headers and row dicts."""
    if not headers and not rows:
        return "_empty region_\n"

    cols = headers if headers else [
        f"Col{i+1}" for i in range(len(rows[0]["values"])) if rows
    ]
    if not cols:
        return "_empty region_\n"

    lines = ["| " + " | ".join(cols) + " |"]
    lines.append("| " + " | ".join(["---"] * len(cols)) + " |")
    for row in rows:
        vals = [str(v) if v is not None else "" for v in row["values"]]
        # Pad or trim to match header count
        while len(vals) < len(cols):
            vals.append("")
        vals = vals[: len(cols)]
        lines.append("| " + " | ".join(vals) + " |")
    return "\n".join(lines) + "\n"


def _md_formulas(formulas: list[dict[str, str]]) -> str:
    if not formulas:
        return ""
    lines = ["\n**Formulas:**\n"]
    for f in formulas:
        cached = f.get("cached_value", "")
        lines.append(f"- `{f['address']}`: `{f['formula']}`  → {cached}")
    return "\n".join(lines) + "\n"


def to_markdown(data: dict[str, Any]) -> str:
    """Convert extracted sheet data dict to a Markdown string."""
    parts = [f"## Sheet: {data['sheet_name']}\n"]
    if data.get("sampled"):
        parts.append(
            f"_Sampled {data['sampled_rows']} of {data['total_rows']} rows._\n"
        )
    for i, reg in enumerate(data["regions"], 1):
        parts.append(f"### Region {i}  (rows {reg['min_row']}–{reg['max_row']}, "
                      f"cols {get_column_letter(reg['min_col'])}–"
                      f"{get_column_letter(reg['max_col'])})\n")
        parts.append(_md_table(reg["headers"], reg["rows"]))
        parts.append(_md_formulas(reg["formulas"]))
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------

def to_json(data: dict[str, Any], pretty: bool = True) -> str:
    """Convert extracted sheet data dict to a JSON string."""
    indent = 2 if pretty else None
    return json.dumps(data, indent=indent, default=str)


# ---------------------------------------------------------------------------
# XML
# ---------------------------------------------------------------------------

def _add_text(parent: ET.Element, tag: str, text: str):
    el = ET.SubElement(parent, tag)
    el.text = str(text)


def to_xml(data: dict[str, Any]) -> str:
    """Convert extracted sheet data dict to an XML string."""
    root = ET.Element("sheet", name=data["sheet_name"])
    if data.get("sampled"):
        root.set("sampled", "true")
        root.set("total_rows", str(data["total_rows"]))
        root.set("sampled_rows", str(data["sampled_rows"]))

    for reg_data in data["regions"]:
        reg_el = ET.SubElement(root, "region",
                                min_row=str(reg_data["min_row"]),
                                max_row=str(reg_data["max_row"]),
                                min_col=str(reg_data["min_col"]),
                                max_col=str(reg_data["max_col"]))

        if reg_data["headers"]:
            hdr_el = ET.SubElement(reg_el, "headers")
            for h in reg_data["headers"]:
                _add_text(hdr_el, "header", h)

        rows_el = ET.SubElement(reg_el, "rows")
        for row in reg_data["rows"]:
            row_el = ET.SubElement(rows_el, "row", number=str(row["row_number"]))
            for v in row["values"]:
                _add_text(row_el, "cell", "" if v is None else str(v))

        if reg_data["formulas"]:
            formulas_el = ET.SubElement(reg_el, "formulas")
            for f in reg_data["formulas"]:
                f_el = ET.SubElement(formulas_el, "formula",
                                      address=f["address"])
                f_el.text = f["formula"]
                f_el.set("cached_value", str(f.get("cached_value", "")))

    return ET.tostring(root, encoding="unicode")
