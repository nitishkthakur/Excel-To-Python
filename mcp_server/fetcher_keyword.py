"""
Keyword Search Fetcher — finds keywords in an Excel workbook and returns
the full rows and columns where they appear.

Given one or more keywords, this fetcher:

1. Scans every sheet in the workbook for cells containing the keyword(s).
2. Identifies which data patches (regions) contain matches.
3. For each match, returns the **entire row** and **entire column** within
   that patch — so the LLM can see the formulas and values surrounding
   the matched cell.

This is useful for an LLM to search for specific terms (e.g. "Revenue",
"COGS") and gather the formulas in the corresponding rows and columns.

The implementation uses **pandas** and **numpy** vectorized operations
via the shared ``load_sheet_frames`` and ``detect_regions`` helpers from
``fetcher_smart_random``.
"""

from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from fetcher_smart_random import (
    load_sheet_frames,
    detect_regions,
)


# ---------------------------------------------------------------------------
# Keyword matching helpers
# ---------------------------------------------------------------------------

def _cell_matches_keyword(value: Any, keyword: str) -> bool:
    """Return True if *value* contains *keyword* (case-insensitive)."""
    if value is None:
        return False
    return keyword.lower() in str(value).lower()


def _find_matches_in_region(
    df_f: pd.DataFrame,
    df_v: pd.DataFrame,
    region: dict,
    keywords: list[str],
) -> list[dict[str, Any]]:
    """Find all cells in *region* that match any of the *keywords*.

    Returns a list of dicts with keys: row, col, keyword, value, formula.
    """
    rmin, rmax = region["min_row"], region["max_row"]
    cmin, cmax = region["min_col"], region["max_col"]
    rows = list(range(rmin, rmax + 1))
    cols = list(range(cmin, cmax + 1))

    matches: list[dict[str, Any]] = []
    for r in rows:
        for c in cols:
            f_val = df_f.loc[r, c] if (r in df_f.index and c in df_f.columns) else None
            v_val = df_v.loc[r, c] if (r in df_v.index and c in df_v.columns) else None

            for kw in keywords:
                if _cell_matches_keyword(f_val, kw) or _cell_matches_keyword(v_val, kw):
                    formula = None
                    if isinstance(f_val, str) and f_val.startswith("="):
                        formula = f_val

                    display_val = v_val
                    if isinstance(display_val, (np.integer,)):
                        display_val = int(display_val)
                    elif isinstance(display_val, (np.floating,)):
                        display_val = float(display_val)
                    if not isinstance(display_val, str) and pd.isna(display_val):
                        display_val = None

                    matches.append({
                        "row": r,
                        "col": c,
                        "keyword": kw,
                        "address": f"{get_column_letter(c)}{r}",
                        "value": display_val,
                        "formula": formula,
                    })
                    break  # avoid duplicate match on same cell for multiple keywords
    return matches


# ---------------------------------------------------------------------------
# Row / column extraction
# ---------------------------------------------------------------------------

def _extract_full_row(df_v: pd.DataFrame, df_f: pd.DataFrame,
                      row: int, cmin: int, cmax: int) -> dict[str, Any]:
    """Extract one full row from *cmin* to *cmax* with values and formulas."""
    cols = list(range(cmin, cmax + 1))
    values: list[Any] = []
    formulas: list[dict[str, str]] = []

    for c in cols:
        v_val = df_v.loc[row, c] if (row in df_v.index and c in df_v.columns) else None
        f_val = df_f.loc[row, c] if (row in df_f.index and c in df_f.columns) else None

        if isinstance(v_val, (np.integer,)):
            v_val = int(v_val)
        elif isinstance(v_val, (np.floating,)):
            v_val = float(v_val)
        if not isinstance(v_val, str) and pd.isna(v_val):
            v_val = None

        values.append(v_val)

        if isinstance(f_val, str) and f_val.startswith("="):
            formulas.append({
                "address": f"{get_column_letter(c)}{row}",
                "formula": f_val,
                "cached_value": v_val,
            })

    return {"row_number": row, "values": values, "formulas": formulas}


def _extract_full_column(df_v: pd.DataFrame, df_f: pd.DataFrame,
                         col: int, rmin: int, rmax: int) -> dict[str, Any]:
    """Extract one full column from *rmin* to *rmax* with values and formulas."""
    rows = list(range(rmin, rmax + 1))
    values: list[Any] = []
    formulas: list[dict[str, str]] = []

    for r in rows:
        v_val = df_v.loc[r, col] if (r in df_v.index and col in df_v.columns) else None
        f_val = df_f.loc[r, col] if (r in df_f.index and col in df_f.columns) else None

        if isinstance(v_val, (np.integer,)):
            v_val = int(v_val)
        elif isinstance(v_val, (np.floating,)):
            v_val = float(v_val)
        if not isinstance(v_val, str) and pd.isna(v_val):
            v_val = None

        values.append(v_val)

        if isinstance(f_val, str) and f_val.startswith("="):
            formulas.append({
                "address": f"{get_column_letter(col)}{r}",
                "formula": f_val,
                "cached_value": v_val,
            })

    return {
        "column_letter": get_column_letter(col),
        "column_index": col,
        "values": values,
        "formulas": formulas,
    }


# ---------------------------------------------------------------------------
# High-level extraction
# ---------------------------------------------------------------------------

def search_keywords(path: str, keywords: list[str],
                    sheet_name: str | None = None) -> dict[str, Any]:
    """Search for *keywords* across the workbook and return context rows/columns.

    Parameters
    ----------
    path : str
        Path to the ``.xlsx`` file.
    keywords : list[str]
        One or more keywords to search for (case-insensitive substring match).
    sheet_name : str or None
        If provided, search only this sheet.  Otherwise search all sheets.

    Returns
    -------
    dict
        A result dict with:
        - ``keywords`` — the keywords searched for
        - ``matches`` — list of per-sheet match results, each containing:
            - ``sheet_name``
            - ``regions`` — list of region results where matches were found,
              each with:
                - ``region`` — region description string
                - ``matched_cells`` — list of matched cell dicts
                - ``rows`` — full row data for each matched row (deduplicated)
                - ``columns`` — full column data for each matched column (deduplicated)
                - ``headers`` — patch headers (if detected)
    """
    wb = load_workbook(path, data_only=False)
    target_sheets = [sheet_name] if sheet_name else wb.sheetnames
    wb.close()

    all_matches: list[dict[str, Any]] = []

    for sn in target_sheets:
        df_v, df_f = load_sheet_frames(path, sn)
        if df_f.empty:
            continue

        regions = detect_regions(df_f)
        sheet_regions: list[dict[str, Any]] = []

        for reg in regions:
            matches = _find_matches_in_region(df_f, df_v, reg, keywords)
            if not matches:
                continue

            rmin, rmax = reg["min_row"], reg["max_row"]
            cmin, cmax = reg["min_col"], reg["max_col"]

            # Deduplicate matched rows and columns
            matched_row_nums = sorted(set(m["row"] for m in matches))
            matched_col_nums = sorted(set(m["col"] for m in matches))

            # Extract full rows
            rows: list[dict[str, Any]] = []
            for r in matched_row_nums:
                rows.append(_extract_full_row(df_v, df_f, r, cmin, cmax))

            # Extract full columns
            columns: list[dict[str, Any]] = []
            for c in matched_col_nums:
                columns.append(_extract_full_column(df_v, df_f, c, rmin, rmax))

            # Extract headers
            headers: list[str] = []
            if reg["header_row"] is not None:
                hdr_cols = list(range(cmin, cmax + 1))
                hdr_vals = df_f.loc[reg["header_row"], hdr_cols]
                headers = [str(v) if pd.notna(v) else "" for v in hdr_vals]

            sheet_regions.append({
                "region": (f"DataRegion(rows={rmin}-{rmax}, "
                           f"cols={cmin}-{cmax}, header={reg['header_row']})"),
                "min_row": rmin,
                "max_row": rmax,
                "min_col": cmin,
                "max_col": cmax,
                "headers": headers,
                "matched_cells": matches,
                "rows": rows,
                "columns": columns,
            })

        if sheet_regions:
            all_matches.append({
                "sheet_name": sn,
                "regions": sheet_regions,
            })

    return {
        "keywords": keywords,
        "matches": all_matches,
    }
