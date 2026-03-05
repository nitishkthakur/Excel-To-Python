from __future__ import annotations

import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .mapping_io import read_mapping_report


def _json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _values_equal(expected: Any, actual: Any) -> bool:
    if isinstance(expected, float) and isinstance(actual, float):
        return math.isclose(expected, actual, rel_tol=1e-9, abs_tol=1e-9)
    return expected == actual


def verify_mapping_against_original(
    mapping_report_path: Path,
    max_mismatches: int = 200,
) -> list[dict[str, Any]]:
    model = read_mapping_report(mapping_report_path)
    wb = load_workbook(model.normalized_workbook, data_only=False)

    mismatches: list[dict[str, Any]] = []

    for sheet in model.sheet_order:
        if sheet not in wb.sheetnames:
            mismatches.append(
                {
                    "stage": "Layer1",
                    "issue": "missing_sheet",
                    "sheet": sheet,
                }
            )
            if len(mismatches) >= max_mismatches:
                return mismatches
            continue

        ws = wb[sheet]
        mapped_coords = {
            (record.row, record.column) for record in model.cells_by_sheet.get(sheet, [])
        }

        original_coords = set(ws._cells.keys())  # pylint: disable=protected-access
        for row, col in sorted(original_coords - mapped_coords):
            cell = ws.cell(row=row, column=col)
            if (
                cell.value is None
                and cell.comment is None
                and cell.hyperlink is None
                and (not cell.has_style or cell.style_id == 0)
            ):
                continue
            mismatches.append(
                {
                    "stage": "Layer1",
                    "issue": "missing_cell_in_mapping",
                    "sheet": sheet,
                    "cell": cell.coordinate,
                }
            )
            if len(mismatches) >= max_mismatches:
                return mismatches

        for record in model.cells_by_sheet.get(sheet, []):
            cell = ws[record.cell]
            if record.cell_type == "Input":
                if not _values_equal(record.value, cell.value):
                    mismatches.append(
                        {
                            "stage": "Layer1",
                            "issue": "input_value_mismatch",
                            "sheet": sheet,
                            "cell": record.cell,
                            "expected": _json_safe(cell.value),
                            "actual": _json_safe(record.value),
                        }
                    )
            else:
                if cell.data_type != "f":
                    mismatches.append(
                        {
                            "stage": "Layer1",
                            "issue": "formula_type_mismatch",
                            "sheet": sheet,
                            "cell": record.cell,
                            "expected": "formula",
                            "actual": cell.data_type,
                        }
                    )
                elif record.formula and record.formula.startswith("="):
                    if cell.value != record.formula:
                        mismatches.append(
                            {
                                "stage": "Layer1",
                                "issue": "formula_text_mismatch",
                                "sheet": sheet,
                                "cell": record.cell,
                                "expected": _json_safe(cell.value),
                                "actual": _json_safe(record.formula),
                            }
                        )

            if len(mismatches) >= max_mismatches:
                return mismatches

    return mismatches


def verify_unstructured_inputs(
    mapping_report_path: Path,
    unstructured_inputs_path: Path,
    max_mismatches: int = 200,
) -> list[dict[str, Any]]:
    model = read_mapping_report(mapping_report_path)
    wb = load_workbook(unstructured_inputs_path, data_only=False)

    mismatches: list[dict[str, Any]] = []

    for sheet in model.sheet_order:
        if sheet not in wb.sheetnames:
            mismatches.append(
                {
                    "stage": "Layer2a",
                    "issue": "missing_sheet",
                    "sheet": sheet,
                }
            )
            if len(mismatches) >= max_mismatches:
                return mismatches
            continue

        ws = wb[sheet]
        for record in model.cells_by_sheet.get(sheet, []):
            cell = ws[record.cell]
            value = cell.value

            if record.cell_type == "Input" and record.include_flag:
                if not _values_equal(record.value, value):
                    mismatches.append(
                        {
                            "stage": "Layer2a",
                            "issue": "input_value_mismatch",
                            "sheet": sheet,
                            "cell": record.cell,
                            "expected": _json_safe(record.value),
                            "actual": _json_safe(value),
                        }
                    )
            elif record.cell_type != "Input":
                if value is not None:
                    mismatches.append(
                        {
                            "stage": "Layer2a",
                            "issue": "formula_not_removed",
                            "sheet": sheet,
                            "cell": record.cell,
                            "expected": None,
                            "actual": _json_safe(value),
                        }
                    )

            if len(mismatches) >= max_mismatches:
                return mismatches

    return mismatches


def verify_structured_input(
    mapping_report_path: Path,
    structured_input_path: Path,
    max_mismatches: int = 200,
) -> list[dict[str, Any]]:
    model = read_mapping_report(mapping_report_path)
    wb = load_workbook(structured_input_path, data_only=False)

    mismatches: list[dict[str, Any]] = []

    if "Index" not in wb.sheetnames:
        return [{"stage": "Layer2b", "issue": "missing_index_sheet"}]

    ws_index = wb["Index"]
    headers = [cell.value for cell in ws_index[1]]
    try:
        idx_source_sheet = headers.index("SourceSheet")
        idx_source_cell = headers.index("SourceCell")
        idx_input_sheet = headers.index("InputSheet")
        idx_input_cell = headers.index("InputCell")
    except ValueError:
        return [{"stage": "Layer2b", "issue": "invalid_index_headers"}]

    expected_inputs: dict[tuple[str, str], Any] = {}
    for sheet in model.sheet_order:
        for record in model.cells_by_sheet.get(sheet, []):
            if record.cell_type == "Input" and record.include_flag and record.value is not None:
                expected_inputs[(sheet, record.cell)] = record.value

    seen_sources: dict[tuple[str, str], int] = {}

    for row in ws_index.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        source_sheet = row[idx_source_sheet]
        source_cell = row[idx_source_cell]
        input_sheet = row[idx_input_sheet]
        input_cell = row[idx_input_cell]

        if not source_sheet or not source_cell or not input_sheet or not input_cell:
            continue

        source_key = (source_sheet, source_cell)
        seen_sources[source_key] = seen_sources.get(source_key, 0) + 1

        if source_key not in expected_inputs:
            mismatches.append(
                {
                    "stage": "Layer2b",
                    "issue": "unexpected_source_mapping",
                    "source_sheet": source_sheet,
                    "source_cell": source_cell,
                }
            )
            if len(mismatches) >= max_mismatches:
                return mismatches
            continue

        if input_sheet not in wb.sheetnames:
            mismatches.append(
                {
                    "stage": "Layer2b",
                    "issue": "missing_input_sheet",
                    "input_sheet": input_sheet,
                }
            )
            if len(mismatches) >= max_mismatches:
                return mismatches
            continue

        value = wb[input_sheet][input_cell].value
        expected_value = expected_inputs[source_key]
        if not _values_equal(expected_value, value):
            mismatches.append(
                {
                    "stage": "Layer2b",
                    "issue": "structured_value_mismatch",
                    "source_sheet": source_sheet,
                    "source_cell": source_cell,
                    "input_sheet": input_sheet,
                    "input_cell": input_cell,
                    "expected": _json_safe(expected_value),
                    "actual": _json_safe(value),
                }
            )

        if len(mismatches) >= max_mismatches:
            return mismatches

    for source_key in expected_inputs:
        if seen_sources.get(source_key, 0) == 0:
            mismatches.append(
                {
                    "stage": "Layer2b",
                    "issue": "missing_source_mapping",
                    "source_sheet": source_key[0],
                    "source_cell": source_key[1],
                }
            )
        elif seen_sources.get(source_key, 0) > 1:
            mismatches.append(
                {
                    "stage": "Layer2b",
                    "issue": "duplicate_source_mapping",
                    "source_sheet": source_key[0],
                    "source_cell": source_key[1],
                    "count": seen_sources[source_key],
                }
            )

        if len(mismatches) >= max_mismatches:
            return mismatches

    return mismatches


def diagnose_stage(
    mapping_mismatches: list[dict[str, Any]],
    stage2_mismatches: list[dict[str, Any]],
    final_mismatches: list[dict[str, Any]],
    mode: str,
) -> str:
    if not final_mismatches:
        return "clean"
    if mapping_mismatches:
        return "Layer1_Mapping"
    if stage2_mismatches:
        return "Layer2a_Unstructured" if mode == "unstructured" else "Layer2b_Structured"
    return "Layer3_CodegenOrRuntime"
