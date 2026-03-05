from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass
class Mismatch:
    sheet: str
    cell: str
    category: str
    expected: Any
    actual: Any


def _normalize_value(value: Any) -> Any:
    if hasattr(value, "t") and value.__class__.__name__.endswith("Formula"):
        try:
            attrs = tuple(sorted(vars(value).items()))
        except Exception:
            attrs = str(value)
        return (value.__class__.__name__, attrs)
    return value


def _json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if hasattr(value, "t") and value.__class__.__name__.endswith("Formula"):
        try:
            return {
                "type": value.__class__.__name__,
                "attrs": dict(vars(value)),
            }
        except Exception:
            return str(value)
    return str(value)


def _style_signature(cell: Any) -> tuple[Any, ...]:
    return (
        cell.number_format,
        cell.font.bold,
        cell.font.italic,
        cell.font.sz,
        cell.border.left.style,
        cell.border.right.style,
        cell.border.top.style,
        cell.border.bottom.style,
    )


def _values_equal(expected: Any, actual: Any) -> bool:
    if isinstance(expected, float) and isinstance(actual, float):
        return math.isclose(expected, actual, rel_tol=1e-9, abs_tol=1e-9)
    return _normalize_value(expected) == _normalize_value(actual)


def compare_workbooks(
    expected_path: Path,
    actual_path: Path,
    max_mismatches: int = 200,
) -> list[Mismatch]:
    expected_wb = load_workbook(expected_path, data_only=False)
    actual_wb = load_workbook(actual_path, data_only=False)

    mismatches: list[Mismatch] = []

    if expected_wb.sheetnames != actual_wb.sheetnames:
        mismatches.append(
            Mismatch(
                sheet="_Workbook",
                cell="",
                category="sheet_names",
                expected=expected_wb.sheetnames,
                actual=actual_wb.sheetnames,
            )
        )
        return mismatches

    for sheet_name in expected_wb.sheetnames:
        ws_expected = expected_wb[sheet_name]
        ws_actual = actual_wb[sheet_name]

        coords = set(ws_expected._cells.keys()) | set(ws_actual._cells.keys())  # pylint: disable=protected-access
        for row, col in sorted(coords):
                exp_cell = ws_expected.cell(row=row, column=col)
                act_cell = ws_actual.cell(row=row, column=col)

                exp_value = exp_cell.value
                act_value = act_cell.value

                exp_is_formula = (
                    (isinstance(exp_value, str) and exp_value.startswith("="))
                    or (hasattr(exp_value, "t") and exp_value.__class__.__name__.endswith("Formula"))
                )
                act_is_formula = (
                    (isinstance(act_value, str) and act_value.startswith("="))
                    or (hasattr(act_value, "t") and act_value.__class__.__name__.endswith("Formula"))
                )

                coord = exp_cell.coordinate

                if exp_is_formula:
                    if not act_is_formula or _normalize_value(exp_value) != _normalize_value(act_value):
                        mismatches.append(
                            Mismatch(
                                sheet=sheet_name,
                                cell=coord,
                                category="formula",
                                expected=_json_safe(exp_value),
                                actual=_json_safe(act_value),
                            )
                        )
                else:
                    if not _values_equal(exp_value, act_value):
                        mismatches.append(
                            Mismatch(
                                sheet=sheet_name,
                                cell=coord,
                                category="value",
                                expected=_json_safe(exp_value),
                                actual=_json_safe(act_value),
                            )
                        )

                if _style_signature(exp_cell) != _style_signature(act_cell):
                    mismatches.append(
                        Mismatch(
                            sheet=sheet_name,
                            cell=coord,
                            category="style",
                            expected=_style_signature(exp_cell),
                            actual=_style_signature(act_cell),
                        )
                    )

                if len(mismatches) >= max_mismatches:
                    return mismatches

    return mismatches
