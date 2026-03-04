from __future__ import annotations

import re
from collections import defaultdict, deque
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter

from .mapping_io import read_mapping_report
from .types import StructuredMappingRow
from .utils import dedupe_sheet_name, maybe_financial_period


INDEX_COLUMNS = [
    "SourceSheet",
    "SourceCell",
    "InputSheet",
    "InputCell",
    "TableName",
    "EntryType",
    "IsTransposed",
    "PatchBounds",
    "MetricLabel",
    "PeriodLabel",
]

CONFIG_COLUMNS = ["Key", "Value", "SourceSheet", "SourceCell"]


class Patch:
    def __init__(self, sheet: str, coords: set[tuple[int, int]]):
        self.sheet = sheet
        self.coords = coords
        self.min_row = min(r for r, _ in coords)
        self.max_row = max(r for r, _ in coords)
        self.min_col = min(c for _, c in coords)
        self.max_col = max(c for _, c in coords)

    @property
    def height(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def width(self) -> int:
        return self.max_col - self.min_col + 1

    @property
    def bounds(self) -> str:
        return (
            f"{get_column_letter(self.min_col)}{self.min_row}:"
            f"{get_column_letter(self.max_col)}{self.max_row}"
        )


def _connected_components(coords: set[tuple[int, int]]) -> list[set[tuple[int, int]]]:
    components: list[set[tuple[int, int]]] = []
    remaining = set(coords)

    while remaining:
        start = remaining.pop()
        queue = deque([start])
        comp = {start}

        while queue:
            row, col = queue.popleft()
            for next_coord in (
                (row - 1, col),
                (row + 1, col),
                (row, col - 1),
                (row, col + 1),
            ):
                if next_coord in remaining:
                    remaining.remove(next_coord)
                    comp.add(next_coord)
                    queue.append(next_coord)

        components.append(comp)

    return components


def _sanitize_key(value: str) -> str:
    key = re.sub(r"\s+", "_", value.strip())
    key = re.sub(r"[^A-Za-z0-9_]+", "", key)
    return key or "Line"


def _lookup_label(
    lookup: dict[tuple[str, int, int], Any],
    sheet: str,
    row: int,
    col: int,
    fallback: str,
) -> str:
    for c in range(col - 1, 0, -1):
        v = lookup.get((sheet, row, c))
        if v is not None and str(v).strip() != "":
            return str(v)

    for r in range(row - 1, 0, -1):
        v = lookup.get((sheet, r, col))
        if v is not None and str(v).strip() != "":
            return str(v)

    return fallback


def _build_patches(
    sheet: str,
    coords: set[tuple[int, int]],
) -> list[Patch]:
    return [Patch(sheet, comp) for comp in _connected_components(coords)]


def create_structured_input(mapping_report_path: Path, output_path: Path) -> Path:
    model = read_mapping_report(mapping_report_path)

    lookup: dict[tuple[str, int, int], Any] = {}
    input_coords_by_sheet: dict[str, set[tuple[int, int]]] = defaultdict(set)

    for sheet, records in model.cells_by_sheet.items():
        for record in records:
            lookup[(sheet, record.row, record.column)] = record.value
            if record.cell_type == "Input" and record.include_flag and record.value is not None:
                input_coords_by_sheet[sheet].add((record.row, record.column))

    wb = Workbook()
    wb.remove(wb.active)

    ws_index = wb.create_sheet("Index")
    ws_index.append(INDEX_COLUMNS)

    ws_config = wb.create_sheet("Config")
    ws_config.append(CONFIG_COLUMNS)

    existing_sheet_names = {"Index", "Config"}
    index_rows: list[StructuredMappingRow] = []

    key_counts: dict[str, int] = defaultdict(int)
    config_row = 2

    for sheet in model.sheet_order:
        coords = input_coords_by_sheet.get(sheet, set())
        if not coords:
            continue

        patches = _build_patches(sheet, coords)
        table_counter = 1

        for patch in sorted(patches, key=lambda p: (p.min_row, p.min_col)):
            is_small_vector = (
                (patch.height == 1 and patch.width <= 2)
                or (patch.width == 1 and patch.height <= 2)
                or len(patch.coords) == 1
            )

            if is_small_vector:
                for row, col in sorted(patch.coords):
                    source_cell = f"{get_column_letter(col)}{row}"
                    fallback = f"{sheet}_{source_cell}"
                    label = _lookup_label(lookup, sheet, row, col, fallback)
                    key = _sanitize_key(label)

                    key_counts[key] += 1
                    if key_counts[key] > 1:
                        key = f"{key}_{key_counts[key]}"

                    value = lookup.get((sheet, row, col))
                    ws_config.append([key, value, sheet, source_cell])

                    input_cell = f"B{config_row}"
                    config_row += 1
                    index_rows.append(
                        StructuredMappingRow(
                            source_sheet=sheet,
                            source_cell=source_cell,
                            input_sheet="Config",
                            input_cell=input_cell,
                            table_name="Config",
                            entry_type="Config",
                            is_transposed=False,
                            patch_bounds=patch.bounds,
                            metric_label=key,
                            period_label=None,
                        )
                    )
                continue

            table_name = f"{sheet}_Table{table_counter}"
            table_counter += 1
            input_sheet_name = dedupe_sheet_name(existing_sheet_names, table_name)
            existing_sheet_names.add(input_sheet_name)

            ws_table = wb.create_sheet(input_sheet_name)

            row_labels: list[str] = []
            for r in range(patch.min_row, patch.max_row + 1):
                left_val = lookup.get((sheet, r, patch.min_col - 1))
                if left_val is None or str(left_val).strip() == "":
                    row_labels.append(f"Line{len(row_labels) + 1}")
                else:
                    row_labels.append(str(left_val))

            col_labels: list[str] = []
            for c in range(patch.min_col, patch.max_col + 1):
                top_val = lookup.get((sheet, patch.min_row - 1, c))
                if top_val is None or str(top_val).strip() == "":
                    col_labels.append(f"Col{len(col_labels) + 1}")
                else:
                    col_labels.append(str(top_val))

            period_hits = sum(1 for label in col_labels if maybe_financial_period(label))
            is_transposed = period_hits >= max(1, len(col_labels) // 2)

            if not is_transposed:
                ws_table.append(["Metric", *col_labels])
                for row_offset, src_row in enumerate(range(patch.min_row, patch.max_row + 1), start=0):
                    row_values = []
                    metric_label = row_labels[row_offset]
                    for col_offset, src_col in enumerate(
                        range(patch.min_col, patch.max_col + 1), start=0
                    ):
                        row_values.append(lookup.get((sheet, src_row, src_col)))
                        source_cell = f"{get_column_letter(src_col)}{src_row}"
                        input_cell = f"{get_column_letter(2 + col_offset)}{2 + row_offset}"
                        index_rows.append(
                            StructuredMappingRow(
                                source_sheet=sheet,
                                source_cell=source_cell,
                                input_sheet=input_sheet_name,
                                input_cell=input_cell,
                                table_name=input_sheet_name,
                                entry_type="Table",
                                is_transposed=False,
                                patch_bounds=patch.bounds,
                                metric_label=metric_label,
                                period_label=col_labels[col_offset],
                            )
                        )

                    ws_table.append([metric_label, *row_values])
            else:
                ws_table.append(["Period", *row_labels])
                for col_offset, src_col in enumerate(range(patch.min_col, patch.max_col + 1), start=0):
                    col_values = []
                    period_label = col_labels[col_offset]
                    for row_offset, src_row in enumerate(
                        range(patch.min_row, patch.max_row + 1), start=0
                    ):
                        col_values.append(lookup.get((sheet, src_row, src_col)))
                        source_cell = f"{get_column_letter(src_col)}{src_row}"
                        input_cell = f"{get_column_letter(2 + row_offset)}{2 + col_offset}"
                        index_rows.append(
                            StructuredMappingRow(
                                source_sheet=sheet,
                                source_cell=source_cell,
                                input_sheet=input_sheet_name,
                                input_cell=input_cell,
                                table_name=input_sheet_name,
                                entry_type="Table",
                                is_transposed=True,
                                patch_bounds=patch.bounds,
                                metric_label=row_labels[row_offset],
                                period_label=period_label,
                            )
                        )

                    ws_table.append([period_label, *col_values])

    for row in index_rows:
        ws_index.append(
            [
                row.source_sheet,
                row.source_cell,
                row.input_sheet,
                row.input_cell,
                row.table_name,
                row.entry_type,
                row.is_transposed,
                row.patch_bounds,
                row.metric_label,
                row.period_label,
            ]
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
