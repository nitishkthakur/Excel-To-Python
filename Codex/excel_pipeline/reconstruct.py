from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

from .types import MappingModel
from .utils import deserialize_style


def _apply_sheet_layout(ws: Any, layout: Any) -> None:
    for merged in layout.merged_ranges:
        ws.merge_cells(merged)

    if layout.freeze_panes:
        ws.freeze_panes = layout.freeze_panes

    if layout.tab_color and ws.sheet_properties is not None:
        ws.sheet_properties.tabColor = layout.tab_color

    for row_dim in layout.row_dimensions:
        row_idx = int(row_dim.get("row"))
        ws.row_dimensions[row_idx].height = row_dim.get("height")
        ws.row_dimensions[row_idx].hidden = bool(row_dim.get("hidden", False))
        ws.row_dimensions[row_idx].outlineLevel = int(row_dim.get("outline_level", 0))

    for col_dim in layout.column_dimensions:
        col_idx = str(col_dim.get("column"))
        ws.column_dimensions[col_idx].width = col_dim.get("width")
        ws.column_dimensions[col_idx].hidden = bool(col_dim.get("hidden", False))
        ws.column_dimensions[col_idx].outlineLevel = int(col_dim.get("outline_level", 0))


def create_workbook_from_mapping(
    model: MappingModel,
    value_overrides: dict[tuple[str, str], Any] | None = None,
    include_formulas: bool = True,
    unstructured_input_mode: bool = False,
) -> Workbook:
    value_overrides = value_overrides or {}
    use_template = model.normalized_workbook.exists()

    if use_template:
        wb = load_workbook(model.normalized_workbook, data_only=False)
    else:
        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name in model.sheet_order:
            ws = wb.create_sheet(sheet_name)
            layout = model.layouts.get(sheet_name)
            if layout:
                _apply_sheet_layout(ws, layout)

    for sheet_name in model.sheet_order:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        records = sorted(
            model.cells_by_sheet.get(sheet_name, []),
            key=lambda r: (r.row, r.column),
        )

        for record in records:
            cell = ws[record.cell]
            if isinstance(cell, MergedCell):
                continue

            if record.cell_type == "Input":
                if unstructured_input_mode:
                    value = record.value if record.include_flag else None
                else:
                    if record.include_flag:
                        value = value_overrides.get((sheet_name, record.cell), record.value)
                    else:
                        value = None
            else:
                if not include_formulas:
                    value = None
                elif record.formula and record.formula.startswith("="):
                    value = record.formula
                elif use_template:
                    # Keep special formula objects (e.g. DataTableFormula) from template.
                    continue
                else:
                    value = None

            cell.value = value

            # Only rehydrate style objects when no template workbook is available.
            if not use_template:
                style = deserialize_style(record.style_json)
                if style:
                    cell.number_format = style["number_format"]
                    cell.font = style["font"]
                    cell.fill = style["fill"]
                    cell.alignment = style["alignment"]
                    cell.border = style["border"]
                    cell.protection = style["protection"]

        if not use_template:
            layout = model.layouts.get(sheet_name)
            if layout:
                _apply_sheet_layout(ws, layout)

    return wb


def save_workbook_from_mapping(
    model: MappingModel,
    output_path: Path,
    value_overrides: dict[tuple[str, str], Any] | None = None,
    include_formulas: bool = True,
    unstructured_input_mode: bool = False,
) -> Path:
    wb = create_workbook_from_mapping(
        model=model,
        value_overrides=value_overrides,
        include_formulas=include_formulas,
        unstructured_input_mode=unstructured_input_mode,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
