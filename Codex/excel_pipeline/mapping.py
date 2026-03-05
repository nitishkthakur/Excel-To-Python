from __future__ import annotations

import json
from collections import defaultdict
from dataclasses import asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import get_column_letter, range_boundaries

from .normalize import normalize_workbook
from .types import CellRecord, MappingModel, SheetLayout
from .utils import (
    color_to_hex,
    excel_bool,
    is_ref_token,
    parse_cell_ref,
    sheet_ref_token,
    split_sheet_ref,
    to_r1c1_token,
)


MAPPING_COLUMNS = [
    "Sheet",
    "Cell",
    "Row",
    "Column",
    "Type",
    "Formula",
    "Value",
    "ValueJSON",
    "NumberFormat",
    "FontBold",
    "FontItalic",
    "FontSize",
    "FontColor",
    "FillColor",
    "HorizontalAlignment",
    "VerticalAlignment",
    "WrapText",
    "IncludeFlag",
    "GroupID",
    "GroupDirection",
    "GroupSize",
    "IsDragged",
    "GroupRange",
    "DragCount",
    "DragSummary",
    "PatternFormula",
    "StyleJSON",
]

METADATA_COLUMNS = ["Section", "Sheet", "Key", "Value"]


def _should_capture_cell(cell: Any) -> bool:
    if cell.value is not None:
        return True
    if cell.comment is not None or cell.hyperlink is not None:
        return True
    if cell.has_style and cell.style_id != 0:
        return True
    return False


def _extract_sheet_layout(ws: Any, index: int) -> SheetLayout:
    # Layout dimensions can be very large; keep metadata compact because
    # reconstruction uses the template workbook path from _Metadata.
    row_dimensions: list[dict[str, Any]] = []
    col_dimensions: list[dict[str, Any]] = []

    tab_color = None
    if ws.sheet_properties and ws.sheet_properties.tabColor is not None:
        tab_color = color_to_hex(ws.sheet_properties.tabColor)

    freeze_panes = str(ws.freeze_panes) if ws.freeze_panes is not None else None

    return SheetLayout(
        title=ws.title,
        index=index,
        merged_ranges=[str(x) for x in ws.merged_cells.ranges],
        freeze_panes=freeze_panes,
        tab_color=tab_color,
        row_dimensions=row_dimensions,
        column_dimensions=col_dimensions,
    )


def _token_to_bounds(ref: str) -> tuple[int | None, int | None, int | None, int | None]:
    ref = ref.replace("$", "")
    if ":" not in ref and ref.isdigit():
        row = int(ref)
        return (None, row, None, row)

    if ":" in ref:
        if ref.replace(":", "").isdigit():
            start, end = ref.split(":", 1)
            return (None, int(start), None, int(end))

        left, right = ref.split(":", 1)
        if left.isalpha() and right.isalpha():
            return (
                parse_cell_ref(f"{left}1")[1],
                None,
                parse_cell_ref(f"{right}1")[1],
                None,
            )

    min_col, min_row, max_col, max_row = range_boundaries(ref)
    return (min_col, min_row, max_col, max_row)


def _is_in_bounds(
    row: int,
    col: int,
    min_col: int | None,
    min_row: int | None,
    max_col: int | None,
    max_row: int | None,
) -> bool:
    if min_col is not None and col < min_col:
        return False
    if max_col is not None and col > max_col:
        return False
    if min_row is not None and row < min_row:
        return False
    if max_row is not None and row > max_row:
        return False
    return True


def _canonicalize_range_token(token: str, sheet: str, row: int, col: int) -> str:
    token = token.strip()
    ref_sheet, ref = split_sheet_ref(token)

    if not is_ref_token(ref):
        return token

    same_sheet = ref_sheet is None or ref_sheet == sheet

    def convert_part(part: str) -> str:
        if part.isdigit() or part.replace("$", "").isdigit():
            return part
        if part.replace("$", "").isalpha():
            return part
        return to_r1c1_token(part, row, col)

    if ":" in ref:
        left, right = ref.split(":", 1)
        if left.replace("$", "").isalpha() and right.replace("$", "").isalpha():
            normalized = f"{left}:{right}"
        elif left.replace("$", "").isdigit() and right.replace("$", "").isdigit():
            normalized = f"{left}:{right}"
        else:
            normalized = f"{convert_part(left)}:{convert_part(right)}"
    else:
        normalized = convert_part(ref)

    if same_sheet:
        return normalized
    return f"{sheet_ref_token(ref_sheet or sheet)}!{normalized}"


def canonical_formula(formula: str, sheet: str, row: int, col: int) -> str:
    tokenizer = Tokenizer(formula)
    parts = ["="]
    for item in tokenizer.items:
        if item.type == "OPERAND" and item.subtype == "RANGE":
            parts.append(_canonicalize_range_token(item.value, sheet, row, col))
        else:
            parts.append(item.value)
    return "".join(parts)


def _extract_formula_references(formula: str, current_sheet: str) -> list[tuple[str, str]]:
    if not formula.startswith("="):
        return []

    refs: list[tuple[str, str]] = []
    tokenizer = Tokenizer(formula)

    for item in tokenizer.items:
        if item.type != "OPERAND" or item.subtype != "RANGE":
            continue

        token = item.value.strip()
        ref_sheet, ref = split_sheet_ref(token)

        if not is_ref_token(ref):
            continue

        refs.append((ref_sheet or current_sheet, ref))

    return refs


def _assign_formula_groups(records_by_sheet: dict[str, list[CellRecord]]) -> None:
    record_lookup: dict[tuple[str, int, int], CellRecord] = {}
    rows: list[dict[str, Any]] = []

    for sheet, records in records_by_sheet.items():
        for record in records:
            record_lookup[(sheet, record.row, record.column)] = record
            if record.formula and record.formula.startswith("="):
                pattern = canonical_formula(record.formula, sheet, record.row, record.column)
                record.pattern_formula = pattern
                rows.append(
                    {
                        "sheet": sheet,
                        "row": record.row,
                        "col": record.column,
                        "pattern": pattern,
                    }
                )

    if not rows:
        return

    df = pd.DataFrame(rows)

    group_counter = 1

    for (sheet, pattern), group_df in df.groupby(["sheet", "pattern"], sort=False):
        working = group_df.copy()
        working["assigned"] = False

        # Vectorized horizontal run detection.
        horizontal = working.sort_values(["row", "col"]).copy()
        horizontal["break"] = (
            (horizontal["row"].diff().fillna(0) != 0)
            | (horizontal["col"].diff().fillna(0) != 1)
        )
        horizontal["run_id"] = horizontal["break"].cumsum()

        for _, run in horizontal.groupby("run_id", sort=False):
            if len(run) <= 1:
                continue
            coords = {(int(r), int(c)) for r, c in run[["row", "col"]].to_numpy()}
            group_id = f"G{group_counter:06d}"
            group_counter += 1
            for row, col in coords:
                record = record_lookup[(sheet, row, col)]
                record.group_id = group_id
                record.group_direction = "Horizontal"
                record.group_size = len(coords)
                record.pattern_formula = pattern
                working.loc[(working["row"] == row) & (working["col"] == col), "assigned"] = True

        # Vectorized vertical run detection on unassigned cells.
        unassigned = working[~working["assigned"]].copy()
        if not unassigned.empty:
            vertical = unassigned.sort_values(["col", "row"]).copy()
            vertical["break"] = (
                (vertical["col"].diff().fillna(0) != 0)
                | (vertical["row"].diff().fillna(0) != 1)
            )
            vertical["run_id"] = vertical["break"].cumsum()

            for _, run in vertical.groupby("run_id", sort=False):
                if len(run) <= 1:
                    continue
                coords = {(int(r), int(c)) for r, c in run[["row", "col"]].to_numpy()}
                group_id = f"G{group_counter:06d}"
                group_counter += 1
                for row, col in coords:
                    record = record_lookup[(sheet, row, col)]
                    record.group_id = group_id
                    record.group_direction = "Vertical"
                    record.group_size = len(coords)
                    record.pattern_formula = pattern
                    working.loc[(working["row"] == row) & (working["col"] == col), "assigned"] = True

        # Remaining fully rectangular ranges become block groups.
        unassigned = working[~working["assigned"]].copy()
        if len(unassigned) > 1:
            unique_rows = sorted(unassigned["row"].astype(int).unique().tolist())
            unique_cols = sorted(unassigned["col"].astype(int).unique().tolist())
            expected = len(unique_rows) * len(unique_cols)
            if len(unassigned) == expected:
                coords = {
                    (int(r), int(c)) for r, c in unassigned[["row", "col"]].to_numpy()
                }
                group_id = f"G{group_counter:06d}"
                group_counter += 1
                for row, col in coords:
                    record = record_lookup[(sheet, row, col)]
                    record.group_id = group_id
                    record.group_direction = "Block"
                    record.group_size = len(coords)
                    record.pattern_formula = pattern


def _formula_group_details(records_by_sheet: dict[str, list[CellRecord]]) -> dict[str, dict[str, Any]]:
    groups: dict[str, list[CellRecord]] = defaultdict(list)
    for records in records_by_sheet.values():
        for record in records:
            if record.group_id:
                groups[record.group_id].append(record)

    details: dict[str, dict[str, Any]] = {}
    for group_id, records in sorted(groups.items()):
        rows_sorted = sorted(records, key=lambda r: (r.row, r.column))
        anchor = rows_sorted[0]
        sheet = anchor.sheet
        direction = anchor.group_direction
        size = anchor.group_size
        pattern = anchor.pattern_formula
        row_nums = [r.row for r in rows_sorted]
        col_nums = [r.column for r in rows_sorted]
        min_row, max_row = min(row_nums), max(row_nums)
        min_col, max_col = min(col_nums), max(col_nums)
        cell_range = (
            f"{sheet}!{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        )

        anchor_formula = anchor.formula
        if isinstance(anchor_formula, str) and anchor_formula.startswith("="):
            anchor_formula_display = f"'{anchor_formula}"
        else:
            anchor_formula_display = anchor_formula

        if isinstance(pattern, str) and pattern.startswith("="):
            pattern_display = f"'{pattern}"
        else:
            pattern_display = pattern

        details[group_id] = {
            "GroupID": group_id,
            "Sheet": sheet,
            "Direction": direction,
            "GroupSize": size,
            "DragCount": size,
            "CellRange": cell_range,
            "AnchorCell": anchor.cell,
            "AnchorFormula": anchor_formula_display,
            "PatternFormula": pattern_display,
            "VectorizationHint": "Vectorizable",
            "DragSummary": (
                f"Dragged formula {size} times across {cell_range} "
                f"(anchor {anchor.cell}: {anchor_formula_display})"
            ),
        }

    return details


def _formula_group_rows(records_by_sheet: dict[str, list[CellRecord]]) -> list[dict[str, Any]]:
    details = _formula_group_details(records_by_sheet)
    return [details[group_id] for group_id in sorted(details)]


def _classify_formula_cells(records_by_sheet: dict[str, list[CellRecord]]) -> None:
    formula_cells_by_sheet: dict[str, set[tuple[int, int]]] = defaultdict(set)
    record_lookup: dict[tuple[str, int, int], CellRecord] = {}

    for sheet, records in records_by_sheet.items():
        for record in records:
            record_lookup[(sheet, record.row, record.column)] = record
            if record.formula:
                formula_cells_by_sheet[sheet].add((record.row, record.column))

    referenced_formula_cells: set[tuple[str, int, int]] = set()

    for sheet, records in records_by_sheet.items():
        for record in records:
            if not record.formula:
                continue

            for ref_sheet, ref in _extract_formula_references(record.formula, sheet):
                target_formula_cells = formula_cells_by_sheet.get(ref_sheet)
                if not target_formula_cells:
                    continue

                min_col, min_row, max_col, max_row = _token_to_bounds(ref)
                for target_row, target_col in target_formula_cells:
                    if (
                        ref_sheet == sheet
                        and target_row == record.row
                        and target_col == record.column
                    ):
                        continue

                    if _is_in_bounds(
                        target_row,
                        target_col,
                        min_col,
                        min_row,
                        max_col,
                        max_row,
                    ):
                        referenced_formula_cells.add((ref_sheet, target_row, target_col))

    for sheet, records in records_by_sheet.items():
        for record in records:
            if not record.formula:
                record.cell_type = "Input"
            elif (sheet, record.row, record.column) in referenced_formula_cells:
                record.cell_type = "Calculation"
            else:
                record.cell_type = "Output"


def build_mapping_model(
    source_workbook: Path,
    cache_dir: Path,
) -> MappingModel:
    normalized_path = normalize_workbook(source_workbook, cache_dir)
    wb_formula = load_workbook(normalized_path, data_only=False)
    wb_values = load_workbook(normalized_path, data_only=True)

    sheet_order = wb_formula.sheetnames.copy()
    layouts: dict[str, SheetLayout] = {}
    records_by_sheet: dict[str, list[CellRecord]] = defaultdict(list)

    for idx, sheet_name in enumerate(sheet_order):
        ws_formula = wb_formula[sheet_name]
        ws_values = wb_values[sheet_name]
        layouts[sheet_name] = _extract_sheet_layout(ws_formula, idx)

        # Sparse iteration is much faster than scanning the full max_row/max_col grid.
        populated_cells = sorted(
            ws_formula._cells.values(),  # pylint: disable=protected-access
            key=lambda c: (c.row, c.column),
        )

        for cell in populated_cells:
            if not _should_capture_cell(cell):
                continue

            value_cell = ws_values[cell.coordinate]
            is_formula = cell.data_type == "f" or (
                isinstance(cell.value, str) and cell.value.startswith("=")
            )

            if is_formula:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
                else:
                    formula = "__SPECIAL_FORMULA__"
            else:
                formula = None
            value = value_cell.value if is_formula else cell.value
            include_flag = value is not None

            records_by_sheet[sheet_name].append(
                CellRecord(
                    sheet=sheet_name,
                    row=cell.row,
                    column=cell.column,
                    cell=cell.coordinate,
                    cell_type="Input",
                    formula=formula,
                    value=value,
                    include_flag=include_flag,
                    number_format=cell.number_format,
                    font_bold=excel_bool(cell.font.bold),
                    font_italic=excel_bool(cell.font.italic),
                    font_size=cell.font.sz,
                    font_color=color_to_hex(cell.font.color),
                    fill_color=color_to_hex(cell.fill.fgColor),
                    horizontal_alignment=cell.alignment.horizontal,
                    vertical_alignment=cell.alignment.vertical,
                    wrap_text=excel_bool(cell.alignment.wrap_text),
                    style_json=None,
                    value_json="",
                )
            )

    _classify_formula_cells(records_by_sheet)
    _assign_formula_groups(records_by_sheet)

    metadata = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "source_ext": source_workbook.suffix.lower(),
    }

    return MappingModel(
        source_workbook=source_workbook.resolve(),
        normalized_workbook=normalized_path.resolve(),
        sheet_order=sheet_order,
        layouts=layouts,
        cells_by_sheet=records_by_sheet,
        metadata=metadata,
    )


def write_mapping_report(model: MappingModel, mapping_report_path: Path) -> Path:
    mapping_report_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    group_details = _formula_group_details(model.cells_by_sheet)

    for sheet_name in model.sheet_order:
        ws = wb.create_sheet(sheet_name)

        records = sorted(
            model.cells_by_sheet.get(sheet_name, []),
            key=lambda x: (x.row, x.column),
        )

        frame_rows = []
        for record in records:
            formula_display = record.formula
            if isinstance(formula_display, str) and formula_display.startswith("="):
                formula_display = f"'{formula_display}"

            pattern_display = record.pattern_formula
            if isinstance(pattern_display, str) and pattern_display.startswith("="):
                pattern_display = f"'{pattern_display}"

            detail = group_details.get(record.group_id or "")
            drag_summary = ""
            if detail and detail.get("AnchorCell") == record.cell and detail.get("Sheet") == record.sheet:
                drag_summary = detail.get("DragSummary", "")

            frame_rows.append(
                {
                    "Sheet": record.sheet,
                    "Cell": record.cell,
                    "Row": record.row,
                    "Column": record.column,
                    "Type": record.cell_type,
                    "Formula": formula_display,
                    "Value": record.value,
                    "ValueJSON": record.value_json,
                    "NumberFormat": record.number_format,
                    "FontBold": record.font_bold,
                    "FontItalic": record.font_italic,
                    "FontSize": record.font_size,
                    "FontColor": record.font_color,
                    "FillColor": record.fill_color,
                    "HorizontalAlignment": record.horizontal_alignment,
                    "VerticalAlignment": record.vertical_alignment,
                    "WrapText": record.wrap_text,
                    "IncludeFlag": record.include_flag,
                    "GroupID": record.group_id,
                    "GroupDirection": record.group_direction,
                    "GroupSize": record.group_size,
                    "IsDragged": bool(record.group_id),
                    "GroupRange": detail.get("CellRange") if detail else "",
                    "DragCount": detail.get("DragCount") if detail else None,
                    "DragSummary": drag_summary,
                    "PatternFormula": pattern_display,
                    "StyleJSON": record.style_json,
                }
            )

        df = pd.DataFrame(frame_rows, columns=MAPPING_COLUMNS)
        for row in dataframe_to_rows(df, index=False, header=True):
            if not any(item is not None for item in row):
                continue
            ws.append(row)

    ws_meta = wb.create_sheet("_Metadata")
    ws_meta.append(METADATA_COLUMNS)

    ws_meta.append(["Workbook", "", "SourceWorkbook", str(model.source_workbook)])
    ws_meta.append(["Workbook", "", "NormalizedWorkbook", str(model.normalized_workbook)])
    ws_meta.append(["Workbook", "", "SheetOrder", json.dumps(model.sheet_order)])

    for key, value in model.metadata.items():
        ws_meta.append(["Workbook", "", key, json.dumps(value)])

    for sheet_name in model.sheet_order:
        layout = model.layouts[sheet_name]
        ws_meta.append(["SheetLayout", sheet_name, "Index", str(layout.index)])
        ws_meta.append(["SheetLayout", sheet_name, "FreezePanes", layout.freeze_panes or ""])
        ws_meta.append(["SheetLayout", sheet_name, "TabColor", layout.tab_color or ""])
        ws_meta.append(
            [
                "SheetLayout",
                sheet_name,
                "MergedRanges",
                json.dumps(layout.merged_ranges),
            ]
        )
        ws_meta.append(
            [
                "SheetLayout",
                sheet_name,
                "RowDimensions",
                json.dumps(layout.row_dimensions),
            ]
        )
        ws_meta.append(
            [
                "SheetLayout",
                sheet_name,
                "ColumnDimensions",
                json.dumps(layout.column_dimensions),
            ]
        )

    formula_group_rows = _formula_group_rows(model.cells_by_sheet)
    ws_groups = wb.create_sheet("_FormulaGroups")
    ws_groups.append(
        [
            "GroupID",
            "Sheet",
            "Direction",
            "GroupSize",
            "DragCount",
            "CellRange",
            "AnchorCell",
            "AnchorFormula",
            "PatternFormula",
            "VectorizationHint",
        ]
    )
    for group_row in formula_group_rows:
        ws_groups.append(
            [
                group_row["GroupID"],
                group_row["Sheet"],
                group_row["Direction"],
                group_row["GroupSize"],
                group_row["DragCount"],
                group_row["CellRange"],
                group_row["AnchorCell"],
                group_row["AnchorFormula"],
                group_row["PatternFormula"],
                group_row["VectorizationHint"],
            ]
        )

    wb.save(mapping_report_path)
    return mapping_report_path


def create_mapping_report(
    source_workbook: Path,
    mapping_report_path: Path,
    cache_dir: Path,
) -> Path:
    model = build_mapping_model(source_workbook=source_workbook, cache_dir=cache_dir)
    return write_mapping_report(model, mapping_report_path)
