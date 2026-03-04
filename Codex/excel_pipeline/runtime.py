from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .mapping_io import read_mapping_report
from .reconstruct import save_workbook_from_mapping


def _extract_unstructured_overrides(mapping_model: Any, input_path: Path) -> dict[tuple[str, str], Any]:
    wb = load_workbook(input_path, data_only=True)
    overrides: dict[tuple[str, str], Any] = {}

    for sheet_name in mapping_model.sheet_order:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for record in mapping_model.cells_by_sheet.get(sheet_name, []):
            if record.cell_type != "Input" or not record.include_flag:
                continue
            overrides[(sheet_name, record.cell)] = ws[record.cell].value

    return overrides


def _extract_structured_overrides(structured_input_path: Path) -> dict[tuple[str, str], Any]:
    wb = load_workbook(structured_input_path, data_only=True)
    if "Index" not in wb.sheetnames:
        raise ValueError("structured_input.xlsx must include Index sheet")

    ws_index = wb["Index"]
    headers = [cell.value for cell in ws_index[1]]
    idx = {name: headers.index(name) for name in headers if name}

    required = ["SourceSheet", "SourceCell", "InputSheet", "InputCell"]
    missing = [name for name in required if name not in idx]
    if missing:
        raise ValueError(f"Index sheet missing columns: {missing}")

    overrides: dict[tuple[str, str], Any] = {}

    for row in ws_index.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        source_sheet = row[idx["SourceSheet"]]
        source_cell = row[idx["SourceCell"]]
        input_sheet = row[idx["InputSheet"]]
        input_cell = row[idx["InputCell"]]

        if not source_sheet or not source_cell or not input_sheet or not input_cell:
            continue
        if input_sheet not in wb.sheetnames:
            continue

        value = wb[input_sheet][input_cell].value
        overrides[(source_sheet, source_cell)] = value

    return overrides


def run_unstructured_calculation(
    mapping_report_path: Path,
    unstructured_input_path: Path,
    output_path: Path,
) -> Path:
    model = read_mapping_report(mapping_report_path)
    overrides = _extract_unstructured_overrides(model, unstructured_input_path)

    return save_workbook_from_mapping(
        model=model,
        output_path=output_path,
        value_overrides=overrides,
        include_formulas=True,
        unstructured_input_mode=False,
    )


def run_structured_calculation(
    mapping_report_path: Path,
    structured_input_path: Path,
    output_path: Path,
) -> Path:
    model = read_mapping_report(mapping_report_path)
    overrides = _extract_structured_overrides(structured_input_path)

    return save_workbook_from_mapping(
        model=model,
        output_path=output_path,
        value_overrides=overrides,
        include_formulas=True,
        unstructured_input_mode=False,
    )
