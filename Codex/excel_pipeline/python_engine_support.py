from __future__ import annotations

import math
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter, range_boundaries

from .mapping_io import read_mapping_report
from .reconstruct import create_workbook_from_mapping

CELL_REF_RE = re.compile(r"^\$?[A-Z]{1,3}\$?\d+$")
CELL_RANGE_RE = re.compile(r"^\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+$")
COL_RANGE_RE = re.compile(r"^\$?[A-Z]{1,3}:\$?[A-Z]{1,3}$")
ROW_RANGE_RE = re.compile(r"^\$?\d+:\$?\d+$")


class ExcelEngineError(Exception):
    """Raised when generated Python engine cannot evaluate a formula path."""


@dataclass
class _SheetBounds:
    min_row: int
    max_row: int
    min_col: int
    max_col: int


class PythonEngineContext:
    def __init__(self, workbook: Any, sheet_bounds: dict[str, _SheetBounds]) -> None:
        self._workbook = workbook
        self._sheet_bounds = sheet_bounds

    def cell(self, sheet: str, cell_ref: str) -> Any:
        if sheet not in self._workbook.sheetnames:
            raise ExcelEngineError(f"Unknown sheet: {sheet}")
        coord = cell_ref.replace("$", "")
        return self._workbook[sheet][coord].value

    def set_cell(self, sheet: str, cell_ref: str, value: Any) -> None:
        if sheet not in self._workbook.sheetnames:
            raise ExcelEngineError(f"Unknown sheet: {sheet}")
        coord = cell_ref.replace("$", "")
        self._workbook[sheet][coord].value = value

    def range(self, sheet: str, range_ref: str) -> list[Any]:
        if sheet not in self._workbook.sheetnames:
            raise ExcelEngineError(f"Unknown sheet: {sheet}")

        ws = self._workbook[sheet]
        ref = range_ref.replace("$", "")

        if CELL_RANGE_RE.fullmatch(ref):
            min_col, min_row, max_col, max_row = range_boundaries(ref)
        elif ROW_RANGE_RE.fullmatch(ref):
            start_row, end_row = [int(x) for x in ref.split(":", 1)]
            bounds = self._sheet_bounds.get(sheet)
            if bounds is None:
                min_col, max_col = 1, max(1, ws.max_column)
            else:
                min_col, max_col = bounds.min_col, bounds.max_col
            min_row, max_row = start_row, end_row
        elif COL_RANGE_RE.fullmatch(ref):
            start_col_text, end_col_text = ref.split(":", 1)
            start_col = column_index_from_string(start_col_text)
            end_col = column_index_from_string(end_col_text)
            bounds = self._sheet_bounds.get(sheet)
            if bounds is None:
                min_row, max_row = 1, max(1, ws.max_row)
            else:
                min_row, max_row = bounds.min_row, bounds.max_row
            min_col, max_col = start_col, end_col
        elif CELL_REF_RE.fullmatch(ref):
            return [ws[ref].value]
        else:
            raise ExcelEngineError(f"Unsupported range reference: {sheet}!{range_ref}")

        values: list[Any] = []
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                coord = f"{get_column_letter(col)}{row}"
                values.append(ws[coord].value)
        return values


# ---------------------------
# Excel-like scalar utilities
# ---------------------------

def _is_blank(value: Any) -> bool:
    return value is None or value == ""


def _is_error(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("#")


def _first_error(values: Any) -> str | None:
    for value in _flatten_values(values):
        if _is_error(value):
            return value
    return None


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _to_number(value: Any, *, strict: bool) -> float:
    if _is_error(value):
        raise ExcelEngineError(f"Cannot coerce Excel error to number: {value}")
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if _is_number(value):
        return float(value)
    if _is_blank(value):
        if strict:
            return 0.0
        return 0.0
    try:
        return float(value)
    except Exception:
        if strict:
            raise ExcelEngineError(f"Cannot coerce to number: {value!r}")
        return 0.0


def _flatten_values(values: Any) -> list[Any]:
    out: list[Any] = []

    def visit(item: Any) -> None:
        if isinstance(item, (list, tuple)):
            for sub in item:
                visit(sub)
        else:
            out.append(item)

    visit(values)
    return out


def xl_error(code: str) -> str:
    return code


def xl_ref(value: Any) -> Any:
    if value is None:
        return 0
    return value


def xl_uplus(value: Any) -> Any:
    if _is_error(value):
        return value
    return _to_number(value, strict=True)


def xl_uminus(value: Any) -> Any:
    if _is_error(value):
        return value
    return -_to_number(value, strict=True)


def xl_percent(value: Any) -> Any:
    if _is_error(value):
        return value
    return _to_number(value, strict=True) / 100.0


def xl_add(left: Any, right: Any) -> Any:
    err = _first_error((left, right))
    if err is not None:
        return err
    return _to_number(left, strict=True) + _to_number(right, strict=True)


def xl_sub(left: Any, right: Any) -> Any:
    err = _first_error((left, right))
    if err is not None:
        return err
    return _to_number(left, strict=True) - _to_number(right, strict=True)


def xl_mul(left: Any, right: Any) -> Any:
    err = _first_error((left, right))
    if err is not None:
        return err
    return _to_number(left, strict=True) * _to_number(right, strict=True)


def xl_div(left: Any, right: Any) -> Any:
    err = _first_error((left, right))
    if err is not None:
        return err
    denominator = _to_number(right, strict=True)
    if math.isclose(denominator, 0.0, abs_tol=1e-12):
        raise ExcelEngineError("Division by zero")
    return _to_number(left, strict=True) / denominator


def xl_pow(left: Any, right: Any) -> Any:
    err = _first_error((left, right))
    if err is not None:
        return err
    return _to_number(left, strict=True) ** _to_number(right, strict=True)


def xl_concat(left: Any, right: Any) -> str:
    err = _first_error((left, right))
    if err is not None:
        return err
    left_text = "" if left is None else str(left)
    right_text = "" if right is None else str(right)
    return left_text + right_text


def _cmp_values(left: Any, right: Any) -> tuple[Any, Any]:
    err = _first_error((left, right))
    if err is not None:
        return err, err
    if _is_number(left) or _is_number(right) or isinstance(left, bool) or isinstance(right, bool):
        return _to_number(left, strict=True), _to_number(right, strict=True)
    left_text = "" if left is None else str(left)
    right_text = "" if right is None else str(right)
    return left_text, right_text


def xl_eq(left: Any, right: Any) -> bool:
    a, b = _cmp_values(left, right)
    return a == b


def xl_ne(left: Any, right: Any) -> bool:
    a, b = _cmp_values(left, right)
    return a != b


def xl_gt(left: Any, right: Any) -> bool:
    a, b = _cmp_values(left, right)
    return a > b


def xl_ge(left: Any, right: Any) -> bool:
    a, b = _cmp_values(left, right)
    return a >= b


def xl_lt(left: Any, right: Any) -> bool:
    a, b = _cmp_values(left, right)
    return a < b


def xl_le(left: Any, right: Any) -> bool:
    a, b = _cmp_values(left, right)
    return a <= b


def _truthy(value: Any) -> bool:
    if _is_blank(value):
        return False
    if isinstance(value, bool):
        return value
    if _is_number(value):
        return not math.isclose(float(value), 0.0, abs_tol=1e-12)
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"true", "false"}:
            return lowered == "true"
    return bool(value)


def xl_if(condition: Any, true_value: Any, false_value: Any = False) -> Any:
    if _is_error(condition):
        return condition
    return true_value if _truthy(condition) else false_value


def xl_iferror(value: Any, fallback: Any) -> Any:
    if isinstance(value, str) and value.startswith("#"):
        return fallback
    return value


def xl_or(*args: Any) -> bool:
    return any(_truthy(v) for v in _flatten_values(args))


def xl_and(*args: Any) -> bool:
    values = _flatten_values(args)
    return bool(values) and all(_truthy(v) for v in values)


def xl_sum(*args: Any) -> Any:
    err = _first_error(args)
    if err is not None:
        return err
    total = 0.0
    for value in _flatten_values(args):
        if _is_blank(value):
            continue
        if isinstance(value, bool):
            total += 1.0 if value else 0.0
            continue
        if _is_number(value):
            total += float(value)
            continue
        try:
            total += float(value)
        except Exception:
            continue

    rounded = round(total)
    if math.isclose(total, float(rounded), rel_tol=1e-12, abs_tol=1e-12):
        return int(rounded)
    return total


def xl_average(*args: Any) -> Any:
    err = _first_error(args)
    if err is not None:
        return err
    numbers: list[float] = []
    for value in _flatten_values(args):
        if _is_blank(value):
            continue
        if isinstance(value, bool):
            numbers.append(1.0 if value else 0.0)
            continue
        if _is_number(value):
            numbers.append(float(value))
            continue
        try:
            numbers.append(float(value))
        except Exception:
            continue

    if not numbers:
        raise ExcelEngineError("AVERAGE with no numeric arguments")
    return sum(numbers) / len(numbers)


def xl_max(*args: Any) -> Any:
    err = _first_error(args)
    if err is not None:
        return err
    numbers = [
        _to_number(value, strict=False)
        for value in _flatten_values(args)
        if not _is_blank(value)
    ]
    if not numbers:
        return 0
    return max(numbers)


def xl_min(*args: Any) -> Any:
    err = _first_error(args)
    if err is not None:
        return err
    numbers = [
        _to_number(value, strict=False)
        for value in _flatten_values(args)
        if not _is_blank(value)
    ]
    if not numbers:
        return 0
    return min(numbers)


def xl_mod(number: Any, divisor: Any) -> Any:
    err = _first_error((number, divisor))
    if err is not None:
        return err
    num = _to_number(number, strict=True)
    div = _to_number(divisor, strict=True)
    if math.isclose(div, 0.0, abs_tol=1e-12):
        raise ExcelEngineError("MOD divisor is zero")
    return num % div


def xl_roundup(number: Any, digits: Any = 0) -> Any:
    err = _first_error((number, digits))
    if err is not None:
        return err
    value = _to_number(number, strict=True)
    d = int(_to_number(digits, strict=True))
    factor = 10 ** d
    if value >= 0:
        return math.ceil(value * factor) / factor
    return math.floor(value * factor) / factor


def xl_rounddown(number: Any, digits: Any = 0) -> Any:
    err = _first_error((number, digits))
    if err is not None:
        return err
    value = _to_number(number, strict=True)
    d = int(_to_number(digits, strict=True))
    factor = 10 ** d
    if value >= 0:
        return math.floor(value * factor) / factor
    return math.ceil(value * factor) / factor


def xl_year(value: Any) -> int:
    if _is_error(value):
        return value
    if isinstance(value, datetime):
        return value.year
    if isinstance(value, date):
        return value.year
    text = str(value)
    parsed = datetime.fromisoformat(text)
    return parsed.year


def xl_month(value: Any) -> int:
    if _is_error(value):
        return value
    if isinstance(value, datetime):
        return value.month
    if isinstance(value, date):
        return value.month
    text = str(value)
    parsed = datetime.fromisoformat(text)
    return parsed.month


def xl_today() -> date:
    return date.today()


def xl_now() -> datetime:
    return datetime.now()


def xl_sumproduct(*args: Any) -> Any:
    err = _first_error(args)
    if err is not None:
        return err
    vectors = [_flatten_values(arg) for arg in args]
    if not vectors:
        return 0
    length = len(vectors[0])
    if any(len(v) != length for v in vectors):
        raise ExcelEngineError("SUMPRODUCT arguments must have equal lengths")

    total = 0.0
    for i in range(length):
        product = 1.0
        for vector in vectors:
            product *= _to_number(vector[i], strict=False)
        total += product
    return total


def xl_npv(rate: Any, *cashflows: Any) -> Any:
    err = _first_error((rate, cashflows))
    if err is not None:
        return err
    r = _to_number(rate, strict=True)
    flows = _flatten_values(cashflows)
    total = 0.0
    period = 1
    for value in flows:
        if _is_blank(value):
            period += 1
            continue
        flow = _to_number(value, strict=False)
        total += flow / ((1.0 + r) ** period)
        period += 1
    return total


def _sheet_bounds_from_model(model: Any) -> dict[str, _SheetBounds]:
    bounds: dict[str, _SheetBounds] = {}
    for sheet in model.sheet_order:
        records = model.cells_by_sheet.get(sheet, [])
        if not records:
            continue

        min_row = min(record.row for record in records)
        max_row = max(record.row for record in records)
        min_col = min(record.column for record in records)
        max_col = max(record.column for record in records)
        bounds[sheet] = _SheetBounds(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        )

    return bounds


def _extract_unstructured_overrides(model: Any, input_path: Path) -> dict[tuple[str, str], Any]:
    wb = load_workbook(input_path, data_only=True)
    overrides: dict[tuple[str, str], Any] = {}

    for sheet_name in model.sheet_order:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for record in model.cells_by_sheet.get(sheet_name, []):
            if record.cell_type != "Input" or not record.include_flag:
                continue
            overrides[(sheet_name, record.cell)] = ws[record.cell].value

    return overrides


def execute_unstructured_python_engine(
    mapping_report_path: Path,
    unstructured_input_path: Path,
    output_path: Path,
    formula_order: list[tuple[str, str]],
    formula_funcs: dict[tuple[str, str], Callable[[PythonEngineContext], Any]],
) -> Path:
    model = read_mapping_report(mapping_report_path)
    overrides = _extract_unstructured_overrides(model, unstructured_input_path)

    wb = create_workbook_from_mapping(
        model=model,
        value_overrides=overrides,
        include_formulas=False,
        unstructured_input_mode=False,
    )

    ctx = PythonEngineContext(
        workbook=wb,
        sheet_bounds=_sheet_bounds_from_model(model),
    )

    for sheet, cell in formula_order:
        func = formula_funcs[(sheet, cell)]
        value = func(ctx)
        ctx.set_cell(sheet, cell, value)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
