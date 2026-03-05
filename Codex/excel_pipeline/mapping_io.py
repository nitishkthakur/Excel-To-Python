from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .types import CellRecord, MappingModel, SheetLayout
from .utils import deserialize_value


MAPPING_SHEET_EXCLUDE = {"_Metadata", "_FormulaGroups", "Index", "Config"}


def _parse_metadata_sheet(ws: Any) -> tuple[dict[str, Any], dict[str, dict[str, Any]]]:
    workbook_meta: dict[str, Any] = {}
    layout_meta: dict[str, dict[str, Any]] = defaultdict(dict)

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for section, sheet, key, value in rows:
        if key is None:
            continue

        if section == "Workbook":
            if key in {"SourceWorkbook", "NormalizedWorkbook"}:
                workbook_meta[key] = Path(value) if value else None
            elif key == "SheetOrder":
                workbook_meta[key] = json.loads(value) if value else []
            else:
                try:
                    workbook_meta[key] = json.loads(value)
                except Exception:
                    workbook_meta[key] = value
        elif section == "SheetLayout" and sheet:
            if key in {"MergedRanges", "RowDimensions", "ColumnDimensions"}:
                if not value:
                    layout_meta[sheet][key] = []
                else:
                    try:
                        layout_meta[sheet][key] = json.loads(value)
                    except Exception:
                        layout_meta[sheet][key] = []
            elif key == "Index":
                layout_meta[sheet][key] = int(value)
            else:
                layout_meta[sheet][key] = value

    return workbook_meta, layout_meta


def read_mapping_report(mapping_report_path: Path) -> MappingModel:
    wb = load_workbook(mapping_report_path, data_only=False)

    if "_Metadata" not in wb.sheetnames:
        raise ValueError("mapping_report.xlsx is missing _Metadata sheet")

    metadata_ws = wb["_Metadata"]
    workbook_meta, layout_meta = _parse_metadata_sheet(metadata_ws)

    source_workbook = workbook_meta.get("SourceWorkbook")
    normalized_workbook = workbook_meta.get("NormalizedWorkbook")

    if source_workbook is None or normalized_workbook is None:
        raise ValueError("_Metadata must include SourceWorkbook and NormalizedWorkbook")

    sheet_order = workbook_meta.get("SheetOrder", [])
    layouts: dict[str, SheetLayout] = {}

    for sheet_name in sheet_order:
        sheet_data = layout_meta.get(sheet_name, {})
        layouts[sheet_name] = SheetLayout(
            title=sheet_name,
            index=int(sheet_data.get("Index", 0)),
            merged_ranges=sheet_data.get("MergedRanges", []),
            freeze_panes=sheet_data.get("FreezePanes") or None,
            tab_color=sheet_data.get("TabColor") or None,
            row_dimensions=sheet_data.get("RowDimensions", []),
            column_dimensions=sheet_data.get("ColumnDimensions", []),
        )

    cells_by_sheet: dict[str, list[CellRecord]] = defaultdict(list)

    for sheet_name in wb.sheetnames:
        if sheet_name in MAPPING_SHEET_EXCLUDE:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        index = {name: i for i, name in enumerate(headers) if name is not None}

        def get_value(row: tuple[Any, ...], key: str, default: Any = None) -> Any:
            idx = index.get(key)
            if idx is None or idx >= len(row):
                return default
            value = row[idx]
            return default if value is None else value

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(x is not None for x in row):
                continue

            record_sheet = get_value(row, "Sheet")
            cell = get_value(row, "Cell")
            row_idx = get_value(row, "Row", 0)
            col_idx = get_value(row, "Column", 0)
            cell_type = get_value(row, "Type", "Input")
            formula = get_value(row, "Formula")
            if isinstance(formula, str) and formula.startswith("'="):
                formula = formula[1:]
            _value_display = get_value(row, "Value")
            value_json = get_value(row, "ValueJSON", "")
            number_format = get_value(row, "NumberFormat")
            font_bold = get_value(row, "FontBold")
            font_italic = get_value(row, "FontItalic")
            font_size = get_value(row, "FontSize")
            font_color = get_value(row, "FontColor")
            fill_color = get_value(row, "FillColor")
            horizontal_alignment = get_value(row, "HorizontalAlignment")
            vertical_alignment = get_value(row, "VerticalAlignment")
            wrap_text = get_value(row, "WrapText")
            include_flag = get_value(row, "IncludeFlag", False)
            group_id = get_value(row, "GroupID")
            group_direction = get_value(row, "GroupDirection")
            group_size = get_value(row, "GroupSize")
            pattern_formula = get_value(row, "PatternFormula")
            style_json = get_value(row, "StyleJSON")

            value = deserialize_value(value_json) if value_json else _value_display

            cells_by_sheet[sheet_name].append(
                CellRecord(
                    sheet=record_sheet,
                    row=int(row_idx),
                    column=int(col_idx),
                    cell=cell,
                    cell_type=cell_type,
                    formula=formula,
                    value=value,
                    include_flag=bool(include_flag),
                    number_format=number_format,
                    font_bold=bool(font_bold) if font_bold is not None else None,
                    font_italic=bool(font_italic) if font_italic is not None else None,
                    font_size=float(font_size) if font_size is not None else None,
                    font_color=font_color,
                    fill_color=fill_color,
                    horizontal_alignment=horizontal_alignment,
                    vertical_alignment=vertical_alignment,
                    wrap_text=bool(wrap_text) if wrap_text is not None else None,
                    style_json=style_json,
                    value_json=value_json,
                    group_id=group_id,
                    group_direction=group_direction,
                    group_size=int(group_size) if group_size is not None else None,
                    pattern_formula=pattern_formula,
                )
            )

    metadata_copy = {
        key: value
        for key, value in workbook_meta.items()
        if key not in {"SourceWorkbook", "NormalizedWorkbook", "SheetOrder"}
    }

    return MappingModel(
        source_workbook=Path(source_workbook),
        normalized_workbook=Path(normalized_workbook),
        sheet_order=sheet_order,
        layouts=layouts,
        cells_by_sheet=cells_by_sheet,
        metadata=metadata_copy,
    )
