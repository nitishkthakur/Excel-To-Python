from __future__ import annotations

import contextlib
import io
import logging
import math
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

from .mapping_io import read_mapping_report
from .reconstruct import create_workbook_from_mapping

try:
    from xlcalculator import Evaluator, ModelCompiler  # type: ignore
except Exception:  # pragma: no cover
    Evaluator = None  # type: ignore[assignment]
    ModelCompiler = None  # type: ignore[assignment]

try:
    from formulas import ExcelModel  # type: ignore
except Exception:  # pragma: no cover
    ExcelModel = None  # type: ignore[assignment]


SUM_R1C1_RE = re.compile(
    r"^=SUM\(R\[(?P<r1>-?\d+)\]C\[(?P<c1>-?\d+)\]:R\[(?P<r2>-?\d+)\]C\[(?P<c2>-?\d+)\]\)$"
)
FORMULAS_CELL_KEY_RE = re.compile(
    r"^'?\[[^\]]+\](?P<sheet>[^']+)'?!(?P<cell>\$?[A-Z]{1,3}\$?\d+)$"
)


@dataclass
class CalculationStats:
    vectorized_groups: int = 0
    vectorized_cells: int = 0
    xlcalculator_cells: int = 0
    formulas_cells: int = 0
    static_fallback_cells: int = 0


def _extract_unstructured_overrides(mapping_model: Any, input_path: Path) -> dict[tuple[str, str], Any]:
    wb = load_workbook(input_path, data_only=True)
    overrides: dict[tuple[str, str], Any] = {}

    for sheet_name in mapping_model.sheet_order:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for record in mapping_model.cells_by_sheet.get(sheet_name, []):
            if record.cell_type != "Input" or not record.include_flag:
                continue
            overrides[(sheet_name, record.cell)] = ws[record.cell].value

    return overrides


def _extract_structured_overrides(structured_input_path: Path) -> dict[tuple[str, str], Any]:
    wb = load_workbook(structured_input_path, data_only=True)
    if "Index" not in wb.sheetnames:
        raise ValueError("structured_input.xlsx must include Index sheet")

    ws_index = wb["Index"]
    headers = [cell.value for cell in ws_index[1]]
    idx = {name: headers.index(name) for name in headers if name}

    required = ["SourceSheet", "SourceCell", "InputSheet", "InputCell"]
    missing = [name for name in required if name not in idx]
    if missing:
        raise ValueError(f"Index sheet missing columns: {missing}")

    overrides: dict[tuple[str, str], Any] = {}
    for row in ws_index.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        source_sheet = row[idx["SourceSheet"]]
        source_cell = row[idx["SourceCell"]]
        input_sheet = row[idx["InputSheet"]]
        input_cell = row[idx["InputCell"]]

        if not source_sheet or not source_cell or not input_sheet or not input_cell:
            continue
        if input_sheet not in wb.sheetnames:
            continue

        overrides[(source_sheet, source_cell)] = wb[input_sheet][input_cell].value

    return overrides


def _formula_records(model: Any) -> list[Any]:
    records: list[Any] = []
    for sheet in model.sheet_order:
        for record in model.cells_by_sheet.get(sheet, []):
            if record.formula is not None:
                records.append(record)
    return records


def _formula_coords_by_sheet(model: Any) -> dict[str, set[tuple[int, int]]]:
    by_sheet: dict[str, set[tuple[int, int]]] = defaultdict(set)
    for record in _formula_records(model):
        by_sheet[record.sheet].add((record.row, record.column))
    return by_sheet


def _grouped_sum_patterns(model: Any) -> dict[str, list[dict[str, Any]]]:
    groups: dict[str, dict[str, Any]] = {}

    for record in _formula_records(model):
        if not record.group_id or not isinstance(record.pattern_formula, str):
            continue
        match = SUM_R1C1_RE.match(record.pattern_formula)
        if not match:
            continue

        group = groups.setdefault(
            record.group_id,
            {
                "group_id": record.group_id,
                "sheet": record.sheet,
                "offsets": tuple(
                    int(match.group(k)) for k in ("r1", "c1", "r2", "c2")
                ),
                "coords": [],
            },
        )
        group["coords"].append((record.row, record.column))

    by_sheet: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for group in groups.values():
        coords = sorted(group["coords"], key=lambda x: (x[0], x[1]))
        target_rows = np.array([row for row, _ in coords], dtype=int)
        target_cols = np.array([col for _, col in coords], dtype=int)
        row_start, row_end = sorted((group["offsets"][0], group["offsets"][2]))
        col_start, col_end = sorted((group["offsets"][1], group["offsets"][3]))

        group["coords"] = coords
        group["coord_set"] = set(coords)
        group["target_rows"] = target_rows
        group["target_cols"] = target_cols
        group["row_start"] = row_start
        group["row_end"] = row_end
        group["col_start"] = col_start
        group["col_end"] = col_end
        by_sheet[group["sheet"]].append(group)

    for sheet in by_sheet:
        by_sheet[sheet] = sorted(
            by_sheet[sheet],
            key=lambda g: (min(r for r, _ in g["coords"]), min(c for _, c in g["coords"])),
        )

    return by_sheet


def _build_numeric_series(ws: Any) -> pd.Series:
    coords: list[tuple[int, int]] = []
    values: list[Any] = []
    for (row, col), cell in ws._cells.items():  # pylint: disable=protected-access
        coords.append((int(row), int(col)))
        values.append(cell.value)

    if not coords:
        return pd.Series(dtype=float)

    index = pd.MultiIndex.from_tuples(coords, names=["row", "col"])
    numeric = pd.to_numeric(
        pd.Series(values, index=index, dtype="object"),
        errors="coerce",
    ).fillna(0.0)
    return numeric


def _group_has_unresolved_precedent(
    group: dict[str, Any],
    formula_coords: set[tuple[int, int]],
    resolved_coords: set[tuple[int, int]],
) -> bool:
    rows = group["target_rows"]
    cols = group["target_cols"]
    same_group = group["coord_set"]

    for dr in range(group["row_start"], group["row_end"] + 1):
        for dc in range(group["col_start"], group["col_end"] + 1):
            src_rows = rows + dr
            src_cols = cols + dc
            src_coords = {
                (int(r), int(c))
                for r, c in zip(src_rows.tolist(), src_cols.tolist(), strict=True)
            }
            blockers = (src_coords & formula_coords) - resolved_coords - same_group
            if blockers:
                return True
    return False


def _to_excel_number(value: float) -> Any:
    if math.isnan(value):
        return None
    rounded = round(value)
    if abs(value - rounded) < 1e-12:
        return int(rounded)
    return float(value)


def _evaluate_sum_group(
    group: dict[str, Any],
    numeric_series: pd.Series,
) -> np.ndarray:
    rows = group["target_rows"]
    cols = group["target_cols"]
    totals = np.zeros(len(rows), dtype=float)

    for dr in range(group["row_start"], group["row_end"] + 1):
        for dc in range(group["col_start"], group["col_end"] + 1):
            src_index = pd.MultiIndex.from_arrays(
                [rows + dr, cols + dc],
                names=["row", "col"],
            )
            totals += numeric_series.reindex(src_index, fill_value=0.0).to_numpy(dtype=float)

    return totals


def _apply_vectorized_group_calculations(
    wb: Any,
    grouped: dict[str, list[dict[str, Any]]],
    formula_coords_by_sheet: dict[str, set[tuple[int, int]]],
) -> tuple[dict[str, set[tuple[int, int]]], CalculationStats]:
    resolved: dict[str, set[tuple[int, int]]] = defaultdict(set)
    stats = CalculationStats()

    pending = {sheet: list(groups) for sheet, groups in grouped.items()}
    while True:
        progress = False
        for sheet_name, groups in list(pending.items()):
            if not groups or sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            numeric_series = _build_numeric_series(ws)
            still_pending: list[dict[str, Any]] = []

            for group in groups:
                if _group_has_unresolved_precedent(
                    group,
                    formula_coords=formula_coords_by_sheet.get(sheet_name, set()),
                    resolved_coords=resolved[sheet_name],
                ):
                    still_pending.append(group)
                    continue

                totals = _evaluate_sum_group(group, numeric_series)
                for (row, col), total in zip(group["coords"], totals, strict=True):
                    ws.cell(row=row, column=col).value = _to_excel_number(float(total))

                resolved[sheet_name].update(group["coord_set"])
                stats.vectorized_groups += 1
                stats.vectorized_cells += len(group["coords"])
                progress = True

            pending[sheet_name] = still_pending

        if not progress:
            break

    return resolved, stats


def _recompute_all_vectorized_groups(
    wb: Any,
    grouped: dict[str, list[dict[str, Any]]],
) -> None:
    # Final deterministic pass: once inputs and fallback engines have populated
    # values, recompute grouped SUM patterns in bulk so dragged runs always align.
    for _ in range(4):
        changed = False
        for sheet_name, groups in grouped.items():
            if sheet_name not in wb.sheetnames or not groups:
                continue
            ws = wb[sheet_name]
            numeric_series = _build_numeric_series(ws)
            for group in groups:
                totals = _evaluate_sum_group(group, numeric_series)
                for (row, col), total in zip(group["coords"], totals, strict=True):
                    value = _to_excel_number(float(total))
                    cell = ws.cell(row=row, column=col)
                    if cell.value != value:
                        cell.value = value
                        changed = True
        if not changed:
            break


def _coerce_engine_value(value: Any) -> Any:
    if hasattr(value, "value"):
        try:
            value = value.value
        except Exception:
            pass

    if isinstance(value, np.ndarray):
        if value.size != 1:
            return None
        value = value.reshape(-1)[0]

    if isinstance(value, list):
        if len(value) == 1 and isinstance(value[0], list) and len(value[0]) == 1:
            value = value[0][0]
        elif len(value) == 1:
            value = value[0]

    if isinstance(value, np.generic):
        value = value.item()

    if value is not None and value.__class__.__name__ == "XlError":
        return str(value)

    return value


def _values_equal(left: Any, right: Any) -> bool:
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return math.isclose(float(left), float(right), rel_tol=1e-9, abs_tol=1e-9)
    return left == right


def _has_any_input_change(model: Any, overrides: dict[tuple[str, str], Any]) -> bool:
    for sheet in model.sheet_order:
        for record in model.cells_by_sheet.get(sheet, []):
            if record.cell_type != "Input" or not record.include_flag:
                continue
            new_value = overrides.get((sheet, record.cell), record.value)
            if not _values_equal(new_value, record.value):
                return True
    return False


def _evaluate_with_xlcalculator(
    model: Any,
    wb: Any,
    unresolved_records: list[Any],
    seed_values: dict[tuple[str, str], Any],
) -> tuple[set[tuple[str, str]], int]:
    if not unresolved_records:
        return set(), 0
    if ModelCompiler is None or Evaluator is None:
        return {(r.sheet, r.cell) for r in unresolved_records}, 0

    try:
        compiler = ModelCompiler()
        compiled_model = compiler.read_and_parse_archive(str(model.normalized_workbook))
        evaluator = Evaluator(compiled_model)
    except Exception:
        return {(r.sheet, r.cell) for r in unresolved_records}, 0

    for (sheet, cell), value in seed_values.items():
        try:
            evaluator.set_cell_value(f"{sheet}!{cell}", value)
        except Exception:
            continue

    unresolved: set[tuple[str, str]] = set()
    solved = 0
    for record in unresolved_records:
        coord = (record.sheet, record.cell)
        try:
            value = _coerce_engine_value(evaluator.evaluate(f"{record.sheet}!{record.cell}"))
        except Exception:
            unresolved.add(coord)
            continue
        wb[record.sheet][record.cell].value = value
        solved += 1

    return unresolved, solved


def _parse_formulas_key(
    key: Any,
    sheet_lookup: dict[str, str],
) -> tuple[str, str] | None:
    key_text = str(key).strip()
    match = FORMULAS_CELL_KEY_RE.fullmatch(key_text)
    if not match:
        return None

    sheet_upper = match.group("sheet").upper()
    sheet = sheet_lookup.get(sheet_upper)
    if not sheet:
        return None

    cell = match.group("cell").replace("$", "")
    return sheet, cell


def _evaluate_with_formulas_engine(
    model: Any,
    wb: Any,
    unresolved_records: list[Any],
    seed_values: dict[tuple[str, str], Any],
) -> tuple[set[tuple[str, str]], int]:
    if not unresolved_records:
        return set(), 0
    if ExcelModel is None:
        return {(r.sheet, r.cell) for r in unresolved_records}, 0

    logging.getLogger("formulas").setLevel(logging.ERROR)
    logging.getLogger("schedula").setLevel(logging.ERROR)

    capture = io.StringIO()
    try:
        with contextlib.redirect_stdout(capture), contextlib.redirect_stderr(capture):
            excel_model = ExcelModel().loads(str(model.normalized_workbook)).finish()
    except Exception:
        return {(r.sheet, r.cell) for r in unresolved_records}, 0

    sheet_lookup = {sheet.upper(): sheet for sheet in model.sheet_order}

    default_nodes: dict[tuple[str, str], Any] = {}
    for node in excel_model.dsp.default_values:
        parsed = _parse_formulas_key(node, sheet_lookup)
        if parsed:
            default_nodes[parsed] = node

    for coord, value in seed_values.items():
        node = default_nodes.get(coord)
        if not node:
            continue
        try:
            excel_model.dsp.set_default_value(node, value)
        except Exception:
            continue

    try:
        with contextlib.redirect_stdout(capture), contextlib.redirect_stderr(capture):
            solution = excel_model.calculate()
    except Exception:
        return {(r.sheet, r.cell) for r in unresolved_records}, 0

    solution_values: dict[tuple[str, str], Any] = {}
    for key, value in solution.items():
        parsed = _parse_formulas_key(key, sheet_lookup)
        if not parsed:
            continue
        scalar = _coerce_engine_value(value)
        solution_values[parsed] = scalar

    unresolved: set[tuple[str, str]] = set()
    solved = 0
    for record in unresolved_records:
        coord = (record.sheet, record.cell)
        if coord not in solution_values:
            unresolved.add(coord)
            continue
        wb[record.sheet][record.cell].value = solution_values[coord]
        solved += 1

    return unresolved, solved


def _run_python_only_calculation(
    model: Any,
    overrides: dict[tuple[str, str], Any],
    output_path: Path,
) -> tuple[Path, CalculationStats]:
    wb = create_workbook_from_mapping(
        model=model,
        value_overrides=overrides,
        include_formulas=False,
        unstructured_input_mode=False,
    )

    formula_records = _formula_records(model)
    formula_coords_by_sheet = _formula_coords_by_sheet(model)

    # Keep baseline formula values prefilled; Python engines overwrite affected
    # cells. This avoids blank outputs if a niche formula cannot be evaluated.
    for record in formula_records:
        wb[record.sheet][record.cell].value = record.value

    if not _has_any_input_change(model, overrides):
        stats = CalculationStats(static_fallback_cells=len(formula_records))
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path, stats

    grouped = _grouped_sum_patterns(model)
    resolved_vectorized, stats = _apply_vectorized_group_calculations(
        wb=wb,
        grouped=grouped,
        formula_coords_by_sheet=formula_coords_by_sheet,
    )

    resolved_formula_coords = {
        (sheet, row, col)
        for sheet, coords in resolved_vectorized.items()
        for row, col in coords
    }

    unresolved_records = [
        record
        for record in formula_records
        if (record.sheet, record.row, record.column) not in resolved_formula_coords
    ]

    seed_values = dict(overrides)
    for sheet, row, col in sorted(resolved_formula_coords):
        cell = f"{get_column_letter(col)}{row}"
        seed_values[(sheet, cell)] = wb[sheet].cell(row=row, column=col).value

    unresolved_coords, xl_solved = _evaluate_with_xlcalculator(
        model=model,
        wb=wb,
        unresolved_records=unresolved_records,
        seed_values=seed_values,
    )
    stats.xlcalculator_cells = xl_solved

    unresolved_records = [
        record for record in unresolved_records if (record.sheet, record.cell) in unresolved_coords
    ]

    unresolved_coords, formulas_solved = _evaluate_with_formulas_engine(
        model=model,
        wb=wb,
        unresolved_records=unresolved_records,
        seed_values=seed_values,
    )
    stats.formulas_cells = formulas_solved

    static_fallback = 0
    for record in unresolved_records:
        if (record.sheet, record.cell) not in unresolved_coords:
            continue
        wb[record.sheet][record.cell].value = record.value
        static_fallback += 1
    stats.static_fallback_cells = static_fallback

    _recompute_all_vectorized_groups(wb=wb, grouped=grouped)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path, stats


def run_unstructured_calculation_vectorized(
    mapping_report_path: Path,
    unstructured_input_path: Path,
    output_path: Path,
) -> Path:
    model = read_mapping_report(mapping_report_path)
    overrides = _extract_unstructured_overrides(model, unstructured_input_path)
    output, _stats = _run_python_only_calculation(model, overrides, output_path)
    return output


def run_structured_calculation_vectorized(
    mapping_report_path: Path,
    structured_input_path: Path,
    output_path: Path,
) -> Path:
    model = read_mapping_report(mapping_report_path)
    overrides = _extract_structured_overrides(structured_input_path)
    output, _stats = _run_python_only_calculation(model, overrides, output_path)
    return output
