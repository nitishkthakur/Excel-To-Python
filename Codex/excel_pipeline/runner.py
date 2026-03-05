from __future__ import annotations

import json
import os
import re
import subprocess
import sys
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path
from typing import Any

from .codegen import generate_structured_calculate, generate_unstructured_calculate
from .compare import compare_workbooks
from .layer2_structured import create_structured_input
from .layer2_unstructured import create_unstructured_inputs
from .mapping import create_mapping_report
from .mapping_io import read_mapping_report
from .runtime import run_structured_calculation, run_unstructured_calculation
from .types import PipelineArtifacts
from .verify import (
    diagnose_stage,
    verify_mapping_against_original,
    verify_structured_input,
    verify_unstructured_inputs,
)


def _safe_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value)
    cleaned = cleaned.strip("._")
    return cleaned or "workbook"


def _inspect_mapping(mapping_report: Path) -> dict[str, Any]:
    model = read_mapping_report(mapping_report)
    summary: dict[str, Any] = {
        "sheets": model.sheet_order,
        "cell_counts": {},
        "type_counts": {},
    }

    for sheet in model.sheet_order:
        records = model.cells_by_sheet.get(sheet, [])
        summary["cell_counts"][sheet] = len(records)

        type_counts = {"Input": 0, "Calculation": 0, "Output": 0}
        for record in records:
            if record.cell_type in type_counts:
                type_counts[record.cell_type] += 1
        summary["type_counts"][sheet] = type_counts

    return summary


def _inspect_structured_input(structured_input_path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(structured_input_path, data_only=True)
    info = {
        "sheetnames": wb.sheetnames,
        "index_rows": 0,
        "config_rows": 0,
        "table_sheets": [],
    }

    if "Index" in wb.sheetnames:
        ws = wb["Index"]
        info["index_rows"] = max(0, ws.max_row - 1)

    if "Config" in wb.sheetnames:
        ws = wb["Config"]
        info["config_rows"] = max(0, ws.max_row - 1)

    info["table_sheets"] = [s for s in wb.sheetnames if s not in {"Index", "Config"}]
    return info


def run_pipeline_for_workbook(
    source_workbook: Path,
    output_root: Path,
    cache_dir: Path,
    python_executable: Path | None = None,
) -> dict[str, Any]:
    output_root.mkdir(parents=True, exist_ok=True)
    cache_dir.mkdir(parents=True, exist_ok=True)

    model_dir = output_root / _safe_name(source_workbook.stem)
    model_dir.mkdir(parents=True, exist_ok=True)

    mapping_report = model_dir / "mapping_report.xlsx"
    unstructured_inputs = model_dir / "unstructured_inputs.xlsx"
    structured_input = model_dir / "structured_input.xlsx"
    unstructured_calculate = model_dir / "unstructured_calculate.py"
    structured_calculate = model_dir / "structured_calculate.py"
    output_unstructured = model_dir / "output_unstructured.xlsx"
    output_structured = model_dir / "output_structured.xlsx"

    create_mapping_report(
        source_workbook=source_workbook,
        mapping_report_path=mapping_report,
        cache_dir=cache_dir,
    )
    mapping_stage_mismatches = verify_mapping_against_original(mapping_report)

    create_unstructured_inputs(mapping_report, unstructured_inputs)
    unstructured_stage_mismatches = verify_unstructured_inputs(
        mapping_report, unstructured_inputs
    )

    create_structured_input(mapping_report, structured_input)
    structured_stage_mismatches = verify_structured_input(mapping_report, structured_input)

    generate_unstructured_calculate(
        mapping_report_path=mapping_report,
        output_script_path=unstructured_calculate,
        default_input_path=unstructured_inputs,
        default_output_path=output_unstructured,
    )

    generate_structured_calculate(
        mapping_report_path=mapping_report,
        output_script_path=structured_calculate,
        default_input_path=structured_input,
        default_output_path=output_structured,
    )

    run_unstructured_calculation(mapping_report, unstructured_inputs, output_unstructured)
    run_structured_calculation(mapping_report, structured_input, output_structured)

    if python_executable:
        subprocess.run(
            [str(python_executable), str(unstructured_calculate)],
            check=True,
            cwd=str(model_dir),
        )
        subprocess.run(
            [str(python_executable), str(structured_calculate)],
            check=True,
            cwd=str(model_dir),
        )

    model = read_mapping_report(mapping_report)
    baseline = model.normalized_workbook

    mismatches_unstructured = compare_workbooks(baseline, output_unstructured)
    mismatches_structured = compare_workbooks(baseline, output_structured)

    diagnosis = {
        "unstructured": diagnose_stage(
            mapping_mismatches=mapping_stage_mismatches,
            stage2_mismatches=unstructured_stage_mismatches,
            final_mismatches=[m.__dict__ for m in mismatches_unstructured],
            mode="unstructured",
        ),
        "structured": diagnose_stage(
            mapping_mismatches=mapping_stage_mismatches,
            stage2_mismatches=structured_stage_mismatches,
            final_mismatches=[m.__dict__ for m in mismatches_structured],
            mode="structured",
        ),
    }

    artifact = PipelineArtifacts(
        mapping_report=mapping_report,
        unstructured_inputs=unstructured_inputs,
        structured_input=structured_input,
        unstructured_calculate=unstructured_calculate,
        structured_calculate=structured_calculate,
        output_unstructured=output_unstructured,
        output_structured=output_structured,
    )

    report = {
        "source_workbook": str(source_workbook),
        "artifacts": {
            "mapping_report": str(artifact.mapping_report),
            "unstructured_inputs": str(artifact.unstructured_inputs),
            "structured_input": str(artifact.structured_input),
            "unstructured_calculate": str(artifact.unstructured_calculate),
            "structured_calculate": str(artifact.structured_calculate),
            "output_unstructured": str(artifact.output_unstructured),
            "output_structured": str(artifact.output_structured),
        },
        "intermediate_checks": {
            "mapping": _inspect_mapping(mapping_report),
            "structured_input": _inspect_structured_input(structured_input),
            "stage_verification": {
                "layer1_mapping_mismatches": mapping_stage_mismatches,
                "layer2a_unstructured_mismatches": unstructured_stage_mismatches,
                "layer2b_structured_mismatches": structured_stage_mismatches,
            },
        },
        "mismatches": {
            "unstructured": [m.__dict__ for m in mismatches_unstructured],
            "structured": [m.__dict__ for m in mismatches_structured],
        },
        "diagnosis": diagnosis,
        "success": (
            not mapping_stage_mismatches
            and not unstructured_stage_mismatches
            and not structured_stage_mismatches
            and not mismatches_unstructured
            and not mismatches_structured
        ),
    }

    report_path = model_dir / "validation_report.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=True), encoding="utf-8")

    return report


def _pipeline_worker(task: tuple[Path, Path, Path, Path | None]) -> dict[str, Any]:
    source_workbook, output_root, cache_dir, python_executable = task
    return run_pipeline_for_workbook(
        source_workbook=source_workbook,
        output_root=output_root,
        cache_dir=cache_dir,
        python_executable=python_executable,
    )


def run_pipeline_for_directory(
    excel_dir: Path,
    output_root: Path,
    cache_dir: Path,
    python_executable: Path | None = None,
    workers: int = 1,
) -> dict[str, Any]:
    files = sorted(
        [
            path
            for path in excel_dir.iterdir()
            if path.suffix.lower() in {".xls", ".xlsx"} and path.is_file()
        ]
    )

    results: list[dict[str, Any]] = []

    if workers <= 1:
        for idx, workbook in enumerate(files, start=1):
            result = run_pipeline_for_workbook(
                source_workbook=workbook,
                output_root=output_root,
                cache_dir=cache_dir,
                python_executable=python_executable,
            )
            print(f"[{idx}/{len(files)}] completed: {workbook.name}")
            results.append(result)
    else:
        tasks = [
            (workbook, output_root, cache_dir, python_executable)
            for workbook in files
        ]
        with ProcessPoolExecutor(max_workers=workers) as executor:
            future_map = {
                executor.submit(_pipeline_worker, task): task[0]
                for task in tasks
            }
            completed = 0
            for future in as_completed(future_map):
                workbook = future_map[future]
                completed += 1
                try:
                    result = future.result()
                except Exception as exc:
                    result = {
                        "source_workbook": str(workbook),
                        "success": False,
                        "error": str(exc),
                        "artifacts": {},
                        "mismatches": {"unstructured": [], "structured": []},
                        "intermediate_checks": {},
                    }
                print(f"[{completed}/{len(files)}] completed: {workbook.name}")
                results.append(result)

        # Keep summary stable and predictable.
        result_lookup = {item.get("source_workbook"): item for item in results}
        results = [result_lookup.get(str(workbook), {}) for workbook in files]

    summary = {
        "excel_dir": str(excel_dir),
        "output_root": str(output_root),
        "total_files": len(files),
        "success_count": sum(1 for item in results if item.get("success")),
        "failure_count": sum(1 for item in results if not item.get("success")),
        "results": results,
    }

    output_root.mkdir(parents=True, exist_ok=True)
    (output_root / "summary.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=True), encoding="utf-8"
    )

    return summary


def main(argv: list[str] | None = None) -> int:
    import argparse

    parser = argparse.ArgumentParser(description="Excel to Python conversion pipeline")
    parser.add_argument("--excel-dir", type=Path, default=Path("ExcelFiles"))
    parser.add_argument("--output-root", type=Path, default=Path("artifacts"))
    parser.add_argument("--cache-dir", type=Path, default=Path(".cache/normalized"))
    parser.add_argument("--single-file", type=Path, default=None)
    parser.add_argument("--verify-generated", action="store_true")
    parser.add_argument(
        "--workers",
        type=int,
        default=max(1, (os.cpu_count() or 2) // 2),
        help="Parallel workers for directory mode",
    )
    args = parser.parse_args(argv)

    python_executable = Path(sys.executable) if args.verify_generated else None

    if args.single_file:
        result = run_pipeline_for_workbook(
            source_workbook=args.single_file,
            output_root=args.output_root,
            cache_dir=args.cache_dir,
            python_executable=python_executable,
        )
        print(json.dumps(result, indent=2, ensure_ascii=True))
        return 0 if result.get("success") else 1

    summary = run_pipeline_for_directory(
        excel_dir=args.excel_dir,
        output_root=args.output_root,
        cache_dir=args.cache_dir,
        python_executable=python_executable,
        workers=max(1, args.workers),
    )
    print(json.dumps(summary, indent=2, ensure_ascii=True))
    return 0 if summary.get("failure_count", 0) == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
