from __future__ import annotations

import json
import math
import re
from datetime import date, datetime, time
from typing import Any

from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Protection, Side
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, get_column_letter


CELL_REF_RE = re.compile(r"^\$?[A-Z]{1,3}\$?\d+$")
CELL_RANGE_RE = re.compile(r"^\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+$")
COL_RANGE_RE = re.compile(r"^\$?[A-Z]{1,3}:\$?[A-Z]{1,3}$")
ROW_RANGE_RE = re.compile(r"^\$?\d+:\$?\d+$")
YEAR_STRING_RE = re.compile(r"^\d{4}[A-Za-z]*$")


def excel_bool(value: Any) -> bool | None:
    if value is None:
        return None
    return bool(value)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value)


def color_to_hex(color: Color | None) -> str | None:
    if color is None:
        return None
    if color.rgb:
        return str(color.rgb)
    if color.indexed is not None:
        return f"INDEXED:{color.indexed}"
    if color.theme is not None:
        tint = color.tint if color.tint is not None else 0
        return f"THEME:{color.theme}:{tint}"
    if color.auto:
        return "AUTO"
    return None


def serialize_color(color: Color | None) -> dict[str, Any] | None:
    if color is None:
        return None

    indexed = color.indexed
    theme = color.theme
    tint = color.tint
    auto = color.auto

    try:
        indexed = int(indexed) if indexed is not None else None
    except Exception:
        indexed = None

    try:
        theme = int(theme) if theme is not None else None
    except Exception:
        theme = None

    try:
        tint = float(tint) if tint is not None else None
    except Exception:
        tint = None

    if auto not in (None, True, False):
        auto = None

    return {
        "type": str(color.type) if color.type is not None else None,
        "rgb": str(color.rgb) if color.rgb is not None else None,
        "indexed": indexed,
        "theme": theme,
        "tint": tint,
        "auto": auto,
    }


def deserialize_color(payload: dict[str, Any] | None) -> Color | None:
    if not payload:
        return None
    indexed = payload.get("indexed")
    theme = payload.get("theme")
    tint = payload.get("tint")
    auto = payload.get("auto")

    try:
        indexed = int(indexed) if indexed is not None else None
    except Exception:
        indexed = None

    try:
        theme = int(theme) if theme is not None else None
    except Exception:
        theme = None

    try:
        tint = float(tint) if tint is not None else None
    except Exception:
        tint = None

    if auto not in (None, True, False):
        auto = None

    return Color(
        type=payload.get("type"),
        rgb=payload.get("rgb"),
        indexed=indexed,
        theme=theme,
        tint=tint,
        auto=auto,
    )


def serialize_side(side: Side | None) -> dict[str, Any] | None:
    if side is None:
        return None
    return {
        "style": side.style,
        "color": serialize_color(side.color),
    }


def deserialize_side(payload: dict[str, Any] | None) -> Side:
    if not payload:
        return Side()
    return Side(
        style=payload.get("style"),
        color=deserialize_color(payload.get("color")),
    )


def serialize_style(cell: Any) -> str:
    payload = {
        "number_format": cell.number_format,
        "font": {
            "name": cell.font.name,
            "size": cell.font.sz,
            "bold": cell.font.bold,
            "italic": cell.font.italic,
            "underline": cell.font.underline,
            "strike": cell.font.strike,
            "color": serialize_color(cell.font.color),
            "vert_align": cell.font.vertAlign,
        },
        "fill": {
            "fill_type": cell.fill.fill_type,
            "fg_color": serialize_color(cell.fill.fgColor),
            "bg_color": serialize_color(cell.fill.bgColor),
        },
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "text_rotation": cell.alignment.text_rotation,
            "wrap_text": cell.alignment.wrap_text,
            "shrink_to_fit": cell.alignment.shrink_to_fit,
            "indent": cell.alignment.indent,
        },
        "border": {
            "left": serialize_side(cell.border.left),
            "right": serialize_side(cell.border.right),
            "top": serialize_side(cell.border.top),
            "bottom": serialize_side(cell.border.bottom),
            "diagonal": serialize_side(cell.border.diagonal),
            "diagonal_down": cell.border.diagonalDown,
            "diagonal_up": cell.border.diagonalUp,
            "outline": cell.border.outline,
            "vertical": serialize_side(cell.border.vertical),
            "horizontal": serialize_side(cell.border.horizontal),
        },
        "protection": {
            "locked": cell.protection.locked,
            "hidden": cell.protection.hidden,
        },
    }
    return json.dumps(payload, ensure_ascii=True, default=str)


def deserialize_style(style_json: str | None) -> dict[str, Any]:
    if not style_json:
        return {}
    payload = json.loads(style_json)

    font_data = payload.get("font", {})
    fill_data = payload.get("fill", {})
    align_data = payload.get("alignment", {})
    border_data = payload.get("border", {})
    protection_data = payload.get("protection", {})

    font = Font(
        name=font_data.get("name"),
        sz=font_data.get("size"),
        bold=font_data.get("bold"),
        italic=font_data.get("italic"),
        underline=font_data.get("underline"),
        strike=font_data.get("strike"),
        color=deserialize_color(font_data.get("color")),
        vertAlign=font_data.get("vert_align"),
    )

    fill = PatternFill(
        fill_type=fill_data.get("fill_type"),
        fgColor=deserialize_color(fill_data.get("fg_color")),
        bgColor=deserialize_color(fill_data.get("bg_color")),
    )

    alignment = Alignment(
        horizontal=align_data.get("horizontal"),
        vertical=align_data.get("vertical"),
        text_rotation=align_data.get("text_rotation"),
        wrap_text=align_data.get("wrap_text"),
        shrink_to_fit=align_data.get("shrink_to_fit"),
        indent=align_data.get("indent"),
    )

    border = Border(
        left=deserialize_side(border_data.get("left")),
        right=deserialize_side(border_data.get("right")),
        top=deserialize_side(border_data.get("top")),
        bottom=deserialize_side(border_data.get("bottom")),
        diagonal=deserialize_side(border_data.get("diagonal")),
        diagonalDown=border_data.get("diagonal_down"),
        diagonalUp=border_data.get("diagonal_up"),
        outline=border_data.get("outline"),
        vertical=deserialize_side(border_data.get("vertical")),
        horizontal=deserialize_side(border_data.get("horizontal")),
    )

    protection = Protection(
        locked=protection_data.get("locked"),
        hidden=protection_data.get("hidden"),
    )

    return {
        "number_format": payload.get("number_format"),
        "font": font,
        "fill": fill,
        "alignment": alignment,
        "border": border,
        "protection": protection,
    }


def serialize_value(value: Any) -> str:
    if value is None:
        payload = {"type": "null", "value": None}
    elif isinstance(value, bool):
        payload = {"type": "bool", "value": value}
    elif isinstance(value, int):
        payload = {"type": "int", "value": value}
    elif isinstance(value, float):
        if math.isnan(value):
            payload = {"type": "nan", "value": None}
        elif math.isinf(value):
            payload = {"type": "inf", "value": value > 0}
        else:
            payload = {"type": "float", "value": value}
    elif isinstance(value, datetime):
        payload = {"type": "datetime", "value": value.isoformat()}
    elif isinstance(value, date):
        payload = {"type": "date", "value": value.isoformat()}
    elif isinstance(value, time):
        payload = {"type": "time", "value": value.isoformat()}
    else:
        payload = {"type": "str", "value": str(value)}

    return json.dumps(payload, ensure_ascii=True)


def deserialize_value(value_json: str) -> Any:
    payload = json.loads(value_json)
    vtype = payload.get("type")
    value = payload.get("value")

    if vtype == "null":
        return None
    if vtype == "bool":
        return bool(value)
    if vtype == "int":
        return int(value)
    if vtype == "float":
        return float(value)
    if vtype == "nan":
        return float("nan")
    if vtype == "inf":
        return float("inf") if value else float("-inf")
    if vtype == "datetime":
        return datetime.fromisoformat(value)
    if vtype == "date":
        return date.fromisoformat(value)
    if vtype == "time":
        return time.fromisoformat(value)
    return value


def parse_cell_ref(coord: str) -> tuple[int, int, bool, bool]:
    match = re.fullmatch(r"(\$?)([A-Z]{1,3})(\$?)(\d+)", coord)
    if not match:
        raise ValueError(f"Invalid cell reference: {coord}")

    col_abs = bool(match.group(1))
    col = column_index_from_string(match.group(2))
    row_abs = bool(match.group(3))
    row = int(match.group(4))
    return row, col, row_abs, col_abs


def to_r1c1_token(coord: str, origin_row: int, origin_col: int) -> str:
    row, col, row_abs, col_abs = parse_cell_ref(coord)

    if row_abs:
        row_token = f"R{row}"
    else:
        row_delta = row - origin_row
        row_token = f"R[{row_delta}]"

    if col_abs:
        col_token = f"C{col}"
    else:
        col_delta = col - origin_col
        col_token = f"C[{col_delta}]"

    return f"{row_token}{col_token}"


def maybe_financial_period(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (datetime, date)):
        return True
    if isinstance(value, int):
        return 1900 <= value <= 2200
    if isinstance(value, float) and value.is_integer():
        year = int(value)
        return 1900 <= year <= 2200
    if isinstance(value, str):
        return bool(YEAR_STRING_RE.fullmatch(value.strip()))
    return False


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", name)
    if len(cleaned) <= 31:
        return cleaned
    return cleaned[:31]


def dedupe_sheet_name(existing: set[str], base: str) -> str:
    candidate = safe_sheet_name(base)
    if candidate not in existing:
        return candidate

    suffix = 1
    while True:
        suffix_text = f"_{suffix}"
        truncated = candidate[: 31 - len(suffix_text)]
        name = f"{truncated}{suffix_text}"
        if name not in existing:
            return name
        suffix += 1


def sheet_cell_key(sheet: str, row: int, col: int) -> str:
    return f"{sheet}!{get_column_letter(col)}{row}"


def is_ref_token(token: str) -> bool:
    token = token.strip()
    return bool(
        CELL_REF_RE.fullmatch(token)
        or CELL_RANGE_RE.fullmatch(token)
        or COL_RANGE_RE.fullmatch(token)
        or ROW_RANGE_RE.fullmatch(token)
    )


def split_sheet_ref(token: str) -> tuple[str | None, str]:
    if "!" not in token:
        return None, token
    sheet_part, ref_part = token.rsplit("!", 1)
    sheet_part = sheet_part.strip()

    if sheet_part.startswith("'") and sheet_part.endswith("'"):
        sheet = sheet_part[1:-1].replace("''", "'")
    else:
        sheet = sheet_part

    if "]" in sheet:
        sheet = sheet.split("]", 1)[-1]

    return sheet, ref_part


def sheet_ref_token(sheet: str) -> str:
    escaped = sheet.replace("'", "''")
    return f"'{escaped}'"


def coordinate_to_row_col(coord: str) -> tuple[int, int]:
    col, row = coordinate_from_string(coord)
    return int(row), int(column_index_from_string(col))
