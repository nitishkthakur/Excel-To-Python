from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from excel_pipeline.python_runner import run_unstructured_python_pipeline_for_workbook


class PythonEnginePipelineTest(unittest.TestCase):
    def test_generated_python_engine_on_simple_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            source = root / "simple.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = 10
            ws["A2"] = 20
            ws["A3"] = "=SUM(A1:A2)"
            ws["B1"] = "=NPV(0.1,A1:A2)"
            wb.save(source)

            output_root = root / "artifacts"
            cache_dir = root / "cache"

            result = run_unstructured_python_pipeline_for_workbook(
                source_workbook=source,
                output_root=output_root,
                cache_dir=cache_dir,
            )
            self.assertIn("artifacts", result)

            output_path = Path(result["artifacts"]["output_unstructured_python"])
            self.assertTrue(output_path.exists())

            out_wb = load_workbook(output_path, data_only=True)
            out_ws = out_wb["Sheet1"]
            self.assertEqual(out_ws["A3"].value, 30)

            expected_npv = 10 / 1.1 + 20 / (1.1**2)
            actual_npv = out_ws["B1"].value
            self.assertIsNotNone(actual_npv)
            self.assertAlmostEqual(float(actual_npv), expected_npv, places=9)


if __name__ == "__main__":
    unittest.main()
