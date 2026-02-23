"""
Code Generator Module
=====================
Generates a standalone Python script that replicates Excel calculations.
The generated script reads inputs from an Excel template, performs all calculations,
and writes formatted output to a new Excel file.
"""

import os
import datetime
import logging
from openpyxl.utils import get_column_letter
from .excel_parser import WorkbookInfo, CellInfo
from .dependency_graph import (
    build_dependency_graph,
    topological_sort,
    find_unreferenced_hardcoded,
)
from .formula_translator import FormulaTranslator

logger = logging.getLogger(__name__)


def _safe_repr(value):
    """Generate a safe Python repr for a cell value, handling datetime objects."""
    if isinstance(value, datetime.datetime):
        # Convert to Excel serial date number
        delta = value - datetime.datetime(1899, 12, 30)
        return repr(delta.days + delta.seconds / 86400)
    elif isinstance(value, datetime.date):
        delta = value - datetime.date(1899, 12, 30)
        return repr(delta.days)
    elif isinstance(value, (int, float, bool, str, type(None))):
        return repr(value)
    else:
        # Fallback: try to convert to string
        try:
            return repr(str(value))
        except Exception:
            return 'None'


def generate_calculator(
    workbook_info: WorkbookInfo,
    output_dir: str,
    delete_unreferenced: bool = True,
) -> str:
    """
    Generate a standalone Python calculator script from the parsed workbook.

    Args:
        workbook_info: Parsed workbook data
        output_dir: Directory to write generated files
        delete_unreferenced: Whether to exclude unreferenced hardcoded values

    Returns:
        Path to the generated calculator script
    """
    os.makedirs(output_dir, exist_ok=True)

    # Build dependency graph
    adjacency, reverse_adj = build_dependency_graph(workbook_info)
    eval_order = topological_sort(workbook_info, adjacency)

    # Find unreferenced hardcoded values
    unreferenced = set()
    if delete_unreferenced:
        unreferenced = find_unreferenced_hardcoded(workbook_info, reverse_adj)
        logger.info(f"Found {len(unreferenced)} unreferenced hardcoded values to exclude")

    # Classify cells
    input_cells = []  # Hardcoded numbers that are inputs
    label_cells = []  # Text labels
    formula_cells = []  # Cells with formulas (in eval order)

    for key, cell in workbook_info.all_cells.items():
        if cell.is_hardcoded_number:
            if key not in unreferenced:
                input_cells.append(cell)
        elif cell.is_label:
            label_cells.append(cell)

    for key in eval_order:
        if key in workbook_info.all_cells:
            formula_cells.append(workbook_info.all_cells[key])

    # Sort input_cells by sheet, row, col for readability
    input_cells.sort(key=lambda c: (workbook_info.sheet_names.index(c.sheet)
                                     if c.sheet in workbook_info.sheet_names else 999,
                                     c.row, c.col))
    label_cells.sort(key=lambda c: (workbook_info.sheet_names.index(c.sheet)
                                     if c.sheet in workbook_info.sheet_names else 999,
                                     c.row, c.col))

    # Generate the script
    lines = []
    lines.append('"""')
    lines.append(f'Auto-generated Excel Calculator')
    lines.append(f'Source: {os.path.basename(workbook_info.file_path)}')
    lines.append(f'')
    lines.append(f'This script reads input values from an Excel template,')
    lines.append(f'performs all calculations, and writes the output to a new Excel file.')
    lines.append(f'')
    lines.append(f'Usage:')
    lines.append(f'    python calculator.py [input_file] [output_file]')
    lines.append(f'')
    lines.append(f'If no input file is specified, uses input_template.xlsx')
    lines.append(f'If no output file is specified, writes to output.xlsx')
    lines.append('"""')
    lines.append('')
    lines.append('import sys')
    lines.append('import os')
    lines.append('import openpyxl')
    lines.append('from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers')
    lines.append('from openpyxl.utils import get_column_letter')
    lines.append('from copy import copy')
    lines.append('import datetime')
    lines.append('import warnings')
    lines.append('warnings.filterwarnings("ignore", category=UserWarning)')
    lines.append('')
    lines.append('# Import Excel function implementations')
    lines.append('from excel_functions import *')
    lines.append('')
    lines.append('')

    # --- DEFAULTS DICT ---
    lines.append('# Default values for all input cells (from the original Excel file)')
    lines.append('# These are used when the input template does not provide a value')
    lines.append('DEFAULT_INPUTS = {')
    for cell in input_cells:
        val_repr = _safe_repr(cell.value)
        lines.append(f'    ({repr(cell.sheet)}, {cell.row}, {cell.col}): {val_repr},  '
                      f'# {cell.coordinate}')
    lines.append('}')
    lines.append('')

    # --- LABELS DICT ---
    lines.append('# Labels and text values')
    lines.append('LABELS = {')
    for cell in label_cells:
        val_repr = _safe_repr(cell.value)
        lines.append(f'    ({repr(cell.sheet)}, {cell.row}, {cell.col}): {val_repr},')
    lines.append('}')
    lines.append('')

    # --- SHEET INFO ---
    lines.append('# Sheet structure information')
    lines.append(f'SHEET_NAMES = {repr(workbook_info.sheet_names)}')
    lines.append('')
    lines.append('SHEET_DIMENSIONS = {')
    for name, sheet in workbook_info.sheets.items():
        lines.append(f'    {repr(name)}: ({sheet.min_row}, {sheet.min_col}, '
                      f'{sheet.max_row}, {sheet.max_col}),')
    lines.append('}')
    lines.append('')

    # --- MERGED CELLS ---
    lines.append('MERGED_CELLS = {')
    for name, sheet in workbook_info.sheets.items():
        if sheet.merged_cells:
            lines.append(f'    {repr(name)}: {repr(sheet.merged_cells)},')
    lines.append('}')
    lines.append('')

    # --- COLUMN WIDTHS ---
    lines.append('COLUMN_WIDTHS = {')
    for name, sheet in workbook_info.sheets.items():
        if sheet.column_widths:
            lines.append(f'    {repr(name)}: {repr(dict(sheet.column_widths))},')
    lines.append('}')
    lines.append('')

    # --- ROW HEIGHTS ---
    lines.append('ROW_HEIGHTS = {')
    for name, sheet in workbook_info.sheets.items():
        if sheet.row_heights:
            lines.append(f'    {repr(name)}: {repr(dict(sheet.row_heights))},')
    lines.append('}')
    lines.append('')

    # --- NUMBER FORMATS ---
    lines.append('# Number formats for all cells')
    lines.append('NUMBER_FORMATS = {')
    for key, cell in workbook_info.all_cells.items():
        if cell.number_format and cell.number_format != 'General':
            lines.append(f'    ({repr(cell.sheet)}, {cell.row}, {cell.col}): '
                          f'{repr(cell.number_format)},')
    lines.append('}')
    lines.append('')

    # --- FONT INFO ---
    lines.append('# Font styling info: (bold, color_rgb, size, name)')
    lines.append('FONT_STYLES = {')
    for key, cell in workbook_info.all_cells.items():
        if cell.font_bold or cell.font_color_rgb or cell.font_size or cell.font_name:
            lines.append(f'    ({repr(cell.sheet)}, {cell.row}, {cell.col}): '
                          f'({cell.font_bold}, {repr(cell.font_color_rgb)}, '
                          f'{repr(cell.font_size)}, {repr(cell.font_name)}),')
    lines.append('}')
    lines.append('')

    # --- FILL COLORS ---
    lines.append('FILL_COLORS = {')
    for key, cell in workbook_info.all_cells.items():
        if cell.fill_color:
            lines.append(f'    ({repr(cell.sheet)}, {cell.row}, {cell.col}): '
                          f'{repr(cell.fill_color)},')
    lines.append('}')
    lines.append('')

    # --- ALIGNMENT ---
    lines.append('ALIGNMENTS = {')
    for key, cell in workbook_info.all_cells.items():
        if cell.alignment_horizontal or cell.alignment_vertical or cell.alignment_wrap:
            lines.append(f'    ({repr(cell.sheet)}, {cell.row}, {cell.col}): '
                          f'({repr(cell.alignment_horizontal)}, '
                          f'{repr(cell.alignment_vertical)}, '
                          f'{cell.alignment_wrap}),')
    lines.append('}')
    lines.append('')
    lines.append('')

    # --- LOAD INPUTS FUNCTION ---
    lines.append('def load_inputs(input_file):')
    lines.append('    """Load input values from the input Excel template."""')
    lines.append('    cells = {}')
    lines.append('    ')
    lines.append('    # Start with defaults')
    lines.append('    cells.update(DEFAULT_INPUTS)')
    lines.append('    cells.update(LABELS)')
    lines.append('    ')
    lines.append('    # Override with values from input file')
    lines.append('    if os.path.exists(input_file):')
    lines.append('        wb = openpyxl.load_workbook(input_file, data_only=True)')
    lines.append('        for sheet_name in wb.sheetnames:')
    lines.append('            ws = wb[sheet_name]')
    lines.append('            for row in ws.iter_rows():')
    lines.append('                for cell in row:')
    lines.append('                    if cell.value is not None:')
    lines.append('                        cells[(sheet_name, cell.row, cell.column)] = cell.value')
    lines.append('        wb.close()')
    lines.append('    else:')
    lines.append('        print(f"Warning: Input file {input_file} not found. Using default values.")')
    lines.append('    ')
    lines.append('    return cells')
    lines.append('')
    lines.append('')

    # --- CALCULATE FUNCTION ---
    lines.append('def calculate(cells):')
    lines.append('    """')
    lines.append('    Perform all calculations in dependency order.')
    lines.append('    Modifies the cells dict in-place with computed values.')
    lines.append('    """')
    lines.append('    ')

    # Generate calculation for each formula cell in order  
    error_count = 0
    for cell in formula_cells:
        if not cell.formula:
            continue
        
        translator = FormulaTranslator(cell.sheet)    
        try:
            python_expr = translator.translate(cell.formula)
        except Exception as e:
            logger.warning(f"Failed to translate formula in {cell.full_address}: "
                           f"{cell.formula} -> {e}")
            error_count += 1
            # Use the precomputed value as fallback
            if cell.value is not None:
                python_expr = _safe_repr(cell.value)
            else:
                python_expr = 'None'

        key = f"({repr(cell.sheet)}, {cell.row}, {cell.col})"
        
        # Generate with error handling
        lines.append(f'    # {cell.full_address}: ={cell.formula[:80]}')
        lines.append(f'    try:')
        
        # Handle INDIRECT specially - needs cells and current_sheet passed
        if 'xl_indirect' in python_expr:
            python_expr = python_expr.replace(
                'xl_indirect(', 
                f'xl_indirect('
            )
            # Add cells and sheet args
            python_expr = _patch_indirect_calls(python_expr, cell.sheet)
        
        lines.append(f'        cells[{key}] = {python_expr}')
        lines.append(f'    except Exception as e:')
        
        # Use precomputed fallback
        if cell.value is not None:
            lines.append(f'        cells[{key}] = {_safe_repr(cell.value)}  # fallback value')
        else:
            lines.append(f'        cells[{key}] = 0  # fallback')
        lines.append(f'')

    if error_count:
        logger.warning(f"Total formula translation errors: {error_count}")

    lines.append('    return cells')
    lines.append('')
    lines.append('')

    # --- OUTPUT WRITER FUNCTION ---
    lines.append('def write_output(cells, output_file):')
    lines.append('    """Write calculated results to an Excel file with formatting."""')
    lines.append('    wb = openpyxl.Workbook()')
    lines.append('    wb.remove(wb.active)')
    lines.append('    ')
    lines.append('    for sheet_name in SHEET_NAMES:')
    lines.append('        ws = wb.create_sheet(title=sheet_name)')
    lines.append('        ')
    lines.append('        # Get sheet dimensions')
    lines.append('        if sheet_name in SHEET_DIMENSIONS:')
    lines.append('            min_r, min_c, max_r, max_c = SHEET_DIMENSIONS[sheet_name]')
    lines.append('        else:')
    lines.append('            continue')
    lines.append('        ')
    lines.append('        # Write all cell values')
    lines.append('        for r in range(min_r, max_r + 1):')
    lines.append('            for c in range(min_c, max_c + 1):')
    lines.append('                key = (sheet_name, r, c)')
    lines.append('                if key in cells and cells[key] is not None:')
    lines.append('                    cell = ws.cell(row=r, column=c, value=cells[key])')
    lines.append('                    ')
    lines.append('                    # Apply number format')
    lines.append('                    if key in NUMBER_FORMATS:')
    lines.append("                        cell.number_format = NUMBER_FORMATS[key]")
    lines.append('                    ')
    lines.append('                    # Apply font style')
    lines.append('                    if key in FONT_STYLES:')
    lines.append('                        bold, color_rgb, size, name = FONT_STYLES[key]')
    lines.append('                        font_kwargs = {}')
    lines.append('                        if bold:')
    lines.append("                            font_kwargs['bold'] = True")
    lines.append('                        if color_rgb:')
    lines.append("                            # Ensure 8-char aRGB format")
    lines.append("                            c = color_rgb.lstrip('#')")
    lines.append("                            if len(c) == 6: c = 'FF' + c")
    lines.append("                            if len(c) == 8:")
    lines.append("                                font_kwargs['color'] = c")
    lines.append('                        if size:')
    lines.append("                            font_kwargs['size'] = size")
    lines.append('                        if name:')
    lines.append("                            font_kwargs['name'] = name")
    lines.append('                        if font_kwargs:')
    lines.append('                            try:')
    lines.append('                                cell.font = Font(**font_kwargs)')
    lines.append('                            except Exception:')
    lines.append('                                pass')
    lines.append('                    ')
    lines.append('                    # Apply fill')
    lines.append('                    if key in FILL_COLORS:')
    lines.append("                        fill_rgb = FILL_COLORS[key]")
    lines.append("                        if fill_rgb:")
    lines.append("                            fc = fill_rgb.lstrip('#')")
    lines.append("                            if len(fc) == 6: fc = 'FF' + fc")
    lines.append("                            if len(fc) == 8:")
    lines.append("                                try:")
    lines.append("                                    cell.fill = PatternFill(start_color=fc, end_color=fc, fill_type='solid')")
    lines.append("                                except Exception:")
    lines.append("                                    pass")
    lines.append('                    ')
    lines.append('                    # Apply alignment')
    lines.append('                    if key in ALIGNMENTS:')
    lines.append('                        h, v, wrap = ALIGNMENTS[key]')
    lines.append('                        cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)')
    lines.append('        ')
    lines.append('        # Apply column widths')
    lines.append('        if sheet_name in COLUMN_WIDTHS:')
    lines.append('            for col_letter, width in COLUMN_WIDTHS[sheet_name].items():')
    lines.append('                ws.column_dimensions[col_letter].width = width')
    lines.append('        ')
    lines.append('        # Apply row heights')
    lines.append('        if sheet_name in ROW_HEIGHTS:')
    lines.append('            for row_num, height in ROW_HEIGHTS[sheet_name].items():')
    lines.append('                ws.row_dimensions[row_num].height = height')
    lines.append('        ')
    lines.append('        # Apply merged cells')
    lines.append('        if sheet_name in MERGED_CELLS:')
    lines.append('            for mc in MERGED_CELLS[sheet_name]:')
    lines.append('                try:')
    lines.append('                    ws.merge_cells(mc)')
    lines.append('                except Exception:')
    lines.append('                    pass')
    lines.append('    ')
    lines.append('    wb.save(output_file)')
    lines.append('    print(f"Output saved to: {output_file}")')
    lines.append('')
    lines.append('')

    # --- MAIN ---
    lines.append('def main():')
    lines.append('    """Main entry point."""')
    lines.append('    input_file = sys.argv[1] if len(sys.argv) > 1 else "input_template.xlsx"')
    lines.append('    output_file = sys.argv[2] if len(sys.argv) > 2 else "output.xlsx"')
    lines.append('    ')
    lines.append('    print(f"Loading inputs from: {input_file}")')
    lines.append('    cells = load_inputs(input_file)')
    lines.append(f'    print(f"Loaded {{len(cells)}} input values")')
    lines.append('    ')
    lines.append('    print("Calculating...")')
    lines.append('    cells = calculate(cells)')
    lines.append(f'    print(f"Calculated {{len(cells)}} cell values")')
    lines.append('    ')
    lines.append('    print(f"Writing output to: {output_file}")')
    lines.append('    write_output(cells, output_file)')
    lines.append('    print("Done!")')
    lines.append('')
    lines.append('')
    lines.append('if __name__ == "__main__":')
    lines.append('    main()')

    # Write the script
    script_path = os.path.join(output_dir, 'calculator.py')
    with open(script_path, 'w') as f:
        f.write('\n'.join(lines))

    logger.info(f"Generated calculator script: {script_path}")
    logger.info(f"  Input cells: {len(input_cells)}")
    logger.info(f"  Label cells: {len(label_cells)}")
    logger.info(f"  Formula cells: {len(formula_cells)}")

    return script_path


def _patch_indirect_calls(expr: str, sheet: str) -> str:
    """
    Patch xl_indirect calls to include cells dict and current_sheet.
    Transforms: xl_indirect(arg) -> xl_indirect(arg, cells=cells, current_sheet='Sheet')
    """
    import re
    
    def replace_indirect(match):
        # Find the matching close paren
        start = match.start()
        paren_count = 0
        i = match.end() - 1  # Position of '('
        while i < len(expr):
            if expr[i] == '(':
                paren_count += 1
            elif expr[i] == ')':
                paren_count -= 1
                if paren_count == 0:
                    # Insert cells and current_sheet args before closing paren
                    inner = expr[match.end():i]
                    return f"xl_indirect({inner}, cells=cells, current_sheet={repr(sheet)})"
            i += 1
        return match.group(0)
    
    # Simple approach: replace xl_indirect( patterns
    result = re.sub(r'xl_indirect\(', lambda m: m.group(0), expr)
    
    # More robust: manually patch
    parts = expr.split('xl_indirect(')
    if len(parts) <= 1:
        return expr
    
    result = parts[0]
    for part in parts[1:]:
        # Find matching close paren
        depth = 1
        i = 0
        while i < len(part) and depth > 0:
            if part[i] == '(':
                depth += 1
            elif part[i] == ')':
                depth -= 1
            i += 1
        
        # i now points past the closing paren
        inner = part[:i-1]
        rest = part[i-1:]
        result += f"xl_indirect({inner}, cells=cells, current_sheet={repr(sheet)}){rest[1:]}"
    
    return result
