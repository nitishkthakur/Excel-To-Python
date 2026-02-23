"""
Input Template Generator Module
================================
Generates an Excel template file that users fill in with input values.
The template mirrors the original sheet structure and includes only input cells
(hardcoded numeric values) with labels for context.
"""

import os
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .excel_parser import WorkbookInfo
from .dependency_graph import build_dependency_graph, find_unreferenced_hardcoded

logger = logging.getLogger(__name__)

# Styling constants
INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # Light yellow
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LABEL_FONT = Font(color="333333", size=10)
INPUT_FONT = Font(color="0000FF", size=10, bold=True)
INSTRUCTION_FONT = Font(color="FF0000", size=10, italic=True)


def generate_input_template(
    workbook_info: WorkbookInfo,
    output_dir: str,
    delete_unreferenced: bool = True,
) -> str:
    """
    Generate an Excel input template for the user to fill in.

    The template has the same sheet names as the original workbook.
    Within each sheet, it shows:
    - Labels in their original positions (read-only context)
    - Input cells highlighted in yellow for the user to modify
    - Default values pre-filled

    Args:
        workbook_info: Parsed workbook data
        output_dir: Directory to write the template
        delete_unreferenced: Whether to exclude unreferenced hardcoded values

    Returns:
        Path to the generated template file
    """
    os.makedirs(output_dir, exist_ok=True)

    # Determine which hardcoded cells to exclude
    adjacency, reverse_adj = build_dependency_graph(workbook_info)
    unreferenced = set()
    if delete_unreferenced:
        unreferenced = find_unreferenced_hardcoded(workbook_info, reverse_adj)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name in workbook_info.sheet_names:
        sheet_info = workbook_info.sheets[sheet_name]
        ws = wb.create_sheet(title=sheet_name)

        input_count = 0

        # Write cells
        for (row, col), cell in sorted(sheet_info.cells.items()):
            if cell.is_label:
                # Write labels for context
                out_cell = ws.cell(row=row, column=col, value=cell.value)
                out_cell.font = LABEL_FONT
                if cell.font_bold:
                    out_cell.font = Font(bold=True, color="333333", size=10)
            elif cell.is_hardcoded_number:
                key = (sheet_name, row, col)
                if key in unreferenced:
                    continue
                # Write input cell with default value, highlighted
                out_cell = ws.cell(row=row, column=col, value=cell.value)
                out_cell.fill = INPUT_FILL
                out_cell.font = INPUT_FONT
                if cell.number_format and cell.number_format != 'General':
                    out_cell.number_format = cell.number_format
                input_count += 1
            # Skip formula cells - they'll be calculated

        # Apply column widths from original
        for col_letter, width in sheet_info.column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Apply row heights from original
        for row_num, height in sheet_info.row_heights.items():
            ws.row_dimensions[row_num].height = height

        logger.info(f"  Sheet '{sheet_name}': {input_count} input cells in template")

    template_path = os.path.join(output_dir, 'input_template.xlsx')
    wb.save(template_path)
    logger.info(f"Generated input template: {template_path}")
    return template_path
