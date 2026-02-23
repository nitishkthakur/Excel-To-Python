"""
Excel Parser Module
===================
Parses an Excel workbook and extracts all cell data, formulas, formatting,
merged cells, and structural information needed for code generation.
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import re
import datetime
import logging
from dataclasses import dataclass, field
from typing import Any, Optional

logger = logging.getLogger(__name__)


@dataclass
class CellInfo:
    """Stores all information about a single cell."""
    sheet: str
    row: int
    col: int
    value: Any = None
    formula: Optional[str] = None  # Raw Excel formula (without leading '=')
    is_formula: bool = False
    is_hardcoded_number: bool = False
    is_label: bool = False
    number_format: str = "General"
    font_bold: bool = False
    font_color_rgb: Optional[str] = None
    font_color_theme: Optional[int] = None
    font_size: Optional[float] = None
    font_name: Optional[str] = None
    fill_color: Optional[str] = None
    alignment_horizontal: Optional[str] = None
    alignment_vertical: Optional[str] = None
    alignment_wrap: bool = False
    border_style: Optional[dict] = None
    col_letter: str = ""

    @property
    def key(self):
        return (self.sheet, self.row, self.col)

    @property
    def coordinate(self):
        return f"{get_column_letter(self.col)}{self.row}"

    @property
    def full_address(self):
        return f"'{self.sheet}'!{self.coordinate}"


@dataclass
class SheetInfo:
    """Stores information about a single worksheet."""
    name: str
    min_row: int = 1
    max_row: int = 1
    min_col: int = 1
    max_col: int = 1
    merged_cells: list = field(default_factory=list)
    cells: dict = field(default_factory=dict)  # (row, col) -> CellInfo
    column_widths: dict = field(default_factory=dict)
    row_heights: dict = field(default_factory=dict)


@dataclass
class WorkbookInfo:
    """Stores all parsed information about the workbook."""
    file_path: str
    sheet_names: list = field(default_factory=list)
    sheets: dict = field(default_factory=dict)  # name -> SheetInfo
    all_cells: dict = field(default_factory=dict)  # (sheet, row, col) -> CellInfo
    defined_names: dict = field(default_factory=dict)


def parse_workbook(file_path: str, skip_sheets: list = None) -> WorkbookInfo:
    """
    Parse an Excel workbook and extract all cell information.

    Args:
        file_path: Path to the .xlsx file
        skip_sheets: List of sheet names to skip

    Returns:
        WorkbookInfo with all parsed data
    """
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning)

    skip_sheets = skip_sheets or []
    logger.info(f"Parsing workbook: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=False)
    wb_data = openpyxl.load_workbook(file_path, data_only=True)

    workbook_info = WorkbookInfo(file_path=file_path)
    workbook_info.sheet_names = [n for n in wb.sheetnames if n not in skip_sheets]

    # Parse defined names
    try:
        for name, defn in wb.defined_names.items():
            workbook_info.defined_names[name] = str(defn.attr_text)
    except Exception:
        pass

    for sheet_name in workbook_info.sheet_names:
        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name]

        sheet_info = SheetInfo(
            name=sheet_name,
            min_row=ws.min_row or 1,
            max_row=ws.max_row or 1,
            min_col=ws.min_column or 1,
            max_col=ws.max_column or 1,
        )

        # Merged cells
        for mc in ws.merged_cells.ranges:
            sheet_info.merged_cells.append(str(mc))

        # Column widths
        for col_letter, dim in ws.column_dimensions.items():
            if dim.width:
                sheet_info.column_widths[col_letter] = dim.width

        # Row heights
        for row_num, dim in ws.row_dimensions.items():
            if dim.height:
                sheet_info.row_heights[row_num] = dim.height

        # Parse all cells
        for row in ws.iter_rows(min_row=sheet_info.min_row,
                                max_row=sheet_info.max_row,
                                min_col=sheet_info.min_col,
                                max_col=sheet_info.max_col):
            for cell in row:
                if cell.value is None:
                    # Still record the cell if it has formatting
                    continue

                cell_info = CellInfo(
                    sheet=sheet_name,
                    row=cell.row,
                    col=cell.column,
                    col_letter=get_column_letter(cell.column),
                    number_format=cell.number_format or "General",
                )

                # Extract value / formula
                val = cell.value
                # Skip special openpyxl objects (DataTableFormula, etc.)
                if not isinstance(val, (str, int, float, bool, datetime.datetime, datetime.date)):
                    # Try to get the computed value instead
                    try:
                        data_cell = ws_data.cell(row=cell.row, column=cell.column)
                        if data_cell.value is not None and isinstance(data_cell.value, (str, int, float, bool, datetime.datetime, datetime.date)):
                            cell_info.is_hardcoded_number = isinstance(data_cell.value, (int, float))
                            cell_info.is_label = isinstance(data_cell.value, str)
                            cell_info.value = data_cell.value
                        else:
                            continue
                    except Exception:
                        continue
                elif isinstance(val, str) and val.startswith('='):
                    cell_info.is_formula = True
                    cell_info.formula = val[1:]  # Remove '='
                    # Get the computed value from data_only workbook
                    try:
                        data_cell = ws_data.cell(row=cell.row, column=cell.column)
                        cell_info.value = data_cell.value
                    except Exception:
                        cell_info.value = None
                elif isinstance(val, (int, float)) and not isinstance(val, bool):
                    cell_info.is_hardcoded_number = True
                    cell_info.value = val
                elif isinstance(val, (datetime.datetime, datetime.date)):
                    cell_info.is_hardcoded_number = True
                    # Convert to Excel serial date for consistency
                    if isinstance(val, datetime.datetime):
                        delta = val - datetime.datetime(1899, 12, 30)
                        cell_info.value = delta.days + delta.seconds / 86400
                    else:
                        delta = val - datetime.date(1899, 12, 30)
                        cell_info.value = delta.days
                else:
                    cell_info.is_label = True
                    cell_info.value = val

                # Extract formatting
                try:
                    cell_info.font_bold = cell.font.bold if cell.font.bold else False
                    if cell.font.size:
                        cell_info.font_size = cell.font.size
                    if cell.font.name:
                        cell_info.font_name = cell.font.name
                    if cell.font.color:
                        try:
                            if cell.font.color.rgb and cell.font.color.rgb != '00000000':
                                cell_info.font_color_rgb = str(cell.font.color.rgb)
                        except (TypeError, ValueError):
                            pass
                        try:
                            if cell.font.color.theme is not None:
                                cell_info.font_color_theme = cell.font.color.theme
                        except (TypeError, ValueError):
                            pass
                except Exception:
                    pass

                try:
                    if cell.fill and cell.fill.start_color:
                        try:
                            rgb = cell.fill.start_color.rgb
                            if rgb and rgb != '00000000':
                                cell_info.fill_color = str(rgb)
                        except (TypeError, ValueError):
                            pass
                except Exception:
                    pass

                try:
                    if cell.alignment:
                        cell_info.alignment_horizontal = cell.alignment.horizontal
                        cell_info.alignment_vertical = cell.alignment.vertical
                        cell_info.alignment_wrap = cell.alignment.wrap_text or False
                except Exception:
                    pass

                sheet_info.cells[(cell.row, cell.column)] = cell_info
                workbook_info.all_cells[(sheet_name, cell.row, cell.column)] = cell_info

        workbook_info.sheets[sheet_name] = sheet_info
        logger.info(f"  Sheet '{sheet_name}': {len(sheet_info.cells)} cells "
                     f"({sum(1 for c in sheet_info.cells.values() if c.is_formula)} formulas, "
                     f"{sum(1 for c in sheet_info.cells.values() if c.is_hardcoded_number)} hardcoded, "
                     f"{sum(1 for c in sheet_info.cells.values() if c.is_label)} labels)")

    wb.close()
    wb_data.close()
    return workbook_info
