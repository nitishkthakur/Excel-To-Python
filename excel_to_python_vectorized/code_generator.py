"""
Generate a vectorised Python calculation script.

The generated script stores every cell value in a dict ``c`` keyed by
``(sheet, col_letter, row_number)`` and uses compact ``for`` loops for
groups of formulas that share the same pattern (dragged formulas).
"""

import re
import sys, os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from formula_converter import (
    FormulaConverter,
    HELPER_FUNCTIONS_CODE,
    cell_to_var_name,
    range_to_var_name,
    table_ref_to_var_name,
    col_letter_to_index,
    index_to_col_letter,
)
from .vectorizer import extract_references

# Excel formatting defaults
_DEFAULT_FONT_SIZE = 11
_ARGB_RE = re.compile(r"^[0-9A-Fa-f]{8}$")


# ------------------------------------------------------------------
# Convert FormulaConverter output (variable names) → dict-access exprs
# ------------------------------------------------------------------

def _expr_to_dict(py_expr, converter, tables):
    """Replace variable-name tokens with ``c[…]`` / ``_rng(…)`` calls."""
    result = py_expr

    # ranges first (longer names avoid partial matches)
    for sheet, c1, r1, c2, r2 in sorted(converter.referenced_ranges,
                                         key=lambda x: -len(range_to_var_name(*x))):
        var = range_to_var_name(sheet, c1, r1, c2, r2)
        repl = f"_rng(c, {repr(sheet)}, {repr(c1)}, {r1}, {repr(c2)}, {r2})"
        result = result.replace(var, repl)

    # table columns
    for tbl_name, col_name in sorted(converter.referenced_tables,
                                     key=lambda x: -len(table_ref_to_var_name(*x))):
        var = table_ref_to_var_name(tbl_name, col_name)
        if tbl_name in tables:
            tbl = tables[tbl_name]
            if col_name in tbl["columns"]:
                ci = tbl["columns"].index(col_name)
                abs_col = index_to_col_letter(col_letter_to_index(tbl["col_start"]) + ci)
                repl = (f"[c.get(({repr(tbl['sheet'])}, {repr(abs_col)}, _tr)) "
                        f"for _tr in range({tbl['data_start_row']}, {tbl['data_end_row'] + 1})]")
                result = result.replace(var, repl)

    # cells – longest name first to avoid partial replacements
    for sheet, col, row in sorted(converter.referenced_cells,
                                  key=lambda x: -len(cell_to_var_name(*x))):
        var = cell_to_var_name(sheet, col, row)
        repl = f"c.get(({repr(sheet)}, {repr(col)}, {row}))"
        result = result.replace(var, repl)

    return result


# ------------------------------------------------------------------
# Vectorised expression helpers
# ------------------------------------------------------------------

def _vec_expr_vertical(py_expr, converter, refs_abs, tables,
                       base_col_idx, base_row, current_sheet):
    """Convert expression for a vertical loop (``_r`` is the loop var)."""
    result = py_expr

    # --- ranges ---
    for sheet, c1, r1, c2, r2 in sorted(converter.referenced_ranges,
                                         key=lambda x: -len(range_to_var_name(*x))):
        var = range_to_var_name(sheet, c1, r1, c2, r2)
        # look up absolute info
        ri = _find_range_ref(refs_abs, sheet, c1, r1, c2, r2)
        r1e = str(r1) if (ri and ri.row_abs) else _row_expr("_r", r1 - base_row)
        r2e = str(r2) if (ri and ri.end_row_abs) else _row_expr("_r", r2 - base_row)
        repl = f"_rng(c, {repr(sheet)}, {repr(c1)}, {r1e}, {repr(c2)}, {r2e})"
        result = result.replace(var, repl)

    # --- table columns (unchanged in vertical loops) ---
    for tbl_name, col_name in sorted(converter.referenced_tables,
                                     key=lambda x: -len(table_ref_to_var_name(*x))):
        var = table_ref_to_var_name(tbl_name, col_name)
        if tbl_name in tables:
            tbl = tables[tbl_name]
            if col_name in tbl["columns"]:
                ci = tbl["columns"].index(col_name)
                abs_col = index_to_col_letter(col_letter_to_index(tbl["col_start"]) + ci)
                repl = (f"[c.get(({repr(tbl['sheet'])}, {repr(abs_col)}, _tr)) "
                        f"for _tr in range({tbl['data_start_row']}, {tbl['data_end_row'] + 1})]")
                result = result.replace(var, repl)

    # --- cells ---
    for sheet, col, row in sorted(converter.referenced_cells,
                                  key=lambda x: -len(cell_to_var_name(*x))):
        var = cell_to_var_name(sheet, col, row)
        ri = _find_cell_ref(refs_abs, sheet, col, row)
        row_e = str(row) if (ri and ri.row_abs) else _row_expr("_r", row - base_row)
        repl = f"c.get(({repr(sheet)}, {repr(col)}, {row_e}))"
        result = result.replace(var, repl)

    return result


def _vec_expr_horizontal(py_expr, converter, refs_abs, tables,
                         base_col_idx, base_row, current_sheet):
    """Convert expression for a horizontal loop (``_ci`` is the loop var)."""
    result = py_expr

    # --- ranges ---
    for sheet, c1, r1, c2, r2 in sorted(converter.referenced_ranges,
                                         key=lambda x: -len(range_to_var_name(*x))):
        var = range_to_var_name(sheet, c1, r1, c2, r2)
        ri = _find_range_ref(refs_abs, sheet, c1, r1, c2, r2)
        c1e = repr(c1) if (ri and ri.col_abs) else _col_expr("_ci", col_letter_to_index(c1) - base_col_idx)
        c2e = repr(c2) if (ri and ri.end_col_abs) else _col_expr("_ci", col_letter_to_index(c2) - base_col_idx)
        repl = f"_rng(c, {repr(sheet)}, {c1e}, {r1}, {c2e}, {r2})"
        result = result.replace(var, repl)

    # --- table columns ---
    for tbl_name, col_name in sorted(converter.referenced_tables,
                                     key=lambda x: -len(table_ref_to_var_name(*x))):
        var = table_ref_to_var_name(tbl_name, col_name)
        if tbl_name in tables:
            tbl = tables[tbl_name]
            if col_name in tbl["columns"]:
                ci = tbl["columns"].index(col_name)
                abs_col = index_to_col_letter(col_letter_to_index(tbl["col_start"]) + ci)
                repl = (f"[c.get(({repr(tbl['sheet'])}, {repr(abs_col)}, _tr)) "
                        f"for _tr in range({tbl['data_start_row']}, {tbl['data_end_row'] + 1})]")
                result = result.replace(var, repl)

    # --- cells ---
    for sheet, col, row in sorted(converter.referenced_cells,
                                  key=lambda x: -len(cell_to_var_name(*x))):
        var = cell_to_var_name(sheet, col, row)
        ri = _find_cell_ref(refs_abs, sheet, col, row)
        col_e = repr(col) if (ri and ri.col_abs) else _col_expr("_ci", col_letter_to_index(col) - base_col_idx)
        repl = f"c.get(({repr(sheet)}, {col_e}, {row}))"
        result = result.replace(var, repl)

    return result


def _row_expr(var, offset):
    if offset == 0:
        return var
    if offset > 0:
        return f"{var} + {offset}"
    return f"{var} - {-offset}"


def _col_expr(var, offset):
    if offset == 0:
        return f"_cl({var})"
    if offset > 0:
        return f"_cl({var} + {offset})"
    return f"_cl({var} - {-offset})"


def _find_cell_ref(refs_abs, sheet, col, row):
    for r in refs_abs:
        if r.kind == "cell" and r.sheet == sheet and r.col == col and r.row == row:
            return r
    return None


def _find_range_ref(refs_abs, sheet, c1, r1, c2, r2):
    for r in refs_abs:
        if (r.kind == "range" and r.sheet == sheet
                and r.col == c1 and r.row == r1
                and r.end_col == c2 and r.end_row == r2):
            return r
    return None


# ------------------------------------------------------------------
# Main code-generation entry point
# ------------------------------------------------------------------

def generate_vectorized_script(sheets, tables, ordered_items,
                               hardcoded_cells, formula_cells, config,
                               external_files):
    """Return the full text of the generated vectorised Python script."""
    lines = _header_lines(external_files)
    lines += _main_open(sheets, hardcoded_cells, external_files)
    lines += _compute_section(ordered_items, tables, formula_cells)
    lines += _write_output(sheets, hardcoded_cells, formula_cells)
    lines += _main_close()
    return "\n".join(lines)


# ------------------------------------------------------------------
# Header (imports + helpers)
# ------------------------------------------------------------------

def _header_lines(external_files):
    L = []
    L.append('"""')
    L.append("Auto-generated *vectorised* Python script from Excel workbook.")
    L.append("")
    L.append("Usage:")
    L.append("    python calculate.py <input_excel> <output_excel>")
    if external_files:
        L.append("")
        L.append("External workbook references are resolved through")
        L.append("input_files_config.json (must sit next to this script).")
    L.append('"""')
    L.append("")
    L.append("import sys")
    L.append("import os")
    L.append("import json")
    L.append("from openpyxl import Workbook, load_workbook")
    L.append("from openpyxl.styles import Font, PatternFill, Alignment")
    L.append("from openpyxl.utils import get_column_letter")
    L.append("")
    L.append("# ---- Excel helper functions ----")
    L.append(HELPER_FUNCTIONS_CODE)
    L.append("")
    # column helpers
    L.append("def _ci(col):")
    L.append('    """Column letter → 1-based index."""')
    L.append("    r = 0")
    L.append("    for ch in col.upper():")
    L.append("        r = r * 26 + (ord(ch) - 64)")
    L.append("    return r")
    L.append("")
    L.append("def _cl(idx):")
    L.append('    """1-based index → column letter."""')
    L.append('    r = ""')
    L.append("    while idx > 0:")
    L.append("        idx, rem = divmod(idx - 1, 26)")
    L.append("        r = chr(65 + rem) + r")
    L.append("    return r")
    L.append("")
    L.append("def _rng(c, sheet, c1, r1, c2, r2):")
    L.append('    """Build a 2-D list from the cell dict."""')
    L.append("    ci1, ci2 = _ci(c1) if isinstance(c1, str) else c1, _ci(c2) if isinstance(c2, str) else c2")
    L.append("    rows = []")
    L.append("    for r in range(int(r1), int(r2) + 1):")
    L.append("        rows.append([c.get((sheet, _cl(ci), r)) for ci in range(ci1, ci2 + 1)])")
    L.append("    return rows")
    L.append("")
    return L


# ------------------------------------------------------------------
# main() – read inputs
# ------------------------------------------------------------------

def _main_open(sheets, hardcoded_cells, external_files):
    L = []
    L.append("def main(input_file, output_file):")
    L.append('    """Read inputs, compute formulas, write output."""')
    L.append("    inp_wb = load_workbook(input_file, data_only=True)")
    L.append("")
    L.append("    c = {}  # cell store: (sheet, col, row) → value")
    L.append("")

    # ---- external workbooks ----
    if external_files:
        L.append("    # ---- Load external workbooks from config ----")
        L.append("    _cfg_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),")
        L.append("                             'input_files_config.json')")
        L.append("    with open(_cfg_path) as _fp:")
        L.append("        _ext_cfg = json.load(_fp)")
        L.append("    _ext_wbs = {}")
        L.append("    for _name, _path in _ext_cfg.items():")
        L.append("        if _path and os.path.exists(_path):")
        L.append("            _ext_wbs[_name] = load_workbook(_path, data_only=True)")
        L.append("")
        # read external cells into c with ext_ prefix
        for ext_file, usages in external_files.items():
            safe = re.sub(r"[^a-zA-Z0-9]", "_", ext_file)
            L.append(f"    if {repr(ext_file)} in _ext_wbs:")
            L.append(f"        _ewb = _ext_wbs[{repr(ext_file)}]")
            # collect sheets referenced
            ext_sheets = set()
            for _s, _c, _r, formula in usages:
                refs = extract_references(formula, _s)
                for ref in refs:
                    if ref.external_file == ext_file:
                        ext_sheets.add(ref.sheet)
            for es in sorted(ext_sheets):
                safe_es = re.sub(r"[^a-zA-Z0-9]", "_", es)
                L.append(f"        if {repr(es)} in _ewb.sheetnames:")
                L.append(f"            _ews = _ewb[{repr(es)}]")
                L.append(f"            for _row in _ews.iter_rows():")
                L.append(f"                for _cell in _row:")
                L.append(f"                    if _cell.value is not None:")
                L.append(f"                        c[({repr(ext_file + '|' + es)}, "
                         f"get_column_letter(_cell.column), _cell.row)] = _cell.value")
            L.append("")

    # ---- read internal hardcoded cells ----
    L.append("    # ---- Read input (hardcoded) values ----")
    input_sheets = sorted({s for s, *_ in hardcoded_cells})
    for sn in input_sheets:
        safe = re.sub(r"[^a-zA-Z0-9]", "_", sn)
        L.append(f"    _ws_{safe} = inp_wb[{repr(sn)}] "
                 f"if {repr(sn)} in inp_wb.sheetnames else None")
    L.append("")

    for sheet, col, row, val, _ci in hardcoded_cells:
        safe = re.sub(r"[^a-zA-Z0-9]", "_", sheet)
        ref = f"{col}{row}"
        default = repr(val)
        L.append(
            f"    c[({repr(sheet)}, {repr(col)}, {row})] = "
            f"_ws_{safe}[{repr(ref)}].value "
            f"if _ws_{safe} is not None and _ws_{safe}[{repr(ref)}].value is not None "
            f"else {default}"
        )
    L.append("")
    return L


# ------------------------------------------------------------------
# Compute section
# ------------------------------------------------------------------

def _compute_section(ordered_items, tables, formula_cells):
    L = []
    L.append("    # ---- Compute formulas (vectorised where possible) ----")
    L.append("")

    for item in ordered_items:
        if isinstance(item, dict):  # group
            L += _emit_group(item, tables)
        else:  # single
            L += _emit_single(item, tables)

    return L


def _emit_single(item, tables):
    """Emit code for a single (non-vectorised) formula."""
    sheet, col, row, formula, _ci = item
    converter = FormulaConverter(sheet, tables)
    py_expr = converter.convert(formula)
    dict_expr = _expr_to_dict(py_expr, converter, tables)

    # Resolve external refs in the expression
    dict_expr = _patch_external_refs(dict_expr, formula, sheet)

    L = []
    L.append(f"    try:")
    L.append(f"        c[({repr(sheet)}, {repr(col)}, {row})] = {dict_expr}")
    L.append(f"    except Exception:")
    L.append(f"        c[({repr(sheet)}, {repr(col)}, {row})] = None")
    return L


def _emit_group(group, tables):
    """Emit a ``for`` loop for a vectorised group."""
    cells = group["cells"]
    direction = group["direction"]
    sheet, col0, row0, formula0, _ci0 = cells[0]
    col0_idx = col_letter_to_index(col0)

    # Use first cell's formula as representative
    converter = FormulaConverter(sheet, tables)
    py_expr = converter.convert(formula0)
    refs_abs = extract_references(formula0, sheet)

    L = []

    if direction == "vertical":
        rows = [r for _s, _c, r, _f, _ci in cells]
        row_list = _compact_range_repr(rows)

        vec_expr = _vec_expr_vertical(
            py_expr, converter, refs_abs, tables,
            col0_idx, row0, sheet)
        vec_expr = _patch_external_refs(vec_expr, formula0, sheet)

        L.append(f"    # Vectorised: {sheet}!{col0}{rows[0]}:{col0}{rows[-1]}")
        L.append(f"    for _r in {row_list}:")
        L.append(f"        try:")
        L.append(f"            c[({repr(sheet)}, {repr(col0)}, _r)] = {vec_expr}")
        L.append(f"        except Exception:")
        L.append(f"            c[({repr(sheet)}, {repr(col0)}, _r)] = None")
    else:  # horizontal
        cols = [c for _s, c, _r, _f, _ci in cells]
        col_indices = [col_letter_to_index(c) for c in cols]
        ci_list = _compact_range_repr(col_indices)

        vec_expr = _vec_expr_horizontal(
            py_expr, converter, refs_abs, tables,
            col0_idx, row0, sheet)
        vec_expr = _patch_external_refs(vec_expr, formula0, sheet)

        L.append(f"    # Vectorised: {sheet}!{cols[0]}{row0}:{cols[-1]}{row0}")
        L.append(f"    for _ci in {ci_list}:")
        L.append(f"        try:")
        L.append(f"            c[({repr(sheet)}, _cl(_ci), {row0})] = {vec_expr}")
        L.append(f"        except Exception:")
        L.append(f"            c[({repr(sheet)}, _cl(_ci), {row0})] = None")

    return L


def _compact_range_repr(nums):
    """Represent a list of consecutive ints as ``range(a, b)`` if possible."""
    if len(nums) >= 2 and nums == list(range(nums[0], nums[-1] + 1)):
        return f"range({nums[0]}, {nums[-1] + 1})"
    return repr(nums)


def _patch_external_refs(expr, formula, current_sheet):
    """Replace external-file reference remnants in the expression.

    The FormulaConverter doesn't understand ``[Book.xlsx]Sheet!A1`` syntax.
    It leaves ``[Book.xlsx]`` as literal text and converts the rest as a
    cross-sheet ref (``s_Sheet_A1``).  After ``_expr_to_dict`` has already
    converted ``s_Sheet_A1`` → ``c.get(('Sheet','A',1))``, the expression
    still contains the ``[Book.xlsx]`` prefix.

    This function:
      1. Strips every ``[filename]`` prefix.
      2. Re-keys the ``c.get(('Sheet', …))`` call to
         ``c.get(('filename|Sheet', …))`` so it looks up the external cell
         store instead.
    """
    refs = extract_references(formula, current_sheet)
    for ref in refs:
        if not ref.external_file:
            continue
        ext_key = f"{ref.external_file}|{ref.sheet}"
        esc_file = re.escape(f"[{ref.external_file}]")

        if ref.kind == "cell":
            # Pattern left after _expr_to_dict:
            #   [ExtData.xlsx]c.get(('Prices', 'A', 1))
            old_dict = f"c.get(({repr(ref.sheet)}, {repr(ref.col)}, {ref.row}))"
            new_dict = f"c.get(({repr(ext_key)}, {repr(ref.col)}, {ref.row}))"

            # Try to replace [file]c.get(…) first
            pat = esc_file + re.escape(old_dict)
            if re.search(pat, expr):
                expr = re.sub(pat, new_dict, expr, count=1)
            else:
                # May also appear with the raw variable name (before _expr_to_dict)
                var = cell_to_var_name(ref.sheet, ref.col, ref.row)
                pat2 = esc_file + re.escape(var)
                repl2 = new_dict
                expr = re.sub(pat2, repl2, expr, count=1)

        elif ref.kind == "range":
            old_rng = (f"_rng(c, {repr(ref.sheet)}, {repr(ref.col)}, "
                       f"{ref.row}, {repr(ref.end_col)}, {ref.end_row})")
            new_rng = (f"_rng(c, {repr(ext_key)}, {repr(ref.col)}, "
                       f"{ref.row}, {repr(ref.end_col)}, {ref.end_row})")
            pat = esc_file + re.escape(old_rng)
            if re.search(pat, expr):
                expr = re.sub(pat, new_rng, expr, count=1)
            else:
                var = range_to_var_name(ref.sheet, ref.col, ref.row,
                                        ref.end_col, ref.end_row)
                pat2 = esc_file + re.escape(var)
                expr = re.sub(pat2, new_rng, expr, count=1)

    # Final safety: strip any remaining [filename] prefixes
    expr = re.sub(r"\[([^\]]+\.xlsx?)\]", "", expr, flags=re.IGNORECASE)
    return expr


# ------------------------------------------------------------------
# Write output workbook
# ------------------------------------------------------------------

def _write_output(sheets, hardcoded_cells, formula_cells):
    L = []
    L.append("")
    L.append("    # ---- Write output workbook ----")
    L.append("    out_wb = Workbook()")
    L.append("    if 'Sheet' in out_wb.sheetnames:")
    L.append("        del out_wb['Sheet']")
    L.append("")

    for sheet_name, sheet_data in sheets.items():
        safe = re.sub(r"[^a-zA-Z0-9]", "_", sheet_name)
        L.append(f"    ws_{safe} = out_wb.create_sheet({repr(sheet_name)})")

        # column widths
        for col_key, w in sheet_data["col_widths"].items():
            if isinstance(col_key, str):
                ck = col_key
            elif isinstance(col_key, int):
                from openpyxl.utils import get_column_letter
                ck = get_column_letter(col_key)
            else:
                continue
            L.append(f"    ws_{safe}.column_dimensions[{repr(ck)}].width = {w}")

        # row heights
        for rk, h in sheet_data["row_heights"].items():
            L.append(f"    ws_{safe}.row_dimensions[{rk}].height = {h}")

        # merged cells
        for mr in sheet_data["merged_cells"]:
            L.append(f"    ws_{safe}.merge_cells({repr(mr)})")

        L.append("")

        # collect all cells for this sheet (hardcoded + formula)
        all_cells = {}
        for s, col, row, val, ci in hardcoded_cells:
            if s == sheet_name:
                all_cells[(col, row)] = ci
        for s, col, row, _f, ci in formula_cells:
            if s == sheet_name:
                all_cells[(col, row)] = ci

        for (col, row), ci in sorted(all_cells.items(), key=lambda x: (x[0][1], x[0][0])):
            ref = f"{col}{row}"
            L.append(f"    ws_{safe}[{repr(ref)}] = c.get(({repr(sheet_name)}, {repr(col)}, {row}))")

            # formatting
            nf = ci.get("number_format")
            if nf and nf != "General":
                L.append(f"    ws_{safe}[{repr(ref)}].number_format = {repr(nf)}")

            fi = ci.get("font")
            if fi:
                fa = []
                if fi.get("bold"):
                    fa.append("bold=True")
                if fi.get("italic"):
                    fa.append("italic=True")
                if fi.get("size") and fi["size"] != _DEFAULT_FONT_SIZE:
                    fa.append(f"size={fi['size']}")
                if fi.get("color"):
                    fa.append(f"color={repr(fi['color'])}")
                if fa:
                    L.append(f"    ws_{safe}[{repr(ref)}].font = Font({', '.join(fa)})")

            fc = ci.get("fill_color")
            if fc:
                L.append(f"    ws_{safe}[{repr(ref)}].fill = PatternFill("
                         f"start_color={repr(fc)}, end_color={repr(fc)}, fill_type='solid')")

            ai = ci.get("alignment")
            if ai:
                aa = []
                if ai.get("horizontal"):
                    aa.append(f"horizontal={repr(ai['horizontal'])}")
                if ai.get("vertical"):
                    aa.append(f"vertical={repr(ai['vertical'])}")
                if ai.get("wrap_text"):
                    aa.append("wrap_text=True")
                if aa:
                    L.append(f"    ws_{safe}[{repr(ref)}].alignment = Alignment({', '.join(aa)})")

        L.append("")

    L.append("    out_wb.save(output_file)")
    L.append("    print(f'Output saved to {output_file}')")
    return L


# ------------------------------------------------------------------
# main() close
# ------------------------------------------------------------------

def _main_close():
    L = []
    L.append("")
    L.append("")
    L.append('if __name__ == "__main__":')
    L.append("    if len(sys.argv) < 3:")
    L.append('        print("Usage: python calculate.py <input_excel> <output_excel>")')
    L.append("        sys.exit(1)")
    L.append("    main(sys.argv[1], sys.argv[2])")
    L.append("")
    return L
