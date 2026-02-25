"""
Orchestrates the full Excel → vectorised-Python conversion pipeline.

Steps:
  1. Parse the Excel workbook (reuses the original parser).
  2. Classify cells as formula / hardcoded.
  3. Detect vectorisable groups.
  4. Order groups + singles by dependency.
  5. Generate vectorised Python script.
  6. Generate input template.
  7. Generate input_files_config.json for external workbook refs.
  8. Generate analysis report.
"""

import json
import os
import re
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import yaml

from excel_to_python import (
    load_config,
    parse_workbook,
    classify_cells,
    find_all_references,
    filter_hardcoded_cells,
    DEFAULT_FONT_SIZE,
)

from .vectorizer import (
    group_formulas,
    order_items,
    discover_external_files,
    analyse_references,
)
from .code_generator import generate_vectorized_script


# ------------------------------------------------------------------
# Input-template generation (same logic as the original converter)
# ------------------------------------------------------------------

def _generate_input_template(sheets, hardcoded_cells, output_path):
    """Create the input-template Excel file."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    sheets_with_inputs = {s for s, *_ in hardcoded_cells}

    for sheet_name in sheets:
        if sheet_name not in sheets_with_inputs:
            continue
        ws = wb.create_sheet(sheet_name)
        sd = sheets[sheet_name]
        for ck, w in sd["col_widths"].items():
            if isinstance(ck, str):
                ws.column_dimensions[ck].width = w
            elif isinstance(ck, int):
                ws.column_dimensions[get_column_letter(ck)].width = w
        for rk, h in sd["row_heights"].items():
            ws.row_dimensions[rk].height = h

    for sheet, col, row, val, ci in hardcoded_cells:
        ws = wb[sheet]
        ref = f"{col}{row}"
        ws[ref] = val
        nf = ci.get("number_format")
        if nf and nf != "General":
            ws[ref].number_format = nf

        fi = ci.get("font")
        if fi:
            kw = {}
            if fi.get("bold"):
                kw["bold"] = True
            if fi.get("italic"):
                kw["italic"] = True
            if fi.get("size") and fi["size"] != DEFAULT_FONT_SIZE:
                kw["size"] = fi["size"]
            if fi.get("color"):
                kw["color"] = fi["color"]
            if kw:
                ws[ref].font = Font(**kw)

        fc = ci.get("fill_color")
        if fc:
            ws[ref].fill = PatternFill(
                start_color=fc, end_color=fc, fill_type="solid"
            )

        ai = ci.get("alignment")
        if ai:
            kw = {}
            if ai.get("horizontal"):
                kw["horizontal"] = ai["horizontal"]
            if ai.get("vertical"):
                kw["vertical"] = ai["vertical"]
            if ai.get("wrap_text"):
                kw["wrap_text"] = True
            if kw:
                ws[ref].alignment = Alignment(**kw)

    wb.save(output_path)
    return output_path


# ------------------------------------------------------------------
# External-files config
# ------------------------------------------------------------------

def _generate_ext_config(external_files, output_path):
    """Write ``input_files_config.json`` listing external workbooks."""
    cfg = {}
    for fname in sorted(external_files):
        cfg[fname] = ""  # user fills in the real path
    with open(output_path, "w") as fp:
        json.dump(cfg, fp, indent=2)
    return output_path


# ------------------------------------------------------------------
# Analysis report
# ------------------------------------------------------------------

def _generate_report(analysis, formula_cells, groups, singles,
                     sheets, external_files, output_path):
    """Create an Excel report with reference analysis."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # --- Summary sheet ---
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "Metric"
    ws_sum["B1"] = "Value"
    ws_sum["A1"].font = Font(bold=True)
    ws_sum["B1"].font = Font(bold=True)

    total_formulas = len(formula_cells)
    total_groups = len(groups)
    total_vectorised = sum(len(g["cells"]) for g in groups)
    total_singles = len(singles)
    total_cross_sheet = len(analysis["cross_sheet"])
    total_external = len(analysis["external"])
    ext_file_count = len(external_files)

    rows = [
        ("Total formula cells", total_formulas),
        ("Vectorised groups", total_groups),
        ("Formulas in vectorised groups", total_vectorised),
        ("Non-vectorised (individual) formulas", total_singles),
        ("Cross-sheet references", total_cross_sheet),
        ("External-file references", total_external),
        ("Distinct external files", ext_file_count),
        ("Sheets in workbook", len(sheets)),
    ]
    for i, (m, v) in enumerate(rows, start=2):
        ws_sum[f"A{i}"] = m
        ws_sum[f"B{i}"] = v
    ws_sum.column_dimensions["A"].width = 40
    ws_sum.column_dimensions["B"].width = 15

    # --- Vectorised Groups sheet ---
    ws_grp = wb.create_sheet("Vectorised Groups")
    headers = ["Group #", "Direction", "Sheet", "Start Cell", "End Cell",
               "Cell Count", "Representative Formula"]
    for ci, h in enumerate(headers, 1):
        ws_grp.cell(row=1, column=ci, value=h).font = Font(bold=True)
    r = 2
    for gi, g in enumerate(groups, 1):
        cells = g["cells"]
        s0, c0, r0, f0, _ = cells[0]
        s1, c1, r1, f1, _ = cells[-1]
        ws_grp.cell(row=r, column=1, value=gi)
        ws_grp.cell(row=r, column=2, value=g["direction"])
        ws_grp.cell(row=r, column=3, value=s0)
        ws_grp.cell(row=r, column=4, value=f"{c0}{r0}")
        ws_grp.cell(row=r, column=5, value=f"{c1}{r1}")
        ws_grp.cell(row=r, column=6, value=len(cells))
        ws_grp.cell(row=r, column=7, value=f0)
        r += 1
    for ci in range(1, len(headers) + 1):
        ws_grp.column_dimensions[get_column_letter(ci)].width = 22

    # --- Cross-Sheet References sheet ---
    ws_cs = wb.create_sheet("Cross-Sheet Refs")
    cs_headers = ["Source Sheet", "Cell", "Formula", "Target Sheet", "Ref Type"]
    for ci, h in enumerate(cs_headers, 1):
        ws_cs.cell(row=1, column=ci, value=h).font = Font(bold=True)
    for i, rec in enumerate(analysis["cross_sheet"], start=2):
        ws_cs.cell(row=i, column=1, value=rec["sheet"])
        ws_cs.cell(row=i, column=2, value=rec["cell"])
        ws_cs.cell(row=i, column=3, value=rec["formula"])
        ws_cs.cell(row=i, column=4, value=rec["target_sheet"])
        ws_cs.cell(row=i, column=5, value=rec["ref_type"])
    for ci in range(1, len(cs_headers) + 1):
        ws_cs.column_dimensions[get_column_letter(ci)].width = 25

    # --- External References sheet ---
    ws_ext = wb.create_sheet("External Refs")
    ext_headers = ["Source Sheet", "Cell", "Formula",
                   "External File", "External Sheet", "Ref Type"]
    for ci, h in enumerate(ext_headers, 1):
        ws_ext.cell(row=1, column=ci, value=h).font = Font(bold=True)
    for i, rec in enumerate(analysis["external"], start=2):
        ws_ext.cell(row=i, column=1, value=rec["sheet"])
        ws_ext.cell(row=i, column=2, value=rec["cell"])
        ws_ext.cell(row=i, column=3, value=rec["formula"])
        ws_ext.cell(row=i, column=4, value=rec["external_file"])
        ws_ext.cell(row=i, column=5, value=rec["external_sheet"])
        ws_ext.cell(row=i, column=6, value=rec["ref_type"])
    for ci in range(1, len(ext_headers) + 1):
        ws_ext.column_dimensions[get_column_letter(ci)].width = 25

    # --- Per-Sheet Breakdown ---
    ws_ps = wb.create_sheet("Per-Sheet Breakdown")
    ps_headers = ["Sheet", "Total Formulas", "Vectorised", "Individual",
                  "Cross-Sheet Refs Out", "External Refs"]
    for ci, h in enumerate(ps_headers, 1):
        ws_ps.cell(row=1, column=ci, value=h).font = Font(bold=True)

    sheet_stats = {}
    for sn in sheets:
        sheet_stats[sn] = {
            "formulas": 0, "vectorised": 0, "individual": 0,
            "cross_sheet": 0, "external": 0,
        }
    for s, *_ in formula_cells:
        if s in sheet_stats:
            sheet_stats[s]["formulas"] += 1
    for g in groups:
        for s, *_ in g["cells"]:
            if s in sheet_stats:
                sheet_stats[s]["vectorised"] += 1
    for s, *_ in singles:
        if s in sheet_stats:
            sheet_stats[s]["individual"] += 1
    for rec in analysis["cross_sheet"]:
        s = rec["sheet"]
        if s in sheet_stats:
            sheet_stats[s]["cross_sheet"] += 1
    for rec in analysis["external"]:
        s = rec["sheet"]
        if s in sheet_stats:
            sheet_stats[s]["external"] += 1

    for i, (sn, st) in enumerate(sheet_stats.items(), start=2):
        ws_ps.cell(row=i, column=1, value=sn)
        ws_ps.cell(row=i, column=2, value=st["formulas"])
        ws_ps.cell(row=i, column=3, value=st["vectorised"])
        ws_ps.cell(row=i, column=4, value=st["individual"])
        ws_ps.cell(row=i, column=5, value=st["cross_sheet"])
        ws_ps.cell(row=i, column=6, value=st["external"])
    for ci in range(1, len(ps_headers) + 1):
        ws_ps.column_dimensions[get_column_letter(ci)].width = 22

    wb.save(output_path)
    return output_path


# ------------------------------------------------------------------
# Main entry point
# ------------------------------------------------------------------

def convert_excel_to_python_vectorized(excel_path, config_path=None,
                                       output_dir=None):
    """Convert an Excel workbook to a vectorised Python script.

    Produces:
      * ``calculate.py``           – the vectorised calculation script
      * ``input_template.xlsx``    – template for user inputs
      * ``input_files_config.json``– config for external workbook paths
                                     (only if external refs exist)
      * ``analysis_report.xlsx``   – reference analysis report

    Returns (script_path, template_path, config_path_or_None, report_path).
    """
    config = load_config(config_path)
    delete_unreferenced = config.get("delete_unreferenced_hardcoded_values", False)

    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(excel_path) or ".", "output")
    os.makedirs(output_dir, exist_ok=True)

    wb = load_workbook(excel_path)
    sheets, tables = parse_workbook(wb)
    formula_cells, hardcoded_cells = classify_cells(sheets, tables)

    # references & filtering
    ref_cells, ref_ranges, ref_tables = find_all_references(formula_cells, tables)
    hardcoded_cells = filter_hardcoded_cells(
        hardcoded_cells, ref_cells, ref_ranges, ref_tables,
        tables, delete_unreferenced,
    )

    # vectorisation
    groups, singles = group_formulas(formula_cells)
    ordered_items = order_items(groups, singles, tables)

    # external files
    external_files = discover_external_files(formula_cells)

    # analysis
    analysis = analyse_references(formula_cells)

    # --- generate script ---
    script_text = generate_vectorized_script(
        sheets, tables, ordered_items, hardcoded_cells,
        formula_cells, config, external_files,
    )
    script_path = os.path.join(output_dir, "calculate.py")
    with open(script_path, "w") as fp:
        fp.write(script_text)

    # --- input template ---
    template_path = os.path.join(output_dir, "input_template.xlsx")
    _generate_input_template(sheets, hardcoded_cells, template_path)

    # --- external config ---
    ext_config_path = None
    if external_files:
        ext_config_path = os.path.join(output_dir, "input_files_config.json")
        _generate_ext_config(external_files, ext_config_path)

    # --- report ---
    report_path = os.path.join(output_dir, "analysis_report.xlsx")
    _generate_report(analysis, formula_cells, groups, singles,
                     sheets, external_files, report_path)

    wb.close()

    print(f"Generated vectorised script : {script_path}")
    print(f"Generated input template    : {template_path}")
    if ext_config_path:
        print(f"Generated external config   : {ext_config_path}")
    print(f"Generated analysis report   : {report_path}")
    print(f"\nTo run: python {script_path} {template_path} <output.xlsx>")

    return script_path, template_path, ext_config_path, report_path
