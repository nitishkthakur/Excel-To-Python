"""generate_coded_excels_for_sampling.py

Takes an Excel file path as input and produces one output Excel file per
sampling algorithm (smart_random, full, column_n).  In each output file the
cells that are included by the corresponding sampling algorithm are highlighted
with a distinct background colour, making it easy to see which rows / columns
were selected by each strategy.

Usage:
    python generate_coded_excels_for_sampling.py <path_to_excel> [output_dir]

Output files are written to the same directory as the input file by default,
named  <basename>_smart_random.xlsx,  <basename>_full.xlsx  and
<basename>_column_n.xlsx.
"""

import os
import sys

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Make mcp_server importable when the script lives in the repo root.
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "mcp_server"))

from excel_reader_smart_sampler import (  # noqa: E402
    detect_regions,
    sample_row_indices,
    DEFAULT_SAMPLE_ROWS,
)
from column_n import _find_label_column, DEFAULT_NUM_COLUMNS  # noqa: E402


# ---------------------------------------------------------------------------
# Highlight colours — one per algorithm
# ---------------------------------------------------------------------------

FILL_SMART_RANDOM = PatternFill(start_color="FFFF99", end_color="FFFF99",
                                fill_type="solid")   # yellow
FILL_FULL         = PatternFill(start_color="99FF99", end_color="99FF99",
                                fill_type="solid")   # green
FILL_COLUMN_N     = PatternFill(start_color="99CCFF", end_color="99CCFF",
                                fill_type="solid")   # blue


# ---------------------------------------------------------------------------
# Helpers: compute the set of (row, col) cells selected by each algorithm
# ---------------------------------------------------------------------------

def _sampled_cells_smart_random(ws,
                                max_rows: int = DEFAULT_SAMPLE_ROWS
                                ) -> set[tuple[int, int]]:
    """Return (row, col) pairs that the smart_random sampler would include."""
    cells: set[tuple[int, int]] = set()
    for reg in detect_regions(ws):
        row_indices = set(sample_row_indices(reg, ws, max_rows))
        # The header row is always shown even if sample_row_indices omits it.
        if reg.header_row is not None:
            row_indices.add(reg.header_row)
        for r in row_indices:
            for c in range(reg.min_col, reg.max_col + 1):
                cells.add((r, c))
    return cells


def _sampled_cells_full(ws) -> set[tuple[int, int]]:
    """Return all (row, col) pairs inside detected data regions (full load)."""
    cells: set[tuple[int, int]] = set()
    for reg in detect_regions(ws):
        for r in range(reg.min_row, reg.max_row + 1):
            for c in range(reg.min_col, reg.max_col + 1):
                cells.add((r, c))
    return cells


def _sampled_cells_column_n(ws,
                            num_columns: int = DEFAULT_NUM_COLUMNS
                            ) -> set[tuple[int, int]]:
    """Return (row, col) pairs in the vertical strip selected by column_n."""
    cells: set[tuple[int, int]] = set()
    for reg in detect_regions(ws):
        label_col = _find_label_column(ws, reg)
        strip_max_col = min(label_col + num_columns, reg.max_col)
        for r in range(reg.min_row, reg.max_row + 1):
            for c in range(label_col, strip_max_col + 1):
                cells.add((r, c))
    return cells


# ---------------------------------------------------------------------------
# Core: produce a colour-coded copy of the workbook for each algorithm
# ---------------------------------------------------------------------------

def generate_coded_excels(input_path: str,
                          output_dir: str | None = None) -> list[str]:
    """Produce one colour-coded Excel file per sampling algorithm.

    Parameters
    ----------
    input_path:
        Path to the source ``.xlsx`` file.
    output_dir:
        Directory where output files are written.  Defaults to the same
        directory as *input_path*.

    Returns
    -------
    list[str]
        Paths of the generated output files.
    """
    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(input_path))
    os.makedirs(output_dir, exist_ok=True)

    base = os.path.splitext(os.path.basename(input_path))[0]

    algorithms = [
        ("smart_random", FILL_SMART_RANDOM, _sampled_cells_smart_random),
        ("full",         FILL_FULL,         _sampled_cells_full),
        ("column_n",     FILL_COLUMN_N,     _sampled_cells_column_n),
    ]

    output_paths: list[str] = []

    for algo_name, fill, cells_fn in algorithms:
        # Load a fresh, independent copy of the workbook for each algorithm.
        # data_only=False keeps formulas intact and is required by detect_regions.
        wb = load_workbook(input_path, data_only=False)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for (r, c) in cells_fn(ws):
                ws.cell(row=r, column=c).fill = fill

        out_path = os.path.join(output_dir, f"{base}_{algo_name}.xlsx")
        wb.save(out_path)
        wb.close()
        output_paths.append(out_path)
        print(f"Saved: {out_path}")

    return output_paths


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python generate_coded_excels_for_sampling.py "
              "<excel_path> [output_dir]")
        sys.exit(1)

    excel_path = sys.argv[1]
    out_dir = sys.argv[2] if len(sys.argv) > 2 else None
    generate_coded_excels(excel_path, out_dir)
