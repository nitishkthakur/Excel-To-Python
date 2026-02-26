"""Tests for the excel_to_mapping module (tabular format)."""

import os
import shutil
import sys
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from tests.create_sample_workbook import create_sample_workbook
from excel_to_mapping.mapper import (
    generate_mapping_report,
    _build_all_referenced_cells,
    _classify_formula_cells,
    _build_sheet_rows,
    COLUMNS,
)
from excel_to_python import parse_workbook, classify_cells


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _create_simple_workbook(path):
    """Workbook with inputs, intermediate calcs, and final outputs."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Price"
    ws["A2"] = 10
    ws["B1"] = "Qty"
    ws["B2"] = 5
    ws["C1"] = "Subtotal"
    ws["C2"] = "=A2*B2"       # intermediate calc (referenced by D2)
    ws["D1"] = "Total"
    ws["D2"] = "=C2+100"      # output (not referenced by anything)
    wb.save(path)
    wb.close()


def _create_vertical_drag_workbook(path):
    """Workbook with a dragged formula column and a SUM output."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws["A1"] = "Revenue"
    ws["B1"] = "Cost"
    ws["C1"] = "Profit"
    for r in range(2, 7):
        ws[f"A{r}"] = 1000 * r
        ws[f"B{r}"] = 600 * r
        ws[f"C{r}"] = f"=A{r}-B{r}"   # dragged formula (intermediate)
    ws["C7"] = "=SUM(C2:C6)"          # output
    wb.save(path)
    wb.close()


def _create_multi_sheet_workbook(path):
    """Workbook with two sheets and cross-sheet references."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Input"
    ws1["A1"] = "Val"
    ws1["A2"] = 42

    ws2 = wb.create_sheet("Calc")
    ws2["A1"] = "Doubled"
    ws2["A2"] = "=Input!A2*2"      # calc (referenced by B2)
    ws2["B1"] = "Final"
    ws2["B2"] = "=A2+10"           # output
    wb.save(path)
    wb.close()


def _create_horizontal_drag_workbook(path):
    """Workbook with horizontally-dragged formulas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast"
    for ci in range(5):
        col = chr(66 + ci)  # B..F
        ws[f"{col}1"] = 100 + ci * 10
        ws[f"{col}2"] = f"={col}1*1.1"  # dragged horizontally
    ws["A1"] = "Base"
    ws["A2"] = "Grown"
    wb.save(path)
    wb.close()


def _create_all_formulas_workbook(path):
    """Workbook where every cell is a formula (no hardcoded inputs)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Formulas"
    ws["A1"] = "=1+1"
    ws["A2"] = "=A1*2"
    wb.save(path)
    wb.close()


def _create_no_formulas_workbook(path):
    """Workbook with no formulas at all."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Static"
    ws["A1"] = "Hello"
    ws["B1"] = 42
    wb.save(path)
    wb.close()


def _read_report_rows(ws):
    """Read all data rows from a tabular report sheet as list of dicts."""
    headers = [ws.cell(row=1, column=c).value for c in range(1, 20)]
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 20)]
        if all(v is None for v in vals):
            continue
        rows.append(dict(zip(headers, vals)))
    return rows


# ---------------------------------------------------------------------------
# Unit tests: classification helpers
# ---------------------------------------------------------------------------

class TestBuildAllReferencedCells(unittest.TestCase):
    def test_simple_references(self):
        formulas = [
            ("Sheet1", "C", 2, "=A2*B2", {}),
            ("Sheet1", "D", 2, "=C2+100", {}),
        ]
        refs = _build_all_referenced_cells(formulas)
        self.assertIn(("Sheet1", "A", 2), refs)
        self.assertIn(("Sheet1", "B", 2), refs)
        self.assertIn(("Sheet1", "C", 2), refs)
        self.assertNotIn(("Sheet1", "D", 2), refs)

    def test_range_references_expanded(self):
        formulas = [
            ("Sheet1", "C", 7, "=SUM(C2:C6)", {}),
        ]
        refs = _build_all_referenced_cells(formulas)
        for r in range(2, 7):
            self.assertIn(("Sheet1", "C", r), refs)

    def test_cross_sheet_reference(self):
        formulas = [
            ("Calc", "A", 2, "=Input!A2*2", {}),
        ]
        refs = _build_all_referenced_cells(formulas)
        self.assertIn(("Input", "A", 2), refs)


class TestClassifyFormulaCells(unittest.TestCase):
    def test_intermediate_vs_output(self):
        formulas = [
            ("S", "C", 2, "=A2*B2", {}),   # referenced by D2 → calculation
            ("S", "D", 2, "=C2+100", {}),   # not referenced → output
        ]
        calcs, outputs = _classify_formula_cells(formulas)
        calc_keys = {(s, c, r) for s, c, r, *_ in calcs}
        out_keys = {(s, c, r) for s, c, r, *_ in outputs}
        self.assertIn(("S", "C", 2), calc_keys)
        self.assertIn(("S", "D", 2), out_keys)

    def test_all_outputs_when_no_inter_deps(self):
        formulas = [
            ("S", "C", 2, "=A2*B2", {}),
            ("S", "D", 2, "=A2+B2", {}),
        ]
        calcs, outputs = _classify_formula_cells(formulas)
        self.assertEqual(len(calcs), 0)
        self.assertEqual(len(outputs), 2)


# ---------------------------------------------------------------------------
# Unit tests: _build_sheet_rows
# ---------------------------------------------------------------------------

class TestBuildSheetRows(unittest.TestCase):
    def test_vertical_group_collapsed(self):
        """Five dragged cells should produce one grouped row."""
        formula_cells = [
            ("Sales", "C", r, f"=A{r}-B{r}", {})
            for r in range(2, 7)
        ]
        rows = _build_sheet_rows("Sales", [], formula_cells)
        grouped = [r for r in rows if r["GroupID"] is not None]
        self.assertEqual(len(grouped), 1)
        self.assertEqual(grouped[0]["GroupSize"], 5)
        self.assertEqual(grouped[0]["GroupDirection"], "vertical")

    def test_single_formula_not_grouped(self):
        formula_cells = [("S", "D", 2, "=SUM(C2:C6)", {})]
        rows = _build_sheet_rows("S", [], formula_cells)
        self.assertEqual(len(rows), 1)
        self.assertIsNone(rows[0]["GroupID"])
        self.assertIn("SUM", rows[0]["Formula"])

    def test_empty_input(self):
        self.assertEqual(_build_sheet_rows("S", [], []), [])


class TestBuildSheetRowsIntegration(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp()
        cls.path = os.path.join(cls.tmpdir, "simple.xlsx")
        _create_simple_workbook(cls.path)

        wb = load_workbook(cls.path)
        cls.sheets, cls.tables = parse_workbook(wb)
        cls.formula_cells, cls.hardcoded_cells = classify_cells(
            cls.sheets, cls.tables)
        wb.close()

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir)

    def test_inputs_present(self):
        rows = _build_sheet_rows(
            "Data", self.hardcoded_cells, self.formula_cells)
        input_rows = [r for r in rows if r["Type"] == "Input"]
        values = [r["Value"] for r in input_rows]
        self.assertIn(10, values)
        self.assertIn(5, values)

    def test_calculations_and_outputs_split(self):
        rows = _build_sheet_rows(
            "Data", self.hardcoded_cells, self.formula_cells)
        calc_cells = [r["Cell"] for r in rows if r["Type"] == "Calculation"]
        out_cells = [r["Cell"] for r in rows if r["Type"] == "Output"]
        self.assertIn("C2", calc_cells)
        self.assertIn("D2", out_cells)


# ---------------------------------------------------------------------------
# End-to-end: report generation
# ---------------------------------------------------------------------------

class TestGenerateMappingReportSimple(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp()
        cls.wb_path = os.path.join(cls.tmpdir, "simple.xlsx")
        _create_simple_workbook(cls.wb_path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir)

    def test_report_created(self):
        rpt = os.path.join(self.tmpdir, "report.xlsx")
        result = generate_mapping_report(self.wb_path, output_path=rpt)
        self.assertTrue(os.path.exists(result))

    def test_report_has_sheet(self):
        rpt = os.path.join(self.tmpdir, "report2.xlsx")
        generate_mapping_report(self.wb_path, output_path=rpt)
        wb = load_workbook(rpt)
        self.assertIn("Data", wb.sheetnames)
        wb.close()

    def test_report_has_tabular_header(self):
        rpt = os.path.join(self.tmpdir, "report3.xlsx")
        generate_mapping_report(self.wb_path, output_path=rpt)
        wb = load_workbook(rpt)
        ws = wb["Data"]
        headers = [ws.cell(row=1, column=c).value
                   for c in range(1, len(COLUMNS) + 1)]
        self.assertEqual(headers, COLUMNS)
        wb.close()

    def test_report_contains_all_types(self):
        rpt = os.path.join(self.tmpdir, "report4.xlsx")
        generate_mapping_report(self.wb_path, output_path=rpt)
        wb = load_workbook(rpt)
        ws = wb["Data"]
        data = _read_report_rows(ws)
        types = {r["Type"] for r in data}
        self.assertIn("Input", types)
        self.assertIn("Calculation", types)
        self.assertIn("Output", types)
        wb.close()

    def test_report_has_metadata_sheet(self):
        rpt = os.path.join(self.tmpdir, "report5.xlsx")
        generate_mapping_report(self.wb_path, output_path=rpt)
        wb = load_workbook(rpt)
        self.assertIn("_Metadata", wb.sheetnames)
        ws = wb["_Metadata"]
        self.assertEqual(ws.cell(row=1, column=1).value, "SheetName")
        self.assertEqual(ws.cell(row=2, column=1).value, "Data")
        wb.close()


class TestGenerateMappingReportVerticalDrag(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp()
        cls.wb_path = os.path.join(cls.tmpdir, "vertical.xlsx")
        _create_vertical_drag_workbook(cls.wb_path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir)

    def test_dragged_formulas_collapsed(self):
        rpt = os.path.join(self.tmpdir, "report.xlsx")
        generate_mapping_report(self.wb_path, output_path=rpt)
        wb = load_workbook(rpt)
        ws = wb["Sales"]
        data = _read_report_rows(ws)
        grouped = [r for r in data if r.get("GroupSize") is not None]
        self.assertEqual(len(grouped), 1)
        self.assertEqual(grouped[0]["GroupSize"], 5)
        self.assertEqual(grouped[0]["GroupDirection"], "vertical")
        wb.close()

    def test_sum_is_output(self):
        rpt = os.path.join(self.tmpdir, "report_out.xlsx")
        generate_mapping_report(self.wb_path, output_path=rpt)
        wb = load_workbook(rpt)
        ws = wb["Sales"]
        data = _read_report_rows(ws)
        output_rows = [r for r in data if r["Type"] == "Output"]
        formulas = [r["Formula"] for r in output_rows if r["Formula"]]
        self.assertTrue(
            any("SUM" in f for f in formulas),
            "SUM formula should appear as an Output row",
        )
        wb.close()


class TestGenerateMappingReportMultiSheet(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp()
        cls.wb_path = os.path.join(cls.tmpdir, "multi.xlsx")
        _create_multi_sheet_workbook(cls.wb_path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir)

    def test_all_sheets_in_report(self):
        rpt = os.path.join(self.tmpdir, "report.xlsx")
        generate_mapping_report(self.wb_path, output_path=rpt)
        wb = load_workbook(rpt)
        self.assertIn("Input", wb.sheetnames)
        self.assertIn("Calc", wb.sheetnames)
        self.assertIn("_Metadata", wb.sheetnames)
        wb.close()

    def test_subset_sheets(self):
        rpt = os.path.join(self.tmpdir, "report_sub.xlsx")
        generate_mapping_report(self.wb_path, sheet_names=["Calc"],
                                output_path=rpt)
        wb = load_workbook(rpt)
        self.assertIn("Calc", wb.sheetnames)
        self.assertNotIn("Input", wb.sheetnames)
        wb.close()

    def test_cross_sheet_calc_vs_output(self):
        rpt = os.path.join(self.tmpdir, "report_cs.xlsx")
        generate_mapping_report(self.wb_path, output_path=rpt)
        wb = load_workbook(rpt)
        ws = wb["Calc"]
        data = _read_report_rows(ws)
        calc_cells = [r["Cell"] for r in data if r["Type"] == "Calculation"]
        out_cells = [r["Cell"] for r in data if r["Type"] == "Output"]
        self.assertIn("A2", calc_cells)
        self.assertIn("B2", out_cells)
        wb.close()


class TestGenerateMappingReportHorizontalDrag(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp()
        cls.wb_path = os.path.join(cls.tmpdir, "horiz.xlsx")
        _create_horizontal_drag_workbook(cls.wb_path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir)

    def test_horizontal_group_detected(self):
        rpt = os.path.join(self.tmpdir, "report.xlsx")
        generate_mapping_report(self.wb_path, output_path=rpt)
        wb = load_workbook(rpt)
        ws = wb["Forecast"]
        data = _read_report_rows(ws)
        grouped = [r for r in data if r.get("GroupSize") is not None]
        self.assertEqual(len(grouped), 1)
        self.assertEqual(grouped[0]["GroupSize"], 5)
        self.assertEqual(grouped[0]["GroupDirection"], "horizontal")
        wb.close()


class TestGenerateMappingReportEdgeCases(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp()

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir)

    def test_no_formulas(self):
        path = os.path.join(self.tmpdir, "nof.xlsx")
        _create_no_formulas_workbook(path)
        rpt = os.path.join(self.tmpdir, "rpt_nof.xlsx")
        generate_mapping_report(path, output_path=rpt)
        wb = load_workbook(rpt)
        ws = wb["Static"]
        data = _read_report_rows(ws)
        types = {r["Type"] for r in data}
        # Only inputs, no calculations or outputs
        self.assertEqual(types, {"Input"})
        wb.close()

    def test_all_formulas(self):
        path = os.path.join(self.tmpdir, "allf.xlsx")
        _create_all_formulas_workbook(path)
        rpt = os.path.join(self.tmpdir, "rpt_allf.xlsx")
        generate_mapping_report(path, output_path=rpt)
        wb = load_workbook(rpt)
        ws = wb["Formulas"]
        data = _read_report_rows(ws)
        types = {r["Type"] for r in data}
        # No inputs; only Calculation and/or Output
        self.assertNotIn("Input", types)
        self.assertTrue(len(data) > 0)
        wb.close()

    def test_nonexistent_sheet_ignored(self):
        path = os.path.join(self.tmpdir, "simple2.xlsx")
        _create_simple_workbook(path)
        rpt = os.path.join(self.tmpdir, "rpt_ne.xlsx")
        generate_mapping_report(path, sheet_names=["NoSuchSheet"],
                                output_path=rpt)
        wb = load_workbook(rpt)
        self.assertNotIn("NoSuchSheet", wb.sheetnames)
        self.assertTrue(len(wb.sheetnames) >= 1)
        wb.close()

    def test_include_flag_defaults_true(self):
        path = os.path.join(self.tmpdir, "simple_inc.xlsx")
        _create_simple_workbook(path)
        rpt = os.path.join(self.tmpdir, "rpt_inc.xlsx")
        generate_mapping_report(path, output_path=rpt)
        wb = load_workbook(rpt)
        ws = wb["Data"]
        data = _read_report_rows(ws)
        for row in data:
            self.assertTrue(row["IncludeFlag"])
        wb.close()


class TestGenerateMappingReportSampleWorkbook(unittest.TestCase):
    """End-to-end with the shared sample workbook used by vectorized tests."""

    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp()
        cls.sample_path = os.path.join(cls.tmpdir, "sample.xlsx")
        create_sample_workbook(cls.sample_path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir)

    def test_all_sheets_present(self):
        rpt = os.path.join(self.tmpdir, "rpt_sample.xlsx")
        generate_mapping_report(self.sample_path, output_path=rpt)
        wb = load_workbook(rpt)
        self.assertIn("Inputs", wb.sheetnames)
        self.assertIn("Summary", wb.sheetnames)
        self.assertIn("Rates", wb.sheetnames)
        self.assertIn("_Metadata", wb.sheetnames)
        wb.close()

    def test_inputs_sheet_has_inputs(self):
        rpt = os.path.join(self.tmpdir, "rpt_sample2.xlsx")
        generate_mapping_report(self.sample_path, output_path=rpt)
        wb = load_workbook(rpt)
        ws = wb["Inputs"]
        data = _read_report_rows(ws)
        input_rows = [r for r in data if r["Type"] == "Input"]
        values = [r["Value"] for r in input_rows]
        self.assertTrue(any(isinstance(v, (int, float)) for v in values))
        wb.close()

    def test_default_output_path(self):
        """When no output_path given, report goes to ./output/."""
        rpt = generate_mapping_report(self.sample_path)
        self.assertTrue(os.path.exists(rpt))
        self.assertIn("mapping_report.xlsx", rpt)
        os.remove(rpt)


if __name__ == "__main__":
    unittest.main()
