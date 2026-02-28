"""
Tests for the keyword search fetcher (fetcher_keyword.py).
"""

import os
import sys
import json
import pytest

# Make the mcp_server package and tests directory importable
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "mcp_server"))
sys.path.insert(0, os.path.dirname(__file__))

from create_sample_workbook import create_sample_workbook
from fetcher_keyword import search_keywords


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def sample_wb(tmp_path_factory):
    """Create the sample workbook once per test module."""
    path = str(tmp_path_factory.mktemp("data") / "sample.xlsx")
    create_sample_workbook(path)
    return path


@pytest.fixture(scope="module")
def multi_patch_wb(tmp_path_factory):
    """Create a workbook with multiple data patches separated by blank rows."""
    from openpyxl import Workbook
    path = str(tmp_path_factory.mktemp("data") / "multi_patch.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Patch 1: rows 1-5 (header + 4 data rows)
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    for i in range(2, 6):
        ws[f"A{i}"] = f"Item{i-1}"
        ws[f"B{i}"] = i * 10

    # blank row 6

    # Patch 2: rows 7-16 (header + 9 data rows)
    ws["A7"] = "Category"
    ws["B7"] = "Amount"
    ws["C7"] = "Tax"
    for i in range(8, 17):
        ws[f"A{i}"] = f"Cat{i-7}"
        ws[f"B{i}"] = i * 100
        ws[f"C{i}"] = f"=B{i}*0.1"

    wb.save(path)
    wb.close()
    return path


# ---------------------------------------------------------------------------
# Basic functionality
# ---------------------------------------------------------------------------

class TestKeywordSearch:

    def test_basic_search_finds_keyword(self, sample_wb):
        """Searching for a known value should return matches."""
        result = search_keywords(sample_wb, ["Widget A"])
        assert "keywords" in result
        assert "matches" in result
        assert len(result["matches"]) > 0

    def test_search_returns_correct_sheet(self, sample_wb):
        """Matches should identify the correct sheet name."""
        result = search_keywords(sample_wb, ["Widget A"])
        sheet_names = [m["sheet_name"] for m in result["matches"]]
        assert "Inputs" in sheet_names

    def test_search_returns_matched_cells(self, sample_wb):
        """Result should include matched_cells with address and keyword."""
        result = search_keywords(sample_wb, ["Widget A"])
        for sheet_match in result["matches"]:
            for reg in sheet_match["regions"]:
                for cell in reg["matched_cells"]:
                    assert "address" in cell
                    assert "keyword" in cell
                    assert "value" in cell

    def test_search_returns_full_rows(self, sample_wb):
        """Each matched row should have full row data with values."""
        result = search_keywords(sample_wb, ["Widget A"])
        for sheet_match in result["matches"]:
            for reg in sheet_match["regions"]:
                assert len(reg["rows"]) > 0
                for row in reg["rows"]:
                    assert "row_number" in row
                    assert "values" in row
                    assert len(row["values"]) > 0

    def test_search_returns_full_columns(self, sample_wb):
        """Each matched column should have full column data with values."""
        result = search_keywords(sample_wb, ["Widget A"])
        for sheet_match in result["matches"]:
            for reg in sheet_match["regions"]:
                assert len(reg["columns"]) > 0
                for col in reg["columns"]:
                    assert "column_letter" in col
                    assert "values" in col
                    assert len(col["values"]) > 0

    def test_search_returns_headers(self, sample_wb):
        """Result should include headers when the patch has a header row."""
        result = search_keywords(sample_wb, ["Widget A"])
        for sheet_match in result["matches"]:
            for reg in sheet_match["regions"]:
                assert "headers" in reg

    def test_search_case_insensitive(self, sample_wb):
        """Keyword matching should be case-insensitive."""
        result_lower = search_keywords(sample_wb, ["widget a"])
        result_upper = search_keywords(sample_wb, ["WIDGET A"])
        assert len(result_lower["matches"]) > 0
        assert len(result_upper["matches"]) > 0

    def test_search_no_match(self, sample_wb):
        """Searching for a nonexistent keyword should return empty matches."""
        result = search_keywords(sample_wb, ["NONEXISTENT_KEYWORD_12345"])
        assert result["matches"] == []

    def test_search_multiple_keywords(self, sample_wb):
        """Searching with multiple keywords should find matches for each."""
        result = search_keywords(sample_wb, ["Widget A", "Widget B"])
        assert len(result["matches"]) > 0
        # All matched cells should reference one of the keywords
        for sheet_match in result["matches"]:
            for reg in sheet_match["regions"]:
                for cell in reg["matched_cells"]:
                    assert cell["keyword"] in ["Widget A", "Widget B"]

    def test_search_specific_sheet(self, sample_wb):
        """Restricting search to a specific sheet should only return that sheet."""
        result = search_keywords(sample_wb, ["Widget A"], sheet_name="Inputs")
        assert len(result["matches"]) > 0
        for m in result["matches"]:
            assert m["sheet_name"] == "Inputs"

    def test_search_wrong_sheet_no_results(self, sample_wb):
        """Searching in a sheet that doesn't contain the keyword returns empty."""
        result = search_keywords(sample_wb, ["Widget A"], sheet_name="Rates")
        assert result["matches"] == []


# ---------------------------------------------------------------------------
# Cross-sheet and formula matching
# ---------------------------------------------------------------------------

class TestKeywordSearchFormulas:

    def test_search_finds_formula_references(self, sample_wb):
        """Keywords should match inside formula strings too."""
        # Summary sheet has formulas like =Inputs!D6
        result = search_keywords(sample_wb, ["Inputs!"])
        assert len(result["matches"]) > 0
        sheet_names = [m["sheet_name"] for m in result["matches"]]
        assert "Summary" in sheet_names

    def test_matched_rows_contain_formulas(self, sample_wb):
        """Rows returned for formula matches should include formula info."""
        result = search_keywords(sample_wb, ["Widget A"])
        for sheet_match in result["matches"]:
            if sheet_match["sheet_name"] == "Inputs":
                for reg in sheet_match["regions"]:
                    # Row with Widget A (row 2) has formula in column D
                    has_formula = False
                    for row in reg["rows"]:
                        if row.get("formulas"):
                            has_formula = True
                    assert has_formula


# ---------------------------------------------------------------------------
# Multi-patch workbook
# ---------------------------------------------------------------------------

class TestKeywordSearchMultiPatch:

    def test_search_correct_patch(self, multi_patch_wb):
        """Search should identify the correct patch for a keyword."""
        result = search_keywords(multi_patch_wb, ["Item1"])
        assert len(result["matches"]) > 0
        for sheet_match in result["matches"]:
            for reg in sheet_match["regions"]:
                for cell in reg["matched_cells"]:
                    assert cell["keyword"] == "Item1"

    def test_search_across_patches(self, multi_patch_wb):
        """Keywords in different patches should return results from both."""
        result = search_keywords(multi_patch_wb, ["Item1", "Cat1"])
        assert len(result["matches"]) > 0
        # Should find matches in regions corresponding to both patches
        all_regions = []
        for sheet_match in result["matches"]:
            all_regions.extend(sheet_match["regions"])
        assert len(all_regions) >= 2


# ---------------------------------------------------------------------------
# Empty / edge cases
# ---------------------------------------------------------------------------

class TestKeywordSearchEdgeCases:

    def test_empty_sheet(self, tmp_path):
        """Keyword search should handle empty sheets gracefully."""
        from openpyxl import Workbook
        path = str(tmp_path / "empty.xlsx")
        wb = Workbook()
        wb.save(path)
        wb.close()

        result = search_keywords(path, ["anything"])
        assert result["matches"] == []

    def test_empty_keywords_list(self, sample_wb):
        """Empty keyword list should return no matches."""
        result = search_keywords(sample_wb, [])
        assert result["matches"] == []

    def test_result_is_json_serializable(self, sample_wb):
        """The result dict should be JSON-serializable."""
        result = search_keywords(sample_wb, ["Widget A"])
        serialized = json.dumps(result, default=str)
        parsed = json.loads(serialized)
        assert parsed["keywords"] == ["Widget A"]


# ---------------------------------------------------------------------------
# Row and column deduplication
# ---------------------------------------------------------------------------

class TestKeywordSearchDedup:

    def test_rows_are_deduplicated(self, sample_wb):
        """If a keyword appears multiple times on the same row,
        the row should only appear once."""
        # "Widget" appears in "Widget A", "Widget B", "Widget C" on different rows
        result = search_keywords(sample_wb, ["Widget"])
        for sheet_match in result["matches"]:
            for reg in sheet_match["regions"]:
                row_nums = [r["row_number"] for r in reg["rows"]]
                assert len(row_nums) == len(set(row_nums)), "Rows should be deduplicated"
