"""Tests for the excel_to_python converter (end-to-end)."""

import os
import sys
import unittest
import tempfile
import shutil

from openpyxl import load_workbook

# Ensure project root is on the path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from excel_to_python import (
    load_config,
    parse_workbook,
    classify_cells,
    find_all_references,
    filter_hardcoded_cells,
    build_dependency_order,
    convert_excel_to_python,
)
from tests.create_sample_workbook import create_sample_workbook


class TestLoadConfig(unittest.TestCase):
    def test_default_config(self):
        config = load_config(None)
        self.assertFalse(config["delete_unreferenced_hardcoded_values"])

    def test_custom_config(self):
        with tempfile.NamedTemporaryFile(mode="w", suffix=".yaml", delete=False) as f:
            f.write("delete_unreferenced_hardcoded_values: true\n")
            f.flush()
            config = load_config(f.name)
        self.assertTrue(config["delete_unreferenced_hardcoded_values"])
        os.unlink(f.name)


class TestParseWorkbook(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp()
        cls.sample_path = os.path.join(cls.tmpdir, "sample.xlsx")
        create_sample_workbook(cls.sample_path)
        cls.wb = load_workbook(cls.sample_path)
        cls.sheets, cls.tables = parse_workbook(cls.wb)

    @classmethod
    def tearDownClass(cls):
        cls.wb.close()
        shutil.rmtree(cls.tmpdir)

    def test_sheets_parsed(self):
        self.assertIn("Inputs", self.sheets)
        self.assertIn("Summary", self.sheets)
        self.assertIn("Rates", self.sheets)

    def test_formula_cells_detected(self):
        cells = self.sheets["Inputs"]["cells"]
        # D2 should be a formula
        d2 = cells.get(("D", 2))
        self.assertIsNotNone(d2)
        self.assertTrue(str(d2["value"]).startswith("="))

    def test_hardcoded_cells_detected(self):
        cells = self.sheets["Inputs"]["cells"]
        # B2 is a hardcoded price
        b2 = cells.get(("B", 2))
        self.assertIsNotNone(b2)
        self.assertEqual(b2["value"], 10.50)


class TestClassifyCells(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp()
        cls.sample_path = os.path.join(cls.tmpdir, "sample.xlsx")
        create_sample_workbook(cls.sample_path)
        cls.wb = load_workbook(cls.sample_path)
        cls.sheets, cls.tables = parse_workbook(cls.wb)

    @classmethod
    def tearDownClass(cls):
        cls.wb.close()
        shutil.rmtree(cls.tmpdir)

    def test_classification(self):
        formula_cells, hardcoded_cells = classify_cells(self.sheets, self.tables)
        # Should have formulas
        self.assertTrue(len(formula_cells) > 0)
        # Should have hardcoded values
        self.assertTrue(len(hardcoded_cells) > 0)

        # Check specific cells
        formula_coords = [(s, c, r) for s, c, r, f, ci in formula_cells]
        self.assertIn(("Inputs", "D", 2), formula_coords)  # =B2*C2
        self.assertIn(("Summary", "B", 3), formula_coords)  # =Inputs!D6

        hardcoded_coords = [(s, c, r) for s, c, r, v, ci in hardcoded_cells]
        self.assertIn(("Inputs", "B", 2), hardcoded_coords)  # 10.50


class TestFilterHardcoded(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp()
        cls.sample_path = os.path.join(cls.tmpdir, "sample.xlsx")
        create_sample_workbook(cls.sample_path)
        cls.wb = load_workbook(cls.sample_path)
        cls.sheets, cls.tables = parse_workbook(cls.wb)

    @classmethod
    def tearDownClass(cls):
        cls.wb.close()
        shutil.rmtree(cls.tmpdir)

    def test_no_filter(self):
        formula_cells, hardcoded_cells = classify_cells(self.sheets, self.tables)
        ref_cells, ref_ranges, ref_tables = find_all_references(formula_cells, self.tables)
        filtered = filter_hardcoded_cells(
            hardcoded_cells, ref_cells, ref_ranges, ref_tables, self.tables, False
        )
        self.assertEqual(len(filtered), len(hardcoded_cells))

    def test_filter_unreferenced(self):
        formula_cells, hardcoded_cells = classify_cells(self.sheets, self.tables)
        ref_cells, ref_ranges, ref_tables = find_all_references(formula_cells, self.tables)
        filtered = filter_hardcoded_cells(
            hardcoded_cells, ref_cells, ref_ranges, ref_tables, self.tables, True
        )
        # Filtered should be a subset
        self.assertLessEqual(len(filtered), len(hardcoded_cells))
        # Labels (A1, A2, etc.) should be filtered out since they're not referenced
        filtered_coords = [(s, c, r) for s, c, r, v, ci in filtered]
        # B2 (price) IS referenced by D2=B2*C2
        self.assertIn(("Inputs", "B", 2), filtered_coords)


class TestEndToEnd(unittest.TestCase):
    """End-to-end test: convert Excel → run generated script → validate output."""

    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp()
        cls.sample_path = os.path.join(cls.tmpdir, "sample.xlsx")
        create_sample_workbook(cls.sample_path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir)

    def test_conversion_produces_files(self):
        output_dir = os.path.join(self.tmpdir, "output")
        script_path, template_path = convert_excel_to_python(
            self.sample_path, output_dir=output_dir
        )
        self.assertTrue(os.path.exists(script_path))
        self.assertTrue(os.path.exists(template_path))
        self.assertTrue(script_path.endswith(".py"))
        self.assertTrue(template_path.endswith(".xlsx"))

    def test_generated_script_is_valid_python(self):
        output_dir = os.path.join(self.tmpdir, "output_syntax")
        script_path, _ = convert_excel_to_python(
            self.sample_path, output_dir=output_dir
        )
        with open(script_path) as f:
            code = f.read()
        # Should compile without syntax errors
        compile(code, script_path, "exec")

    def test_generated_script_runs(self):
        output_dir = os.path.join(self.tmpdir, "output_run")
        script_path, template_path = convert_excel_to_python(
            self.sample_path, output_dir=output_dir
        )
        result_path = os.path.join(output_dir, "result.xlsx")

        # Execute the generated script
        import subprocess
        proc = subprocess.run(
            [sys.executable, script_path, template_path, result_path],
            capture_output=True, text=True, timeout=30,
        )
        self.assertEqual(proc.returncode, 0, f"Script failed:\n{proc.stderr}")
        self.assertTrue(os.path.exists(result_path))

        # Validate output
        result_wb = load_workbook(result_path, data_only=True)
        self.assertIn("Inputs", result_wb.sheetnames)
        self.assertIn("Summary", result_wb.sheetnames)

        # Check calculated values
        ws_inputs = result_wb["Inputs"]
        # D2 = B2 * C2 = 10.50 * 5 = 52.50
        d2_val = ws_inputs["D2"].value
        self.assertIsNotNone(d2_val)
        self.assertAlmostEqual(float(d2_val), 52.50, places=2)

        # D6 = SUM(D2:D4) = 52.50 + 75.00 + 72.50 = 200.00
        d6_val = ws_inputs["D6"].value
        self.assertIsNotNone(d6_val)
        self.assertAlmostEqual(float(d6_val), 200.00, places=2)

        # D9 = D6 * B8 = 200 * 0.08 = 16.00
        d9_val = ws_inputs["D9"].value
        self.assertIsNotNone(d9_val)
        self.assertAlmostEqual(float(d9_val), 16.00, places=2)

        # D10 = D6 + D9 = 216.00
        d10_val = ws_inputs["D10"].value
        self.assertIsNotNone(d10_val)
        self.assertAlmostEqual(float(d10_val), 216.00, places=2)

        # Summary sheet
        ws_summary = result_wb["Summary"]
        # B5 = Inputs!D10 = 216.00
        b5_val = ws_summary["B5"].value
        self.assertIsNotNone(b5_val)
        self.assertAlmostEqual(float(b5_val), 216.00, places=2)

        # B9 = B5 - B8 = 216 - 21.6 = 194.40
        b9_val = ws_summary["B9"].value
        self.assertIsNotNone(b9_val)
        self.assertAlmostEqual(float(b9_val), 194.40, places=2)

        result_wb.close()

    def test_input_template_has_correct_sheets(self):
        output_dir = os.path.join(self.tmpdir, "output_template")
        _, template_path = convert_excel_to_python(
            self.sample_path, output_dir=output_dir
        )
        wb = load_workbook(template_path)
        # Template should have sheets with input values
        self.assertIn("Inputs", wb.sheetnames)
        self.assertIn("Summary", wb.sheetnames)

        # Check that input values are in the template
        ws = wb["Inputs"]
        self.assertEqual(ws["B2"].value, 10.50)
        self.assertEqual(ws["C2"].value, 5)
        wb.close()

    def test_delete_unreferenced_config(self):
        """Test that the delete_unreferenced_hardcoded_values config works."""
        config_path = os.path.join(self.tmpdir, "test_config.yaml")
        with open(config_path, "w") as f:
            f.write("delete_unreferenced_hardcoded_values: true\n")

        output_dir = os.path.join(self.tmpdir, "output_filtered")
        script_path, template_path = convert_excel_to_python(
            self.sample_path, config_path=config_path, output_dir=output_dir
        )

        # The template should have fewer cells (labels removed)
        wb = load_workbook(template_path)
        ws = wb["Inputs"]
        # B2 (price) should still be there (referenced by formula)
        self.assertIsNotNone(ws["B2"].value)
        # A1 (label "Item") should NOT be there (not referenced by any formula)
        self.assertIsNone(ws["A1"].value)
        wb.close()


if __name__ == "__main__":
    unittest.main()
