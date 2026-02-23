"""
Create a sample Excel workbook for testing the Excel-to-Python converter.

This workbook has:
- Sheet1: Input values and basic calculations
- Sheet2: Calculations referencing Sheet1
- Sheet3: Table-based calculations
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


def create_sample_workbook(output_path):
    """Create a multi-sheet workbook with formulas, cross-sheet refs, and tables."""
    wb = Workbook()

    # ---- Sheet1: Inputs and basic arithmetic ----
    ws1 = wb.active
    ws1.title = "Inputs"

    # Headers
    ws1["A1"] = "Item"
    ws1["B1"] = "Price"
    ws1["C1"] = "Quantity"
    ws1["D1"] = "Subtotal"
    ws1["A1"].font = Font(bold=True)
    ws1["B1"].font = Font(bold=True)
    ws1["C1"].font = Font(bold=True)
    ws1["D1"].font = Font(bold=True)

    # Data rows (hardcoded inputs)
    ws1["A2"] = "Widget A"
    ws1["B2"] = 10.50
    ws1["C2"] = 5
    ws1["D2"] = "=B2*C2"  # Formula: 52.50

    ws1["A3"] = "Widget B"
    ws1["B3"] = 25.00
    ws1["C3"] = 3
    ws1["D3"] = "=B3*C3"  # Formula: 75.00

    ws1["A4"] = "Widget C"
    ws1["B4"] = 7.25
    ws1["C4"] = 10
    ws1["D4"] = "=B4*C4"  # Formula: 72.50

    # Summary
    ws1["A6"] = "Total"
    ws1["A6"].font = Font(bold=True)
    ws1["D6"] = "=SUM(D2:D4)"  # Formula: 200.00
    ws1["D6"].number_format = '#,##0.00'

    ws1["A7"] = "Average"
    ws1["D7"] = "=AVERAGE(D2:D4)"

    ws1["A8"] = "Tax Rate"
    ws1["B8"] = 0.08  # 8% tax rate (hardcoded input)
    ws1["B8"].number_format = '0%'

    ws1["A9"] = "Tax Amount"
    ws1["D9"] = "=D6*B8"  # Formula using another cell

    ws1["A10"] = "Grand Total"
    ws1["A10"].font = Font(bold=True, size=12)
    ws1["D10"] = "=D6+D9"
    ws1["D10"].number_format = '#,##0.00'
    ws1["D10"].font = Font(bold=True)

    # Formatting
    ws1.column_dimensions["A"].width = 15
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 15

    # ---- Sheet2: Cross-sheet references ----
    ws2 = wb.create_sheet("Summary")

    ws2["A1"] = "Summary Report"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.merge_cells("A1:C1")

    ws2["A3"] = "Subtotal from Inputs"
    ws2["B3"] = "=Inputs!D6"  # Cross-sheet ref

    ws2["A4"] = "Tax"
    ws2["B4"] = "=Inputs!D9"

    ws2["A5"] = "Grand Total"
    ws2["B5"] = "=Inputs!D10"
    ws2["B5"].font = Font(bold=True)
    ws2["B5"].number_format = '#,##0.00'

    ws2["A7"] = "Discount Rate"
    ws2["B7"] = 0.10  # 10% discount (hardcoded)
    ws2["B7"].number_format = '0%'

    ws2["A8"] = "Discount Amount"
    ws2["B8"] = "=B5*B7"

    ws2["A9"] = "Final Total"
    ws2["B9"] = "=B5-B8"
    ws2["B9"].font = Font(bold=True, size=12)
    ws2["B9"].number_format = '#,##0.00'

    ws2["A11"] = "Item Count"
    ws2["B11"] = "=SUM(Inputs!C2:C4)"

    ws2["A12"] = "Max Price"
    ws2["B12"] = "=MAX(Inputs!B2:B4)"

    ws2["A13"] = "Conditional"
    ws2["B13"] = '=IF(B5>100,"High","Low")'

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 15

    # ---- Sheet3: Table-based calculations ----
    ws3 = wb.create_sheet("Rates")

    ws3["A1"] = "Category"
    ws3["B1"] = "Rate"
    ws3["C1"] = "Multiplier"
    ws3["A1"].font = Font(bold=True)
    ws3["B1"].font = Font(bold=True)
    ws3["C1"].font = Font(bold=True)

    ws3["A2"] = "Standard"
    ws3["B2"] = 1.0
    ws3["C2"] = "=B2*2"

    ws3["A3"] = "Premium"
    ws3["B3"] = 1.5
    ws3["C3"] = "=B3*2"

    ws3["A4"] = "Enterprise"
    ws3["B4"] = 2.0
    ws3["C4"] = "=B4*2"

    ws3["A6"] = "Total Rates"
    ws3["B6"] = "=SUM(B2:B4)"
    ws3["C6"] = "=SUM(C2:C4)"

    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 12

    # Save
    wb.save(output_path)
    wb.close()
    return output_path


if __name__ == "__main__":
    path = os.path.join(os.path.dirname(__file__), "test_data", "sample.xlsx")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    create_sample_workbook(path)
    print(f"Created: {path}")
