"""Tests for the excel_to_python_vectorized converter."""

import json
import os
import subprocess
import sys
import tempfile
import shutil
import unittest

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

# Ensure project root is on the path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from tests.create_sample_workbook import create_sample_workbook
from excel_to_python_vectorized.vectorizer import (
    extract_references,
    compute_pattern,
    group_formulas,
    order_items,
    discover_external_files,
    analyse_references,
)
from excel_to_python_vectorized.converter import convert_excel_to_python_vectorized


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _create_workbook_with_external_refs(path):
    """Create a workbook whose formulas mention an external file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws["A1"] = "Local Value"
    ws["B1"] = 100
    # Simulate external reference (openpyxl stores as string, not live link)
    ws["C1"] = "=[ExtData.xlsx]Prices!A1*2"
    ws["C2"] = "=[ExtData.xlsx]Prices!A2*2"
    ws["D1"] = "=B1+10"
    wb.save(path)
    wb.close()


def _create_workbook_with_vertical_drag(path):
    """Workbook with a clearly dragged formula down a column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Year"
    ws["B1"] = "Revenue"
    ws["C1"] = "Cost"
    ws["D1"] = "Profit"
    for r in range(2, 12):  # 10 rows of data
        ws[f"A{r}"] = 2020 + r - 2
        ws[f"B{r}"] = 1000 * r
        ws[f"C{r}"] = 600 * r
        ws[f"D{r}"] = f"=B{r}-C{r}"  # dragged formula
    ws["D12"] = "=SUM(D2:D11)"
    ws["B12"] = "=SUM(B2:B11)"
    ws["C12"] = "=SUM(C2:C11)"
    wb.save(path)
    wb.close()


def _create_workbook_with_horizontal_drag(path):
    """Workbook with a formula dragged across columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast"
    ws["A1"] = "Item"
    ws["A2"] = "Base"
    ws["A3"] = "Growth"
    # Years in columns B..F
    for ci, year in enumerate(range(2024, 2029)):
        col_letter = chr(66 + ci)  # B, C, D, E, F
        ws[f"{col_letter}1"] = year
        ws[f"{col_letter}2"] = 100 + ci * 10  # base values
        ws[f"{col_letter}3"] = f"={col_letter}2*1.05"  # growth = base * 1.05
    wb.save(path)
    wb.close()


def _create_workbook_with_cross_sheet(path):
    """Workbook with cross-sheet references."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Revenue"
    ws1["A1"] = "Q1"
    ws1["A2"] = 1000
    ws1["B1"] = "Q2"
    ws1["B2"] = 1500
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Total Revenue"
    ws2["B1"] = "=Revenue!A2+Revenue!B2"
    ws2["A2"] = "Double"
    ws2["B2"] = "=B1*2"
    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# Test: reference extraction
# ---------------------------------------------------------------------------

class TestExtractReferences(unittest.TestCase):
    def test_simple_cell_ref(self):
        refs = extract_references("=A1+B2", "Sheet1")
        self.assertEqual(len(refs), 2)
        self.assertEqual(refs[0].col, "A")
        self.assertEqual(refs[0].row, 1)
        self.assertEqual(refs[1].col, "B")
        self.assertEqual(refs[1].row, 2)

    def test_absolute_markers(self):
        refs = extract_references("=$A$1+B2", "Sheet1")
        self.assertTrue(refs[0].col_abs)
        self.assertTrue(refs[0].row_abs)
        self.assertFalse(refs[1].col_abs)
        self.assertFalse(refs[1].row_abs)

    def test_cross_sheet_ref(self):
        refs = extract_references("=Sheet2!A1", "Sheet1")
        self.assertEqual(len(refs), 1)
        self.assertEqual(refs[0].sheet, "Sheet2")
        self.assertIsNone(refs[0].external_file)

    def test_quoted_sheet_ref(self):
        refs = extract_references("='My Sheet'!B5", "Sheet1")
        self.assertEqual(len(refs), 1)
        self.assertEqual(refs[0].sheet, "My Sheet")

    def test_external_file_ref(self):
        refs = extract_references("=[Book.xlsx]Data!C3", "Sheet1")
        self.assertEqual(len(refs), 1)
        self.assertEqual(refs[0].external_file, "Book.xlsx")
        self.assertEqual(refs[0].sheet, "Data")
        self.assertEqual(refs[0].col, "C")
        self.assertEqual(refs[0].row, 3)

    def test_range_ref(self):
        refs = extract_references("=SUM(A1:A10)", "Sheet1")
        self.assertEqual(len(refs), 1)
        self.assertEqual(refs[0].kind, "range")
        self.assertEqual(refs[0].col, "A")
        self.assertEqual(refs[0].end_col, "A")
        self.assertEqual(refs[0].row, 1)
        self.assertEqual(refs[0].end_row, 10)

    def test_no_refs_in_string(self):
        refs = extract_references('=IF(A1>0,"A1","B2")', "Sheet1")
        # Only A1 should be found as an actual ref
        cell_refs = [r for r in refs if r.kind == "cell"]
        self.assertEqual(len(cell_refs), 1)
        self.assertEqual(cell_refs[0].col, "A")


# ---------------------------------------------------------------------------
# Test: pattern computation
# ---------------------------------------------------------------------------

class TestComputePattern(unittest.TestCase):
    def test_dragged_vertical_same_pattern(self):
        """=B2*C2 at D2 and =B3*C3 at D3 should have the same pattern."""
        from formula_converter import col_letter_to_index
        p1, _ = compute_pattern("=B2*C2", "Sheet", col_letter_to_index("D"), 2)
        p2, _ = compute_pattern("=B3*C3", "Sheet", col_letter_to_index("D"), 3)
        self.assertEqual(p1, p2)

    def test_different_formula_different_pattern(self):
        from formula_converter import col_letter_to_index
        p1, _ = compute_pattern("=B2*C2", "Sheet", col_letter_to_index("D"), 2)
        p2, _ = compute_pattern("=B2+C2", "Sheet", col_letter_to_index("D"), 2)
        self.assertNotEqual(p1, p2)

    def test_absolute_ref_same_pattern(self):
        """=$A$1*B2 at D2 and =$A$1*B3 at D3 should match."""
        from formula_converter import col_letter_to_index
        p1, _ = compute_pattern("=$A$1*B2", "Sheet", col_letter_to_index("D"), 2)
        p2, _ = compute_pattern("=$A$1*B3", "Sheet", col_letter_to_index("D"), 3)
        self.assertEqual(p1, p2)


# ---------------------------------------------------------------------------
# Test: grouping
# ---------------------------------------------------------------------------

class TestGroupFormulas(unittest.TestCase):
    def test_vertical_group(self):
        """Three dragged formulas should form one vertical group."""
        cells = [
            ("Sheet", "D", 2, "=B2*C2", {}),
            ("Sheet", "D", 3, "=B3*C3", {}),
            ("Sheet", "D", 4, "=B4*C4", {}),
        ]
        groups, singles = group_formulas(cells)
        self.assertEqual(len(groups), 1)
        self.assertEqual(groups[0]["direction"], "vertical")
        self.assertEqual(len(groups[0]["cells"]), 3)
        self.assertEqual(len(singles), 0)

    def test_horizontal_group(self):
        cells = [
            ("Sheet", "B", 3, "=B2*1.05", {}),
            ("Sheet", "C", 3, "=C2*1.05", {}),
            ("Sheet", "D", 3, "=D2*1.05", {}),
        ]
        groups, singles = group_formulas(cells)
        self.assertEqual(len(groups), 1)
        self.assertEqual(groups[0]["direction"], "horizontal")
        self.assertEqual(len(groups[0]["cells"]), 3)

    def test_mixed_leaves_singles(self):
        """A unique formula should become a single."""
        cells = [
            ("Sheet", "D", 2, "=B2*C2", {}),
            ("Sheet", "D", 3, "=B3*C3", {}),
            ("Sheet", "D", 6, "=SUM(D2:D5)", {}),
        ]
        groups, singles = group_formulas(cells)
        self.assertTrue(len(groups) >= 1)
        self.assertTrue(any(s[2] == 6 for s in singles))


# ---------------------------------------------------------------------------
# Test: external file discovery
# ---------------------------------------------------------------------------

class TestExternalFileDiscovery(unittest.TestCase):
    def test_discovers_external_files(self):
        cells = [
            ("Sheet1", "C", 1, "=[ExtData.xlsx]Prices!A1*2", {}),
            ("Sheet1", "C", 2, "=[ExtData.xlsx]Prices!A2*2", {}),
            ("Sheet1", "D", 1, "=[Other.xlsx]Summary!B1", {}),
        ]
        ext = discover_external_files(cells)
        self.assertIn("ExtData.xlsx", ext)
        self.assertIn("Other.xlsx", ext)
        self.assertEqual(len(ext["ExtData.xlsx"]), 2)

    def test_no_external(self):
        cells = [("Sheet1", "D", 2, "=B2*C2", {})]
        ext = discover_external_files(cells)
        self.assertEqual(len(ext), 0)


# ---------------------------------------------------------------------------
# Test: analysis
# ---------------------------------------------------------------------------

class TestAnalyseReferences(unittest.TestCase):
    def test_cross_sheet(self):
        cells = [
            ("Summary", "B", 1, "=Revenue!A2+Revenue!B2", {}),
        ]
        analysis = analyse_references(cells)
        self.assertTrue(len(analysis["cross_sheet"]) >= 1)
        self.assertEqual(analysis["cross_sheet"][0]["target_sheet"], "Revenue")

    def test_external(self):
        cells = [
            ("Sheet1", "C", 1, "=[ExtData.xlsx]Prices!A1*2", {}),
        ]
        analysis = analyse_references(cells)
        self.assertEqual(len(analysis["external"]), 1)
        self.assertEqual(analysis["external"][0]["external_file"], "ExtData.xlsx")


# ---------------------------------------------------------------------------
# End-to-end tests
# ---------------------------------------------------------------------------

class TestEndToEndVectorized(unittest.TestCase):
    """Full pipeline: Excel → vectorised script → run → validate."""

    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp()
        cls.sample_path = os.path.join(cls.tmpdir, "sample.xlsx")
        create_sample_workbook(cls.sample_path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir)

    def test_produces_all_files(self):
        out = os.path.join(self.tmpdir, "out_files")
        script, template, ext_cfg, report = convert_excel_to_python_vectorized(
            self.sample_path, output_dir=out)
        self.assertTrue(os.path.exists(script))
        self.assertTrue(os.path.exists(template))
        self.assertTrue(os.path.exists(report))
        # No external refs in sample → no config
        self.assertIsNone(ext_cfg)

    def test_generated_script_syntax(self):
        out = os.path.join(self.tmpdir, "out_syntax")
        script, *_ = convert_excel_to_python_vectorized(
            self.sample_path, output_dir=out)
        with open(script) as f:
            code = f.read()
        compile(code, script, "exec")

    def test_generated_script_runs_correctly(self):
        out = os.path.join(self.tmpdir, "out_run")
        script, template, *_ = convert_excel_to_python_vectorized(
            self.sample_path, output_dir=out)
        result_path = os.path.join(out, "result.xlsx")
        proc = subprocess.run(
            [sys.executable, script, template, result_path],
            capture_output=True, text=True, timeout=60)
        self.assertEqual(proc.returncode, 0, f"Script failed:\n{proc.stderr}")
        self.assertTrue(os.path.exists(result_path))

        wb = load_workbook(result_path, data_only=True)
        ws = wb["Inputs"]
        self.assertAlmostEqual(float(ws["D2"].value), 52.5, places=2)
        self.assertAlmostEqual(float(ws["D6"].value), 200.0, places=2)
        self.assertAlmostEqual(float(ws["D9"].value), 16.0, places=2)
        self.assertAlmostEqual(float(ws["D10"].value), 216.0, places=2)

        ws2 = wb["Summary"]
        self.assertAlmostEqual(float(ws2["B5"].value), 216.0, places=2)
        self.assertAlmostEqual(float(ws2["B9"].value), 194.4, places=2)
        self.assertEqual(ws2["B13"].value, "High")

        ws3 = wb["Rates"]
        self.assertAlmostEqual(float(ws3["C2"].value), 2.0, places=2)
        self.assertAlmostEqual(float(ws3["B6"].value), 4.5, places=2)
        self.assertAlmostEqual(float(ws3["C6"].value), 9.0, places=2)
        wb.close()

    def test_script_is_vectorised(self):
        """The generated script should contain ``for _r in`` loops."""
        out = os.path.join(self.tmpdir, "out_vec")
        script, *_ = convert_excel_to_python_vectorized(
            self.sample_path, output_dir=out)
        with open(script) as f:
            code = f.read()
        self.assertIn("for _r in", code)
        self.assertIn("# Vectorised:", code)

    def test_report_has_expected_sheets(self):
        out = os.path.join(self.tmpdir, "out_report")
        *_, report = convert_excel_to_python_vectorized(
            self.sample_path, output_dir=out)
        wb = load_workbook(report)
        self.assertIn("Summary", wb.sheetnames)
        self.assertIn("Vectorised Groups", wb.sheetnames)
        self.assertIn("Cross-Sheet Refs", wb.sheetnames)
        self.assertIn("External Refs", wb.sheetnames)
        self.assertIn("Per-Sheet Breakdown", wb.sheetnames)
        wb.close()


class TestEndToEndVerticalDrag(unittest.TestCase):
    """Test with a workbook that has 10 dragged rows."""

    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp()
        cls.path = os.path.join(cls.tmpdir, "vertical.xlsx")
        _create_workbook_with_vertical_drag(cls.path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir)

    def test_runs_and_computes_correctly(self):
        out = os.path.join(self.tmpdir, "out")
        script, template, *_ = convert_excel_to_python_vectorized(
            self.path, output_dir=out)
        result = os.path.join(out, "result.xlsx")
        proc = subprocess.run(
            [sys.executable, script, template, result],
            capture_output=True, text=True, timeout=60)
        self.assertEqual(proc.returncode, 0, f"Failed:\n{proc.stderr}")

        wb = load_workbook(result)
        ws = wb["Data"]
        # D2 = B2 - C2 = 2000 - 1200 = 800
        self.assertAlmostEqual(float(ws["D2"].value), 800, places=2)
        # D11 = B11 - C11 = 11000 - 6600 = 4400
        self.assertAlmostEqual(float(ws["D11"].value), 4400, places=2)
        # D12 = SUM(D2:D11)
        total = sum(1000 * r - 600 * r for r in range(2, 12))  # = 400 * sum(2..11)
        self.assertAlmostEqual(float(ws["D12"].value), total, places=2)
        wb.close()

    def test_vectorisation_reduces_lines(self):
        out = os.path.join(self.tmpdir, "out_lines")
        script, *_ = convert_excel_to_python_vectorized(
            self.path, output_dir=out)
        with open(script) as f:
            code = f.read()
        # 10 dragged formulas should be a single loop, not 10 separate blocks
        self.assertIn("for _r in", code)
        # Count occurrences of the loop pattern
        self.assertLessEqual(code.count("for _r in"), 3)


class TestEndToEndHorizontalDrag(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp()
        cls.path = os.path.join(cls.tmpdir, "horizontal.xlsx")
        _create_workbook_with_horizontal_drag(cls.path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir)

    def test_runs_correctly(self):
        out = os.path.join(self.tmpdir, "out")
        script, template, *_ = convert_excel_to_python_vectorized(
            self.path, output_dir=out)
        result = os.path.join(out, "result.xlsx")
        proc = subprocess.run(
            [sys.executable, script, template, result],
            capture_output=True, text=True, timeout=60)
        self.assertEqual(proc.returncode, 0, f"Failed:\n{proc.stderr}")

        wb = load_workbook(result)
        ws = wb["Forecast"]
        # B3 = B2 * 1.05 = 100 * 1.05 = 105
        self.assertAlmostEqual(float(ws["B3"].value), 105.0, places=2)
        # C3 = C2 * 1.05 = 110 * 1.05 = 115.5
        self.assertAlmostEqual(float(ws["C3"].value), 115.5, places=2)
        wb.close()

    def test_has_horizontal_loop(self):
        out = os.path.join(self.tmpdir, "out_h")
        script, *_ = convert_excel_to_python_vectorized(
            self.path, output_dir=out)
        with open(script) as f:
            code = f.read()
        self.assertIn("for _ci in", code)


class TestEndToEndCrossSheet(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp()
        cls.path = os.path.join(cls.tmpdir, "cross_sheet.xlsx")
        _create_workbook_with_cross_sheet(cls.path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir)

    def test_cross_sheet_values(self):
        out = os.path.join(self.tmpdir, "out")
        script, template, *_ = convert_excel_to_python_vectorized(
            self.path, output_dir=out)
        result = os.path.join(out, "result.xlsx")
        proc = subprocess.run(
            [sys.executable, script, template, result],
            capture_output=True, text=True, timeout=60)
        self.assertEqual(proc.returncode, 0, f"Failed:\n{proc.stderr}")

        wb = load_workbook(result)
        ws = wb["Summary"]
        # B1 = Revenue!A2 + Revenue!B2 = 1000 + 1500 = 2500
        self.assertAlmostEqual(float(ws["B1"].value), 2500, places=2)
        # B2 = B1 * 2 = 5000
        self.assertAlmostEqual(float(ws["B2"].value), 5000, places=2)
        wb.close()

    def test_report_shows_cross_sheet(self):
        out = os.path.join(self.tmpdir, "out_report")
        *_, report = convert_excel_to_python_vectorized(
            self.path, output_dir=out)
        wb = load_workbook(report)
        ws = wb["Cross-Sheet Refs"]
        # Should have at least one cross-sheet entry
        self.assertIsNotNone(ws["A2"].value)
        wb.close()


class TestEndToEndExternalRefs(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp()
        cls.path = os.path.join(cls.tmpdir, "with_ext.xlsx")
        _create_workbook_with_external_refs(cls.path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir)

    def test_ext_config_generated(self):
        out = os.path.join(self.tmpdir, "out")
        _, _, ext_cfg, _ = convert_excel_to_python_vectorized(
            self.path, output_dir=out)
        self.assertIsNotNone(ext_cfg)
        self.assertTrue(os.path.exists(ext_cfg))
        with open(ext_cfg) as f:
            cfg = json.load(f)
        self.assertIn("ExtData.xlsx", cfg)

    def test_report_shows_external(self):
        out = os.path.join(self.tmpdir, "out_report")
        *_, report = convert_excel_to_python_vectorized(
            self.path, output_dir=out)
        wb = load_workbook(report)
        ws = wb["External Refs"]
        self.assertIsNotNone(ws["A2"].value)
        wb.close()

    def test_script_syntax_ok(self):
        out = os.path.join(self.tmpdir, "out_syntax")
        script, *_ = convert_excel_to_python_vectorized(
            self.path, output_dir=out)
        with open(script) as f:
            code = f.read()
        compile(code, script, "exec")


if __name__ == "__main__":
    unittest.main()
