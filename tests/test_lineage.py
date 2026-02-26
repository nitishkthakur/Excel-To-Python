"""Tests for the lineage module — builder and graph generation."""

import os
import sys
import tempfile

import pytest

# Ensure the repo root is on the path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from tests.create_sample_workbook import create_sample_workbook
from lineage.lineage_builder import (
    build_simple_lineage,
    build_complex_lineage,
    write_simple_lineage,
    write_complex_lineage,
)
from lineage.lineage_graph import render_simple_graph, render_complex_graph


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def sample_wb_path(tmp_path_factory):
    """Create a sample workbook once per module."""
    d = tmp_path_factory.mktemp("lineage")
    path = str(d / "sample.xlsx")
    create_sample_workbook(path)
    return path


@pytest.fixture(scope="module")
def simple_lineage(sample_wb_path):
    return build_simple_lineage(sample_wb_path)


@pytest.fixture(scope="module")
def complex_lineage(sample_wb_path):
    return build_complex_lineage(sample_wb_path)


# ---------------------------------------------------------------------------
# Simple lineage builder
# ---------------------------------------------------------------------------

class TestSimpleLineage:
    def test_returns_dict_with_sheets(self, simple_lineage):
        assert "sheets" in simple_lineage
        assert isinstance(simple_lineage["sheets"], list)
        assert len(simple_lineage["sheets"]) > 0

    def test_each_sheet_has_required_keys(self, simple_lineage):
        for s in simple_lineage["sheets"]:
            assert "sheet_name" in s
            assert "inputs" in s
            assert "outputs" in s
            assert "calculations" in s

    def test_has_cross_sheet_edges(self, simple_lineage):
        assert "cross_sheet_edges" in simple_lineage
        assert isinstance(simple_lineage["cross_sheet_edges"], list)

    def test_inputs_have_column(self, simple_lineage):
        for s in simple_lineage["sheets"]:
            for inp in s["inputs"]:
                assert "column" in inp
                assert "count" in inp

    def test_calculations_have_pattern(self, simple_lineage):
        for s in simple_lineage["sheets"]:
            for calc in s["calculations"]:
                assert "pattern" in calc
                assert "example_formula" in calc

    def test_sample_workbook_has_formulas(self, simple_lineage):
        """The sample workbook should have at least one sheet with calcs."""
        total_calcs = sum(len(s["calculations"]) for s in simple_lineage["sheets"])
        assert total_calcs > 0


# ---------------------------------------------------------------------------
# Complex lineage builder
# ---------------------------------------------------------------------------

class TestComplexLineage:
    def test_returns_dict_with_sheets(self, complex_lineage):
        assert "sheets" in complex_lineage
        assert isinstance(complex_lineage["sheets"], list)

    def test_has_dependency_edges(self, complex_lineage):
        assert "dependency_edges" in complex_lineage
        assert isinstance(complex_lineage["dependency_edges"], list)

    def test_patterns_have_dependencies(self, complex_lineage):
        for s in complex_lineage["sheets"]:
            for pat in s["patterns"]:
                assert "dependencies" in pat
                assert isinstance(pat["dependencies"], list)

    def test_has_cross_sheet_refs(self, complex_lineage):
        assert "cross_sheet_refs" in complex_lineage

    def test_has_external_refs(self, complex_lineage):
        assert "external_refs" in complex_lineage


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

class TestExcelOutput:
    def test_write_simple_lineage(self, simple_lineage, tmp_path):
        out = str(tmp_path / "simple.xlsx")
        write_simple_lineage(simple_lineage, out)
        assert os.path.exists(out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        assert "Overview" in wb.sheetnames
        assert "Inputs" in wb.sheetnames
        assert "Calculations" in wb.sheetnames
        assert "Outputs" in wb.sheetnames
        assert "Cross-Sheet Edges" in wb.sheetnames
        wb.close()

    def test_write_complex_lineage(self, complex_lineage, tmp_path):
        out = str(tmp_path / "complex.xlsx")
        write_complex_lineage(complex_lineage, out)
        assert os.path.exists(out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        assert "Summary" in wb.sheetnames
        assert "All Patterns" in wb.sheetnames
        assert "Dependency Edges" in wb.sheetnames
        wb.close()


# ---------------------------------------------------------------------------
# Graph rendering
# ---------------------------------------------------------------------------

class TestGraphRendering:
    def test_render_simple_graph(self, simple_lineage, tmp_path):
        xlsx = str(tmp_path / "simple.xlsx")
        write_simple_lineage(simple_lineage, xlsx)
        png = str(tmp_path / "simple.png")
        render_simple_graph(xlsx, png)
        assert os.path.exists(png)
        assert os.path.getsize(png) > 0

    def test_render_complex_graph(self, complex_lineage, tmp_path):
        xlsx = str(tmp_path / "complex.xlsx")
        write_complex_lineage(complex_lineage, xlsx)
        png = str(tmp_path / "complex.png")
        render_complex_graph(xlsx, png)
        assert os.path.exists(png)
        assert os.path.getsize(png) > 0
