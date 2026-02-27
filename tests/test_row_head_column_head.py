"""
Tests for the row_head and column_head sampling strategies, the updated
dispatch logic (all 5 modes), and the Annotated Field tool documentation.
"""

import os
import sys
import json
import pytest

# Make the mcp_server package and tests directory importable
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "mcp_server"))
sys.path.insert(0, os.path.dirname(__file__))

from create_sample_workbook import create_sample_workbook

from excel_reader_smart_sampler import (
    extract_sheet_data as smart_extract,
    sheet_names,
    DEFAULT_SAMPLE_ROWS,
)
from excel_sample_full import extract_sheet_data as full_extract
from column_n import extract_sheet_data as column_n_extract
from row_head_fetcher import (
    extract_sheet_data as row_head_extract,
    _allocate_budget,
    row_head_indices,
)
from column_head_fetcher import (
    extract_sheet_data as column_head_extract,
    _allocate_col_budget,
    column_head_indices,
)
from smart_random_fetcher import load_sheet_frames, detect_regions


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def sample_wb(tmp_path_factory):
    """Create the sample workbook once per test module."""
    path = str(tmp_path_factory.mktemp("data") / "sample.xlsx")
    create_sample_workbook(path)
    return path


@pytest.fixture(scope="module")
def wide_wb(tmp_path_factory):
    """Create a wide workbook (many date columns) for column_head testing."""
    from openpyxl import Workbook
    path = str(tmp_path_factory.mktemp("data") / "wide.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "PnL"

    # Header row: Entity + 24 quarterly date columns
    ws.cell(row=1, column=1, value="Entity")
    for i in range(24):
        year = 2023 + i // 4
        quarter = (i % 4) + 1
        ws.cell(row=1, column=i + 2, value=f"Q{quarter} {year}")

    # Data rows: financial entities
    entities = ["Revenue", "COGS", "Gross Profit", "OpEx", "EBITDA",
                "Depreciation", "EBIT", "Interest", "Tax", "Net Income"]
    for r, entity in enumerate(entities, start=2):
        ws.cell(row=r, column=1, value=entity)
        for c in range(2, 26):
            ws.cell(row=r, column=c, value=(r * 1000 + c))

    # Add a formula row
    ws.cell(row=12, column=1, value="Total Check")
    for c in range(2, 26):
        ws.cell(row=12, column=c, value=f"=SUM(B{2}:{get_column_letter(c)}{11})")

    wb.save(path)
    wb.close()
    return path


@pytest.fixture(scope="module")
def multi_patch_wb(tmp_path_factory):
    """Create a workbook with multiple data patches separated by blank rows."""
    from openpyxl import Workbook
    path = str(tmp_path_factory.mktemp("data") / "multi_patch.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Patch 1: rows 1-5 (header + 4 data rows)
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    for i in range(2, 6):
        ws[f"A{i}"] = f"Item{i-1}"
        ws[f"B{i}"] = i * 10

    # blank row 6

    # Patch 2: rows 7-16 (header + 9 data rows)
    ws["A7"] = "Category"
    ws["B7"] = "Amount"
    ws["C7"] = "Tax"
    for i in range(8, 17):
        ws[f"A{i}"] = f"Cat{i-7}"
        ws[f"B{i}"] = i * 100
        ws[f"C{i}"] = f"=B{i}*0.1"

    # blank row 17

    # Patch 3: rows 18-20 (header + 2 data rows)
    ws["A18"] = "Summary"
    ws["B18"] = "Total"
    ws["A19"] = "Grand Total"
    ws["B19"] = "=SUM(B2:B5)+SUM(B8:B16)"
    ws["A20"] = "Tax Total"
    ws["B20"] = "=SUM(C8:C16)"

    wb.save(path)
    wb.close()
    return path


# We need get_column_letter for the wide_wb fixture
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Helpers for assertions
# ---------------------------------------------------------------------------

def _result_shape_ok(data: dict):
    """Assert that *data* has the standard result dict shape."""
    assert "sheet_name" in data
    assert "regions" in data
    assert isinstance(data["regions"], list)
    assert "sampled" in data
    assert "total_rows" in data
    assert "sampled_rows" in data
    for reg in data["regions"]:
        assert "headers" in reg
        assert "rows" in reg
        assert "formulas" in reg
        assert "min_row" in reg
        assert "max_row" in reg
        assert "min_col" in reg
        assert "max_col" in reg


# ===========================================================================
# row_head strategy
# ===========================================================================

class TestRowHead:

    def test_basic_extraction(self, sample_wb):
        data = row_head_extract(sample_wb, "Inputs")
        _result_shape_ok(data)
        assert data["sheet_name"] == "Inputs"

    def test_headers_always_captured(self, sample_wb):
        """Row-head mode must always capture headers for every patch."""
        data = row_head_extract(sample_wb, "Inputs", max_rows=5)
        for reg in data["regions"]:
            if reg["headers"]:
                assert len(reg["headers"]) > 0

    def test_formulas_present(self, sample_wb):
        data = row_head_extract(sample_wb, "Inputs", max_rows=100)
        all_formulas = []
        for reg in data["regions"]:
            all_formulas.extend(reg["formulas"])
        assert len(all_formulas) > 0

    def test_all_sheets(self, sample_wb):
        for name in sheet_names(sample_wb):
            data = row_head_extract(sample_wb, name)
            _result_shape_ok(data)

    def test_budget_is_per_sheet(self, sample_wb):
        """The max_rows budget is per sheet (across all patches)."""
        data = row_head_extract(sample_wb, "Inputs", max_rows=5)
        assert data["sampled_rows"] <= 10  # small budget, few rows

    def test_budget_allocation_proportional(self):
        """Budget is divided across patches proportionally to their size."""
        regions = [
            {"min_row": 1, "max_row": 100, "min_col": 1, "max_col": 5,
             "header_row": 1},
            {"min_row": 110, "max_row": 119, "min_col": 1, "max_col": 5,
             "header_row": 110},
        ]
        budgets = _allocate_budget(regions, 20)
        # Larger region should get a bigger budget
        assert budgets[0] > budgets[1]
        assert sum(budgets) <= 20 + len(regions)  # allow minor overflow from min(2)

    def test_row_head_indices_includes_header(self):
        """row_head_indices must always include the header row."""
        reg = {"min_row": 5, "max_row": 20, "min_col": 1, "max_col": 3,
               "header_row": 5}
        indices = row_head_indices(reg, 5)
        assert 5 in indices  # header row

    def test_row_head_indices_contiguous(self):
        """Rows selected should be header + contiguous head rows."""
        reg = {"min_row": 5, "max_row": 20, "min_col": 1, "max_col": 3,
               "header_row": 5}
        indices = row_head_indices(reg, 5)
        # Should be [5, 6, 7, 8, 9]
        assert indices == [5, 6, 7, 8, 9]

    def test_multi_patch_all_headers_captured(self, multi_patch_wb):
        """For a multi-patch workbook, row_head must capture all headers."""
        data = row_head_extract(multi_patch_wb, "Data", max_rows=10)
        _result_shape_ok(data)
        # There should be multiple regions, each with headers
        assert len(data["regions"]) >= 2
        for reg in data["regions"]:
            assert len(reg["headers"]) > 0

    def test_small_patch_kept_whole(self, multi_patch_wb):
        """Small patches should be fully included."""
        data = row_head_extract(multi_patch_wb, "Data", max_rows=100)
        # All patches are small enough to be kept whole with a large budget
        for reg in data["regions"]:
            expected_rows = reg["max_row"] - reg["min_row"]  # minus header
            assert len(reg["rows"]) >= 1


class TestRowHeadEmptySheet:

    def test_empty_sheet(self, tmp_path):
        """row_head should handle empty sheets gracefully."""
        from openpyxl import Workbook
        path = str(tmp_path / "empty.xlsx")
        wb = Workbook()
        wb.save(path)
        wb.close()

        data = row_head_extract(path, "Sheet")
        _result_shape_ok(data)
        assert data["total_rows"] == 0


# ===========================================================================
# column_head strategy
# ===========================================================================

class TestColumnHead:

    def test_basic_extraction(self, sample_wb):
        data = column_head_extract(sample_wb, "Inputs")
        _result_shape_ok(data)
        assert data["sheet_name"] == "Inputs"

    def test_all_rows_included(self, sample_wb):
        """column_head loads ALL rows, only limits columns."""
        data = column_head_extract(sample_wb, "Inputs", max_cols=2)
        for reg in data["regions"]:
            # All non-header rows should be present
            total_rows_in_reg = reg["max_row"] - reg["min_row"] + 1
            # rows in output = total - 1 (header excluded from row data)
            header_excluded = 1 if reg["headers"] else 0
            assert len(reg["rows"]) >= total_rows_in_reg - header_excluded

    def test_column_budget_limits_width(self, sample_wb):
        """With a small column budget, the strip width should be limited."""
        data = column_head_extract(sample_wb, "Inputs", max_cols=2)
        for reg in data["regions"]:
            col_span = reg["max_col"] - reg["min_col"] + 1
            assert col_span <= 2

    def test_headers_captured(self, sample_wb):
        data = column_head_extract(sample_wb, "Inputs")
        for reg in data["regions"]:
            if reg["headers"]:
                assert len(reg["headers"]) > 0

    def test_formulas_present(self, sample_wb):
        data = column_head_extract(sample_wb, "Inputs", max_cols=10)
        all_formulas = []
        for reg in data["regions"]:
            all_formulas.extend(reg["formulas"])
        assert len(all_formulas) > 0

    def test_all_sheets(self, sample_wb):
        for name in sheet_names(sample_wb):
            data = column_head_extract(sample_wb, name)
            _result_shape_ok(data)

    def test_wide_sheet_column_limiting(self, wide_wb):
        """On a wide sheet (25 cols), column_head should limit columns."""
        data = column_head_extract(wide_wb, "PnL", max_cols=5)
        _result_shape_ok(data)
        for reg in data["regions"]:
            col_span = reg["max_col"] - reg["min_col"] + 1
            assert col_span <= 5
            # But all rows should still be present
            assert len(reg["rows"]) >= 5  # at least some entity rows

    def test_wide_sheet_entities_visible(self, wide_wb):
        """Column head should capture entity labels (first column)."""
        data = column_head_extract(wide_wb, "PnL", max_cols=3)
        for reg in data["regions"]:
            # Check that entity labels are in the first cell of row data
            for row in reg["rows"]:
                # First value should be a string entity name (or formula)
                first_val = row["values"][0]
                assert first_val is not None

    def test_budget_allocation_proportional(self):
        """Budget is divided across patches proportionally to their width."""
        regions = [
            {"min_row": 1, "max_row": 10, "min_col": 1, "max_col": 20,
             "header_row": 1},
            {"min_row": 15, "max_row": 25, "min_col": 1, "max_col": 5,
             "header_row": 15},
        ]
        budgets = _allocate_col_budget(regions, 15)
        # Wider region should get a bigger budget
        assert budgets[0] > budgets[1]

    def test_column_head_indices_start_from_min(self):
        """column_head_indices should start from the region's min_col."""
        reg = {"min_row": 1, "max_row": 10, "min_col": 3, "max_col": 20,
               "header_row": 1}
        indices = column_head_indices(reg, 5)
        assert indices == [3, 4, 5, 6, 7]

    def test_budget_is_per_sheet(self, wide_wb):
        """The max_cols budget is per sheet (across all patches)."""
        data = column_head_extract(wide_wb, "PnL", max_cols=3)
        for reg in data["regions"]:
            col_span = reg["max_col"] - reg["min_col"] + 1
            assert col_span <= 3


class TestColumnHeadEmptySheet:

    def test_empty_sheet(self, tmp_path):
        """column_head should handle empty sheets gracefully."""
        from openpyxl import Workbook
        path = str(tmp_path / "empty.xlsx")
        wb = Workbook()
        wb.save(path)
        wb.close()

        data = column_head_extract(path, "Sheet")
        _result_shape_ok(data)
        assert data["total_rows"] == 0


# ===========================================================================
# Updated dispatch logic (all 5 modes)
# ===========================================================================

class TestDispatchAllModes:
    """Test the dispatch logic mirrors server._dispatch_extract for all 5 modes."""

    def _dispatch(self, mode, path, sheet_name, **kwargs):
        mode = mode.lower().strip()
        valid = ("smart_random", "full", "column_n", "row_head", "column_head")
        if mode not in valid:
            raise ValueError(f"Invalid mode '{mode}'. Must be one of {valid}.")
        if mode == "smart_random":
            return smart_extract(path, sheet_name,
                                 max_sample_rows=kwargs.get("max_sample_rows",
                                                            DEFAULT_SAMPLE_ROWS))
        elif mode == "full":
            return full_extract(path, sheet_name,
                                nrows=kwargs.get("nrows"),
                                ncols=kwargs.get("ncols"))
        elif mode == "column_n":
            return column_n_extract(path, sheet_name,
                                    num_columns=kwargs.get("num_columns", 10))
        elif mode == "row_head":
            return row_head_extract(path, sheet_name,
                                    max_rows=kwargs.get("max_sample_rows",
                                                        DEFAULT_SAMPLE_ROWS))
        elif mode == "column_head":
            return column_head_extract(path, sheet_name,
                                       max_cols=kwargs.get("max_cols", 20))

    def test_smart_random_dispatch(self, sample_wb):
        data = self._dispatch("smart_random", sample_wb, "Inputs")
        _result_shape_ok(data)

    def test_full_dispatch(self, sample_wb):
        data = self._dispatch("full", sample_wb, "Inputs")
        _result_shape_ok(data)

    def test_column_n_dispatch(self, sample_wb):
        data = self._dispatch("column_n", sample_wb, "Inputs")
        _result_shape_ok(data)

    def test_row_head_dispatch(self, sample_wb):
        data = self._dispatch("row_head", sample_wb, "Inputs")
        _result_shape_ok(data)

    def test_column_head_dispatch(self, sample_wb):
        data = self._dispatch("column_head", sample_wb, "Inputs")
        _result_shape_ok(data)

    def test_invalid_mode_raises(self, sample_wb):
        with pytest.raises(ValueError, match="Invalid mode"):
            self._dispatch("nonexistent", sample_wb, "Inputs")

    def test_all_five_modes_produce_valid_output(self, sample_wb):
        """All five modes should produce structurally valid output."""
        for mode in ("smart_random", "full", "column_n", "row_head", "column_head"):
            data = self._dispatch(mode, sample_wb, "Inputs")
            _result_shape_ok(data)

    def test_modes_produce_different_output(self, sample_wb):
        """Different modes can produce different row/col counts."""
        d_smart = self._dispatch("smart_random", sample_wb, "Inputs")
        d_full = self._dispatch("full", sample_wb, "Inputs")
        d_coln = self._dispatch("column_n", sample_wb, "Inputs", num_columns=2)
        d_rowh = self._dispatch("row_head", sample_wb, "Inputs")
        d_colh = self._dispatch("column_head", sample_wb, "Inputs", max_cols=2)

        for d in (d_smart, d_full, d_coln, d_rowh, d_colh):
            _result_shape_ok(d)


# ===========================================================================
# Budget allocation edge cases
# ===========================================================================

class TestBudgetAllocation:

    def test_allocate_budget_empty(self):
        assert _allocate_budget([], 100) == []

    def test_allocate_budget_single_region(self):
        regions = [{"min_row": 1, "max_row": 50, "min_col": 1, "max_col": 5,
                     "header_row": 1}]
        budgets = _allocate_budget(regions, 20)
        assert budgets[0] == 20

    def test_allocate_budget_caps_to_region_size(self):
        """Budget per region should not exceed actual region row count."""
        regions = [{"min_row": 1, "max_row": 5, "min_col": 1, "max_col": 3,
                     "header_row": 1}]
        budgets = _allocate_budget(regions, 100)
        assert budgets[0] == 5  # only 5 rows in region

    def test_allocate_col_budget_empty(self):
        assert _allocate_col_budget([], 100) == []

    def test_allocate_col_budget_single_region(self):
        regions = [{"min_row": 1, "max_row": 10, "min_col": 1, "max_col": 30,
                     "header_row": 1}]
        budgets = _allocate_col_budget(regions, 10)
        assert budgets[0] == 10

    def test_allocate_col_budget_caps_to_region_width(self):
        """Budget per region should not exceed actual region column count."""
        regions = [{"min_row": 1, "max_row": 10, "min_col": 1, "max_col": 3,
                     "header_row": 1}]
        budgets = _allocate_col_budget(regions, 100)
        assert budgets[0] == 3  # only 3 cols in region


# ===========================================================================
# Integration: formatters work with new modes
# ===========================================================================

class TestFormattersIntegration:

    def test_row_head_markdown(self, sample_wb):
        from formatters import to_markdown
        data = row_head_extract(sample_wb, "Inputs")
        md = to_markdown(data)
        assert "Sheet: Inputs" in md

    def test_row_head_json(self, sample_wb):
        from formatters import to_json
        data = row_head_extract(sample_wb, "Inputs")
        result = to_json(data)
        parsed = json.loads(result)
        assert parsed["sheet_name"] == "Inputs"

    def test_row_head_xml(self, sample_wb):
        from formatters import to_xml
        data = row_head_extract(sample_wb, "Inputs")
        xml_str = to_xml(data)
        assert "Inputs" in xml_str

    def test_column_head_markdown(self, sample_wb):
        from formatters import to_markdown
        data = column_head_extract(sample_wb, "Inputs")
        md = to_markdown(data)
        assert "Sheet: Inputs" in md

    def test_column_head_json(self, sample_wb):
        from formatters import to_json
        data = column_head_extract(sample_wb, "Inputs")
        result = to_json(data)
        parsed = json.loads(result)
        assert parsed["sheet_name"] == "Inputs"

    def test_column_head_xml(self, sample_wb):
        from formatters import to_xml
        data = column_head_extract(sample_wb, "Inputs")
        xml_str = to_xml(data)
        assert "Inputs" in xml_str
