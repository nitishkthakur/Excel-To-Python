"""
Tests for the vectorized smart_random_fetcher module.
"""

import os
import sys

import pytest

# Make the mcp_server package and tests directory importable
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "mcp_server"))
sys.path.insert(0, os.path.dirname(__file__))

from create_sample_workbook import create_sample_workbook
from fetcher_smart_random import (
    extract_sheet_data,
    load_sheet_frames,
    detect_regions,
    sampled_cells,
    highlight_workbook,
    sample_row_indices,
    DEFAULT_SAMPLE_ROWS,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def sample_wb(tmp_path_factory):
    """Create the sample workbook once per test module."""
    path = str(tmp_path_factory.mktemp("data") / "sample.xlsx")
    create_sample_workbook(path)
    return path


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _result_shape_ok(data: dict):
    """Assert standard result dict shape."""
    assert "sheet_name" in data
    assert "regions" in data
    assert isinstance(data["regions"], list)
    assert "sampled" in data
    assert "total_rows" in data
    assert "sampled_rows" in data
    for reg in data["regions"]:
        assert "headers" in reg
        assert "rows" in reg
        assert "formulas" in reg
        assert "min_row" in reg
        assert "max_row" in reg
        assert "min_col" in reg
        assert "max_col" in reg


# ---------------------------------------------------------------------------
# DataFrame loading
# ---------------------------------------------------------------------------

class TestLoadFrames:

    def test_loads_non_empty(self, sample_wb):
        df_v, df_f = load_sheet_frames(sample_wb, "Inputs")
        assert not df_v.empty
        assert not df_f.empty

    def test_formulas_in_frame(self, sample_wb):
        _, df_f = load_sheet_frames(sample_wb, "Inputs")
        # D2 should contain the formula =B2*C2
        val = df_f.loc[2, 4]  # row 2, col 4 (D)
        assert isinstance(val, str) and val.startswith("=")

    def test_values_frame_no_formulas(self, sample_wb):
        df_v, _ = load_sheet_frames(sample_wb, "Inputs")
        # D2 should hold a cached numeric value, not a formula string
        val = df_v.loc[2, 4]
        assert not (isinstance(val, str) and val.startswith("="))


# ---------------------------------------------------------------------------
# Region detection
# ---------------------------------------------------------------------------

class TestDetectRegions:

    def test_detects_regions(self, sample_wb):
        _, df_f = load_sheet_frames(sample_wb, "Inputs")
        regions = detect_regions(df_f)
        assert len(regions) >= 1

    def test_region_has_header(self, sample_wb):
        _, df_f = load_sheet_frames(sample_wb, "Inputs")
        regions = detect_regions(df_f)
        # First region should have a header (row 1 with Item, Price, …)
        assert regions[0]["header_row"] is not None

    def test_region_bounds(self, sample_wb):
        _, df_f = load_sheet_frames(sample_wb, "Inputs")
        regions = detect_regions(df_f)
        for reg in regions:
            assert reg["min_row"] <= reg["max_row"]
            assert reg["min_col"] <= reg["max_col"]


# ---------------------------------------------------------------------------
# Extraction
# ---------------------------------------------------------------------------

class TestExtraction:

    def test_basic_extraction(self, sample_wb):
        data = extract_sheet_data(sample_wb, "Inputs")
        _result_shape_ok(data)
        assert data["sheet_name"] == "Inputs"

    def test_formulas_present(self, sample_wb):
        data = extract_sheet_data(sample_wb, "Inputs")
        all_formulas = []
        for reg in data["regions"]:
            all_formulas.extend(reg["formulas"])
        assert len(all_formulas) > 0

    def test_values_present(self, sample_wb):
        data = extract_sheet_data(sample_wb, "Inputs")
        has_value = False
        for reg in data["regions"]:
            for row in reg["rows"]:
                for v in row["values"]:
                    if v is not None:
                        has_value = True
                        break
        assert has_value

    def test_small_sheet_not_sampled(self, sample_wb):
        data = extract_sheet_data(sample_wb, "Inputs")
        assert data["sampled"] is False

    def test_all_sheets(self, sample_wb):
        from openpyxl import load_workbook
        wb = load_workbook(sample_wb)
        names = wb.sheetnames
        wb.close()
        for name in names:
            data = extract_sheet_data(sample_wb, name)
            _result_shape_ok(data)

    def test_headers_detected(self, sample_wb):
        data = extract_sheet_data(sample_wb, "Inputs")
        first_reg = data["regions"][0]
        assert len(first_reg["headers"]) > 0
        assert "Item" in first_reg["headers"]

    def test_cross_sheet_formulas(self, sample_wb):
        data = extract_sheet_data(sample_wb, "Summary")
        all_formulas = []
        for reg in data["regions"]:
            all_formulas.extend(reg["formulas"])
        # Summary has formulas referencing Inputs sheet
        cross = [f for f in all_formulas if "Inputs!" in f["formula"]]
        assert len(cross) > 0

    def test_isolated_regions_captured(self, sample_wb):
        """Isolated tiny regions (like single-row summaries) are kept whole."""
        data = extract_sheet_data(sample_wb, "Rates")
        # Rates has a region at row 6 with just one row (Total Rates)
        has_single_row_region = False
        for reg in data["regions"]:
            if reg["max_row"] - reg["min_row"] + 1 <= 3:
                has_single_row_region = True
                assert len(reg["rows"]) >= 1
        assert has_single_row_region


# ---------------------------------------------------------------------------
# Sampled cells & highlighting
# ---------------------------------------------------------------------------

class TestHighlighting:

    def test_sampled_cells_non_empty(self, sample_wb):
        cells = sampled_cells(sample_wb, "Inputs")
        assert len(cells) > 0

    def test_highlight_creates_file(self, sample_wb, tmp_path):
        out = str(tmp_path / "highlighted.xlsx")
        result = highlight_workbook(sample_wb, output_path=out)
        assert os.path.isfile(result)
        assert result == out

    def test_highlight_applies_fill(self, sample_wb, tmp_path):
        out = str(tmp_path / "highlighted.xlsx")
        highlight_workbook(sample_wb, sheet_name="Inputs", output_path=out)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb["Inputs"]
        highlighted = 0
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                fill = ws.cell(row=r, column=c).fill
                if fill and fill.start_color and fill.start_color.rgb == "00FFFF00":
                    highlighted += 1
        wb.close()
        assert highlighted > 0
