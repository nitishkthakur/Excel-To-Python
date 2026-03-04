"""Tests for excel_to_mapping.structured_input_generator."""

import os
import re
import sys
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from excel_to_mapping.mapper import generate_mapping_report
from excel_to_mapping.structured_input_generator import (
    generate_structured_input,
    _group_into_vectors_and_scalars,
    _extract_header_vector,
    _split_label_from_vector,
    _find_row_label,
    _find_col_headers_in_source,
    _is_financial_date,
    _are_date_headers,
)


# ──────────────────────────────────────────────────────────────────
# Fixtures
# ──────────────────────────────────────────────────────────────────

def _make_input_cells(spec):
    """Create flat cell dicts from a list of (col_letter, row, value) tuples."""
    from excel_to_python import col_letter_to_index
    return [
        {
            "col":     col,
            "row":     row,
            "col_idx": col_letter_to_index(col),
            "value":   val,
            "nf":      None,
            "cell":    f"{col}{row}",
        }
        for col, row, val in spec
    ]


def _build_timeseries_excel(path):
    """Create a minimal financial-model-style workbook for integration tests.

    Layout (sheet "Revenue"):
        Row 1: A1=(blank)  B1=2021  C1=2022  D1=2023
        Row 2: A2="Sales"  B2=100   C2=120   D2=140
        Row 3: A3="Cost"   B3=60    C3=72    D3=84
        Row 4: A4="Profit" B4=calc  C4=calc  D4=calc  (formulae)

    Layout (sheet "Config"):
        A1="TaxRate"  B1=0.25  (scalar – no formula)
        A2="Discount" B2=0.05  (scalar)
    """
    wb = Workbook()
    ws_rev = wb.active
    ws_rev.title = "Revenue"
    ws_rev["A1"] = None
    ws_rev["B1"] = 2021
    ws_rev["C1"] = 2022
    ws_rev["D1"] = 2023
    ws_rev["A2"] = "Sales"
    ws_rev["B2"] = 100
    ws_rev["C2"] = 120
    ws_rev["D2"] = 140
    ws_rev["A3"] = "Cost"
    ws_rev["B3"] = 60
    ws_rev["C3"] = 72
    ws_rev["D3"] = 84
    ws_rev["A4"] = "Profit"
    ws_rev["B4"] = "=B2-B3"
    ws_rev["C4"] = "=C2-C3"
    ws_rev["D4"] = "=D2-D3"

    ws_cfg = wb.create_sheet("Config")
    ws_cfg["A1"] = "TaxRate"
    ws_cfg["B1"] = 0.25
    ws_cfg["A2"] = "Discount"
    ws_cfg["B2"] = 0.05
    wb.save(path)
    wb.close()


def _build_dated_no_labels_excel(path):
    """Year headers in row 1; data rows have NO labels in col A.

    Sheet "Series":  B1=2021 C1=2022 D1=2023, rows 2-4 = pure numeric data.
    Expected: transposed=True (year headers), metric headers = Line1/Line2/Line3.
    """
    from openpyxl import Workbook as _WB
    wb = _WB()
    ws = wb.active
    ws.title = "Series"
    ws["B1"] = 2021
    ws["C1"] = 2022
    ws["D1"] = 2023
    for r, base in enumerate([12, 22, 32], start=2):
        ws.cell(r, 2, base)
        ws.cell(r, 3, base + 1)
        ws.cell(r, 4, base + 2)
    wb.save(path)
    wb.close()


def _build_categorical_no_labels_excel(path):
    """Non-date column headers; data rows have NO labels in col A.

    Sheet "Categories": B1="Region_A" C1="Region_B" D1="Region_C", rows 2-4 numeric.
    Expected: transposed=False (no dates), row labels in col A = Line1/Line2/Line3.
    """
    from openpyxl import Workbook as _WB
    wb = _WB()
    ws = wb.active
    ws.title = "Categories"
    ws["B1"] = "Region_A"
    ws["C1"] = "Region_B"
    ws["D1"] = "Region_C"
    for r, base in enumerate([10, 40, 70], start=2):
        ws.cell(r, 2, base)
        ws.cell(r, 3, base + 10)
        ws.cell(r, 4, base + 20)
    wb.save(path)
    wb.close()


# ──────────────────────────────────────────────────────────────────
# Unit tests
# ──────────────────────────────────────────────────────────────────

class TestGroupVectorsAndScalars(unittest.TestCase):
    """Tests for _group_into_vectors_and_scalars."""

    def test_contiguous_horizontal_run_becomes_vector(self):
        inputs = _make_input_cells([
            ("C", 2, 100), ("D", 2, 200), ("E", 2, 300),
        ])
        vecs, scalars = _group_into_vectors_and_scalars(inputs)
        self.assertEqual(len(vecs), 1)
        self.assertEqual(len(vecs[0]), 3)
        self.assertEqual(len(scalars), 0)

    def test_isolated_cell_becomes_scalar(self):
        inputs = _make_input_cells([("B", 5, "hello")])
        vecs, scalars = _group_into_vectors_and_scalars(inputs)
        self.assertEqual(len(vecs), 0)
        self.assertEqual(len(scalars), 1)

    def test_gap_breaks_run(self):
        # Cols B, C, E (D is missing) → two scalars + (B,C) as vector
        inputs = _make_input_cells([
            ("B", 1, 10), ("C", 1, 20), ("E", 1, 40),
        ])
        vecs, scalars = _group_into_vectors_and_scalars(inputs)
        self.assertEqual(len(vecs), 1)
        self.assertEqual(len(vecs[0]), 2)   # B–C
        self.assertEqual(len(scalars), 1)   # E alone

    def test_mixed_rows(self):
        # A cells are string labels that happen to be contiguous with B/C/D;
        # _group_into_vectors_and_scalars groups by position, not value type.
        inputs = _make_input_cells([
            ("A", 1, "label"),               # contiguous with B1 → same vector
            ("B", 1, 10), ("C", 1, 20),
            ("A", 2, "label2"),              # contiguous with B2,C2,D2
            ("B", 2, 30), ("C", 2, 40), ("D", 2, 50),
        ])
        vecs, scalars = _group_into_vectors_and_scalars(inputs)
        # All cells in each row are contiguous → 2 vectors, no isolated scalars
        self.assertEqual(len(vecs), 2)
        self.assertEqual(len(scalars), 0)
        lengths = sorted(len(v) for v in vecs)
        self.assertEqual(lengths, [3, 4])  # A+B+C and A+B+C+D

    def test_empty_input(self):
        vecs, scalars = _group_into_vectors_and_scalars([])
        self.assertEqual(vecs, [])
        self.assertEqual(scalars, [])


class TestSplitLabelFromVector(unittest.TestCase):
    """Tests for _split_label_from_vector."""

    def test_first_cell_string_becomes_label(self):
        vec = _make_input_cells([("B", 2, "Sales"), ("C", 2, 100), ("D", 2, 200)])
        label, data = _split_label_from_vector(vec)
        self.assertEqual(label, "Sales")
        self.assertEqual(len(data), 2)

    def test_first_cell_number_no_split(self):
        vec = _make_input_cells([("C", 2, 100), ("D", 2, 200)])
        label, data = _split_label_from_vector(vec)
        self.assertIsNone(label)
        self.assertEqual(len(data), 2)

    def test_empty_vector(self):
        label, data = _split_label_from_vector([])
        self.assertIsNone(label)
        self.assertEqual(data, [])

    def test_whitespace_string_not_treated_as_label(self):
        vec = _make_input_cells([("B", 2, "  "), ("C", 2, 100)])
        label, data = _split_label_from_vector(vec)
        self.assertIsNone(label)
        self.assertEqual(len(data), 2)


class TestExtractHeaderVector(unittest.TestCase):
    """Tests for _extract_header_vector."""

    def test_integer_year_vector_extracted(self):
        vectors = [
            _make_input_cells([("C", 1, 2020), ("D", 1, 2021), ("E", 1, 2022)]),
            _make_input_cells([("C", 2, 100), ("D", 2, 200), ("E", 2, 300)]),
        ]
        forced, remaining = _extract_header_vector(vectors)
        self.assertIsNotNone(forced)
        self.assertIn(3, forced)   # col C index = 3
        self.assertEqual(forced[3], "2020")
        self.assertEqual(len(remaining), 1)

    def test_string_year_vector_extracted(self):
        """'2018E', '2019E', '2020E' should be treated as headers."""
        vectors = [
            _make_input_cells([("K", 2, "2018E"), ("L", 2, "2019E"), ("M", 2, "2020E")]),
            _make_input_cells([("K", 3, 0.5), ("L", 3, 0.6), ("M", 3, 0.7)]),
        ]
        year_vec = vectors[0]
        # Inject a label so _split_label_from_vector strips it
        year_vec_with_label = _make_input_cells([("J", 2, "Year")]) + year_vec
        forced, remaining = _extract_header_vector([year_vec_with_label, vectors[1]])
        self.assertIsNotNone(forced)
        self.assertEqual(len(remaining), 1)

    def test_data_vector_not_extracted(self):
        """Vectors with actual financial data should NOT be extracted."""
        vectors = [
            _make_input_cells([("B", 2, "Revenue"), ("C", 2, 100.5), ("D", 2, 200.5)]),
        ]
        forced, remaining = _extract_header_vector(vectors)
        self.assertIsNone(forced)
        self.assertEqual(len(remaining), 1)

    def test_no_vectors(self):
        forced, remaining = _extract_header_vector([])
        self.assertIsNone(forced)
        self.assertEqual(remaining, [])


class TestFindRowLabel(unittest.TestCase):
    """Tests for _find_row_label."""

    def _make_ws(self, cell_values):
        """Create an in-memory worksheet (just the first row)."""
        wb = Workbook()
        ws = wb.active
        for (r, c), val in cell_values.items():
            ws.cell(r, c, val)
        return ws

    def test_finds_string_label_to_left(self):
        ws = self._make_ws({(2, 1): "Sales", (2, 2): None, (2, 3): None})
        label = _find_row_label(ws, 2, 4)  # start at col 4 (D)
        self.assertEqual(label, "Sales")

    def test_skips_numeric_non_label(self):
        # Col A has a number, col B has text – text should be returned
        ws = self._make_ws({(2, 1): "Label", (2, 2): 999, (2, 3): None})
        # Start at col D (idx 4) – col C is None, col B is 999, col A is "Label"
        label = _find_row_label(ws, 2, 4)
        self.assertEqual(label, "Label")

    def test_none_when_no_text_found(self):
        ws = self._make_ws({(2, 1): 42, (2, 2): 100})
        label = _find_row_label(ws, 2, 3)
        self.assertIsNone(label)

    def test_none_for_none_worksheet(self):
        label = _find_row_label(None, 1, 5)
        self.assertIsNone(label)


class TestFindColHeaders(unittest.TestCase):
    """Tests for _find_col_headers_in_source."""

    def _make_ws(self, rows_cols):
        wb = Workbook()
        ws = wb.active
        for r, c_vals in rows_cols.items():
            for c, v in c_vals.items():
                ws.cell(r, c, v)
        return ws

    def test_picks_year_row_over_data_row(self):
        # Row 1: years; Row 2: large numbers
        ws = self._make_ws({
            1: {3: 2020, 4: 2021, 5: 2022},
            2: {3: 99999, 4: 88888, 5: 77777},
        })
        headers = _find_col_headers_in_source(ws, [3, 4, 5], max_data_row=3)
        self.assertEqual(headers.get(3), "2020")

    def test_falls_back_to_col_letters_if_no_header(self):
        ws = self._make_ws({})
        from excel_to_python import index_to_col_letter
        headers = _find_col_headers_in_source(ws, [3, 4, 5], max_data_row=1)
        # max_data_row=1 means search rows < 1 → nothing → fallback
        self.assertEqual(headers[3], index_to_col_letter(3))  # "C"

    def test_prefers_year_like_integers(self):
        # Row 1 has generic text, row 2 has year integers
        ws = self._make_ws({
            1: {3: "Item A", 4: "Item B"},
            2: {3: 2021, 4: 2022},
            3: {3: 100, 4: 200},
        })
        # Search rows < 3 (max_data_row=3)
        headers = _find_col_headers_in_source(ws, [3, 4], max_data_row=3)
        self.assertEqual(headers[3], "2021")


# ──────────────────────────────────────────────────────────────────
# Integration tests
# ──────────────────────────────────────────────────────────────────

class TestGenerateStructuredInput(unittest.TestCase):
    """End-to-end integration tests for generate_structured_input."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.excel_path   = os.path.join(self.tmpdir, "source.xlsx")
        self.mapping_path = os.path.join(self.tmpdir, "mapping_report.xlsx")
        self.output_path  = os.path.join(self.tmpdir, "structured_input.xlsx")
        _build_timeseries_excel(self.excel_path)
        generate_mapping_report(
            self.excel_path,
            output_path=self.mapping_path,
        )

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _run(self):
        return generate_structured_input(
            self.mapping_path,
            excel_path=self.excel_path,
            output_path=self.output_path,
        )

    # ── sheet ordering ──────────────────────────────────────────────

    def test_output_file_created(self):
        self._run()
        self.assertTrue(os.path.exists(self.output_path))

    def test_sheet_order_index_first_config_second(self):
        self._run()
        wb = load_workbook(self.output_path)
        sheets = wb.sheetnames
        self.assertEqual(sheets[0], "Index", "Index must be the first sheet")
        self.assertEqual(sheets[1], "Config", "Config must be the second sheet")

    def test_revenue_sheet_present(self):
        self._run()
        wb = load_workbook(self.output_path)
        self.assertIn("Revenue", wb.sheetnames)

    # ── Revenue vector sheet ─────────────────────────────────────────

    def test_revenue_headers_contain_year_labels(self):
        """Year labels end up in col A (transposed layout: dates as rows)."""
        self._run()
        wb = load_workbook(self.output_path)
        ws = wb["Revenue"]
        # In transposed mode col A (row 1) is the period header, and rows 2+ hold year values.
        col_a_vals = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        year_pattern = re.compile(r"202[123]")
        self.assertTrue(
            any(v and year_pattern.search(str(v)) for v in col_a_vals),
            f"Expected year labels in col A, got: {col_a_vals}",
        )

    def test_revenue_sales_row_has_correct_values(self):
        """In transposed layout 'Sales' is a column header; values run down that column."""
        self._run()
        wb = load_workbook(self.output_path)
        ws = wb["Revenue"]
        # Transposed: find the column whose row-1 header is 'Sales'
        sales_col = None
        for c in range(2, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h and str(h).strip() == "Sales":
                sales_col = c
                break
        self.assertIsNotNone(sales_col, "Sales column not found in Revenue sheet")
        # Collect numeric values from that column (skipping the header row)
        col_values = [
            ws.cell(r, sales_col).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, sales_col).value is not None
        ]
        self.assertIn(100, col_values, "Expected Sales=100 (2021)")
        self.assertIn(120, col_values, "Expected Sales=120 (2022)")
        self.assertIn(140, col_values, "Expected Sales=140 (2023)")

    def test_profit_formula_rows_excluded_from_vector_sheet(self):
        """Profit (formula cell) must NOT appear anywhere in the Revenue sheet."""
        self._run()
        wb = load_workbook(self.output_path)
        ws = wb["Revenue"]
        # In transposed layout metric names are in row 1; also check col A for safety.
        row1_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_a_labels = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
        self.assertNotIn("Profit", row1_headers,
                         "Profit is a formula cell and must not be a column header")
        self.assertNotIn("Profit", col_a_labels,
                         "Profit is a formula cell and must not be a row label")

    # ── Config sheet ─────────────────────────────────────────────────

    def test_config_has_correct_headers(self):
        self._run()
        wb = load_workbook(self.output_path)
        ws = wb["Config"]
        headers = [ws.cell(1, c).value for c in range(1, 5)]
        self.assertEqual(headers, ["Source Sheet", "Cell Ref", "Label", "Value"])

    def test_config_contains_tax_rate(self):
        self._run()
        wb = load_workbook(self.output_path)
        ws = wb["Config"]
        found = False
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 4).value == 0.25:
                found = True
                break
        self.assertTrue(found, "TaxRate=0.25 not found in Config")

    def test_config_source_sheet_column_populated(self):
        self._run()
        wb = load_workbook(self.output_path)
        ws = wb["Config"]
        for r in range(2, ws.max_row + 1):
            sheet = ws.cell(r, 1).value
            cell_ref = ws.cell(r, 2).value
            if cell_ref and "──" not in str(cell_ref):
                self.assertIsNotNone(
                    sheet,
                    f"Source Sheet is None for row {r}",
                )

    # ── Index sheet ───────────────────────────────────────────────────

    def test_index_has_correct_column_headers(self):
        self._run()
        wb = load_workbook(self.output_path)
        ws = wb["Index"]
        headers = [ws.cell(1, c).value for c in range(1, 8)]
        expected = [
            "Input File Sheet",
            "Table",
            "Column Name",
            "Source Sheet",
            "Source Range",
            "Description",
            "Vector Length",
        ]
        self.assertEqual(headers, expected)

    def test_index_has_data_rows(self):
        self._run()
        wb = load_workbook(self.output_path)
        ws = wb["Index"]
        self.assertGreater(ws.max_row, 1, "Index sheet must have at least one data row")

    def test_index_source_sheet_always_populated(self):
        self._run()
        wb = load_workbook(self.output_path)
        ws = wb["Index"]
        for r in range(2, ws.max_row + 1):
            source_sheet = ws.cell(r, 4).value
            self.assertIsNotNone(
                source_sheet,
                f"Index row {r} is missing Source Sheet",
            )

    def test_index_references_config_and_revenue(self):
        self._run()
        wb = load_workbook(self.output_path)
        ws = wb["Index"]
        tables = {ws.cell(r, 2).value for r in range(2, ws.max_row + 1)}
        self.assertIn("Config", tables)
        self.assertIn("Revenue", tables)

    # ── fill colours ─────────────────────────────────────────────────

    def test_revenue_sheet_data_cells_have_fill(self):
        """Input value cells in vector sheets must carry a fill colour."""
        self._run()
        wb = load_workbook(self.output_path)
        ws = wb["Revenue"]
        # In transposed layout 'Sales' is a column header in row 1.
        sales_col = None
        for c in range(2, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h and str(h).strip() == "Sales":
                sales_col = c
                break
        if sales_col is not None:
            fill_rgb = ws.cell(2, sales_col).fill.fgColor.rgb
            self.assertNotEqual(fill_rgb, "00000000",
                                "Data cells must have a fill colour")

    # ── no-source-excel fallback ──────────────────────────────────────

    def test_runs_without_source_excel(self):
        """generate_structured_input must succeed without the source Excel."""
        out = os.path.join(self.tmpdir, "si_no_src.xlsx")
        generate_structured_input(self.mapping_path, output_path=out)
        self.assertTrue(os.path.exists(out))
        wb = load_workbook(out)
        self.assertIn("Index", wb.sheetnames)
        self.assertIn("Config", wb.sheetnames)


# ──────────────────────────────────────────────────────────────────
# Integration test with real Indigo.xlsx (skipped if not available)
# ──────────────────────────────────────────────────────────────────

@unittest.skipUnless(
    os.path.exists(
        os.path.join(os.path.dirname(__file__), "..", "output", "mapping_report.xlsx")
    )
    and os.path.exists(
        os.path.join(os.path.dirname(__file__), "..", "Indigo.xlsx")
    ),
    "Indigo.xlsx or mapping_report.xlsx not available",
)
class TestIndigoStructuredInput(unittest.TestCase):
    """Smoke tests against the real Indigo workbook."""

    WORKSPACE = os.path.join(os.path.dirname(__file__), "..")
    MAPPING   = os.path.join(WORKSPACE, "output", "mapping_report.xlsx")
    EXCEL     = os.path.join(WORKSPACE, "Indigo.xlsx")

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.output = os.path.join(self.tmpdir, "structured_input_indigo.xlsx")

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_generates_without_error(self):
        result = generate_structured_input(
            self.MAPPING, excel_path=self.EXCEL, output_path=self.output
        )
        self.assertTrue(os.path.exists(result))

    def test_sheet_order(self):
        generate_structured_input(
            self.MAPPING, excel_path=self.EXCEL, output_path=self.output
        )
        wb = load_workbook(self.output)
        sheets = wb.sheetnames
        self.assertEqual(sheets[0], "Index")
        self.assertEqual(sheets[1], "Config")

    def test_assumptions_sheet_has_year_headers(self):
        """In transposed layout year labels live in col A (rows 2+)."""
        generate_structured_input(
            self.MAPPING, excel_path=self.EXCEL, output_path=self.output
        )
        wb = load_workbook(self.output)
        self.assertIn("Assumptions Sheet", wb.sheetnames)
        ws = wb["Assumptions Sheet"]
        year_pattern = re.compile(r"\b20\d{2}")
        # col A rows 2+ hold the period labels when the sheet is transposed
        col_a_vals = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)
                      if ws.cell(r, 1).value is not None]
        year_labels = [v for v in col_a_vals if year_pattern.search(str(v))]
        self.assertGreater(
            len(year_labels), 3,
            f"Expected multiple year labels in col A, got: {col_a_vals[:10]}",
        )

    def test_assumptions_ask_row_correct(self):
        """In transposed layout 'ASK (in million)' is a column header in row 1."""
        generate_structured_input(
            self.MAPPING, excel_path=self.EXCEL, output_path=self.output
        )
        wb = load_workbook(self.output)
        ws = wb["Assumptions Sheet"]
        # Find the column whose row-1 header is 'ASK (in million)'
        ask_col = None
        for c in range(2, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h and "ASK" in str(h):
                ask_col = c
                break
        self.assertIsNotNone(ask_col, "ASK (in million) column not found")
        col_vals = [
            ws.cell(r, ask_col).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, ask_col).value is not None
        ]
        self.assertIn(9286, col_vals, "Expected ASK=9286 for 2010")

    def test_income_statement_labels_correct(self):
        """In transposed layout metric labels are column headers in row 1."""
        generate_structured_input(
            self.MAPPING, excel_path=self.EXCEL, output_path=self.output
        )
        wb = load_workbook(self.output)
        self.assertIn("Income statement", wb.sheetnames)
        ws = wb["Income statement"]
        # Transposed: metric names are in row 1 (cols 2+)
        metric_headers = [
            ws.cell(1, c).value
            for c in range(2, ws.max_column + 1)
            if ws.cell(1, c).value is not None
        ]
        self.assertIn("Revenue from operations", metric_headers)
        self.assertIn("Other Income", metric_headers)

    def test_config_has_many_scalars(self):
        generate_structured_input(
            self.MAPPING, excel_path=self.EXCEL, output_path=self.output
        )
        wb = load_workbook(self.output)
        ws = wb["Config"]
        data_rows = sum(
            1 for r in range(2, ws.max_row + 1)
            if ws.cell(r, 2).value and "──" not in str(ws.cell(r, 2).value)
        )
        self.assertGreater(data_rows, 50, "Expected many scalar config rows")

    def test_index_has_source_references(self):
        generate_structured_input(
            self.MAPPING, excel_path=self.EXCEL, output_path=self.output
        )
        wb = load_workbook(self.output)
        ws = wb["Index"]
        source_sheets = {
            ws.cell(r, 4).value
            for r in range(2, ws.max_row + 1)
        }
        self.assertIn("Assumptions Sheet", source_sheets)
        self.assertIn("Income statement", source_sheets)


if __name__ == "__main__":
    unittest.main()


# ──────────────────────────────────────────────────────────────────
# Tests for Line-N fallback labels
# ──────────────────────────────────────────────────────────────────

class TestLineNFallbackLabel(unittest.TestCase):
    """Verify unlabelled rows get 'Line1', 'Line2', ... not 'Row N'."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _run(self, excel_builder):
        excel_path   = os.path.join(self.tmpdir, "src.xlsx")
        mapping_path = os.path.join(self.tmpdir, "map.xlsx")
        output_path  = os.path.join(self.tmpdir, "si.xlsx")
        excel_builder(excel_path)
        generate_mapping_report(excel_path, output_path=mapping_path)
        generate_structured_input(mapping_path, excel_path=excel_path,
                                  output_path=output_path)
        return load_workbook(output_path)

    def _assert_no_row_n(self, wb):
        for sn in wb.sheetnames:
            if sn in ("Index", "Config"):
                continue
            ws = wb[sn]
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    self.assertFalse(
                        isinstance(v, str) and v.startswith("Row ") and v[4:].isdigit(),
                        f"Found 'Row N' fallback in {sn}!R{r}C{c}: {v!r}",
                    )

    # ── transposed layout (year headers, no row labels) ─────────────────────

    def test_transposed_unlabelled_metrics_get_line_n(self):
        wb = self._run(_build_dated_no_labels_excel)
        ws = wb["Series"]
        self.assertIn("Period", str(ws.cell(1, 1).value or ""),
                      "Expected transposed layout (Period in A1)")
        metric_headers = [
            ws.cell(1, c).value
            for c in range(2, ws.max_column + 1)
            if ws.cell(1, c).value is not None
        ]
        line_labels = [h for h in metric_headers
                       if isinstance(h, str) and h.startswith("Line")]
        self.assertGreaterEqual(len(line_labels), 3,
                                f"Expected Line1/2/3 metric headers, got: {metric_headers}")
        self.assertIn("Line1", line_labels)
        self.assertIn("Line2", line_labels)
        self.assertIn("Line3", line_labels)

    def test_transposed_no_row_n_labels(self):
        self._assert_no_row_n(self._run(_build_dated_no_labels_excel))

    # ── original layout (non-date headers, no row labels) ────────────────────

    def test_original_unlabelled_rows_get_line_n(self):
        wb = self._run(_build_categorical_no_labels_excel)
        ws = wb["Categories"]
        self.assertIn("Metric", str(ws.cell(1, 1).value or ""),
                      "Expected original layout (Metric in A1)")
        col_a = [
            ws.cell(r, 1).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, 1).value is not None
        ]
        line_labels = [v for v in col_a
                       if isinstance(v, str) and v.startswith("Line")]
        self.assertGreaterEqual(len(line_labels), 3,
                                f"Expected Line1/2/3 row labels, got: {col_a}")
        self.assertIn("Line1", line_labels)
        self.assertIn("Line2", line_labels)
        self.assertIn("Line3", line_labels)

    def test_original_no_row_n_labels(self):
        self._assert_no_row_n(self._run(_build_categorical_no_labels_excel))

    def test_line_counter_starts_at_one(self):
        """The first unlabelled row in a sheet must be 'Line1', not 'Line2' or higher."""
        wb = self._run(_build_categorical_no_labels_excel)
        ws = wb["Categories"]
        col_a = [
            ws.cell(r, 1).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, 1).value is not None
        ]
        line_labels = [v for v in col_a
                       if isinstance(v, str) and v.startswith("Line")]
        self.assertTrue(line_labels,
                        "Expected at least one Line-N label in col A")
        self.assertEqual(line_labels[0], "Line1",
                         f"First Line-N label must be 'Line1', got {line_labels[0]!r}")


@unittest.skipUnless(
    os.path.exists(
        os.path.join(os.path.dirname(__file__), "..", "output", "mapping_report.xlsx")
    )
    and os.path.exists(
        os.path.join(os.path.dirname(__file__), "..", "Indigo.xlsx")
    ),
    "Indigo.xlsx or mapping_report.xlsx not available",
)
class TestIndigoNoRowNLabel(unittest.TestCase):
    """Regression: Indigo structured_input.xlsx must contain no 'Row N' labels."""

    WORKSPACE = os.path.join(os.path.dirname(__file__), "..")
    MAPPING   = os.path.join(WORKSPACE, "output", "mapping_report.xlsx")
    EXCEL     = os.path.join(WORKSPACE, "Indigo.xlsx")

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.output = os.path.join(self.tmpdir, "si.xlsx")
        generate_structured_input(self.MAPPING, excel_path=self.EXCEL,
                                  output_path=self.output)

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_no_row_n_label_anywhere(self):
        wb = load_workbook(self.output)
        violations = []
        for sn in wb.sheetnames:
            if sn in ("Index", "Config"):
                continue
            ws = wb[sn]
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and v.startswith("Row ") and v[4:].isdigit():
                        violations.append(f"{sn}!R{r}C{c}={v!r}")
        self.assertEqual(violations, [],
                         f"Found 'Row N' fallbacks: {violations[:10]}")
