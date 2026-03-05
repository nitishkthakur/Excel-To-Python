"""Generate unstructured input file from mapping report.

Layer 2a creates unstructured_inputs.xlsx with the same layout as the original
Excel file, but containing only Input cells. All Calculation and Output cells
are removed, leaving a clean template that users can edit with new values.
"""

from pathlib import Path
from typing import Dict, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
from excel_pipeline.core.excel_io import load_workbook, save_workbook
from excel_pipeline.utils.logging_setup import get_logger
from excel_pipeline.utils.config import config

logger = get_logger(__name__)


class UnstructuredInputGenerator:
    """Generate unstructured input file preserving original layout."""

    def __init__(self, mapping_report_path: str):
        """
        Initialize generator.

        Args:
            mapping_report_path: Path to mapping_report.xlsx
        """
        self.mapping_report_path = mapping_report_path
        self.mapping_wb = None
        self.input_cells: Dict[str, List[Dict]] = {}  # sheet_name -> list of input cells

    def generate(self, output_path: str) -> None:
        """
        Generate unstructured_inputs.xlsx from mapping report.

        Process:
        1. Load mapping_report.xlsx
        2. Extract all Input cells (Type == "Input" AND IncludeFlag == TRUE)
        3. Create new workbook matching original structure
        4. Write values and formatting to same cell positions
        5. Save as unstructured_inputs.xlsx

        Args:
            output_path: Path to save unstructured_inputs.xlsx

        Example:
            >>> generator = UnstructuredInputGenerator("mapping_report.xlsx")
            >>> generator.generate("unstructured_inputs.xlsx")
        """
        logger.info("=" * 80)
        logger.info("LAYER 2a: Generating Unstructured Input File")
        logger.info("=" * 80)
        logger.info(f"Mapping report: {self.mapping_report_path}")
        logger.info(f"Output: {output_path}")

        # Step 1: Load mapping report
        logger.info("\n[Step 1/4] Loading mapping report...")
        self.mapping_wb = load_workbook(self.mapping_report_path, data_only=True, read_only=True)
        logger.info(f"Loaded {len(self.mapping_wb.sheetnames)} sheets")

        # Step 2: Extract Input cells
        logger.info("\n[Step 2/4] Extracting Input cells...")
        self._extract_input_cells()

        total_inputs = sum(len(cells) for cells in self.input_cells.values())
        logger.info(f"Found {total_inputs} Input cells across {len(self.input_cells)} sheets")

        # Step 3: Build output workbook
        logger.info("\n[Step 3/4] Building unstructured input workbook...")
        output_wb = self._build_output_workbook()

        # Step 4: Save
        logger.info("\n[Step 4/4] Saving unstructured input file...")
        save_workbook(output_wb, output_path)

        # Summary
        logger.info("\n" + "=" * 80)
        logger.info("LAYER 2a: Complete!")
        logger.info("=" * 80)
        logger.info(f"Output: {output_path}")
        logger.info(f"Total Input cells: {total_inputs}")
        logger.info("=" * 80)

    def _extract_input_cells(self) -> None:
        """Extract all Input cells from mapping report."""
        # Iterate through all sheets except _Metadata
        for sheet_name in self.mapping_wb.sheetnames:
            if sheet_name == "_Metadata":
                continue

            sheet = self.mapping_wb[sheet_name]
            logger.debug(f"Processing sheet: {sheet_name}")

            # Read header row to find column indices
            headers = {}
            for col_idx, cell in enumerate(sheet[1], start=1):
                if cell.value:
                    headers[cell.value] = col_idx

            # Extract Input cells
            sheet_inputs = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                # Get cell values by column name
                cell_type = row[headers["Type"] - 1].value
                include_flag = row[headers["IncludeFlag"] - 1].value

                # Only include Input cells with IncludeFlag == TRUE
                if cell_type == "Input" and include_flag:
                    cell_data = {
                        'row_num': row[headers["RowNum"] - 1].value,
                        'col_num': row[headers["ColNum"] - 1].value,
                        'cell': row[headers["Cell"] - 1].value,
                        'value': row[headers["Value"] - 1].value,
                        'number_format': row[headers["NumberFormat"] - 1].value,
                        'font_bold': row[headers["FontBold"] - 1].value,
                        'font_italic': row[headers["FontItalic"] - 1].value,
                        'font_size': row[headers["FontSize"] - 1].value,
                        'font_color': row[headers["FontColor"] - 1].value,
                        'fill_color': row[headers["FillColor"] - 1].value,
                        'alignment': row[headers["Alignment"] - 1].value,
                        'wrap_text': row[headers["WrapText"] - 1].value,
                    }
                    sheet_inputs.append(cell_data)

            if sheet_inputs:
                self.input_cells[sheet_name] = sheet_inputs
                logger.debug(f"  Found {len(sheet_inputs)} Input cells in {sheet_name}")

    def _build_output_workbook(self) -> Workbook:
        """
        Build output workbook with Input cells only.

        Returns:
            Workbook with Input cells in original positions
        """
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create sheets and populate with Input cells
        for sheet_name, input_cells in self.input_cells.items():
            logger.debug(f"Creating sheet: {sheet_name}")
            ws = wb.create_sheet(title=sheet_name)

            # Write each Input cell to its original position
            for cell_data in input_cells:
                self._write_input_cell(ws, cell_data)

            logger.debug(f"  Wrote {len(input_cells)} cells to {sheet_name}")

        return wb

    def _write_input_cell(self, ws, cell_data: Dict) -> None:
        """
        Write a single Input cell with formatting.

        Args:
            ws: Worksheet to write to
            cell_data: Dictionary with cell metadata
        """
        # Handle consolidated ranges (e.g., "5-10" or "D[-F]")
        # For unstructured, we need individual cells, so we'll parse the first value
        row_num_str = str(cell_data['row_num'])
        col_num_str = str(cell_data['col_num'])

        # Extract first row number
        if '-' in row_num_str:
            row_num = int(row_num_str.split('-')[0])
        else:
            row_num = int(row_num_str)

        # Extract first column letter
        if '[-' in col_num_str:
            col_letter = col_num_str.split('[-')[0]
        else:
            col_letter = col_num_str

        # Get cell
        cell = ws[f"{col_letter}{row_num}"]

        # Write value
        cell.value = cell_data['value']

        # Apply number format
        if cell_data['number_format']:
            cell.number_format = cell_data['number_format']

        # Apply font
        font_kwargs = {}
        if cell_data['font_bold']:
            font_kwargs['bold'] = True
        if cell_data['font_italic']:
            font_kwargs['italic'] = True
        if cell_data['font_size']:
            font_kwargs['size'] = cell_data['font_size']
        if cell_data['font_color']:
            # Remove # prefix if present
            color = cell_data['font_color'].replace('#', '')
            # Ensure ARGB format (8 hex digits) - prepend FF if RGB (6 digits)
            if len(color) == 6:
                color = 'FF' + color
            if len(color) == 8:
                font_kwargs['color'] = color

        if font_kwargs:
            cell.font = Font(**font_kwargs)

        # Apply fill
        if cell_data['fill_color']:
            color = cell_data['fill_color'].replace('#', '')
            # Ensure ARGB format (8 hex digits) - prepend FF if RGB (6 digits)
            if len(color) == 6:
                color = 'FF' + color
            if len(color) == 8:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        # Apply alignment
        if cell_data['alignment']:
            cell.alignment = Alignment(horizontal=cell_data['alignment'],
                                      wrap_text=cell_data['wrap_text'] or False)


def generate_unstructured_inputs(mapping_report_path: str, output_path: str) -> None:
    """
    Generate unstructured input file from mapping report.

    This is the main entry point for Layer 2a.

    Args:
        mapping_report_path: Path to mapping_report.xlsx
        output_path: Path to save unstructured_inputs.xlsx

    Example:
        >>> generate_unstructured_inputs("mapping_report.xlsx", "unstructured_inputs.xlsx")
    """
    generator = UnstructuredInputGenerator(mapping_report_path)
    generator.generate(output_path)


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 3:
        print("Usage: python -m excel_pipeline.layer2.unstructured_generator <mapping_report.xlsx> <unstructured_inputs.xlsx>")
        sys.exit(1)

    generate_unstructured_inputs(sys.argv[1], sys.argv[2])
