"""Generate structured input file from mapping report.

Layer 2b creates structured_input.xlsx with clean tabular format, automatically
detecting input patches, identifying headers, and transposing tables with
financial dates. Produces Config sheet for scalars and Index sheet for metadata.
"""

from pathlib import Path
from typing import Dict, List, Tuple, Set, Optional, Any
from dataclasses import dataclass
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from excel_pipeline.core.excel_io import load_workbook, save_workbook
from excel_pipeline.utils.helpers import is_financial_date
from excel_pipeline.utils.logging_setup import get_logger
from excel_pipeline.utils.config import config

logger = get_logger(__name__)


@dataclass
class InputPatch:
    """Represents a contiguous rectangle of Input cells."""
    sheet_name: str
    cells: List[Tuple[int, int]]  # List of (row, col) tuples
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    patch_type: str  # "scalar", "vector", "table"
    row_headers: List[Any]  # Row labels (leftmost column)
    col_headers: List[Any]  # Column labels (top row)
    data: List[List[Any]]  # 2D data grid (excluding headers)
    should_transpose: bool  # True if table should be transposed
    table_id: str  # Unique identifier for this patch

    def get_cell_range(self) -> str:
        """Get Excel-style cell range (e.g., 'B5:F20')."""
        from openpyxl.utils import get_column_letter
        start = f"{get_column_letter(self.min_col)}{self.min_row}"
        end = f"{get_column_letter(self.max_col)}{self.max_row}"
        return f"{start}:{end}"


class StructuredInputGenerator:
    """Generate structured input file with tabular format."""

    def __init__(self, mapping_report_path: str):
        """
        Initialize generator.

        Args:
            mapping_report_path: Path to mapping_report.xlsx
        """
        self.mapping_report_path = mapping_report_path
        self.mapping_wb = None
        self.input_cells: Dict[str, List[Dict]] = {}  # sheet_name -> list of input cells
        self.patches: List[InputPatch] = []
        self.scalars: List[Dict] = []  # Scalar values for Config sheet
        self.next_table_id = 1

    def generate(self, output_path: str) -> None:
        """
        Generate structured_input.xlsx from mapping report.

        Process:
        1. Load mapping_report.xlsx
        2. Extract all Input cells
        3. Find contiguous patches using flood-fill
        4. Classify patches (scalar/vector/table)
        5. Detect headers and apply auto-transpose
        6. Separate scalars to Config sheet
        7. Create per-sheet tables
        8. Write Index sheet
        9. Save workbook

        Args:
            output_path: Path to save structured_input.xlsx
        """
        logger.info("=" * 80)
        logger.info("LAYER 2b: Generating Structured Input File")
        logger.info("=" * 80)
        logger.info(f"Mapping report: {self.mapping_report_path}")
        logger.info(f"Output: {output_path}")

        # Step 1: Load mapping report
        logger.info("\n[Step 1/7] Loading mapping report...")
        self.mapping_wb = load_workbook(self.mapping_report_path, data_only=True, read_only=True)
        logger.info(f"Loaded {len(self.mapping_wb.sheetnames)} sheets")

        # Step 2: Extract Input cells
        logger.info("\n[Step 2/7] Extracting Input cells...")
        self._extract_input_cells()
        total_inputs = sum(len(cells) for cells in self.input_cells.values())
        logger.info(f"Found {total_inputs} Input cells")

        # Step 3: Find contiguous patches
        logger.info("\n[Step 3/7] Detecting input patches...")
        self._find_input_patches()
        logger.info(f"Found {len(self.patches)} patches "
                   f"({len(self.scalars)} scalars, {len(self.patches) - len(self.scalars)} tables)")

        # Step 4: Apply auto-transpose logic
        logger.info("\n[Step 4/7] Applying auto-transpose to tables with financial dates...")
        transposed_count = self._apply_auto_transpose()
        logger.info(f"Transposed {transposed_count} tables")

        # Step 5: Build output workbook
        logger.info("\n[Step 5/7] Building structured input workbook...")
        output_wb = self._build_output_workbook()

        # Step 6: Write Index sheet
        logger.info("\n[Step 6/7] Writing Index sheet...")
        self._write_index_sheet(output_wb)

        # Step 7: Save
        logger.info("\n[Step 7/7] Saving structured input file...")
        save_workbook(output_wb, output_path)

        # Summary
        logger.info("\n" + "=" * 80)
        logger.info("LAYER 2b: Complete!")
        logger.info("=" * 80)
        logger.info(f"Output: {output_path}")
        logger.info(f"Total Input cells: {total_inputs}")
        logger.info(f"Patches: {len(self.patches)} "
                   f"({len(self.scalars)} scalars, {len(self.patches) - len(self.scalars)} tables)")
        logger.info(f"Transposed tables: {transposed_count}")
        logger.info("=" * 80)

    def _extract_input_cells(self) -> None:
        """Extract all Input cells from mapping report."""
        for sheet_name in self.mapping_wb.sheetnames:
            if sheet_name == "_Metadata":
                continue

            sheet = self.mapping_wb[sheet_name]

            # Read header row
            headers = {}
            for col_idx, cell in enumerate(sheet[1], start=1):
                if cell.value:
                    headers[cell.value] = col_idx

            # Extract Input cells
            sheet_inputs = []
            for row in sheet.iter_rows(min_row=2):
                cell_type = row[headers["Type"] - 1].value
                include_flag = row[headers["IncludeFlag"] - 1].value

                if cell_type == "Input" and include_flag:
                    # Parse row/col numbers (handle consolidated ranges)
                    row_num_str = str(row[headers["RowNum"] - 1].value)
                    col_num_str = str(row[headers["ColNum"] - 1].value)

                    # Extract first row/col for consolidated ranges
                    if '-' in row_num_str:
                        row_num = int(row_num_str.split('-')[0])
                    else:
                        row_num = int(row_num_str)

                    if '[-' in col_num_str:
                        col_letter = col_num_str.split('[-')[0]
                    else:
                        col_letter = col_num_str

                    # Convert column letter to number
                    from openpyxl.utils import column_index_from_string
                    col_num = column_index_from_string(col_letter)

                    cell_data = {
                        'row': row_num,
                        'col': col_num,
                        'value': row[headers["Value"] - 1].value,
                    }
                    sheet_inputs.append(cell_data)

            if sheet_inputs:
                self.input_cells[sheet_name] = sheet_inputs
                logger.debug(f"  {sheet_name}: {len(sheet_inputs)} Input cells")

    def _find_input_patches(self) -> None:
        """Find contiguous rectangles of Input cells using flood-fill."""
        for sheet_name, cells in self.input_cells.items():
            # Create grid of Input cell positions
            cell_grid = {(c['row'], c['col']): c['value'] for c in cells}
            visited = set()

            for cell_pos in cell_grid.keys():
                if cell_pos in visited:
                    continue

                # Flood-fill to find contiguous patch
                patch_cells = self._flood_fill(cell_pos, cell_grid, visited)

                if patch_cells:
                    # Create InputPatch object
                    patch = self._create_patch(sheet_name, patch_cells, cell_grid)

                    if patch.patch_type == "scalar":
                        # Add to scalars list
                        self.scalars.append({
                            'parameter': f"{sheet_name}_{self.next_table_id}",
                            'value': patch.data[0][0] if patch.data else None,
                            'source_sheet': sheet_name,
                            'source_cell': patch.get_cell_range(),
                        })

                    self.patches.append(patch)
                    self.next_table_id += 1

    def _flood_fill(self, start_pos: Tuple[int, int],
                   cell_grid: Dict[Tuple[int, int], Any],
                   visited: Set[Tuple[int, int]]) -> List[Tuple[int, int]]:
        """
        Flood-fill algorithm to find contiguous rectangle.

        Returns list of (row, col) positions in the patch.
        """
        patch_cells = []
        queue = [start_pos]

        while queue:
            pos = queue.pop(0)
            if pos in visited or pos not in cell_grid:
                continue

            visited.add(pos)
            patch_cells.append(pos)

            # Check 4 neighbors (up, down, left, right)
            row, col = pos
            neighbors = [
                (row - 1, col),
                (row + 1, col),
                (row, col - 1),
                (row, col + 1),
            ]

            for neighbor in neighbors:
                if neighbor in cell_grid and neighbor not in visited:
                    queue.append(neighbor)

        return patch_cells

    def _create_patch(self, sheet_name: str,
                     patch_cells: List[Tuple[int, int]],
                     cell_grid: Dict[Tuple[int, int], Any]) -> InputPatch:
        """Create InputPatch object from list of cell positions."""
        # Find bounding box
        rows = [pos[0] for pos in patch_cells]
        cols = [pos[1] for pos in patch_cells]
        min_row, max_row = min(rows), max(rows)
        min_col, max_col = min(cols), max(cols)

        # Determine patch type
        num_cells = len(patch_cells)
        if num_cells == 1:
            patch_type = "scalar"
        elif num_cells == 2:
            patch_type = "vector"
        else:
            patch_type = "table"

        # Extract headers and data
        row_headers = []
        col_headers = []
        data = []

        if patch_type == "scalar":
            # Single cell
            data = [[cell_grid[patch_cells[0]]]]
        elif patch_type == "vector":
            # Two cells - assume one is header, one is value
            sorted_cells = sorted(patch_cells)
            row_headers = [cell_grid[sorted_cells[0]]]
            data = [[cell_grid[sorted_cells[1]]]]
        else:
            # Table - assume top row is column headers, leftmost column is row headers
            # Get top row (column headers)
            col_headers = [cell_grid.get((min_row, c), "") for c in range(min_col, max_col + 1)]

            # Get leftmost column (row headers) excluding top-left cell
            row_headers = [cell_grid.get((r, min_col), "") for r in range(min_row + 1, max_row + 1)]

            # Get data (excluding header row and column)
            data = []
            for r in range(min_row + 1, max_row + 1):
                row_data = []
                for c in range(min_col + 1, max_col + 1):
                    row_data.append(cell_grid.get((r, c), None))
                data.append(row_data)

        # Create patch
        patch = InputPatch(
            sheet_name=sheet_name,
            cells=patch_cells,
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            patch_type=patch_type,
            row_headers=row_headers,
            col_headers=col_headers,
            data=data,
            should_transpose=False,
            table_id=f"{sheet_name}_{self.next_table_id}"
        )

        return patch

    def _apply_auto_transpose(self) -> int:
        """
        Apply auto-transpose to tables with financial dates as column headers.

        Returns count of transposed tables.
        """
        transposed_count = 0

        for patch in self.patches:
            if patch.patch_type != "table":
                continue

            # Check if column headers are financial dates
            if self._should_transpose(patch.col_headers):
                patch.should_transpose = True
                transposed_count += 1
                logger.debug(f"  Transposing {patch.table_id}: {len(patch.col_headers)} date columns")

        return transposed_count

    def _should_transpose(self, col_headers: List[Any]) -> bool:
        """
        Determine if table should be transposed based on column headers.

        Returns True if >50% of headers are financial dates.
        """
        if not col_headers:
            return False

        date_count = sum(1 for h in col_headers if is_financial_date(h))
        return date_count / len(col_headers) > 0.5

    def _build_output_workbook(self) -> Workbook:
        """Build structured output workbook."""
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create Config sheet for scalars
        if self.scalars:
            self._write_config_sheet(wb)

        # Group non-scalar patches by sheet
        table_patches = [p for p in self.patches if p.patch_type != "scalar"]
        patches_by_sheet = {}
        for patch in table_patches:
            if patch.sheet_name not in patches_by_sheet:
                patches_by_sheet[patch.sheet_name] = []
            patches_by_sheet[patch.sheet_name].append(patch)

        # Create sheet for each source sheet with tables
        for sheet_name, patches in patches_by_sheet.items():
            # If multiple patches from same source sheet, number them
            for idx, patch in enumerate(patches, start=1):
                table_name = f"{sheet_name}_{idx}" if len(patches) > 1 else sheet_name
                self._write_table_sheet(wb, table_name, patch)

        return wb

    def _write_config_sheet(self, wb: Workbook) -> None:
        """Write Config sheet with scalar values."""
        ws = wb.create_sheet(title="Config")

        # Headers
        headers = ["Parameter", "Value", "SourceSheet", "SourceCell"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Write scalars
        for row_idx, scalar in enumerate(self.scalars, start=2):
            ws.cell(row=row_idx, column=1, value=scalar['parameter'])
            ws.cell(row=row_idx, column=2, value=scalar['value'])
            ws.cell(row=row_idx, column=3, value=scalar['source_sheet'])
            ws.cell(row=row_idx, column=4, value=scalar['source_cell'])

        # Auto-width columns
        for col_idx in range(1, 5):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        logger.debug(f"  Config sheet: {len(self.scalars)} scalars")

    def _write_table_sheet(self, wb: Workbook, table_name: str, patch: InputPatch) -> None:
        """Write a table to its own sheet."""
        ws = wb.create_sheet(title=table_name[:31])  # Excel sheet name limit

        # Apply transpose if needed
        if patch.should_transpose:
            # Transpose: rows become columns, columns become rows
            # Column headers become row labels (first column)
            # Row headers become column labels (first row)
            # Data is transposed

            # Write row labels (original column headers) in first column
            for row_idx, header in enumerate(patch.col_headers, start=2):
                ws.cell(row=row_idx, column=1, value=header)
                ws.cell(row=row_idx, column=1).font = Font(bold=True)

            # Write column labels (original row headers) in first row
            ws.cell(row=1, column=1, value="Period")  # Top-left label
            ws.cell(row=1, column=1).font = Font(bold=True)
            for col_idx, header in enumerate(patch.row_headers, start=2):
                ws.cell(row=1, column=col_idx, value=header)
                ws.cell(row=1, column=col_idx).font = Font(bold=True)

            # Write transposed data
            for col_idx, data_row in enumerate(patch.data, start=2):
                for row_idx, value in enumerate(data_row, start=2):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        else:
            # No transpose - write as-is

            # Write column headers in first row
            ws.cell(row=1, column=1, value="")  # Top-left empty
            for col_idx, header in enumerate(patch.col_headers, start=2):
                ws.cell(row=1, column=col_idx, value=header)
                ws.cell(row=1, column=col_idx).font = Font(bold=True)

            # Write row headers in first column
            for row_idx, header in enumerate(patch.row_headers, start=2):
                ws.cell(row=row_idx, column=1, value=header)
                ws.cell(row=row_idx, column=1).font = Font(bold=True)

            # Write data
            for row_idx, data_row in enumerate(patch.data, start=2):
                for col_idx, value in enumerate(data_row, start=2):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-width columns
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        logger.debug(f"  {table_name}: {len(patch.row_headers)}x{len(patch.col_headers)} "
                    f"{'(transposed)' if patch.should_transpose else ''}")

    def _write_index_sheet(self, wb: Workbook) -> None:
        """Write Index sheet with metadata about all tables."""
        ws = wb.create_sheet(title="Index", index=0)  # First sheet

        # Headers
        headers = ["StructuredTable", "SourceSheet", "CellRange", "TableType", "Transposed", "Notes"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Write table metadata
        row_idx = 2

        # Config scalars
        if self.scalars:
            ws.cell(row=row_idx, column=1, value="Config")
            ws.cell(row=row_idx, column=2, value="(multiple)")
            ws.cell(row=row_idx, column=3, value="(multiple)")
            ws.cell(row=row_idx, column=4, value="Scalar")
            ws.cell(row=row_idx, column=5, value=False)
            ws.cell(row=row_idx, column=6, value=f"{len(self.scalars)} scalar parameters")
            row_idx += 1

        # Tables
        table_patches = [p for p in self.patches if p.patch_type != "scalar"]
        patches_by_sheet = {}
        for patch in table_patches:
            if patch.sheet_name not in patches_by_sheet:
                patches_by_sheet[patch.sheet_name] = []
            patches_by_sheet[patch.sheet_name].append(patch)

        for sheet_name, patches in patches_by_sheet.items():
            for idx, patch in enumerate(patches, start=1):
                table_name = f"{sheet_name}_{idx}" if len(patches) > 1 else sheet_name

                ws.cell(row=row_idx, column=1, value=table_name[:31])
                ws.cell(row=row_idx, column=2, value=patch.sheet_name)
                ws.cell(row=row_idx, column=3, value=patch.get_cell_range())
                ws.cell(row=row_idx, column=4, value=patch.patch_type.capitalize())
                ws.cell(row=row_idx, column=5, value=patch.should_transpose)

                # Generate notes
                notes = f"{len(patch.row_headers)} rows × {len(patch.col_headers)} columns"
                if patch.should_transpose:
                    notes += " (transposed for time-series)"
                ws.cell(row=row_idx, column=6, value=notes)

                row_idx += 1

        # Auto-width columns
        for col_idx in range(1, 7):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        logger.debug(f"  Index sheet: {row_idx - 2} entries")


def generate_structured_inputs(mapping_report_path: str, output_path: str) -> None:
    """
    Generate structured input file from mapping report.

    This is the main entry point for Layer 2b.

    Args:
        mapping_report_path: Path to mapping_report.xlsx
        output_path: Path to save structured_input.xlsx

    Example:
        >>> generate_structured_inputs("mapping_report.xlsx", "structured_input.xlsx")
    """
    generator = StructuredInputGenerator(mapping_report_path)
    generator.generate(output_path)


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 3:
        print("Usage: python -m excel_pipeline.layer2.structured_generator <mapping_report.xlsx> <structured_input.xlsx>")
        sys.exit(1)

    generate_structured_inputs(sys.argv[1], sys.argv[2])
