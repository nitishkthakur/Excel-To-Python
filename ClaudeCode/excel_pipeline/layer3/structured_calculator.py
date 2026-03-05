"""Layer 3b: Structured Calculator.

Calculates output workbook from structured inputs by mapping tables back to original cells.
"""

from pathlib import Path
from typing import Dict, Tuple, Any, List
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from excel_pipeline.core.excel_io import save_workbook
from excel_pipeline.utils.logging_setup import get_logger

logger = get_logger(__name__)


class StructuredCalculator:
    """Calculate output from structured inputs."""

    def __init__(self, input_path: str, mapping_path: str):
        """
        Initialize calculator.

        Args:
            input_path: Path to structured_input.xlsx
            mapping_path: Path to mapping_report.xlsx
        """
        self.input_path = input_path
        self.mapping_path = mapping_path
        self.cell_values: Dict[Tuple[str, str], Any] = {}
        self.mapping_data: Dict[Tuple[str, str], Dict] = {}  # (sheet, coord) -> metadata

    def calculate(self, output_path: str) -> None:
        """
        Calculate and generate output workbook.

        Process:
        1. Load structured inputs and map to original cells
        2. Load mapping metadata for formatting
        3. Build output workbook with values and formulas
        4. Apply formatting from mapping report
        5. Save output.xlsx

        Args:
            output_path: Path to save output.xlsx
        """
        logger.info("=" * 80)
        logger.info("LAYER 3b: Structured Calculator")
        logger.info("=" * 80)
        logger.info(f"Inputs: {self.input_path}")
        logger.info(f"Mapping: {self.mapping_path}")
        logger.info(f"Output: {output_path}")

        # Step 1: Load structured inputs and map to original cells
        logger.info("\n[Step 1/4] Loading structured inputs and mapping to original cells...")
        self._load_structured_inputs()

        # Step 2: Load mapping metadata for formatting
        logger.info("\n[Step 2/4] Loading formatting from mapping report...")
        self._load_mapping_metadata()

        # Step 3: Build output workbook
        logger.info("\n[Step 3/4] Building output workbook...")
        output_wb = self._build_output_workbook()

        # Step 4: Save
        logger.info("\n[Step 4/4] Saving output...")
        save_workbook(output_wb, output_path)

        # Summary
        logger.info("\n" + "=" * 80)
        logger.info("LAYER 3b: Complete!")
        logger.info("=" * 80)
        logger.info(f"Output: {output_path}")
        logger.info(f"Total cells: {len(self.cell_values)}")
        logger.info("=" * 80)

    def _load_structured_inputs(self) -> None:
        """Load structured inputs and map back to original cell coordinates."""
        wb = load_workbook(self.input_path, data_only=True)

        # Read Index sheet to understand table mappings
        if 'Index' not in wb.sheetnames:
            logger.warning("No Index sheet found in structured inputs")
            return

        index_sheet = wb['Index']
        index_data = self._parse_index_sheet(index_sheet)

        # Read Config sheet for scalars
        if 'Config' in wb.sheetnames:
            logger.info("  Loading Config sheet...")
            self._load_config_sheet(wb['Config'])

        # Read each table and map to original cells
        for table_info in index_data:
            if table_info['table_type'] == 'Scalar':
                continue  # Already loaded from Config

            table_name = table_info['structured_table']
            if table_name not in wb.sheetnames:
                logger.warning(f"  Table {table_name} not found in workbook")
                continue

            logger.info(f"  Loading table: {table_name}")
            self._load_table_sheet(
                wb[table_name],
                table_info
            )

        logger.info(f"Loaded {len(self.cell_values)} cell values from structured inputs")

    def _parse_index_sheet(self, index_sheet) -> List[Dict]:
        """Parse Index sheet to get table metadata."""
        index_data = []

        # Read headers
        headers = {}
        for col_idx, cell in enumerate(index_sheet[1], start=1):
            if cell.value:
                headers[cell.value] = col_idx

        # Read table information
        for row in index_sheet.iter_rows(min_row=2):
            if not row[0].value:
                continue

            table_info = {
                'structured_table': row[headers['StructuredTable'] - 1].value,
                'source_sheet': row[headers['SourceSheet'] - 1].value,
                'cell_range': row[headers['CellRange'] - 1].value,
                'table_type': row[headers['TableType'] - 1].value,
                'transposed': row[headers['Transposed'] - 1].value,
            }
            index_data.append(table_info)

        logger.debug(f"  Found {len(index_data)} tables in Index")
        return index_data

    def _load_config_sheet(self, config_sheet) -> None:
        """Load scalar values from Config sheet."""
        # Read headers
        headers = {}
        for col_idx, cell in enumerate(config_sheet[1], start=1):
            if cell.value:
                headers[cell.value] = col_idx

        # Read scalar values
        scalar_count = 0
        for row in config_sheet.iter_rows(min_row=2):
            source_sheet = row[headers['SourceSheet'] - 1].value
            source_cell = row[headers['SourceCell'] - 1].value
            value = row[headers['Value'] - 1].value

            if source_sheet and source_cell:
                key = (source_sheet, source_cell)
                self.cell_values[key] = value
                scalar_count += 1

        logger.debug(f"    Loaded {scalar_count} scalars")

    def _load_table_sheet(self, table_sheet, table_info: Dict) -> None:
        """Load table data and map to original cells."""
        source_sheet = table_info['source_sheet']
        cell_range = table_info['cell_range']
        transposed = table_info['transposed']

        # Parse cell range (e.g., "B5:Q40")
        range_match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', cell_range)
        if not range_match:
            logger.warning(f"    Could not parse cell range: {cell_range}")
            return

        start_col = range_match.group(1)
        start_row = int(range_match.group(2))
        end_col = range_match.group(3)
        end_row = int(range_match.group(4))

        # Read table data (skip headers in row 1)
        table_data = []
        for row in table_sheet.iter_rows(min_row=2):
            row_data = [cell.value for cell in row]
            if any(v is not None for v in row_data):  # Skip empty rows
                table_data.append(row_data)

        if not table_data:
            logger.debug(f"    No data in table")
            return

        # If transposed, reverse the transpose
        if transposed:
            # Table has: Period (col A) | Metrics (cols B, C, D, ...)
            # Original had: Periods as columns, metrics as rows
            # Need to transpose back
            logger.debug(f"    Reversing transpose")

            # Skip first column (period labels) and transpose the rest
            data_cols = []
            for row in table_data:
                data_cols.append(row[1:] if len(row) > 1 else [])

            # Transpose
            if data_cols and data_cols[0]:
                num_metrics = len(data_cols[0])
                transposed_data = []
                for metric_idx in range(num_metrics):
                    metric_row = [row[metric_idx] if metric_idx < len(row) else None
                                 for row in data_cols]
                    transposed_data.append(metric_row)
                table_data = transposed_data

        # Map to original cells
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)

        cell_count = 0
        for row_offset, row_data in enumerate(table_data):
            orig_row = start_row + row_offset + 1  # +1 because we skipped header row

            for col_offset, value in enumerate(row_data):
                orig_col_idx = start_col_idx + col_offset + 1  # +1 because first col might be labels

                if orig_col_idx > end_col_idx or orig_row > end_row:
                    break  # Out of range

                if value is not None:
                    orig_col = get_column_letter(orig_col_idx)
                    coord = f"{orig_col}{orig_row}"
                    key = (source_sheet, coord)
                    self.cell_values[key] = value
                    cell_count += 1

        logger.debug(f"    Mapped {cell_count} cells to {source_sheet}")

    def _load_mapping_metadata(self) -> None:
        """Load cell metadata from mapping report (same as Layer 3a)."""
        mapping_wb = load_workbook(self.mapping_path, data_only=True)

        for sheet_name in mapping_wb.sheetnames:
            if sheet_name == "_Metadata":
                continue

            sheet = mapping_wb[sheet_name]

            # Read headers
            headers = {}
            for col_idx, cell in enumerate(sheet[1], start=1):
                if cell.value:
                    headers[cell.value] = col_idx

            # Read cell metadata
            for row in sheet.iter_rows(min_row=2):
                cell_coord = row[headers["Cell"] - 1].value
                if not cell_coord:
                    continue

                # Check if this is a consolidated range
                if ':' in str(cell_coord):
                    # Expand consolidated range
                    pattern_formula = row[headers.get("PatternFormula", 0) - 1].value if "PatternFormula" in headers else None
                    group_direction = row[headers.get("GroupDirection", 0) - 1].value if "GroupDirection" in headers else None

                    # Parse range
                    start_coord, end_coord = cell_coord.split(':')

                    # Expand range
                    expanded_cells = self._expand_range(
                        start_coord, end_coord, pattern_formula, group_direction
                    )

                    # Create metadata for each cell
                    base_metadata = self._extract_metadata(row, headers)

                    for coord, formula in expanded_cells:
                        key = (sheet_name, coord)
                        metadata = base_metadata.copy()
                        metadata['formula'] = formula
                        self.mapping_data[key] = metadata

                else:
                    # Individual cell
                    key = (sheet_name, cell_coord)
                    metadata = self._extract_metadata(row, headers)
                    self.mapping_data[key] = metadata

        logger.info(f"Loaded metadata for {len(self.mapping_data)} cells")

    def _extract_metadata(self, row, headers: Dict) -> Dict:
        """Extract metadata from a mapping report row."""
        return {
            'cell_type': row[headers.get("Type", 1) - 1].value if "Type" in headers else None,
            'formula': row[headers.get("Formula", 1) - 1].value if "Formula" in headers else None,
            'number_format': row[headers.get("NumberFormat", 1) - 1].value if "NumberFormat" in headers else None,
            'font_bold': row[headers.get("FontBold", 1) - 1].value if "FontBold" in headers else None,
            'font_italic': row[headers.get("FontItalic", 1) - 1].value if "FontItalic" in headers else None,
            'font_size': row[headers.get("FontSize", 1) - 1].value if "FontSize" in headers else None,
            'font_color': row[headers.get("FontColor", 1) - 1].value if "FontColor" in headers else None,
            'fill_color': row[headers.get("FillColor", 1) - 1].value if "FillColor" in headers else None,
            'alignment': row[headers.get("Alignment", 1) - 1].value if "Alignment" in headers else None,
            'wrap_text': row[headers.get("WrapText", 1) - 1].value if "WrapText" in headers else None,
        }

    def _expand_range(self, start_coord: str, end_coord: str,
                     pattern_formula: str, direction: str) -> List[Tuple[str, str]]:
        """Expand consolidated range into individual cells (same as Layer 3a)."""
        start_col_match = re.match(r'([A-Z]+)(\d+)', start_coord)
        end_col_match = re.match(r'([A-Z]+)(\d+)', end_coord)

        if not start_col_match or not end_col_match:
            return []

        start_col_letter = start_col_match.group(1)
        start_row = int(start_col_match.group(2))
        end_col_letter = end_col_match.group(1)
        end_row = int(end_col_match.group(2))

        start_col_idx = column_index_from_string(start_col_letter)
        end_col_idx = column_index_from_string(end_col_letter)

        results = []

        if direction == "horizontal":
            for col_idx in range(start_col_idx, end_col_idx + 1):
                col_letter = get_column_letter(col_idx)
                coord = f"{col_letter}{start_row}"

                if pattern_formula:
                    formula = pattern_formula.replace('{col}', col_letter).lstrip("'")
                    results.append((coord, formula))

        elif direction == "vertical":
            for row_num in range(start_row, end_row + 1):
                coord = f"{start_col_letter}{row_num}"

                if pattern_formula:
                    formula = pattern_formula.replace('{row}', str(row_num)).lstrip("'")
                    results.append((coord, formula))

        return results

    def _build_output_workbook(self) -> Workbook:
        """Build output workbook (same as Layer 3a)."""
        wb = Workbook()

        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Group metadata by sheet
        cells_by_sheet = {}
        for (sheet_name, coord), metadata in self.mapping_data.items():
            if sheet_name not in cells_by_sheet:
                cells_by_sheet[sheet_name] = []

            value = self.cell_values.get((sheet_name, coord), None)
            cells_by_sheet[sheet_name].append((coord, value, metadata))

        # Create sheets
        for sheet_name, cells in cells_by_sheet.items():
            logger.debug(f"Creating sheet: {sheet_name}")
            ws = wb.create_sheet(title=sheet_name)

            for coord, value, metadata in cells:
                self._write_cell_with_metadata(ws, sheet_name, coord, value, metadata)

            logger.debug(f"  Wrote {len(cells)} cells")

        return wb

    def _write_cell_with_metadata(self, ws, sheet_name: str, coord: str, value: Any, metadata: Dict) -> None:
        """Write cell with formatting (same as Layer 3a)."""
        if ':' in str(coord):
            return

        excel_cell = ws[coord]

        formula = metadata.get('formula', '')

        if formula and metadata.get('cell_type') in ['Calculation', 'Output']:
            formula = formula.lstrip("'")
            if formula.startswith('='):
                excel_cell.value = formula
            else:
                excel_cell.value = value or formula
        else:
            excel_cell.value = value

        # Apply formatting (number format, font, fill, alignment)
        number_format = metadata.get('number_format')
        if number_format and number_format != 'General':
            try:
                excel_cell.number_format = number_format
            except:
                pass

        font_kwargs = {}
        if metadata.get('font_bold'):
            font_kwargs['bold'] = True
        if metadata.get('font_italic'):
            font_kwargs['italic'] = True
        if metadata.get('font_size'):
            font_kwargs['size'] = metadata['font_size']
        if metadata.get('font_color'):
            color = metadata['font_color'].replace('#', '')
            if len(color) == 6:
                color = 'FF' + color
            if len(color) == 8:
                font_kwargs['color'] = color

        if font_kwargs:
            excel_cell.font = Font(**font_kwargs)

        fill_color = metadata.get('fill_color')
        if fill_color:
            color = fill_color.replace('#', '')
            if len(color) == 6:
                color = 'FF' + color
            if len(color) == 8:
                excel_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        alignment = metadata.get('alignment')
        wrap_text = metadata.get('wrap_text', False)
        if alignment or wrap_text:
            excel_cell.alignment = Alignment(
                horizontal=alignment if alignment else 'general',
                wrap_text=wrap_text
            )


def calculate_structured(input_path: str, mapping_path: str, output_path: str) -> None:
    """
    Calculate output from structured inputs.

    This is the main entry point for Layer 3b.

    Args:
        input_path: Path to structured_input.xlsx
        mapping_path: Path to mapping_report.xlsx
        output_path: Path to save output.xlsx
    """
    calculator = StructuredCalculator(input_path, mapping_path)
    calculator.calculate(output_path)


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 4:
        print("Usage: python -m excel_pipeline.layer3.structured_calculator "
              "<structured_input.xlsx> <mapping_report.xlsx> <output.xlsx>")
        sys.exit(1)

    calculate_structured(sys.argv[1], sys.argv[2], sys.argv[3])
