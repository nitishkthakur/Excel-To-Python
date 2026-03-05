"""Layer 3a: Unstructured Calculator.

Calculates output workbook from unstructured inputs using runtime formula engine.
"""

from pathlib import Path
from typing import Dict, Tuple, Any, List
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from excel_pipeline.runtime.formula_engine import calculate_workbook
from excel_pipeline.core.excel_io import save_workbook
from excel_pipeline.utils.logging_setup import get_logger

logger = get_logger(__name__)


class UnstructuredCalculator:
    """Calculate output from unstructured inputs."""

    def __init__(self, input_path: str, mapping_path: str):
        """
        Initialize calculator.

        Args:
            input_path: Path to unstructured_inputs.xlsx
            mapping_path: Path to mapping_report.xlsx
        """
        self.input_path = input_path
        self.mapping_path = mapping_path
        self.cell_values: Dict[Tuple[str, str], Any] = {}
        self.mapping_data: Dict[str, Dict] = {}  # (sheet, coord) -> metadata

    def calculate(self, output_path: str) -> None:
        """
        Calculate and generate output workbook.

        Process:
        1. Use formula engine to calculate all values
        2. Load mapping report for formatting
        3. Build output workbook with values and formulas
        4. Apply formatting from mapping report
        5. Save output.xlsx

        Args:
            output_path: Path to save output.xlsx
        """
        logger.info("=" * 80)
        logger.info("LAYER 3a: Unstructured Calculator")
        logger.info("=" * 80)
        logger.info(f"Inputs: {self.input_path}")
        logger.info(f"Mapping: {self.mapping_path}")
        logger.info(f"Output: {output_path}")

        # Step 1: Calculate all values using formula engine
        logger.info("\n[Step 1/4] Calculating formulas...")
        self.cell_values = calculate_workbook(self.input_path, self.mapping_path)

        # Step 2: Load mapping metadata for formatting
        logger.info("\n[Step 2/4] Loading formatting from mapping report...")
        self._load_mapping_metadata()

        # Step 3: Build output workbook
        logger.info("\n[Step 3/4] Building output workbook...")
        output_wb = self._build_output_workbook()

        # Step 4: Save
        logger.info("\n[Step 4/4] Saving output...")
        save_workbook(output_wb, output_path)

        # Summary
        logger.info("\n" + "=" * 80)
        logger.info("LAYER 3a: Complete!")
        logger.info("=" * 80)
        logger.info(f"Output: {output_path}")
        logger.info(f"Total cells: {len(self.cell_values)}")
        logger.info("=" * 80)

    def _load_mapping_metadata(self) -> None:
        """Load cell metadata from mapping report."""
        mapping_wb = load_workbook(self.mapping_path, data_only=True)

        for sheet_name in mapping_wb.sheetnames:
            if sheet_name == "_Metadata":
                continue

            sheet = mapping_wb[sheet_name]

            # Read headers
            headers = {}
            for col_idx, cell in enumerate(sheet[1], start=1):
                if cell.value:
                    headers[cell.value] = col_idx

            # Read cell metadata
            for row in sheet.iter_rows(min_row=2):
                cell_coord = row[headers["Cell"] - 1].value
                if not cell_coord:
                    continue

                # Check if this is a consolidated range
                if ':' in str(cell_coord):
                    # Expand consolidated range
                    pattern_formula = row[headers["PatternFormula"] - 1].value
                    group_direction = row[headers["GroupDirection"] - 1].value

                    # Parse range (e.g., "D5:O5")
                    start_coord, end_coord = cell_coord.split(':')

                    # Expand range and generate formulas
                    expanded_cells = self._expand_range(
                        start_coord, end_coord, pattern_formula, group_direction
                    )

                    # Create metadata for each cell in range
                    base_metadata = {
                        'cell_type': row[headers["Type"] - 1].value,
                        'number_format': row[headers["NumberFormat"] - 1].value,
                        'font_bold': row[headers["FontBold"] - 1].value,
                        'font_italic': row[headers["FontItalic"] - 1].value,
                        'font_size': row[headers["FontSize"] - 1].value,
                        'font_color': row[headers["FontColor"] - 1].value,
                        'fill_color': row[headers["FillColor"] - 1].value,
                        'alignment': row[headers["Alignment"] - 1].value,
                        'wrap_text': row[headers["WrapText"] - 1].value,
                    }

                    for coord, formula in expanded_cells:
                        key = (sheet_name, coord)
                        metadata = base_metadata.copy()
                        metadata['formula'] = formula
                        self.mapping_data[key] = metadata

                else:
                    # Individual cell
                    key = (sheet_name, cell_coord)

                    metadata = {
                        'cell_type': row[headers["Type"] - 1].value,
                        'formula': row[headers["Formula"] - 1].value,
                        'number_format': row[headers["NumberFormat"] - 1].value,
                        'font_bold': row[headers["FontBold"] - 1].value,
                        'font_italic': row[headers["FontItalic"] - 1].value,
                        'font_size': row[headers["FontSize"] - 1].value,
                        'font_color': row[headers["FontColor"] - 1].value,
                        'fill_color': row[headers["FillColor"] - 1].value,
                        'alignment': row[headers["Alignment"] - 1].value,
                        'wrap_text': row[headers["WrapText"] - 1].value,
                    }

                    self.mapping_data[key] = metadata

        logger.info(f"Loaded metadata for {len(self.mapping_data)} cells")

    def _expand_range(self, start_coord: str, end_coord: str,
                     pattern_formula: str, direction: str) -> List[Tuple[str, str]]:
        """
        Expand a consolidated cell range into individual cells with formulas.

        Args:
            start_coord: Start cell (e.g., "D5")
            end_coord: End cell (e.g., "O5")
            pattern_formula: Pattern formula with {col} or {row} placeholders
            direction: "horizontal" or "vertical"

        Returns:
            List of (coordinate, formula) tuples
        """
        from openpyxl.utils import column_index_from_string, get_column_letter
        import re

        # Parse start and end coordinates
        start_col_match = re.match(r'([A-Z]+)(\d+)', start_coord)
        end_col_match = re.match(r'([A-Z]+)(\d+)', end_coord)

        if not start_col_match or not end_col_match:
            logger.warning(f"Could not parse range: {start_coord}:{end_coord}")
            return []

        start_col_letter = start_col_match.group(1)
        start_row = int(start_col_match.group(2))
        end_col_letter = end_col_match.group(1)
        end_row = int(end_col_match.group(2))

        start_col_idx = column_index_from_string(start_col_letter)
        end_col_idx = column_index_from_string(end_col_letter)

        results = []

        if direction == "horizontal":
            # Expand horizontally across columns
            for col_idx in range(start_col_idx, end_col_idx + 1):
                col_letter = get_column_letter(col_idx)
                coord = f"{col_letter}{start_row}"

                # Generate formula by replacing {col} placeholder
                if pattern_formula:
                    formula = pattern_formula.replace('{col}', col_letter)
                    # Remove leading apostrophe if present
                    formula = formula.lstrip("'")
                    results.append((coord, formula))

        elif direction == "vertical":
            # Expand vertically across rows
            for row_num in range(start_row, end_row + 1):
                coord = f"{start_col_letter}{row_num}"

                # Generate formula by replacing {row} placeholder
                if pattern_formula:
                    formula = pattern_formula.replace('{row}', str(row_num))
                    # Remove leading apostrophe if present
                    formula = formula.lstrip("'")
                    results.append((coord, formula))

        return results

    def _build_output_workbook(self) -> Workbook:
        """Build output workbook with all values and formulas."""
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Group metadata by sheet
        cells_by_sheet = {}
        for (sheet_name, coord), metadata in self.mapping_data.items():
            if sheet_name not in cells_by_sheet:
                cells_by_sheet[sheet_name] = []

            # Get value (from calculated values or cell_values dict)
            value = self.cell_values.get((sheet_name, coord), None)

            cells_by_sheet[sheet_name].append((coord, value, metadata))

        # Create sheets and populate
        for sheet_name, cells in cells_by_sheet.items():
            logger.debug(f"Creating sheet: {sheet_name}")
            ws = wb.create_sheet(title=sheet_name)

            for coord, value, metadata in cells:
                self._write_cell_with_metadata(ws, sheet_name, coord, value, metadata)

            logger.debug(f"  Wrote {len(cells)} cells")

        return wb

    def _write_cell_with_metadata(self, ws, sheet_name: str, coord: str, value: Any, metadata: Dict) -> None:
        """Write a single cell with value, formula, and formatting."""
        # Skip if coord is a range (consolidated cells)
        if ':' in str(coord):
            logger.debug(f"Skipping consolidated range: {coord}")
            return

        excel_cell = ws[coord]

        # Determine what to write based on cell type
        formula = metadata.get('formula', '')

        if formula and metadata.get('cell_type') in ['Calculation', 'Output']:
            # Write formula (let Excel calculate it)
            formula = formula.lstrip("'")
            if formula.startswith('='):
                excel_cell.value = formula
            else:
                # No leading =, write as value
                excel_cell.value = value or formula
        else:
            # Input cell - write value
            excel_cell.value = value

        # Apply number format
        number_format = metadata.get('number_format')
        if number_format and number_format != 'General':
            try:
                excel_cell.number_format = number_format
            except:
                pass

        # Apply font
        font_kwargs = {}
        if metadata.get('font_bold'):
            font_kwargs['bold'] = True
        if metadata.get('font_italic'):
            font_kwargs['italic'] = True
        if metadata.get('font_size'):
            font_kwargs['size'] = metadata['font_size']
        if metadata.get('font_color'):
            color = metadata['font_color'].replace('#', '')
            if len(color) == 6:
                color = 'FF' + color
            if len(color) == 8:
                font_kwargs['color'] = color

        if font_kwargs:
            excel_cell.font = Font(**font_kwargs)

        # Apply fill
        fill_color = metadata.get('fill_color')
        if fill_color:
            color = fill_color.replace('#', '')
            if len(color) == 6:
                color = 'FF' + color
            if len(color) == 8:
                excel_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        # Apply alignment
        alignment = metadata.get('alignment')
        wrap_text = metadata.get('wrap_text', False)
        if alignment or wrap_text:
            excel_cell.alignment = Alignment(
                horizontal=alignment if alignment else 'general',
                wrap_text=wrap_text
            )


def calculate_unstructured(input_path: str, mapping_path: str, output_path: str) -> None:
    """
    Calculate output from unstructured inputs.

    This is the main entry point for Layer 3a.

    Args:
        input_path: Path to unstructured_inputs.xlsx
        mapping_path: Path to mapping_report.xlsx
        output_path: Path to save output.xlsx

    Example:
        >>> calculate_unstructured(
        ...     "unstructured_inputs.xlsx",
        ...     "mapping_report.xlsx",
        ...     "output.xlsx"
        ... )
    """
    calculator = UnstructuredCalculator(input_path, mapping_path)
    calculator.calculate(output_path)


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 4:
        print("Usage: python -m excel_pipeline.layer3.unstructured_calculator "
              "<unstructured_inputs.xlsx> <mapping_report.xlsx> <output.xlsx>")
        sys.exit(1)

    calculate_unstructured(sys.argv[1], sys.argv[2], sys.argv[3])
