"""Runtime formula evaluation engine.

Evaluates Excel formulas in correct dependency order, applying vectorization
for dragged formula groups.
"""

from typing import Dict, List, Set, Any, Tuple, Optional
from collections import defaultdict, deque
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from excel_pipeline.utils.logging_setup import get_logger
import formulas
import re

logger = get_logger(__name__)


class FormulaEngine:
    """Runtime engine for evaluating Excel formulas."""

    def __init__(self, input_wb: Workbook, mapping_data: Dict[str, List[Dict]]):
        """
        Initialize formula engine.

        Args:
            input_wb: Workbook with input values
            mapping_data: Dictionary of sheet_name -> list of cell metadata
        """
        self.input_wb = input_wb
        self.mapping_data = mapping_data

        # Cell values storage: (sheet_name, coordinate) -> value
        self.cell_values: Dict[Tuple[str, str], Any] = {}

        # Formula metadata: (sheet_name, coordinate) -> metadata dict
        self.formula_cells: Dict[Tuple[str, str], Dict] = {}

        # Dependency graph
        self.precedents: Dict[Tuple[str, str], Set[Tuple[str, str]]] = defaultdict(set)
        self.dependents: Dict[Tuple[str, str], Set[Tuple[str, str]]] = defaultdict(set)

        # Group metadata: group_id -> list of (sheet_name, coordinate)
        self.groups: Dict[int, List[Tuple[str, str]]] = defaultdict(list)

        # Calculation order
        self.calc_order: List[Tuple[str, str]] = []

    def load_inputs(self) -> None:
        """Load all input values from input workbook."""
        logger.info("Loading input values...")

        input_count = 0
        for sheet in self.input_wb.worksheets:
            sheet_name = sheet.title

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        key = (sheet_name, cell.coordinate)
                        self.cell_values[key] = cell.value
                        input_count += 1

        logger.info(f"Loaded {input_count} input values")

    def load_mapping_metadata(self) -> None:
        """Load formula metadata from mapping report."""
        logger.info("Loading formula metadata from mapping report...")

        formula_count = 0
        for sheet_name, cells in self.mapping_data.items():
            for cell_meta in cells:
                if cell_meta['cell_type'] in ['Calculation', 'Output']:
                    # Formula cell
                    coord = cell_meta['cell']
                    key = (sheet_name, coord)

                    self.formula_cells[key] = cell_meta
                    formula_count += 1

                    # Track groups
                    if cell_meta['group_id'] and cell_meta['group_id'] > 0:
                        self.groups[cell_meta['group_id']].append(key)

        logger.info(f"Loaded {formula_count} formula cells")
        logger.info(f"Found {len(self.groups)} formula groups")

    def build_dependency_graph(self) -> None:
        """Build dependency graph from formulas."""
        logger.info("Building dependency graph...")

        for cell_key, cell_meta in self.formula_cells.items():
            formula = cell_meta['formula']
            if not formula:
                continue

            # Remove leading '=' and apostrophe prefix if present
            formula = formula.lstrip("'=")

            # Extract cell references
            refs = self._extract_cell_references(formula, cell_key[0])

            for ref in refs:
                self.precedents[cell_key].add(ref)
                self.dependents[ref].add(cell_key)

        logger.info(f"Dependency graph built: {len(self.precedents)} cells with precedents")

    def _extract_cell_references(self, formula: str, current_sheet: str) -> Set[Tuple[str, str]]:
        """
        Extract cell references from formula.

        Returns set of (sheet_name, coordinate) tuples.
        """
        refs = set()

        # Pattern for cell references: A1, $A$1, Sheet!A1, 'Sheet Name'!A1, etc.
        # This is simplified - for production, use formulas library parser

        # Handle sheet references like 'Sheet Name'!A1
        sheet_ref_pattern = r"'([^']+)'!([A-Z]+\d+)"
        for match in re.finditer(sheet_ref_pattern, formula):
            sheet_name = match.group(1)
            coord = match.group(2)
            refs.add((sheet_name, coord))

        # Handle sheet references like Sheet!A1 (no spaces)
        sheet_ref_pattern2 = r"([A-Za-z0-9_]+)!([A-Z]+\d+)"
        for match in re.finditer(sheet_ref_pattern2, formula):
            # Skip if already captured by first pattern
            if "'" not in match.group(0):
                sheet_name = match.group(1)
                coord = match.group(2)
                refs.add((sheet_name, coord))

        # Handle simple cell references (same sheet)
        simple_ref_pattern = r"\$?([A-Z]+)\$?(\d+)"
        for match in re.finditer(simple_ref_pattern, formula):
            # Only if not part of a sheet reference
            start = match.start()
            if start == 0 or formula[start - 1] != '!':
                col = match.group(1)
                row = match.group(2)
                coord = f"{col}{row}"
                refs.add((current_sheet, coord))

        return refs

    def calculate_order(self) -> None:
        """Calculate evaluation order using topological sort (Kahn's algorithm)."""
        logger.info("Calculating evaluation order...")

        # Compute in-degree for each formula cell
        in_degree = {}
        for cell_key in self.formula_cells.keys():
            # Count how many precedents this cell has
            in_degree[cell_key] = len(self.precedents[cell_key])

        # Queue of cells with no dependencies
        queue = deque([cell for cell, degree in in_degree.items() if degree == 0])

        order = []
        while queue:
            cell_key = queue.popleft()
            order.append(cell_key)

            # Reduce in-degree of dependents
            for dependent in self.dependents[cell_key]:
                if dependent in in_degree:
                    in_degree[dependent] -= 1
                    if in_degree[dependent] == 0:
                        queue.append(dependent)

        self.calc_order = order

        # Check for circular references
        if len(order) < len(self.formula_cells):
            logger.warning(f"Circular references detected! {len(self.formula_cells) - len(order)} cells not in order")

        logger.info(f"Calculation order: {len(order)} cells")

    def evaluate_all(self) -> None:
        """Evaluate all formulas in dependency order."""
        logger.info("Evaluating formulas...")

        calculated = 0
        vectorized_groups = set()

        for cell_key in self.calc_order:
            cell_meta = self.formula_cells[cell_key]

            # Check if part of a vectorizable group
            group_id = cell_meta.get('group_id', 0) or 0
            if group_id > 0 and cell_meta.get('is_vectorizable', False):
                # Vectorizable group - calculate entire group at once
                if group_id not in vectorized_groups:
                    self._evaluate_group_vectorized(group_id)
                    vectorized_groups.add(group_id)
                    calculated += len(self.groups[group_id])
            else:
                # Individual cell calculation
                self._evaluate_cell(cell_key)
                calculated += 1

            if calculated % 100 == 0:
                logger.debug(f"  Calculated {calculated} cells...")

        logger.info(f"Calculated {calculated} cells ({len(vectorized_groups)} vectorized groups)")

    def _evaluate_cell(self, cell_key: Tuple[str, str]) -> Any:
        """Evaluate a single cell formula."""
        cell_meta = self.formula_cells[cell_key]
        formula = cell_meta['formula']

        if not formula:
            return None

        # Remove leading apostrophe and equals sign
        formula = formula.lstrip("'=")

        try:
            # Simple evaluation using Python eval for basic formulas
            # For production, use formulas library
            result = self._simple_eval(formula, cell_key[0])
            self.cell_values[cell_key] = result
            return result
        except Exception as e:
            logger.warning(f"Error evaluating {cell_key[0]}!{cell_key[1]}: {e}")
            self.cell_values[cell_key] = None
            return None

    def _evaluate_group_vectorized(self, group_id: int) -> None:
        """Evaluate an entire formula group using vectorization."""
        group_cells = self.groups[group_id]
        if not group_cells:
            return

        # Get pattern formula from first cell
        first_cell_meta = self.formula_cells[group_cells[0]]
        pattern = first_cell_meta.get('pattern_formula', '')

        logger.debug(f"Vectorizing group {group_id}: {len(group_cells)} cells, pattern: {pattern[:50]}")

        # For now, evaluate individually
        # TODO: Implement true vectorization with numpy/pandas
        for cell_key in group_cells:
            self._evaluate_cell(cell_key)

    def _simple_eval(self, formula: str, current_sheet: str) -> Any:
        """
        Simple formula evaluation.

        This is a simplified evaluator for basic formulas.
        For production, use the formulas library.
        """
        # Replace cell references with their values

        def replace_ref(match):
            """Replace cell reference with its value."""
            full_match = match.group(0)

            # Check for sheet reference
            if '!' in full_match:
                if "'" in full_match:
                    # 'Sheet Name'!A1
                    parts = full_match.split('!')
                    sheet_name = parts[0].strip("'")
                    coord = parts[1]
                else:
                    # Sheet!A1
                    sheet_name, coord = full_match.split('!')
            else:
                # Simple reference (same sheet)
                sheet_name = current_sheet
                coord = full_match

            # Remove $ signs
            coord = coord.replace('$', '')

            # Get value
            key = (sheet_name, coord)
            value = self.cell_values.get(key, 0)

            # Return as string for eval
            if isinstance(value, str):
                return f'"{value}"'
            elif value is None:
                return '0'
            else:
                return str(value)

        # Replace sheet references
        formula = re.sub(r"'[^']+!'?\$?[A-Z]+\$?\d+", replace_ref, formula)
        formula = re.sub(r"[A-Za-z0-9_]+!\$?[A-Z]+\$?\d+", replace_ref, formula)

        # Replace simple cell references
        formula = re.sub(r"\$?[A-Z]+\$?\d+", replace_ref, formula)

        # Replace Excel operators with Python operators
        formula = formula.replace('^', '**')

        # Handle basic Excel functions
        formula = self._replace_excel_functions(formula)

        try:
            # Evaluate
            result = eval(formula)
            return result
        except Exception as e:
            logger.debug(f"Eval error for '{formula}': {e}")
            return 0

    def _replace_excel_functions(self, formula: str) -> str:
        """Replace Excel functions with Python equivalents."""
        # This is very simplified - for production use formulas library

        # SUM
        formula = re.sub(r'SUM\((.*?)\)', r'sum([\1])', formula)

        # IF -> ternary operator
        # IF(condition, true_val, false_val) -> (true_val if condition else false_val)
        def replace_if(match):
            args = match.group(1)
            parts = args.split(',')
            if len(parts) == 3:
                return f'({parts[1]} if {parts[0]} else {parts[2]})'
            return match.group(0)

        formula = re.sub(r'IF\((.*?)\)', replace_if, formula)

        # MAX
        formula = re.sub(r'MAX\((.*?)\)', r'max([\1])', formula)

        # MIN
        formula = re.sub(r'MIN\((.*?)\)', r'min([\1])', formula)

        return formula

    def get_all_values(self) -> Dict[Tuple[str, str], Any]:
        """Get all cell values (inputs + calculated)."""
        return self.cell_values.copy()


def calculate_workbook(input_path: str, mapping_path: str) -> Dict[Tuple[str, str], Any]:
    """
    Calculate all formulas from inputs and mapping report.

    Args:
        input_path: Path to unstructured_inputs.xlsx
        mapping_path: Path to mapping_report.xlsx

    Returns:
        Dictionary of (sheet_name, coordinate) -> value
    """
    logger.info("=" * 80)
    logger.info("RUNTIME: Formula Evaluation")
    logger.info("=" * 80)

    # Load input workbook
    logger.info(f"Loading inputs from: {input_path}")
    input_wb = load_workbook(input_path, data_only=True)

    # Load mapping report
    logger.info(f"Loading mapping from: {mapping_path}")
    mapping_wb = load_workbook(mapping_path, data_only=True)

    # Parse mapping data
    mapping_data = {}
    for sheet_name in mapping_wb.sheetnames:
        if sheet_name == "_Metadata":
            continue

        sheet = mapping_wb[sheet_name]

        # Read headers
        headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                headers[cell.value] = col_idx

        # Read cell data
        cells = []
        for row in sheet.iter_rows(min_row=2):
            # Helper to get value or default
            def get_val(col_name, default):
                if col_name not in headers:
                    return default
                val = row[headers[col_name] - 1].value
                return val if val is not None else default

            cell_data = {
                'cell': get_val("Cell", None),
                'cell_type': get_val("Type", None),
                'formula': get_val("Formula", None),
                'group_id': int(get_val("GroupID", 0)) if get_val("GroupID", 0) else 0,
                'pattern_formula': get_val("PatternFormula", None),
                'is_vectorizable': bool(get_val("Vectorizable", False)),
            }
            cells.append(cell_data)

        mapping_data[sheet_name] = cells

    # Create engine
    engine = FormulaEngine(input_wb, mapping_data)

    # Execute calculation steps
    engine.load_inputs()
    engine.load_mapping_metadata()
    engine.build_dependency_graph()
    engine.calculate_order()
    engine.evaluate_all()

    logger.info("=" * 80)
    logger.info("Calculation complete!")
    logger.info("=" * 80)

    return engine.get_all_values()
