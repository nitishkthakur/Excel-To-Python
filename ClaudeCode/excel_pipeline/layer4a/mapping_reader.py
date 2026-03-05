"""MappingReader: Parse mapping_report.xlsx to extract formula metadata."""

from dataclasses import dataclass
from typing import Dict, List, Optional, Any
from openpyxl import load_workbook
from excel_pipeline.utils.logging_setup import get_logger

logger = get_logger(__name__)


@dataclass
class CellMetadata:
    """Metadata for a single cell or group of cells."""
    sheet: str
    cell: str  # "A1" or "C2:C10" for groups
    cell_type: str  # "Input", "Calculation", "Output"
    formula: Optional[str]
    value: Any
    group_id: Optional[int]
    group_direction: Optional[str]  # "vertical" | "horizontal"
    group_size: Optional[int]
    pattern_formula: Optional[str]  # "=A{row}+B{row}"
    vectorizable: bool  # Computed: group_size >= 10

    # Formatting
    number_format: Optional[str] = None
    font_bold: Optional[bool] = None
    font_italic: Optional[bool] = None
    font_size: Optional[int] = None
    font_color: Optional[str] = None
    fill_color: Optional[str] = None
    alignment: Optional[str] = None
    wrap_text: Optional[bool] = None


@dataclass
class GroupMetadata:
    """Metadata for a vectorizable group of formulas."""
    group_id: int
    sheet: str
    direction: str  # "vertical" or "horizontal"
    cells: List[str]  # ["C2", "C3", "C4", ...]
    base_formula: str  # First cell's formula
    pattern_formula: str  # "=A{row}+B{row}"
    group_size: int
    vectorizable: bool


class MappingReader:
    """Parse mapping_report.xlsx to extract cell metadata and formula groups."""

    VECTORIZATION_THRESHOLD = 10  # Minimum group size for vectorization

    def __init__(self, mapping_report_path: str):
        """
        Initialize MappingReader.

        Args:
            mapping_report_path: Path to mapping_report.xlsx
        """
        self.mapping_report_path = mapping_report_path
        self.cells_by_sheet: Dict[str, List[CellMetadata]] = {}
        self.groups: Dict[int, GroupMetadata] = {}

    def read_mapping_report(self) -> Dict[str, List[CellMetadata]]:
        """
        Read mapping report and parse all cell metadata.

        Returns:
            Dictionary mapping sheet names to lists of CellMetadata
        """
        logger.info(f"Reading mapping report: {self.mapping_report_path}")

        wb = load_workbook(self.mapping_report_path, data_only=True)

        for sheet_name in wb.sheetnames:
            if sheet_name == "_Metadata":
                continue  # Skip metadata sheet

            logger.debug(f"  Processing sheet: {sheet_name}")
            sheet = wb[sheet_name]

            # Read headers
            headers = {}
            for col_idx, cell in enumerate(sheet[1], start=1):
                if cell.value:
                    headers[cell.value] = col_idx

            # Parse each row
            sheet_cells = []
            for row in sheet.iter_rows(min_row=2):
                cell_meta = self._parse_row(row, headers, sheet_name)
                if cell_meta:
                    sheet_cells.append(cell_meta)

            self.cells_by_sheet[sheet_name] = sheet_cells
            logger.debug(f"    Loaded {len(sheet_cells)} cells")

        total_cells = sum(len(cells) for cells in self.cells_by_sheet.values())
        logger.info(f"Loaded {total_cells} cells from {len(self.cells_by_sheet)} sheets")

        return self.cells_by_sheet

    def _parse_row(self, row, headers: Dict[str, int], sheet_name: str) -> Optional[CellMetadata]:
        """Parse a single row from mapping report."""
        # Get cell coordinate
        cell_coord = self._get_value(row, headers, "Cell")
        if not cell_coord:
            return None

        # Get basic metadata
        cell_type = self._get_value(row, headers, "Type")
        formula = self._get_value(row, headers, "Formula")
        value = self._get_value(row, headers, "Value")

        # Get grouping information
        group_id = self._get_int_value(row, headers, "GroupID")
        group_direction = self._get_value(row, headers, "GroupDirection")
        group_size = self._get_int_value(row, headers, "GroupSize")
        pattern_formula = self._get_value(row, headers, "PatternFormula")

        # Get vectorizable flag from mapping report (already computed by Layer 1)
        vectorizable = self._get_bool_value(row, headers, "Vectorizable")

        # Get formatting
        number_format = self._get_value(row, headers, "NumberFormat")
        font_bold = self._get_bool_value(row, headers, "FontBold")
        font_italic = self._get_bool_value(row, headers, "FontItalic")
        font_size = self._get_int_value(row, headers, "FontSize")
        font_color = self._get_value(row, headers, "FontColor")
        fill_color = self._get_value(row, headers, "FillColor")
        alignment = self._get_value(row, headers, "Alignment")
        wrap_text = self._get_bool_value(row, headers, "WrapText")

        return CellMetadata(
            sheet=sheet_name,
            cell=cell_coord,
            cell_type=cell_type,
            formula=formula,
            value=value,
            group_id=group_id,
            group_direction=group_direction,
            group_size=group_size,
            pattern_formula=pattern_formula,
            vectorizable=vectorizable,
            number_format=number_format,
            font_bold=font_bold,
            font_italic=font_italic,
            font_size=font_size,
            font_color=font_color,
            fill_color=fill_color,
            alignment=alignment,
            wrap_text=wrap_text,
        )

    def _get_value(self, row, headers: Dict[str, int], column_name: str) -> Optional[Any]:
        """Get value from row by column name."""
        if column_name not in headers:
            return None
        col_idx = headers[column_name] - 1
        if col_idx >= len(row):
            return None
        value = row[col_idx].value
        return value if value not in (None, '') else None

    def _get_int_value(self, row, headers: Dict[str, int], column_name: str) -> Optional[int]:
        """Get integer value from row by column name."""
        value = self._get_value(row, headers, column_name)
        if value is None:
            return None
        try:
            return int(value)
        except (ValueError, TypeError):
            return None

    def _get_bool_value(self, row, headers: Dict[str, int], column_name: str) -> Optional[bool]:
        """Get boolean value from row by column name."""
        value = self._get_value(row, headers, column_name)
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.upper() in ('TRUE', 'YES', '1')
        return bool(value)

    def identify_vectorizable_groups(self) -> List[GroupMetadata]:
        """
        Identify and collect vectorizable formula groups.

        Returns:
            List of GroupMetadata for groups where vectorizable=True
        """
        logger.info("Identifying formula groups...")

        # Collect all cells by group_id (both vectorizable and non-vectorizable)
        groups_dict: Dict[int, List[CellMetadata]] = {}

        for sheet_name, cells in self.cells_by_sheet.items():
            for cell in cells:
                if cell.group_id:  # Collect ALL groups, not just vectorizable
                    if cell.group_id not in groups_dict:
                        groups_dict[cell.group_id] = []
                    groups_dict[cell.group_id].append(cell)

        # Convert to GroupMetadata
        vectorizable_groups = []
        all_groups = 0

        for group_id, cells in groups_dict.items():
            if not cells:
                continue

            # Use first cell as representative
            first_cell = cells[0]

            # Expand cell ranges to individual coordinates
            expanded_cells = []
            for c in cells:
                if ':' in c.cell:
                    # This is a range like "C2:C13" - expand it
                    expanded = self._expand_range(c)
                    expanded_cells.extend([ec.cell for ec in expanded])
                else:
                    # Individual cell
                    expanded_cells.append(c.cell)

            group = GroupMetadata(
                group_id=group_id,
                sheet=first_cell.sheet,
                direction=first_cell.group_direction or "vertical",
                cells=expanded_cells,  # Now contains ["C2", "C3", ..., "C13"]
                base_formula=first_cell.formula or "",
                pattern_formula=first_cell.pattern_formula or first_cell.formula or "",
                group_size=first_cell.group_size or len(expanded_cells),
                vectorizable=first_cell.vectorizable  # Keep original vectorizable status
            )

            # Store ALL groups in self.groups (not just vectorizable)
            self.groups[group_id] = group
            all_groups += 1

            # Only add to return list if vectorizable
            if group.vectorizable:
                vectorizable_groups.append(group)

        logger.info(f"Found {all_groups} total groups, {len(vectorizable_groups)} vectorizable (size >= {self.VECTORIZATION_THRESHOLD})")

        return vectorizable_groups

    def expand_groups(self) -> List[CellMetadata]:
        """
        Expand consolidated cell ranges into individual cells.

        For ranges like "C2:C10", expands to individual cells C2, C3, ..., C10.

        Returns:
            Flattened list of all cells (expanded from ranges)
        """
        logger.info("Expanding grouped cell ranges...")

        expanded_cells = []

        for sheet_name, cells in self.cells_by_sheet.items():
            for cell in cells:
                if ':' in cell.cell:
                    # This is a range - expand it
                    expanded = self._expand_range(cell)
                    expanded_cells.extend(expanded)
                else:
                    # Individual cell
                    expanded_cells.append(cell)

        logger.info(f"Expanded to {len(expanded_cells)} individual cells")

        return expanded_cells

    def _expand_range(self, cell_meta: CellMetadata) -> List[CellMetadata]:
        """Expand a cell range like 'C2:C10' into individual cells."""
        import re
        from openpyxl.utils import get_column_letter, column_index_from_string

        cell_range = cell_meta.cell
        match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', cell_range)

        if not match:
            # Can't parse - return as-is
            return [cell_meta]

        start_col = match.group(1)
        start_row = int(match.group(2))
        end_col = match.group(3)
        end_row = int(match.group(4))

        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)

        expanded = []

        if cell_meta.group_direction == "horizontal":
            # Horizontal expansion (same row, different columns)
            for col_idx in range(start_col_idx, end_col_idx + 1):
                col_letter = get_column_letter(col_idx)
                coord = f"{col_letter}{start_row}"

                # Generate formula if pattern exists
                formula = cell_meta.formula
                if cell_meta.pattern_formula and '{col}' in cell_meta.pattern_formula:
                    formula = cell_meta.pattern_formula.replace('{col}', col_letter).lstrip("'")

                expanded_cell = CellMetadata(
                    sheet=cell_meta.sheet,
                    cell=coord,
                    cell_type=cell_meta.cell_type,
                    formula=formula,
                    value=cell_meta.value,
                    group_id=cell_meta.group_id,
                    group_direction=cell_meta.group_direction,
                    group_size=cell_meta.group_size,
                    pattern_formula=cell_meta.pattern_formula,
                    vectorizable=cell_meta.vectorizable,
                    number_format=cell_meta.number_format,
                    font_bold=cell_meta.font_bold,
                    font_italic=cell_meta.font_italic,
                    font_size=cell_meta.font_size,
                    font_color=cell_meta.font_color,
                    fill_color=cell_meta.fill_color,
                    alignment=cell_meta.alignment,
                    wrap_text=cell_meta.wrap_text,
                )
                expanded.append(expanded_cell)

        else:  # vertical
            # Vertical expansion (same column, different rows)
            for row_num in range(start_row, end_row + 1):
                coord = f"{start_col}{row_num}"

                # Generate formula if pattern exists
                formula = cell_meta.formula
                if cell_meta.pattern_formula and '{row}' in cell_meta.pattern_formula:
                    formula = cell_meta.pattern_formula.replace('{row}', str(row_num)).lstrip("'")

                expanded_cell = CellMetadata(
                    sheet=cell_meta.sheet,
                    cell=coord,
                    cell_type=cell_meta.cell_type,
                    formula=formula,
                    value=cell_meta.value,
                    group_id=cell_meta.group_id,
                    group_direction=cell_meta.group_direction,
                    group_size=cell_meta.group_size,
                    pattern_formula=cell_meta.pattern_formula,
                    vectorizable=cell_meta.vectorizable,
                    number_format=cell_meta.number_format,
                    font_bold=cell_meta.font_bold,
                    font_italic=cell_meta.font_italic,
                    font_size=cell_meta.font_size,
                    font_color=cell_meta.font_color,
                    fill_color=cell_meta.fill_color,
                    alignment=cell_meta.alignment,
                    wrap_text=cell_meta.wrap_text,
                )
                expanded.append(expanded_cell)

        return expanded
