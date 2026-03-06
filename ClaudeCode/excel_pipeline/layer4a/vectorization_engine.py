"""VectorizationEngine: Generate pandas/numpy vectorized code for formula groups."""

import re
from typing import List, Dict, Set, Tuple, Optional
from openpyxl.utils import get_column_letter, column_index_from_string
from excel_pipeline.layer4a.mapping_reader import GroupMetadata
from excel_pipeline.layer4a.formula_translator import FormulaTranslator
from excel_pipeline.utils.logging_setup import get_logger

logger = get_logger(__name__)


class VectorizationEngine:
    """Generate optimized pandas/numpy code for vectorizable formula groups."""

    def __init__(self, translator: FormulaTranslator):
        """
        Initialize VectorizationEngine.

        Args:
            translator: FormulaTranslator instance for formula conversion
        """
        self.translator = translator

    def generate_vectorized_code(self, group: GroupMetadata) -> List[str]:
        """
        Generate pandas code for a vectorizable group.

        Args:
            group: GroupMetadata for the formula group

        Returns:
            List of code lines (Python statements)
        """
        if not group.vectorizable:
            logger.warning(f"Group {group.group_id} is not vectorizable, falling back to loop")
            return self._generate_loop_code(group)

        # Analyze formula complexity
        complexity = self.translator.detect_vectorization_complexity(group.pattern_formula)

        if complexity == "SIMPLE":
            return self._generate_simple_vectorized_code(group)
        elif complexity == "MODERATE":
            return self._generate_moderate_vectorized_code(group)
        else:
            # Too complex - fall back to efficient loop
            logger.info(f"Group {group.group_id} too complex for vectorization, using loop")
            return self._generate_loop_code(group)

    def _generate_simple_vectorized_code(self, group: GroupMetadata) -> List[str]:
        """
        Generate simple pandas DataFrame operations.

        For patterns like =A{row}+B{row}, generates:
            df['C'] = df['A'] + df['B']
        """
        code = []

        # Determine if vertical or horizontal
        if group.direction == "vertical":
            code.extend(self._generate_vertical_vectorized(group))
        else:
            code.extend(self._generate_horizontal_vectorized(group))

        return code

    def _generate_vertical_vectorized(self, group: GroupMetadata) -> List[str]:
        """
        Generate code for vertical (row-wise) vectorization.

        Example:
            Input: C2:C26 = =A2+B2 (pattern: =A{row}+B{row})
            Output:
                df_Sheet1 = pd.DataFrame({
                    'A': [c.get(('Sheet1', 'A', r), 0) for r in range(2, 27)],
                    'B': [c.get(('Sheet1', 'B', r), 0) for r in range(2, 27)],
                }, index=range(2, 27))
                df_Sheet1['C'] = df_Sheet1['A'] + df_Sheet1['B']
                for r in df_Sheet1.index:
                    c[('Sheet1', 'C', r)] = df_Sheet1.loc[r, 'C']
        """
        code = []

        # Parse cell range
        first_cell, last_cell = self._parse_cell_range(group.cells)
        col = self._get_column_from_cell(first_cell)
        start_row, end_row = self._get_row_range(group.cells)

        # Extract dependencies from pattern formula
        dependencies = self._extract_column_dependencies(group.pattern_formula)

        if not dependencies:
            logger.warning(f"No dependencies found in pattern: {group.pattern_formula}")
            return self._generate_loop_code(group)

        # Generate DataFrame creation
        sheet_var = self._sanitize_sheet_name(group.sheet)
        code.append(f"# Vectorized group {group.group_id}: {len(group.cells)} cells")
        code.append(f"df_{sheet_var} = pd.DataFrame({{")

        for dep_col in dependencies:
            code.append(f"    '{dep_col}': [c.get(('{group.sheet}', '{dep_col}', r), 0) for r in range({start_row}, {end_row + 1})],")

        code.append(f"}}, index=range({start_row}, {end_row + 1}))")

        # Generate vectorized operation
        pandas_expr = self.translator.translate_formula(
            group.pattern_formula,
            group.sheet,
            is_vectorized=True
        )

        code.append(f"df_{sheet_var}['{col}'] = {pandas_expr}")

        # Write back to cell store
        code.append(f"for r in df_{sheet_var}.index:")
        code.append(f"    c[('{group.sheet}', '{col}', r)] = df_{sheet_var}.loc[r, '{col}']")
        code.append("")  # Empty line for readability

        return code

    def _generate_horizontal_vectorized(self, group: GroupMetadata) -> List[str]:
        """
        Generate code for horizontal (column-wise) vectorization.

        Example:
            Input: D5:O5 = =$C5*D$4 (horizontal drag)
            Output:
                base_val = c.get(('Sheet1', 'C', 5), 0)
                for col in ['D','E','F','G','H','I','J','K','L','M','N','O']:
                    multiplier = c.get(('Sheet1', col, 4), 0)
                    c[('Sheet1', col, 5)] = base_val * multiplier
        """
        code = []

        # Parse cell range
        first_cell, last_cell = self._parse_cell_range(group.cells)
        row = self._get_row_from_cell(first_cell)
        start_col, end_col = self._get_column_range(group.cells)

        # Generate column list
        columns = [get_column_letter(i) for i in range(
            column_index_from_string(start_col),
            column_index_from_string(end_col) + 1
        )]

        code.append(f"# Vectorized group {group.group_id}: {len(group.cells)} cells (horizontal)")
        code.append(f"for col in {columns}:")

        # Translate using the first column, then swap that column letter for `col`.
        # (The old string-concat trick produced broken Python for function calls like YEAR.)
        formula_for_first = group.pattern_formula.replace('{col}', columns[0])
        python_expr = self.translator.translate_formula(
            formula_for_first,
            group.sheet,
            is_vectorized=False
        )
        python_expr = python_expr.replace(f"'{columns[0]}'", "col")

        code.append(f"    c[('{group.sheet}', col, {row})] = {python_expr}")
        code.append("")

        return code

    def _generate_moderate_vectorized_code(self, group: GroupMetadata) -> List[str]:
        """
        Generate code for moderately complex formulas (IF, VLOOKUP, etc.).

        Uses np.where() for conditionals or .apply() for complex operations.
        """
        code = []

        # Check if contains IF function
        if 'IF(' in group.pattern_formula:
            return self._generate_conditional_vectorized(group)

        # Otherwise, fall back to loop
        return self._generate_loop_code(group)

    def _generate_conditional_vectorized(self, group: GroupMetadata) -> List[str]:
        """
        Generate np.where() code for IF functions.

        Example:
            Input: =IF(D2>100, "High", "Low")
            Output: df['E'] = np.where(df['D'] > 100, "High", "Low")
        """
        code = []

        col = self._get_column_from_cell(group.cells[0])
        start_row, end_row = self._get_row_range(group.cells)

        # Extract dependencies
        dependencies = self._extract_column_dependencies(group.pattern_formula)

        sheet_var = self._sanitize_sheet_name(group.sheet)
        code.append(f"# Vectorized group {group.group_id}: {len(group.cells)} cells (conditional)")
        code.append(f"df_{sheet_var} = pd.DataFrame({{")

        for dep_col in dependencies:
            code.append(f"    '{dep_col}': [c.get(('{group.sheet}', '{dep_col}', r), 0) for r in range({start_row}, {end_row + 1})],")

        code.append(f"}}, index=range({start_row}, {end_row + 1}))")

        # Translate IF to np.where
        pandas_expr = self.translator.translate_formula(
            group.pattern_formula,
            group.sheet,
            is_vectorized=True
        )

        code.append(f"df_{sheet_var}['{col}'] = {pandas_expr}")

        # Write back
        code.append(f"for r in df_{sheet_var}.index:")
        code.append(f"    c[('{group.sheet}', '{col}', r)] = df_{sheet_var}.loc[r, '{col}']")
        code.append("")

        return code

    def _generate_loop_code(self, group: GroupMetadata) -> List[str]:
        """
        Generate efficient loop code for non-vectorizable groups.

        Falls back to cell-by-cell calculation but in a clean loop.
        """
        code = []

        code.append(f"# Group {group.group_id}: {len(group.cells)} cells (loop)")

        if group.direction == "vertical":
            col = self._get_column_from_cell(group.cells[0])
            start_row, end_row = self._get_row_range(group.cells)

            code.append(f"for r in range({start_row}, {end_row + 1}):")

            # Replace {row} with r
            formula = group.pattern_formula.replace('{row}', str(start_row))
            python_expr = self.translator.translate_formula(formula, group.sheet, is_vectorized=False)

            # Replace specific row number with r
            python_expr = python_expr.replace(f", {start_row})", ", r)")

            code.append(f"    c[('{group.sheet}', '{col}', r)] = {python_expr}")

        else:  # horizontal
            row = self._get_row_from_cell(group.cells[0])
            start_col, end_col = self._get_column_range(group.cells)

            columns = [get_column_letter(i) for i in range(
                column_index_from_string(start_col),
                column_index_from_string(end_col) + 1
            )]

            code.append(f"for col in {columns}:")

            # Replace {col} in pattern
            formula_template = group.pattern_formula.replace('{col}', columns[0])
            python_expr = self.translator.translate_formula(formula_template, group.sheet, is_vectorized=False)

            # Replace specific column with 'col'
            python_expr = python_expr.replace(f"'{columns[0]}'", "col")

            code.append(f"    c[('{group.sheet}', col, {row})] = {python_expr}")

        code.append("")

        return code

    # Helper methods

    def _parse_cell_range(self, cells: List[str]) -> Tuple[str, str]:
        """Get first and last cell from cell list."""
        return cells[0], cells[-1]

    def _get_column_from_cell(self, cell: str) -> str:
        """Extract column letter from cell coordinate."""
        match = re.match(r'([A-Z]+)(\d+)', cell)
        return match.group(1) if match else cell

    def _get_row_from_cell(self, cell: str) -> int:
        """Extract row number from cell coordinate."""
        match = re.match(r'([A-Z]+)(\d+)', cell)
        return int(match.group(2)) if match else 0

    def _get_row_range(self, cells: List[str]) -> Tuple[int, int]:
        """Get start and end row numbers from cell list."""
        rows = [self._get_row_from_cell(cell) for cell in cells]
        return min(rows), max(rows)

    def _get_column_range(self, cells: List[str]) -> Tuple[str, str]:
        """Get start and end column letters from cell list."""
        cols = [self._get_column_from_cell(cell) for cell in cells]
        return min(cols), max(cols)

    def _extract_column_dependencies(self, pattern_formula: str) -> Set[str]:
        """
        Extract column letters referenced in pattern formula.

        Example: =A{row}+B{row}*C{row} → {'A', 'B', 'C'}
        """
        # Remove leading = and apostrophe
        formula = pattern_formula.lstrip("'").lstrip("=")

        # Find all column references with {row} placeholder
        columns = set(re.findall(r'([A-Z]+)\{row\}', formula))

        # Also find absolute column references like $A
        abs_columns = set(re.findall(r'\$([A-Z]+)(?:\d+|\{row\})', formula))

        return columns | abs_columns

    def _sanitize_sheet_name(self, sheet_name: str) -> str:
        """Convert sheet name to valid Python variable name."""
        # Replace spaces and special chars with underscores
        sanitized = re.sub(r'[^a-zA-Z0-9_]', '_', sheet_name)
        # Ensure doesn't start with number
        if sanitized and sanitized[0].isdigit():
            sanitized = '_' + sanitized
        return sanitized
