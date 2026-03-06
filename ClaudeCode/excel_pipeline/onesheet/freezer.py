"""
Workbook Freezer: creates the intermediate workbook for single-sheet processing.

The target sheet keeps all its formulas intact.
Every other sheet has its formula cells replaced with their last-cached values
(the values Excel stored when the file was last saved/calculated).

This preserves cross-sheet reference resolution: when Layer 4a generates
  c.get(('OtherSheet', 'A', 1), 0)
the frozen value for OtherSheet!A1 ends up in unstructured_inputs.xlsx and
is loaded into the cell store c before any calculations run.
"""

import datetime
import openpyxl
from openpyxl.cell.cell import MergedCell
from pathlib import Path
from typing import Dict
from excel_pipeline.utils.logging_setup import get_logger

# Types that are safe to write back as literal cell values.
# openpyxl can return DataTableFormula / ArrayFormula objects for cached
# values of data-table cells; we treat those as None.
_SIMPLE_TYPES = (int, float, str, bool, datetime.datetime, datetime.date, type(None))

logger = get_logger(__name__)


def freeze_other_sheets(
    input_path: str,
    target_sheet: str,
    frozen_path: str,
) -> Dict[str, int]:
    """
    Create an intermediate workbook where non-target sheets are value-only.

    For the target sheet: formulas and structure are left completely unchanged.
    For every other sheet: any cell whose value starts with '=' is replaced
    with its last-cached value (from openpyxl data_only=True).  Cells that
    never had a cached value (file saved before calculation) are left as None.

    Args:
        input_path:   Path to the original Excel workbook (.xlsx).
        target_sheet: Name of the sheet whose formulas must be preserved.
        frozen_path:  Destination path for the frozen intermediate workbook.

    Returns:
        Stats dict: {other_sheets, frozen_cells, null_cached_values}

    Raises:
        ValueError: If target_sheet does not exist in the workbook.
        FileNotFoundError: If input_path does not exist.
    """
    input_path = str(input_path)
    frozen_path = str(frozen_path)

    if not Path(input_path).exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    # Load the same file twice:
    #   wb_f  – formulas intact, used as the workbook we will mutate and save
    #   wb_v  – data_only, supplies the cached values for formula replacement
    logger.info(f"Loading workbook (formulas): {input_path}")
    wb_f = openpyxl.load_workbook(input_path, data_only=False, keep_vba=False)
    logger.info(f"Loading workbook (cached values): {input_path}")
    wb_v = openpyxl.load_workbook(input_path, data_only=True,  keep_vba=False)

    if target_sheet not in wb_f.sheetnames:
        raise ValueError(
            f"Sheet '{target_sheet}' not found. "
            f"Available: {', '.join(wb_f.sheetnames)}"
        )

    stats = {"other_sheets": 0, "frozen_cells": 0, "null_cached_values": 0}

    for sheet_name in wb_f.sheetnames:
        if sheet_name == target_sheet:
            logger.info(f"  [{sheet_name}]  kept with formulas (target)")
            continue

        ws_f = wb_f[sheet_name]
        ws_v = wb_v[sheet_name]
        frozen_in_sheet = 0
        null_in_sheet = 0

        for row in ws_f.iter_rows():
            for cell in row:
                # MergedCell proxy objects cannot have their .value set
                if isinstance(cell, MergedCell):
                    continue
                # Replace formula strings AND special formula objects
                # (openpyxl represents DataTableFormula / ArrayFormula cells as
                # non-string objects in data_only=False mode, so we must handle
                # both cases).
                is_formula = (
                    (isinstance(cell.value, str) and cell.value.startswith("="))
                    or (cell.value is not None and not isinstance(cell.value, _SIMPLE_TYPES))
                )
                if is_formula:
                    cached = ws_v[cell.coordinate].value
                    # DataTableFormula objects can also appear in data_only mode;
                    # treat anything that isn't a plain value as None.
                    if not isinstance(cached, _SIMPLE_TYPES):
                        cached = None
                    cell.value = cached
                    frozen_in_sheet += 1
                    if cached is None:
                        null_in_sheet += 1

        stats["other_sheets"] += 1
        stats["frozen_cells"] += frozen_in_sheet
        stats["null_cached_values"] += null_in_sheet
        logger.info(
            f"  [{sheet_name}]  frozen  "
            f"({frozen_in_sheet} formula cells → cached values, "
            f"{null_in_sheet} were None)"
        )

    if stats["null_cached_values"] > 0:
        logger.warning(
            f"{stats['null_cached_values']} formula cell(s) had no cached value. "
            "This happens when the workbook was never opened and saved in Excel "
            "after the formulas were entered. Those cells will appear as None/0 "
            "in the generated code. Open and save the file in Excel to fix this."
        )

    Path(frozen_path).parent.mkdir(parents=True, exist_ok=True)
    wb_f.save(frozen_path)
    logger.info(
        f"Frozen workbook saved: {frozen_path}  "
        f"({stats['other_sheets']} sheets frozen, "
        f"{stats['frozen_cells']} cells hardcoded)"
    )
    return stats
