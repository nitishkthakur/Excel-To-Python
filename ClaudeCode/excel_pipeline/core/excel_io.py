"""Central module for Excel I/O operations."""

import os
import tempfile
from pathlib import Path
from typing import Any, Dict, Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Alignment
from excel_pipeline.utils.logging_setup import get_logger
from excel_pipeline.utils.helpers import rgb_to_hex, hex_to_rgb, safe_str

logger = get_logger(__name__)


def load_workbook(filepath: str, data_only: bool = False, read_only: bool = False) -> Workbook:
    """
    Load Excel workbook with error handling and logging.

    Args:
        filepath: Path to Excel file
        data_only: If True, load cell values instead of formulas
        read_only: If True, open in read-only mode for better performance

    Returns:
        Loaded workbook

    Raises:
        FileNotFoundError: If file doesn't exist
        Exception: If workbook cannot be loaded

    Examples:
        >>> wb = load_workbook("model.xlsx")
        >>> wb_values = load_workbook("model.xlsx", data_only=True)
    """
    path = Path(filepath)

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    logger.info(f"Loading workbook: {filepath} (data_only={data_only}, read_only={read_only})")

    try:
        wb = openpyxl.load_workbook(
            filepath,
            data_only=data_only,
            read_only=read_only,
            keep_vba=False  # Don't load VBA to save memory
        )
        logger.info(f"Successfully loaded workbook with {len(wb.sheetnames)} sheets")
        return wb

    except Exception as e:
        logger.error(f"Failed to load workbook {filepath}: {e}")
        raise


def save_workbook(wb: Workbook, filepath: str, atomic: bool = True) -> None:
    """
    Save workbook with atomic write (temp file + rename) for safety.

    Args:
        wb: Workbook to save
        filepath: Destination file path
        atomic: If True, use atomic write (default: True)

    Raises:
        Exception: If save fails

    Examples:
        >>> save_workbook(wb, "output.xlsx")
    """
    path = Path(filepath)
    path.parent.mkdir(parents=True, exist_ok=True)

    logger.info(f"Saving workbook to: {filepath}")

    try:
        if atomic:
            # Write to temporary file first, then rename
            with tempfile.NamedTemporaryFile(
                mode='wb',
                delete=False,
                dir=path.parent,
                suffix='.xlsx'
            ) as tmp_file:
                temp_path = tmp_file.name

            wb.save(temp_path)

            # Atomic rename
            if os.path.exists(filepath):
                os.remove(filepath)
            os.rename(temp_path, filepath)
        else:
            wb.save(filepath)

        logger.info(f"Successfully saved workbook to {filepath}")

    except Exception as e:
        logger.error(f"Failed to save workbook to {filepath}: {e}")
        # Clean up temp file if it exists
        if atomic and 'temp_path' in locals() and os.path.exists(temp_path):
            os.remove(temp_path)
        raise


def get_cell_info(cell: Cell) -> Dict[str, Any]:
    """
    Extract comprehensive cell metadata: value, formula, format, style.

    Args:
        cell: openpyxl Cell object

    Returns:
        Dictionary containing all cell metadata

    Example:
        >>> info = get_cell_info(sheet['A1'])
        >>> print(info['value'], info['formula'], info['font_bold'])
    """
    # Check if this is a merged cell
    is_merged = isinstance(cell, openpyxl.cell.cell.MergedCell)

    # For merged cells, get column letter from column index
    if is_merged:
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(cell.column)
    else:
        column_letter = cell.column_letter

    info = {
        # Position
        'row': cell.row,
        'column': cell.column,
        'column_letter': column_letter,
        'coordinate': cell.coordinate,

        # Content
        'value': cell.value,
        'formula': cell.value if hasattr(cell, 'data_type') and cell.data_type == 'f' else None,
        'data_type': cell.data_type if hasattr(cell, 'data_type') else 'n',

        # Number format
        'number_format': cell.number_format if hasattr(cell, 'number_format') and cell.number_format else 'General',

        # Font (merged cells may not have formatting)
        'font_name': cell.font.name if hasattr(cell, 'font') and cell.font else None,
        'font_size': cell.font.size if hasattr(cell, 'font') and cell.font else None,
        'font_bold': cell.font.bold if hasattr(cell, 'font') and cell.font else False,
        'font_italic': cell.font.italic if hasattr(cell, 'font') and cell.font else False,
        'font_color': rgb_to_hex(cell.font.color.rgb) if hasattr(cell, 'font') and cell.font and cell.font.color and hasattr(cell.font.color, 'rgb') else None,

        # Fill
        'fill_color': rgb_to_hex(cell.fill.fgColor.rgb) if hasattr(cell, 'fill') and cell.fill and cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb') else None,

        # Alignment
        'alignment_horizontal': cell.alignment.horizontal if hasattr(cell, 'alignment') and cell.alignment else None,
        'alignment_vertical': cell.alignment.vertical if hasattr(cell, 'alignment') and cell.alignment else None,
        'wrap_text': cell.alignment.wrap_text if hasattr(cell, 'alignment') and cell.alignment else False,

        # Merged cell
        'is_merged': is_merged,
    }

    return info


def copy_cell_format(source_cell: Cell, target_cell: Cell) -> None:
    """
    Copy all formatting from source cell to target cell.

    Args:
        source_cell: Cell to copy format from
        target_cell: Cell to copy format to

    Note:
        This copies: font, fill, alignment, number format, border
        Does NOT copy: value, formula
    """
    # Copy font
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            color=source_cell.font.color,
            underline=source_cell.font.underline,
            strike=source_cell.font.strike
        )

    # Copy fill
    if source_cell.fill:
        target_cell.fill = PatternFill(
            patternType=source_cell.fill.patternType,
            fgColor=source_cell.fill.fgColor,
            bgColor=source_cell.fill.bgColor
        )

    # Copy alignment
    if source_cell.alignment:
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            text_rotation=source_cell.alignment.text_rotation,
            wrap_text=source_cell.alignment.wrap_text,
            shrink_to_fit=source_cell.alignment.shrink_to_fit,
            indent=source_cell.alignment.indent
        )

    # Copy number format
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format

    # Copy border
    if source_cell.border:
        target_cell.border = source_cell.border.copy()


def apply_cell_format(cell: Cell, format_dict: Dict[str, Any]) -> None:
    """
    Apply formatting to cell from dictionary.

    Args:
        cell: Cell to format
        format_dict: Dictionary with format properties (from get_cell_info)

    Example:
        >>> format_dict = {'font_bold': True, 'fill_color': '#FFFF00'}
        >>> apply_cell_format(sheet['A1'], format_dict)
    """
    # Font
    font_kwargs = {}
    if format_dict.get('font_name'):
        font_kwargs['name'] = format_dict['font_name']
    if format_dict.get('font_size'):
        font_kwargs['size'] = format_dict['font_size']
    if format_dict.get('font_bold'):
        font_kwargs['bold'] = True
    if format_dict.get('font_italic'):
        font_kwargs['italic'] = True
    if format_dict.get('font_color'):
        font_kwargs['color'] = hex_to_rgb(format_dict['font_color'])

    if font_kwargs:
        cell.font = Font(**font_kwargs)

    # Fill
    if format_dict.get('fill_color'):
        cell.fill = PatternFill(
            patternType='solid',
            fgColor=hex_to_rgb(format_dict['fill_color'])
        )

    # Alignment
    alignment_kwargs = {}
    if format_dict.get('alignment_horizontal'):
        alignment_kwargs['horizontal'] = format_dict['alignment_horizontal']
    if format_dict.get('alignment_vertical'):
        alignment_kwargs['vertical'] = format_dict['alignment_vertical']
    if format_dict.get('wrap_text'):
        alignment_kwargs['wrap_text'] = True

    if alignment_kwargs:
        cell.alignment = Alignment(**alignment_kwargs)

    # Number format
    if format_dict.get('number_format'):
        cell.number_format = format_dict['number_format']


def create_workbook_from_template(template_wb: Workbook, include_data: bool = False) -> Workbook:
    """
    Create a new workbook with same structure as template.

    Args:
        template_wb: Template workbook
        include_data: If True, copy cell values (default: False)

    Returns:
        New workbook with same sheets and structure

    Example:
        >>> new_wb = create_workbook_from_template(original_wb)
    """
    new_wb = Workbook()

    # Remove default sheet
    if 'Sheet' in new_wb.sheetnames:
        new_wb.remove(new_wb['Sheet'])

    # Create sheets matching template
    for sheet_name in template_wb.sheetnames:
        new_sheet = new_wb.create_sheet(title=sheet_name)

        if include_data:
            template_sheet = template_wb[sheet_name]

            # Copy dimensions
            new_sheet.column_dimensions = template_sheet.column_dimensions.copy()
            new_sheet.row_dimensions = template_sheet.row_dimensions.copy()

            # Copy cells
            for row in template_sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet[cell.coordinate]
                    new_cell.value = cell.value
                    copy_cell_format(cell, new_cell)

    return new_wb


def get_used_range(sheet) -> tuple:
    """
    Get the actual used range of a sheet (min/max row/col).

    Args:
        sheet: Worksheet object

    Returns:
        Tuple of (min_row, max_row, min_col, max_col)

    Example:
        >>> min_row, max_row, min_col, max_col = get_used_range(sheet)
    """
    return (
        sheet.min_row,
        sheet.max_row,
        sheet.min_column,
        sheet.max_column
    )
