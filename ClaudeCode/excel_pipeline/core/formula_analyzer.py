"""Formula analysis for detecting dragged patterns and groups for vectorization."""

import re
from typing import List, Set, Dict, Tuple, Optional
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.cell import Cell
from excel_pipeline.utils.logging_setup import get_logger

logger = get_logger(__name__)


@dataclass
class FormulaGroup:
    """
    Represents a group of dragged formulas.

    Detects all dragged formulas (2+ cells with same pattern).
    Large groups (>= vectorization_threshold) will be vectorized for performance.
    Small groups are still consolidated for readability.
    """
    group_id: int
    direction: str  # "horizontal" or "vertical"
    cells: List[str]  # List of cell coordinates in the group
    pattern: str  # Pattern formula (e.g., "=B{row}*C{row}")
    size: int  # Number of cells in group
    sheet_name: str
    is_vectorizable: bool  # True if size >= vectorization_threshold

    def __repr__(self) -> str:
        vec_marker = "VECTORIZABLE" if self.is_vectorizable else "DRAGGED"
        return (f"FormulaGroup(id={self.group_id}, {vec_marker}, dir={self.direction}, "
                f"size={self.size}, pattern={self.pattern[:50]}...)")


class FormulaAnalyzer:
    """
    Analyze formulas to detect patterns and groups for vectorization.

    The analyzer identifies rectangular groups of dragged formulas that
    can be vectorized for performance improvements on large workbooks (100MB+).
    """

    def __init__(self, vectorization_threshold: int = 10):
        """
        Initialize formula analyzer.

        Args:
            vectorization_threshold: Minimum cells in group to vectorize (default: 10)
        """
        self.vectorization_threshold = vectorization_threshold
        self.groups: List[FormulaGroup] = []
        self.next_group_id = 1

    def analyze_sheet(self, sheet) -> List[FormulaGroup]:
        """
        Analyze a worksheet to detect ALL dragged formula groups (2+ cells).

        Detects all patterns regardless of size, marking which are vectorizable.

        Args:
            sheet: openpyxl Worksheet object

        Returns:
            List of FormulaGroup objects (all dragged formulas)

        Example:
            >>> analyzer = FormulaAnalyzer()
            >>> groups = analyzer.analyze_sheet(sheet)
            >>> print(f"Found {len(groups)} dragged formula groups")
        """
        logger.info(f"Analyzing formulas in sheet: {sheet.title}")

        # Get all formula cells
        formula_cells = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    formula_cells.append(cell)

        # Sort by row, then column for consistent grouping
        formula_cells.sort(key=lambda c: (c.row, c.column))

        # Track which cells have been grouped
        visited = set()
        sheet_groups = []
        vectorizable_count = 0

        for cell in formula_cells:
            if cell.coordinate in visited:
                continue

            # Try to find horizontal and vertical groups
            h_group = self._expand_horizontal(sheet, cell, formula_cells, visited)
            v_group = self._expand_vertical(sheet, cell, formula_cells, visited)

            # Choose the larger group
            if h_group and v_group:
                group = h_group if h_group.size > v_group.size else v_group
            elif h_group:
                group = h_group
            elif v_group:
                group = v_group
            else:
                continue

            # Include ALL groups with 2+ cells (dragged formulas)
            if group.size >= 2:
                sheet_groups.append(group)
                visited.update(group.cells)
                if group.is_vectorizable:
                    vectorizable_count += 1
                    logger.debug(f"Found vectorizable group: {group}")
                else:
                    logger.debug(f"Found dragged group (not vectorizable): {group}")

        logger.info(f"Sheet {sheet.title}: {len(sheet_groups)} dragged formula groups "
                   f"({vectorizable_count} vectorizable, threshold: {self.vectorization_threshold} cells)")

        self.groups.extend(sheet_groups)
        return sheet_groups

    def _expand_horizontal(self, sheet, start_cell: Cell,
                          formula_cells: List[Cell], visited: Set[str]) -> Optional[FormulaGroup]:
        """
        Expand horizontally to find dragged formula group.

        Args:
            sheet: Worksheet
            start_cell: Starting cell
            formula_cells: All formula cells
            visited: Set of already-visited cells

        Returns:
            FormulaGroup or None
        """
        cells = [start_cell]
        row = start_cell.row
        col = start_cell.column

        # Get pattern from start cell
        base_formula = str(start_cell.value)

        # Try to expand right
        current_col = col + 1
        while True:
            next_cell = sheet.cell(row=row, column=current_col)

            if next_cell.data_type != 'f':
                break

            # Check if formula matches pattern
            if self._is_horizontally_dragged(base_formula, str(next_cell.value), col, current_col):
                cells.append(next_cell)
                current_col += 1
            else:
                break

        if len(cells) < 2:
            return None

        # Generate pattern formula
        pattern = self._generate_horizontal_pattern(base_formula, col)

        group = FormulaGroup(
            group_id=self.next_group_id,
            direction="horizontal",
            cells=[c.coordinate for c in cells],
            pattern=pattern,
            size=len(cells),
            sheet_name=sheet.title,
            is_vectorizable=(len(cells) >= self.vectorization_threshold)
        )

        self.next_group_id += 1
        return group

    def _expand_vertical(self, sheet, start_cell: Cell,
                        formula_cells: List[Cell], visited: Set[str]) -> Optional[FormulaGroup]:
        """
        Expand vertically to find dragged formula group.

        Args:
            sheet: Worksheet
            start_cell: Starting cell
            formula_cells: All formula cells
            visited: Set of already-visited cells

        Returns:
            FormulaGroup or None
        """
        cells = [start_cell]
        row = start_cell.row
        col = start_cell.column

        # Get pattern from start cell
        base_formula = str(start_cell.value)

        # Try to expand down
        current_row = row + 1
        while True:
            next_cell = sheet.cell(row=current_row, column=col)

            if next_cell.data_type != 'f':
                break

            # Check if formula matches pattern
            if self._is_vertically_dragged(base_formula, str(next_cell.value), row, current_row):
                cells.append(next_cell)
                current_row += 1
            else:
                break

        if len(cells) < 2:
            return None

        # Generate pattern formula
        pattern = self._generate_vertical_pattern(base_formula, row)

        group = FormulaGroup(
            group_id=self.next_group_id,
            direction="vertical",
            cells=[c.coordinate for c in cells],
            pattern=pattern,
            size=len(cells),
            sheet_name=sheet.title,
            is_vectorizable=(len(cells) >= self.vectorization_threshold)
        )

        self.next_group_id += 1
        return group

    def _is_horizontally_dragged(self, formula1: str, formula2: str,
                                  col1: int, col2: int) -> bool:
        """
        Check if two formulas are horizontally dragged versions of each other.

        Args:
            formula1: First formula
            formula2: Second formula
            col1: Column index of first formula
            col2: Column index of second formula

        Returns:
            True if formulas match horizontal drag pattern
        """
        # Simple heuristic: Replace column references and compare
        col_diff = col2 - col1

        # Convert formula1 by shifting columns
        shifted = self._shift_formula_columns(formula1, col_diff)

        return shifted == formula2

    def _is_vertically_dragged(self, formula1: str, formula2: str,
                               row1: int, row2: int) -> bool:
        """
        Check if two formulas are vertically dragged versions of each other.

        Args:
            formula1: First formula
            formula2: Second formula
            row1: Row index of first formula
            row2: Row index of second formula

        Returns:
            True if formulas match vertical drag pattern
        """
        # Simple heuristic: Replace row references and compare
        row_diff = row2 - row1

        # Convert formula1 by shifting rows
        shifted = self._shift_formula_rows(formula1, row_diff)

        return shifted == formula2

    def _shift_formula_columns(self, formula: str, col_shift: int) -> str:
        """Shift all column references in formula by col_shift."""
        def shift_col(match):
            abs_marker = match.group(1)  # $ or empty
            col_letter = match.group(2)

            # Don't shift absolute references
            if abs_marker == '$':
                return match.group(0)

            # Convert to index, shift, convert back
            from openpyxl.utils import column_index_from_string, get_column_letter
            col_index = column_index_from_string(col_letter)
            new_col = get_column_letter(col_index + col_shift)
            return f"{abs_marker}{new_col}"

        pattern = r'(\$?)([A-Z]+)(?=\d)'
        return re.sub(pattern, shift_col, formula)

    def _shift_formula_rows(self, formula: str, row_shift: int) -> str:
        """Shift all row references in formula by row_shift."""
        def shift_row(match):
            abs_marker = match.group(1)  # $ or empty
            row_num = int(match.group(2))

            # Don't shift absolute references
            if abs_marker == '$':
                return match.group(0)

            new_row = row_num + row_shift
            return f"{abs_marker}{new_row}"

        pattern = r'(\$?)(\d+)'
        return re.sub(pattern, shift_row, formula)

    def _generate_horizontal_pattern(self, formula: str, base_col: int) -> str:
        """
        Generate pattern formula for horizontal drag.

        Example: "=B2*C2" with base_col=2 becomes "=B{col}*C{col}"
        """
        # Replace column letters with {col} placeholder
        pattern = re.sub(r'(\$?)([A-Z]+)(?=\d)', r'\1{col}', formula)
        return pattern

    def _generate_vertical_pattern(self, formula: str, base_row: int) -> str:
        """
        Generate pattern formula for vertical drag.

        Example: "=B2*C2" with base_row=2 becomes "=B{row}*C{row}"
        """
        # Replace row numbers with {row} placeholder (preserve absolute refs)
        def replace_row(match):
            abs_marker = match.group(1)
            return f"{abs_marker}{{row}}"

        pattern = re.sub(r'(\$?)(\d+)', replace_row, formula)
        return pattern

    def get_group_by_id(self, group_id: int) -> Optional[FormulaGroup]:
        """Get formula group by ID."""
        for group in self.groups:
            if group.group_id == group_id:
                return group
        return None

    def get_stats(self) -> Dict[str, int]:
        """Get analyzer statistics."""
        total_cells = sum(g.size for g in self.groups)
        vectorizable_groups = [g for g in self.groups if g.is_vectorizable]
        vectorizable_cells = sum(g.size for g in vectorizable_groups)
        horizontal = sum(1 for g in self.groups if g.direction == "horizontal")
        vertical = sum(1 for g in self.groups if g.direction == "vertical")

        return {
            'total_groups': len(self.groups),
            'total_cells_in_groups': total_cells,
            'vectorizable_groups': len(vectorizable_groups),
            'vectorizable_cells': vectorizable_cells,
            'horizontal_groups': horizontal,
            'vertical_groups': vertical,
            'avg_group_size': total_cells // len(self.groups) if self.groups else 0,
        }
