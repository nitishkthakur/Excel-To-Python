"""Write mapping report Excel file with cell metadata."""

from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from excel_pipeline.layer1.cell_extractor import CellMetadata
from excel_pipeline.core.excel_io import save_workbook
from excel_pipeline.utils.logging_setup import get_logger
from excel_pipeline.utils.config import config

logger = get_logger(__name__)


class MappingWriter:
    """Write mapping report with all cell metadata."""

    # Column headers for mapping report
    HEADERS = [
        "RowNum", "ColNum", "Cell", "Type", "Formula", "Value",
        "NumberFormat", "FontBold", "FontItalic", "FontSize",
        "FontColor", "FillColor", "Alignment", "WrapText",
        "GroupID", "GroupDirection", "GroupSize", "PatternFormula",
        "Vectorizable", "IncludeFlag"
    ]

    def __init__(self, original_workbook_name: str):
        """
        Initialize mapping writer.

        Args:
            original_workbook_name: Name of original Excel file
        """
        self.original_workbook_name = original_workbook_name
        self.wb = Workbook()

    def write(self, cells: List[CellMetadata], output_path: str,
             dep_graph_stats: dict, analyzer_stats: dict) -> None:
        """
        Write mapping report to Excel file.

        Args:
            cells: List of CellMetadata objects
            output_path: Path to save mapping_report.xlsx
            dep_graph_stats: Statistics from DependencyGraph
            analyzer_stats: Statistics from FormulaAnalyzer

        Example:
            >>> writer = MappingWriter("model.xlsx")
            >>> writer.write(cells, "mapping_report.xlsx", stats)
        """
        logger.info(f"Writing mapping report to: {output_path}")

        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])

        # Group cells by sheet
        cells_by_sheet = {}
        for cell in cells:
            if cell.sheet_name not in cells_by_sheet:
                cells_by_sheet[cell.sheet_name] = []
            cells_by_sheet[cell.sheet_name].append(cell)

        # Write each sheet (with consolidation of vectorizable groups)
        for sheet_name, sheet_cells in cells_by_sheet.items():
            self._write_sheet_consolidated(sheet_name, sheet_cells)

        # Write metadata sheet
        self._write_metadata_sheet(cells, dep_graph_stats, analyzer_stats)

        # Save workbook
        save_workbook(self.wb, output_path)
        logger.info(f"Mapping report saved successfully")

    def _write_sheet_consolidated(self, sheet_name: str, cells: List[CellMetadata]) -> None:
        """
        Write cells from one source sheet to mapping report.

        Consolidates vectorizable groups into single rows showing ranges
        for better readability and to highlight where vectorization applies.
        """
        logger.debug(f"Writing mapping sheet: {sheet_name}")

        ws = self.wb.create_sheet(title=sheet_name)

        # Write headers
        for col_idx, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Group cells by GroupID for consolidation
        grouped_cells = {}  # group_id -> list of cells
        ungrouped_cells = []  # cells with group_id = 0

        for cell_meta in cells:
            if cell_meta.group_id > 0:
                if cell_meta.group_id not in grouped_cells:
                    grouped_cells[cell_meta.group_id] = []
                grouped_cells[cell_meta.group_id].append(cell_meta)
            else:
                ungrouped_cells.append(cell_meta)

        # Write data rows
        row_idx = 2

        # First, write ungrouped cells (individual rows)
        for cell_meta in ungrouped_cells:
            self._write_cell_row(ws, row_idx, cell_meta, is_group=False)
            row_idx += 1

        # Then, write grouped cells (consolidated rows showing ranges)
        for group_id in sorted(grouped_cells.keys()):
            group_cells = grouped_cells[group_id]
            self._write_group_row(ws, row_idx, group_cells)
            row_idx += 1

        # Auto-adjust column widths
        for col_idx in range(1, len(self.HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # Make Formula and PatternFormula columns wider
        ws.column_dimensions['E'].width = 40  # Formula
        ws.column_dimensions['R'].width = 40  # PatternFormula

        # Make Cell column wider for ranges
        ws.column_dimensions['C'].width = 20

    def _write_cell_row(self, ws, row_idx: int, cell_meta: CellMetadata, is_group: bool = False) -> None:
        """Write a single cell row."""
        # Prefix formula with apostrophe to make it display as text
        formula_display = f"'{cell_meta.formula}" if cell_meta.formula else ""

        ws.cell(row=row_idx, column=1, value=cell_meta.row_num)
        ws.cell(row=row_idx, column=2, value=cell_meta.col_letter)
        ws.cell(row=row_idx, column=3, value=cell_meta.cell_coordinate)
        ws.cell(row=row_idx, column=4, value=cell_meta.cell_type)
        ws.cell(row=row_idx, column=5, value=formula_display)
        ws.cell(row=row_idx, column=6, value=cell_meta.value)
        ws.cell(row=row_idx, column=7, value=cell_meta.number_format)
        ws.cell(row=row_idx, column=8, value=cell_meta.font_bold)
        ws.cell(row=row_idx, column=9, value=cell_meta.font_italic)
        ws.cell(row=row_idx, column=10, value=cell_meta.font_size)
        ws.cell(row=row_idx, column=11, value=cell_meta.font_color)
        ws.cell(row=row_idx, column=12, value=cell_meta.fill_color)
        ws.cell(row=row_idx, column=13, value=cell_meta.alignment)
        ws.cell(row=row_idx, column=14, value=cell_meta.wrap_text)
        ws.cell(row=row_idx, column=15, value=cell_meta.group_id if cell_meta.group_id > 0 else "")
        ws.cell(row=row_idx, column=16, value=cell_meta.group_direction)
        ws.cell(row=row_idx, column=17, value=cell_meta.group_size if cell_meta.group_size > 0 else "")
        ws.cell(row=row_idx, column=18, value=cell_meta.pattern_formula)
        ws.cell(row=row_idx, column=19, value=cell_meta.is_vectorizable)
        ws.cell(row=row_idx, column=20, value=cell_meta.include_flag)

    def _write_group_row(self, ws, row_idx: int, group_cells: List[CellMetadata]) -> None:
        """
        Write a consolidated row for a dragged formula group.

        Shows the cell range (e.g., B2:B50) instead of individual cells,
        making it clear where formulas were dragged.
        Vectorizable groups (large) are highlighted differently from small dragged groups.
        """
        if not group_cells:
            return

        # Sort cells by position
        group_cells.sort(key=lambda c: (c.row_num, c.col_letter))

        first_cell = group_cells[0]
        last_cell = group_cells[-1]

        # Create range string (e.g., "B2:B50")
        if len(group_cells) == 1:
            cell_range = first_cell.cell_coordinate
        else:
            cell_range = f"{first_cell.cell_coordinate}:{last_cell.cell_coordinate}"

        # Prefix pattern formula with apostrophe
        pattern_display = f"'{first_cell.pattern_formula}" if first_cell.pattern_formula else ""

        # Value display depends on whether vectorizable
        value_display = "[VECTORIZED]" if first_cell.is_vectorizable else "[DRAGGED]"

        # Write consolidated row
        ws.cell(row=row_idx, column=1, value=f"{first_cell.row_num}-{last_cell.row_num}")
        ws.cell(row=row_idx, column=2, value=f"{first_cell.col_letter}[-{last_cell.col_letter}]" if first_cell.col_letter != last_cell.col_letter else first_cell.col_letter)
        ws.cell(row=row_idx, column=3, value=cell_range)
        ws.cell(row=row_idx, column=4, value=first_cell.cell_type)
        ws.cell(row=row_idx, column=5, value=pattern_display)  # Pattern formula, not individual formula
        ws.cell(row=row_idx, column=6, value=value_display)
        ws.cell(row=row_idx, column=7, value=first_cell.number_format)
        ws.cell(row=row_idx, column=8, value=first_cell.font_bold)
        ws.cell(row=row_idx, column=9, value=first_cell.font_italic)
        ws.cell(row=row_idx, column=10, value=first_cell.font_size)
        ws.cell(row=row_idx, column=11, value=first_cell.font_color)
        ws.cell(row=row_idx, column=12, value=first_cell.fill_color)
        ws.cell(row=row_idx, column=13, value=first_cell.alignment)
        ws.cell(row=row_idx, column=14, value=first_cell.wrap_text)
        ws.cell(row=row_idx, column=15, value=first_cell.group_id)
        ws.cell(row=row_idx, column=16, value=first_cell.group_direction)
        ws.cell(row=row_idx, column=17, value=first_cell.group_size)
        ws.cell(row=row_idx, column=18, value=pattern_display)  # Same as formula column
        ws.cell(row=row_idx, column=19, value=first_cell.is_vectorizable)
        ws.cell(row=row_idx, column=20, value=first_cell.include_flag)

        # Highlight rows: green for vectorizable, yellow for dragged but not vectorizable
        if first_cell.is_vectorizable:
            # Light green for vectorizable groups
            color = "E8F5E9"
        else:
            # Light yellow for dragged (but not vectorizable) groups
            color = "FFF9C4"

        for col_idx in range(1, len(self.HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _write_metadata_sheet(self, cells: List[CellMetadata],
                              dep_graph_stats: dict, analyzer_stats: dict) -> None:
        """
        Write metadata summary sheet.

        This sheet provides high-level statistics about the workbook.
        """
        logger.debug("Writing metadata sheet")

        ws = self.wb.create_sheet(title="_Metadata", index=0)  # First sheet

        # Count cells by type
        type_counts = {'Input': 0, 'Calculation': 0, 'Output': 0}
        for cell in cells:
            type_counts[cell.cell_type] += 1

        # Count grouped and vectorizable cells
        grouped = sum(1 for c in cells if c.group_id > 0)
        vectorizable = sum(1 for c in cells if c.is_vectorizable)

        # Write metadata
        metadata = [
            ("Original Workbook", self.original_workbook_name),
            ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Pipeline Version", config.version),
            ("", ""),
            ("Total Cells", len(cells)),
            ("Input Cells", type_counts['Input']),
            ("Calculation Cells", type_counts['Calculation']),
            ("Output Cells", type_counts['Output']),
            ("", ""),
            ("Formula Cells", dep_graph_stats.get('total_formula_cells', 0)),
            ("Cells with Dependencies", dep_graph_stats.get('cells_with_precedents', 0)),
            ("Circular References", dep_graph_stats.get('circular_references', 0)),
            ("", ""),
            ("Dragged Formula Groups", analyzer_stats.get('total_groups', 0)),
            ("Dragged Formula Cells", grouped),
            ("Vectorizable Groups", analyzer_stats.get('vectorizable_groups', 0)),
            ("Vectorizable Cells", vectorizable),
            ("Horizontal Groups", analyzer_stats.get('horizontal_groups', 0)),
            ("Vertical Groups", analyzer_stats.get('vertical_groups', 0)),
            ("Avg Group Size", analyzer_stats.get('avg_group_size', 0)),
        ]

        # Write to sheet
        for row_idx, (key, value) in enumerate(metadata, start=1):
            cell_key = ws.cell(row=row_idx, column=1, value=key)
            cell_val = ws.cell(row=row_idx, column=2, value=value)

            if key == "":
                continue  # Empty row

            cell_key.font = Font(bold=True)

        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

        logger.info(f"Metadata: {len(cells)} cells, {grouped} in dragged groups "
                   f"({vectorizable} vectorizable in {analyzer_stats.get('vectorizable_groups', 0)} groups)")
