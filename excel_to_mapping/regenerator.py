"""
Regenerate an Excel workbook from an intermediate mapping report.

Reads the tabular mapping Excel file produced by :func:`mapper.generate_mapping_report`
and reconstructs the original workbook (formulas, values, formatting, and layout).
Optionally accepts an input-template override so users can supply new hardcoded
values while keeping all formulas intact.
"""

import json
import os
import re
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

from excel_to_python import col_letter_to_index, index_to_col_letter
from excel_to_python_vectorized.vectorizer import extract_references

from .mapper import COLUMNS


# ------------------------------------------------------------------
# Formula shifting
# ------------------------------------------------------------------

def _shift_formula(formula, current_sheet, row_offset, col_offset):
    """Shift all non-absolute references in *formula* by the given offsets.

    Parameters
    ----------
    formula : str
        The Excel formula (may or may not start with ``=``).
    current_sheet : str
        Sheet name used to resolve same-sheet references.
    row_offset : int
        Number of rows to shift relative row references.
    col_offset : int
        Number of columns to shift relative column references.

    Returns
    -------
    str
        The formula with shifted references.
    """
    if row_offset == 0 and col_offset == 0:
        return formula

    prefix = ""
    body = formula
    if formula.startswith("="):
        prefix = "="
        body = formula[1:]

    refs = extract_references(formula, current_sheet)
    if not refs:
        return formula

    # Process from right to left so earlier replacements don't shift positions
    refs_sorted = sorted(refs, key=lambda r: r.start, reverse=True)

    for ref in refs_sorted:
        if ref.kind == "cell":
            new_col_idx = ref.col_idx + (0 if ref.col_abs else col_offset)
            new_row = ref.row + (0 if ref.row_abs else row_offset)
            if new_col_idx < 1 or new_row < 1:
                continue
            col_dollar = "$" if ref.col_abs else ""
            row_dollar = "$" if ref.row_abs else ""
            new_ref_str = f"{col_dollar}{index_to_col_letter(new_col_idx)}{row_dollar}{new_row}"

            # Preserve sheet prefix if present in raw
            raw = ref.raw
            # Find the cell portion in raw to know what prefix (sheet!) exists
            cell_portion = f"{'$' if ref.col_abs else ''}{ref.col}{'$' if ref.row_abs else ''}{ref.row}"
            if raw.endswith(cell_portion):
                sheet_prefix = raw[: len(raw) - len(cell_portion)]
            else:
                sheet_prefix = ""
            replacement = f"{sheet_prefix}{new_ref_str}"

        elif ref.kind == "range":
            new_col_idx = ref.col_idx + (0 if ref.col_abs else col_offset)
            new_row = ref.row + (0 if ref.row_abs else row_offset)
            new_end_col_idx = ref.end_col_idx + (0 if ref.end_col_abs else col_offset)
            new_end_row = ref.end_row + (0 if ref.end_row_abs else row_offset)
            if new_col_idx < 1 or new_row < 1 or new_end_col_idx < 1 or new_end_row < 1:
                continue
            col_dollar = "$" if ref.col_abs else ""
            row_dollar = "$" if ref.row_abs else ""
            end_col_dollar = "$" if ref.end_col_abs else ""
            end_row_dollar = "$" if ref.end_row_abs else ""
            new_start = f"{col_dollar}{index_to_col_letter(new_col_idx)}{row_dollar}{new_row}"
            new_end = f"{end_col_dollar}{index_to_col_letter(new_end_col_idx)}{end_row_dollar}{new_end_row}"

            raw = ref.raw
            # Derive sheet prefix from raw text
            start_cell = (
                f"{'$' if ref.col_abs else ''}{ref.col}"
                f"{'$' if ref.row_abs else ''}{ref.row}"
            )
            idx = raw.find(start_cell)
            sheet_prefix = raw[:idx] if idx > 0 else ""
            replacement = f"{sheet_prefix}{new_start}:{new_end}"
        else:
            continue

        # Replace in body using character positions (extract_references offsets
        # are relative to body without the leading '=')
        start = ref.start
        end = ref.end
        body = body[:start] + replacement + body[end:]

    return prefix + body


# ------------------------------------------------------------------
# Group expansion
# ------------------------------------------------------------------

def _expand_group(row_dict):
    """Expand a group row into individual ``(cell_address, formula)`` tuples.

    Parameters
    ----------
    row_dict : dict
        A single row from the mapping report with keys from :data:`COLUMNS`.

    Returns
    -------
    list[tuple[str, str]]
        One ``(cell_address, formula)`` entry per cell in the group.
    """
    cell_range = str(row_dict["Cell"])
    formula = row_dict["Formula"]
    direction = row_dict["GroupDirection"]
    sheet = row_dict["Sheet"]

    m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", cell_range)
    if not m:
        return [(cell_range, formula)]

    start_col, start_row = m.group(1), int(m.group(2))
    end_col, end_row = m.group(3), int(m.group(4))
    start_col_idx = col_letter_to_index(start_col)
    end_col_idx = col_letter_to_index(end_col)

    results = []

    if direction == "vertical":
        for r in range(start_row, end_row + 1):
            offset = r - start_row
            shifted = _shift_formula(formula, sheet, row_offset=offset, col_offset=0)
            results.append((f"{start_col}{r}", shifted))
    else:  # horizontal
        for c in range(start_col_idx, end_col_idx + 1):
            offset = c - start_col_idx
            col_letter = index_to_col_letter(c)
            shifted = _shift_formula(formula, sheet, row_offset=0, col_offset=offset)
            results.append((f"{col_letter}{start_row}", shifted))

    return results


# ------------------------------------------------------------------
# Formatting helpers
# ------------------------------------------------------------------

_ARGB_RE = re.compile(r"^[0-9A-Fa-f]{8}$")


def _is_valid_argb(color):
    """Return True if *color* is a valid 8-character ARGB hex string."""
    return bool(color and _ARGB_RE.match(str(color)))


def _read_sheet_rows(ws):
    """Read a worksheet into a list of dicts keyed by the header row.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Worksheet whose first row contains column headers.

    Returns
    -------
    list[dict]
        One dict per data row (row 2 onwards).
    """
    headers = [ws.cell(row=1, column=c).value
               for c in range(1, ws.max_column + 1)]
    col_map = {h: i for i, h in enumerate(headers)}
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value
                for c in range(1, ws.max_column + 1)]
        rows.append({h: vals[i] for h, i in col_map.items()})
    return rows


def _apply_formatting(ws, cell_addr, row_dict):
    """Apply formatting attributes from *row_dict* to a worksheet cell.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Target worksheet.
    cell_addr : str
        Cell address such as ``"A1"``.
    row_dict : dict
        Row from the mapping report.
    """
    cell = ws[cell_addr]

    # Number format
    nf = row_dict.get("NumberFormat")
    if nf:
        cell.number_format = nf

    # Font
    font_kw = {}
    if row_dict.get("FontBold"):
        font_kw["bold"] = True
    if row_dict.get("FontItalic"):
        font_kw["italic"] = True
    if row_dict.get("FontSize"):
        font_kw["size"] = row_dict["FontSize"]
    fc = row_dict.get("FontColor")
    if _is_valid_argb(fc):
        font_kw["color"] = str(fc)
    if font_kw:
        cell.font = Font(**font_kw)

    # Fill
    fill_color = row_dict.get("FillColor")
    if _is_valid_argb(fill_color):
        cell.fill = PatternFill(
            start_color=str(fill_color),
            end_color=str(fill_color),
            fill_type="solid",
        )

    # Alignment
    align_kw = {}
    if row_dict.get("HorizAlign"):
        align_kw["horizontal"] = row_dict["HorizAlign"]
    if row_dict.get("VertAlign"):
        align_kw["vertical"] = row_dict["VertAlign"]
    if row_dict.get("WrapText"):
        align_kw["wrap_text"] = True
    if align_kw:
        cell.alignment = Alignment(**align_kw)


# ------------------------------------------------------------------
# Input template generation
# ------------------------------------------------------------------

def generate_input_template(mapping_path, output_path):
    """Generate an input-template Excel with only the Input cells.

    Users can fill in or override values in the template and later pass it
    to :func:`regenerate_workbook` via the *input_values_path* parameter.

    Parameters
    ----------
    mapping_path : str
        Path to the intermediate mapping report Excel file.
    output_path : str
        Where to write the input template.

    Returns
    -------
    str
        Path to the generated template file.
    """
    wb_map = load_workbook(mapping_path)
    wb_out = Workbook()
    has_default = "Sheet" in wb_out.sheetnames

    for sn in wb_map.sheetnames:
        if sn == "_Metadata":
            continue

        ws_src = wb_map[sn]
        src_rows = _read_sheet_rows(ws_src)

        ws_out = wb_out.create_sheet(sn)
        ws_out.cell(row=1, column=1, value="Cell")
        ws_out.cell(row=1, column=2, value="Value")

        out_row = 2
        for row_dict in src_rows:
            if row_dict.get("Type") != "Input" or not row_dict.get("IncludeFlag"):
                continue
            ws_out.cell(row=out_row, column=1, value=row_dict.get("Cell"))
            ws_out.cell(row=out_row, column=2, value=row_dict.get("Value"))
            out_row += 1

    if has_default and "Sheet" in wb_out.sheetnames and len(wb_out.sheetnames) > 1:
        del wb_out["Sheet"]

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb_out.save(output_path)
    wb_map.close()
    print(f"Generated input template: {output_path}")
    return output_path


# ------------------------------------------------------------------
# Workbook regeneration
# ------------------------------------------------------------------

def regenerate_workbook(mapping_path, output_path, input_values_path=None):
    """Regenerate an Excel workbook from the intermediate mapping report.

    Parameters
    ----------
    mapping_path : str
        Path to the intermediate mapping report Excel file produced by
        :func:`mapper.generate_mapping_report`.
    output_path : str
        Path for the regenerated output Excel workbook.
    input_values_path : str or None
        Optional path to an input-template Excel whose values override the
        Input rows in the mapping file.

    Returns
    -------
    str
        Path to the regenerated workbook.
    """
    wb_map = load_workbook(mapping_path)

    # ------ read metadata ------
    meta = {}
    if "_Metadata" in wb_map.sheetnames:
        ws_meta = wb_map["_Metadata"]
        meta_headers = [ws_meta.cell(row=1, column=c).value
                        for c in range(1, ws_meta.max_column + 1)]
        mh = {h: i for i, h in enumerate(meta_headers)}
        for r in range(2, ws_meta.max_row + 1):
            vals = [ws_meta.cell(row=r, column=c).value
                    for c in range(1, ws_meta.max_column + 1)]
            sn = vals[mh["SheetName"]]
            merged_raw = vals[mh.get("MergedCells", 1)] or ""
            merged = [s.strip() for s in merged_raw.split(";") if s.strip()]
            col_widths = json.loads(vals[mh.get("ColWidths", 2)] or "{}")
            row_heights = json.loads(vals[mh.get("RowHeights", 3)] or "{}")
            meta[sn] = {
                "merged_cells": merged,
                "col_widths": col_widths,
                "row_heights": row_heights,
            }

    # ------ read input overrides ------
    input_overrides = {}  # {sheet_name: {cell_addr: value}}
    if input_values_path:
        wb_inp = load_workbook(input_values_path)
        for sn in wb_inp.sheetnames:
            ws_inp = wb_inp[sn]
            inp_headers = [ws_inp.cell(row=1, column=c).value
                           for c in range(1, ws_inp.max_column + 1)]
            ih = {h: i for i, h in enumerate(inp_headers)}
            cell_col = ih.get("Cell", 0)
            val_col = ih.get("Value", 1)
            overrides = {}
            for r in range(2, ws_inp.max_row + 1):
                ca = ws_inp.cell(row=r, column=cell_col + 1).value
                v = ws_inp.cell(row=r, column=val_col + 1).value
                if ca is not None:
                    overrides[str(ca)] = v
            input_overrides[sn] = overrides
        wb_inp.close()

    # ------ build output workbook ------
    wb_out = Workbook()
    has_default = "Sheet" in wb_out.sheetnames

    for sn in wb_map.sheetnames:
        if sn == "_Metadata":
            continue

        ws_src = wb_map[sn]
        data_rows = _read_sheet_rows(ws_src)

        # Filter to included rows
        data_rows = [d for d in data_rows if d.get("IncludeFlag")]

        ws_out = wb_out.create_sheet(sn)

        for row_dict in data_rows:
            rtype = row_dict.get("Type")
            cell_addr = str(row_dict.get("Cell") or "")
            formula = row_dict.get("Formula")
            value = row_dict.get("Value")
            group_id = row_dict.get("GroupID")

            if rtype == "Input":
                # Use override value if available
                override = input_overrides.get(sn, {}).get(cell_addr)
                ws_out[cell_addr] = override if override is not None else value
                _apply_formatting(ws_out, cell_addr, row_dict)

            elif group_id:
                # Formula group: expand and write each cell
                expanded = _expand_group(row_dict)
                for addr, shifted_formula in expanded:
                    ws_out[addr] = shifted_formula
                    _apply_formatting(ws_out, addr, row_dict)

            elif formula is not None:
                # Single formula cell
                ws_out[cell_addr] = formula
                _apply_formatting(ws_out, cell_addr, row_dict)

        # ------ apply metadata ------
        sheet_meta = meta.get(sn, {})

        for mc in sheet_meta.get("merged_cells", []):
            try:
                ws_out.merge_cells(mc)
            except (ValueError, TypeError):
                pass

        for col_key, width in sheet_meta.get("col_widths", {}).items():
            try:
                col_idx = int(col_key)
                ws_out.column_dimensions[get_column_letter(col_idx)].width = width
            except (ValueError, TypeError):
                pass

        for row_key, height in sheet_meta.get("row_heights", {}).items():
            try:
                ws_out.row_dimensions[int(row_key)].height = height
            except (ValueError, TypeError):
                pass

    # Remove the default "Sheet" if real sheets were added
    if has_default and "Sheet" in wb_out.sheetnames and len(wb_out.sheetnames) > 1:
        del wb_out["Sheet"]

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb_out.save(output_path)
    wb_map.close()

    print(f"Regenerated workbook: {output_path}")
    return output_path
