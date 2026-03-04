"""
Generate a **structured input** Excel workbook from a mapping report.

The structured input file is designed for *new users* who need to supply
hardcoded (Input) values for a model without having to navigate the raw
source Excel.  It contains:

* **Index** – overview sheet mapping every column in this file back to the
  source sheet / cell range in the original workbook.
* **Config** – scalar / isolated input cells; 4 columns:
  ``Source Sheet | Cell | Label | Value``
* **One sheet per source sheet** that contains vector inputs (contiguous
  horizontal runs of ≥ 2 same-row Input cells).  Each sheet has:

  - Row 1: column headers of the form ``<period_label> [<col_letter>]``
    where *period_label* is inferred from the source workbook's header row.
  - Column A: row label (metric name from the source workbook).
  - Remaining columns: one column per time period / year.

Usage
-----
    from excel_to_mapping.structured_input_generator import generate_structured_input

    generate_structured_input(
        mapping_path="output/mapping_report.xlsx",
        excel_path="Indigo.xlsx",
        output_path="output/structured_input.xlsx",
    )
"""

import os
import re
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from excel_to_python import col_letter_to_index, index_to_col_letter


# ──────────────────────────────────────────────────────────────────
# Shared style constants
# ──────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4",
                           fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

_INPUT_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                           fill_type="solid")  # light green – same as mapper

_ALT_FILL    = PatternFill(start_color="EBF3FB", end_color="EBF3FB",
                           fill_type="solid")  # light blue alternating row

_SCALAR_NOTE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                                fill_type="solid")

_CELL_RE   = re.compile(r"^([A-Z]+)(\d+)$")
_RANGE_RE  = re.compile(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$")


# ──────────────────────────────────────────────────────────────────
# Financial date / period detection
# ──────────────────────────────────────────────────────────────────

import datetime as _dt

# Ordered from most-specific to least-specific so that the first full
# match wins.  All patterns use re.IGNORECASE where case matters.
_FINANCIAL_DATE_PATTERNS = [
    # Quarter – number first: 1Q2021, 4Q21, 1Q-2024
    re.compile(r'^[1-4]Q[-]?\d{2,4}$', re.IGNORECASE),
    # Quarter – Q prefix: Q12023, Q421, Q1-2024
    re.compile(r'^Q[1-4][-]?\d{2,4}$', re.IGNORECASE),
    # Quarter – year first: 2024Q1, 24Q4
    re.compile(r'^\d{2,4}Q[1-4]$', re.IGNORECASE),
    # Quarter – year first, number after: 20241Q, 2024-1Q, 24-4Q
    re.compile(r'^\d{2,4}[-]?[1-4]Q$', re.IGNORECASE),
    # Half-year – H first: H12024, H1-24
    re.compile(r'^H[1-2][-]?\d{2,4}$', re.IGNORECASE),
    # Half-year – year first: 2024H1, 24-H2
    re.compile(r'^\d{2,4}[-]?H[1-2]$', re.IGNORECASE),
    # Fiscal year: FY2024, FYE2024, FY24, FY24E, FYE24A
    re.compile(r'^FYE?\d{2,4}[EABFP]?$', re.IGNORECASE),
    # Fiscal year – year first: 2024FY, 24FYE
    re.compile(r'^\d{2,4}FYE?[EABFP]?$', re.IGNORECASE),
    # Calendar year with fiscal tag: CY2024, CY24E
    re.compile(r'^CY\d{2,4}[EABFP]?$', re.IGNORECASE),
    # Year with single-letter financial suffix: 2024E, 2024A, 2024F, 2024B, 2024P
    re.compile(r'^\d{4}[EABFP]$', re.IGNORECASE),
    # Suffix-first: E2024, A2024
    re.compile(r'^[EABFP]\d{4}$', re.IGNORECASE),
    # Plain 4-digit year (most common): 2023
    re.compile(r'^\d{4}$'),
    # Month abbreviation + year: Jan-24, Jan-2024, Jan 2024, Jan'24
    re.compile(
        r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[- \'\u2019]?\d{2,4}$',
        re.IGNORECASE,
    ),
    # Year + month abbreviation: 2024-Jan, 24-Dec
    re.compile(
        r'^\d{2,4}[-/]?(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)$',
        re.IGNORECASE,
    ),
    # Year + numeric month: 2024-03, 2024/3
    re.compile(r'^\d{4}[-/]\d{1,2}$'),
    # Day-Month-Year or Month-Day-Year: 02-01-2024, 2/1/24, 01/02/2024
    re.compile(r'^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$'),
    # ISO date: 2024-01-02
    re.compile(r'^\d{4}[-/]\d{2}[-/]\d{2}$'),
    # Relative period markers (optionally with year): LTM, NTM, TTM, YTD
    re.compile(r'^(LTM|NTM|TTM|YTD)(\s?\d{4})?$', re.IGNORECASE),
]


def _is_financial_date(val):
    """Return True if *val* looks like a financial date or period label.

    Handles:
    * ``int`` / ``float`` — plain years 1900–2100
    * ``datetime.datetime`` / ``datetime.date`` — any real date
    * ``str`` — any of the patterns in ``_FINANCIAL_DATE_PATTERNS``:
      quarter labels (1Q2021, Q12023, 2024Q1, 2024-1Q, 20241Q),
      half-year (H12024, 2024H1), fiscal-year (FY2024, FYE24),
      year-with-suffix (2024E/A/F/B/P), plain year (2023),
      month-year (Jan-24, Jan-2024), date strings (02-01-2024,
      2024-03-31), relative labels (LTM, NTM, TTM, YTD), and
      calendar-year tags (CY2024).
    """
    if isinstance(val, (_dt.datetime, _dt.date)):
        return True
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return 1900 <= int(val) <= 2100
    if not isinstance(val, str):
        return False
    s = val.strip()
    return any(pat.match(s) for pat in _FINANCIAL_DATE_PATTERNS)


def _are_date_headers(col_headers):
    """Return True if the majority of *col_headers* values look like financial dates.

    Parameters
    ----------
    col_headers : dict[int, str]
        Mapping from column index → header string (as returned by
        :func:`_find_col_headers_in_source`).

    Returns
    -------
    bool
        True when ≥ 50 % of the header values are recognised financial dates.
    """
    if not col_headers:
        return False
    values = list(col_headers.values())
    date_count = sum(1 for v in values if _is_financial_date(v))
    return date_count / len(values) >= 0.5


# ──────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────

def _apply_header(ws, row, col, value):
    """Write a styled header cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.font  = _HEADER_FONT
    c.fill  = _HEADER_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    return c


def _read_mapping_inputs(wb_map):
    """Read all Input rows from every data sheet in the mapping report.

    Returns
    -------
    dict[str, list[dict]]
        Mapping of *sheet_name* → list of cell dicts:
        ``{col, row, col_idx, value, nf, cell}``
    """
    CELL_COL_NAME = "Cell"
    TYPE_COL_NAME = "Type"
    VAL_COL_NAME  = "Value"
    NF_COL_NAME   = "NumberFormat"

    result = {}
    for sn in wb_map.sheetnames:
        if sn == "_Metadata":
            continue
        ws = wb_map[sn]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        if TYPE_COL_NAME not in headers:
            continue
        tc  = headers.index(TYPE_COL_NAME) + 1
        cc  = headers.index(CELL_COL_NAME) + 1
        vc  = headers.index(VAL_COL_NAME)  + 1
        nfc = (headers.index(NF_COL_NAME)  + 1) if NF_COL_NAME in headers else None

        inputs = []
        for r in range(2, ws.max_row + 1):
            t = ws.cell(r, tc).value
            if t != "Input":
                continue
            ca  = ws.cell(r, cc).value
            val = ws.cell(r, vc).value
            nf  = ws.cell(r, nfc).value if nfc else None
            m   = _CELL_RE.match(str(ca or ""))
            if not m:
                continue
            inputs.append({
                "col":     m.group(1),
                "row":     int(m.group(2)),
                "col_idx": col_letter_to_index(m.group(1)),
                "value":   val,
                "nf":      nf,
                "cell":    str(ca),
            })
        result[sn] = inputs
    return result


def _group_into_vectors_and_scalars(inputs):
    """Split a flat list of input cells into horizontal vectors and scalars.

    A **vector** is a run of ≥ 2 cells in the *same row* with consecutive
    column indices.  Isolated cells are **scalars**.

    Returns
    -------
    tuple[list[list[dict]], list[dict]]
        ``(vectors, scalars)``
    """
    by_row = {}
    for inp in inputs:
        by_row.setdefault(inp["row"], []).append(inp)

    vectors = []
    scalars = []

    for _row, cells in sorted(by_row.items()):
        cells.sort(key=lambda x: x["col_idx"])
        run = [cells[0]]
        for cell in cells[1:]:
            if cell["col_idx"] == run[-1]["col_idx"] + 1:
                run.append(cell)
            else:
                if len(run) >= 2:
                    vectors.append(run[:])
                else:
                    scalars.extend(run)
                run = [cell]
        if len(run) >= 2:
            vectors.append(run[:])
        else:
            scalars.extend(run)

    return vectors, scalars


def _find_col_headers_in_source(ws_src, col_indices, max_data_row):
    """Find the header row for *col_indices* in *ws_src*.

    Looks at rows **above** *max_data_row* (i.e. pure header rows), picking
    the row that has the most non-null values in *col_indices*.  Prefers rows
    with integer year values or short strings over datetime objects.  Falls
    back to column letters if nothing is found.

    Parameters
    ----------
    max_data_row : int
        The smallest row number that contains actual data values among the
        vectors.  Headers must be in a row strictly less than this.

    Returns
    -------
    dict[int, str]
        Mapping of *col_index* → header string.
    """
    import datetime as _dt

    if ws_src is None:
        return {ci: index_to_col_letter(ci) for ci in col_indices}

    def _score_row(hr):
        """Return (non_null_count, is_year_like) for ranking."""
        vals = [ws_src.cell(hr, ci).value for ci in col_indices]
        non_null = sum(1 for v in vals if v is not None)
        year_like = sum(
            1 for v in vals
            if isinstance(v, int) and 1900 <= v <= 2100
        ) + sum(
            1 for v in vals
            if isinstance(v, str) and re.search(r'\b(19|20)\d{2}\b', v)
        )
        is_datetime = sum(1 for v in vals if isinstance(v, _dt.datetime))
        return non_null, year_like, -is_datetime  # prefer year-like, penalise datetimes

    best_row   = None
    best_score = (-1, -1, -1)

    search_up_to = min(max_data_row, ws_src.max_row + 1)
    for hr in range(1, search_up_to):
        score = _score_row(hr)
        if score > best_score:
            best_score = score
            best_row   = hr

    if best_row is None or best_score[0] == 0:
        return {ci: index_to_col_letter(ci) for ci in col_indices}

    import datetime as _dt
    result = {}
    for ci in col_indices:
        v = ws_src.cell(best_row, ci).value
        if v is None:
            result[ci] = index_to_col_letter(ci)
        elif isinstance(v, _dt.datetime):
            result[ci] = str(v.year)
        else:
            result[ci] = str(v)
    return result


def _find_row_label(ws_src, row, start_col_idx, label_search_depth=None):
    """Look for a **text** label to the *left* of *start_col_idx* in *ws_src*.

    Only returns string values (skips pure numbers, which are data values).
    Searches all the way to column 1 by default (``label_search_depth=None``).

    Returns
    -------
    str or None
    """
    if ws_src is None:
        return None
    max_depth = start_col_idx - 1 if label_search_depth is None else label_search_depth
    for offset in range(1, max_depth + 1):
        ci = start_col_idx - offset
        if ci < 1:
            break
        v = ws_src.cell(row, ci).value
        if v is not None and isinstance(v, str) and v.strip():
            return v.strip()
    return None


def _extract_header_vector(vectors):
    """Detect and remove ALL "column-header" vectors from *vectors*.

    A vector qualifies as a column-header row when *after* splitting off an
    optional leading label cell (via :func:`_split_label_from_vector`):

    * The remaining data cells are all year-like values (integers 1900–2100
      or strings matching a year pattern such as "2018E"), AND
    * The vector is in one of the first 6 rows of the sheet (for the main
      header) OR all values are year-like strings/ints.

    Returns
    -------
    tuple[dict[int,str], list[list[dict]]]
        ``(forced_col_headers, remaining_vectors)``
        ``forced_col_headers`` merges all detected header vectors.
    """
    import re as _re

    def _is_year_like(v):
        if isinstance(v, int):
            return 1900 <= v <= 2100
        if isinstance(v, float) and v == int(v):
            return 1900 <= int(v) <= 2100
        if isinstance(v, str):
            return bool(_re.fullmatch(r'(FY)?\d{4}[A-Z]*', v.strip()))
        return False

    forced        = {}
    remaining     = []

    for vec in vectors:
        # Split off potential string label
        _label, data_cells = _split_label_from_vector(vec)
        if not data_cells:
            remaining.append(vec)
            continue

        row_num   = vec[0]["row"]
        all_years = all(_is_year_like(inp["value"]) for inp in data_cells)

        if all_years and (row_num <= 6 or _label is not None):
            # This is a year-header row → add to forced headers, don't emit as data
            for inp in data_cells:
                forced[inp["col_idx"]] = str(inp["value"])
        else:
            remaining.append(vec)

    return forced if forced else None, remaining


def _split_label_from_vector(vec):
    """Detect if the first cell of *vec* is a metric-name label.

    When the first cell contains a non-numeric string it is treated as
    the row label (metric name) rather than a data point.

    Returns
    -------
    tuple[str | None, list[dict]]
        ``(label_or_None, data_cells)``
    """
    if not vec:
        return None, vec
    first = vec[0]
    val   = first["value"]
    if isinstance(val, str) and val.strip():
        # First cell is text → label column embedded in the vector
        return val.strip(), vec[1:]
    return None, vec


# ──────────────────────────────────────────────────────────────────
# Sheet builders
# ──────────────────────────────────────────────────────────────────

def _build_config_sheet(ws_config, sheet_scalars, wb_src):
    """Populate the Config sheet with scalar inputs."""
    _apply_header(ws_config, 1, 1, "Source Sheet")
    _apply_header(ws_config, 1, 2, "Cell Ref")
    _apply_header(ws_config, 1, 3, "Label")
    _apply_header(ws_config, 1, 4, "Value")

    ws_config.column_dimensions["A"].width = 22
    ws_config.column_dimensions["B"].width = 10
    ws_config.column_dimensions["C"].width = 45
    ws_config.column_dimensions["D"].width = 18

    # Freeze header row
    ws_config.freeze_panes = "A2"

    out_row    = 2
    index_rows = []
    prev_sheet = None

    for sn, scalars in sheet_scalars.items():
        if not scalars:
            continue

        ws_src = (wb_src[sn]
                  if wb_src is not None and sn in wb_src.sheetnames
                  else None)

        if sn != prev_sheet:
            # Section separator row
            sep = ws_config.cell(out_row, 1, f"── {sn} ──")
            sep.font = Font(bold=True, italic=True, color="4472C4")
            ws_config.merge_cells(
                start_row=out_row, start_column=1,
                end_row=out_row, end_column=4
            )
            out_row   += 1
            prev_sheet = sn

        for sc in scalars:
            label = _find_row_label(ws_src, sc["row"], sc["col_idx"])

            for col, val in enumerate([sn, sc["cell"], label, sc["value"]], 1):
                cell = ws_config.cell(out_row, col, val)
                cell.fill = _INPUT_FILL

            index_rows.append({
                "Input File Sheet": "Config",
                "Table": "Config",
                "Column Name": f"Value [{sc['cell']}]",
                "Source Sheet": sn,
                "Source Range": sc["cell"],
                "Description": label or "",
                "Vector Length": 1,
            })
            out_row += 1

    return index_rows


def _build_vector_sheet(wb_out, sn, vectors, wb_src, forced_col_headers=None):
    """Create and populate a vector-input sheet for one source sheet.

    Returns
    -------
    list[dict]
        Index rows describing every column written.
    """
    ws_src = (wb_src[sn]
              if wb_src is not None and sn in wb_src.sheetnames
              else None)

    # ---- Split each vector into (label, data_cells) ----
    label_data_pairs = []
    for vec in vectors:
        label, data_cells = _split_label_from_vector(vec)
        if not data_cells:          # vector was only 1 label cell → skip
            continue
        label_data_pairs.append((label, data_cells, vec))

    if not label_data_pairs:
        return []

    # ---- Determine the union of DATA column indices ----
    all_data_col_indices = sorted({
        inp["col_idx"]
        for _lbl, data_cells, _vec in label_data_pairs
        for inp in data_cells
    })

    # ---- Find column headers from source (rows above first data row) ----
    min_data_row = min(vec[0]["row"] for vec in vectors)
    if forced_col_headers:
        col_headers = {
            ci: forced_col_headers.get(ci, index_to_col_letter(ci))
            for ci in all_data_col_indices
        }
    else:
        col_headers = _find_col_headers_in_source(
            ws_src, all_data_col_indices, max_data_row=min_data_row
        )

    # ---- Detect whether column headers are financial dates → transpose ----
    transposed = _are_date_headers(col_headers)

    # ---- Create the sheet ----
    sheet_title = sn[:31]
    ws_out      = wb_out.create_sheet(sheet_title)

    index_rows = []

    if transposed:
        # ── TRANSPOSED LAYOUT ──────────────────────────────────────────────
        # Financial date/period labels recognised in column headers.
        # Layout: col A = period label (rows grow downward as new periods are
        # added); row 1 = metric names (one column per metric).
        # Users add new time periods by appending rows – no schema changes.

        period_order = [
            (ci, col_headers.get(ci, index_to_col_letter(ci)))
            for ci in all_data_col_indices
        ]

        # Pre-resolve metric labels (may require a source-sheet look-up)
        metric_labels = []
        for label, data_cells, vec in label_data_pairs:
            row_num       = vec[0]["row"]
            start_col_idx = data_cells[0]["col_idx"] if data_cells else vec[0]["col_idx"]
            if label is None:
                label = _find_row_label(ws_src, row_num, start_col_idx)
            metric_labels.append(label or f"Row {row_num}")

        # Pre-build per-metric data lookup: col_idx → value
        metric_data_lookup = [
            {inp["col_idx"]: inp["value"] for inp in data_cells}
            for _lbl, data_cells, _vec in label_data_pairs
        ]

        # ---- Header row ----
        _apply_header(ws_out, 1, 1, f"Period\n[Source: '{sn}']")
        ws_out.column_dimensions["A"].width = 18

        for metric_idx, (metric_lbl, (_lbl, data_cells, vec)) in enumerate(
            zip(metric_labels, label_data_pairs)
        ):
            out_col  = metric_idx + 2
            row_num  = vec[0]["row"]
            start_ci = data_cells[0]["col_idx"] if data_cells else vec[0]["col_idx"]
            end_ci   = data_cells[-1]["col_idx"] if data_cells else start_ci
            _apply_header(ws_out, 1, out_col, metric_lbl)
            ws_out.column_dimensions[get_column_letter(out_col)].width = 16
            index_rows.append({
                "Input File Sheet": sheet_title,
                "Table": sheet_title,
                "Column Name": metric_lbl,
                "Source Sheet": sn,
                "Source Range": (
                    f"{index_to_col_letter(start_ci)}{row_num}"
                    f":{index_to_col_letter(end_ci)}{row_num}"
                ),
                "Description": metric_lbl,
                "Vector Length": len(period_order),
            })

        ws_out.freeze_panes = "B2"

        # ---- Data rows: one row per period ----
        for out_row, (period_ci, period_label) in enumerate(period_order, start=2):
            row_fill = _INPUT_FILL if out_row % 2 == 0 else _ALT_FILL

            period_cell = ws_out.cell(out_row, 1, period_label)
            period_cell.font      = Font(bold=True)
            period_cell.alignment = Alignment(wrap_text=False)
            period_cell.fill      = row_fill

            for metric_idx in range(len(label_data_pairs)):
                out_col = metric_idx + 2
                val     = metric_data_lookup[metric_idx].get(period_ci)
                c       = ws_out.cell(out_row, out_col, val)
                c.fill  = row_fill

        # ---- Source note at bottom ----
        if label_data_pairs:
            nrow      = len(period_order) + 2
            first_vec = label_data_pairs[0][2]
            last_vec  = label_data_pairs[-1][2]
            start_ref = (f"{index_to_col_letter(first_vec[0]['col_idx'])}"
                         f"{first_vec[0]['row']}")
            end_ref   = (f"{index_to_col_letter(last_vec[-1]['col_idx'])}"
                         f"{last_vec[-1]['row']}")
            note = ws_out.cell(
                nrow, 1,
                f"Source: '{sn}'  |  cells {start_ref}\u2013{end_ref}"
                f"  [layout: dates as rows, metrics as columns]"
            )
            note.font = Font(italic=True, color="808080", size=9)
            ws_out.merge_cells(
                start_row=nrow, start_column=1,
                end_row=nrow, end_column=len(label_data_pairs) + 1
            )

    else:
        # ── ORIGINAL LAYOUT ────────────────────────────────────────────────
        # Column headers are not dates.  Metrics run DOWN rows (col A = metric
        # label); non-date period values run ACROSS columns.

        # Map data col_idx → output column number (B onwards = col 2+)
        col_idx_to_out_col = {
            ci: i + 2
            for i, ci in enumerate(all_data_col_indices)
        }

        # ---- Header row ----
        _apply_header(ws_out, 1, 1, f"Metric\n[Source: '{sn}']")
        ws_out.column_dimensions["A"].width = 42

        for ci, out_col in col_idx_to_out_col.items():
            col_letter = index_to_col_letter(ci)
            period_lbl = col_headers.get(ci, col_letter)
            col_name   = f"{period_lbl} [{col_letter}]"
            _apply_header(ws_out, 1, out_col, col_name)
            ws_out.column_dimensions[get_column_letter(out_col)].width = 14
            index_rows.append({
                "Input File Sheet": sheet_title,
                "Table": sheet_title,
                "Column Name": col_name,
                "Source Sheet": sn,
                "Source Range": f"{col_letter}* (row varies)",
                "Description": str(period_lbl),
                "Vector Length": len(label_data_pairs),
            })

        ws_out.freeze_panes = "B2"

        # ---- Data rows ----
        for out_row, (label, data_cells, vec) in enumerate(label_data_pairs, start=2):
            row_num       = vec[0]["row"]
            start_col_idx = data_cells[0]["col_idx"] if data_cells else vec[0]["col_idx"]

            # If no embedded label, try looking left in source
            if label is None:
                label = _find_row_label(ws_src, row_num, start_col_idx)

            label_cell = ws_out.cell(out_row, 1, label or f"Row {row_num}")
            label_cell.font      = Font(bold=True)
            label_cell.alignment = Alignment(wrap_text=True)

            # Alternate row shading
            row_fill          = _INPUT_FILL if out_row % 2 == 0 else _ALT_FILL
            label_cell.fill   = row_fill

            for inp in data_cells:
                oc = col_idx_to_out_col.get(inp["col_idx"])
                if oc is None:
                    continue
                c = ws_out.cell(out_row, oc, inp["value"])
                c.fill = row_fill

        # ---- Source note at bottom ----
        if label_data_pairs:
            nrow      = len(label_data_pairs) + 2
            first_vec = label_data_pairs[0][2]
            last_vec  = label_data_pairs[-1][2]
            start_ref = (f"{index_to_col_letter(first_vec[0]['col_idx'])}"
                         f"{first_vec[0]['row']}")
            end_ref   = (f"{index_to_col_letter(last_vec[-1]['col_idx'])}"
                         f"{last_vec[-1]['row']}")
            note = ws_out.cell(
                nrow, 1,
                f"Source: '{sn}'  |  input cells from {start_ref} to {end_ref}"
            )
            note.font = Font(italic=True, color="808080", size=9)
            ws_out.merge_cells(
                start_row=nrow, start_column=1,
                end_row=nrow, end_column=len(all_data_col_indices) + 1
            )

    return index_rows


def _build_index_sheet(ws_idx, index_rows):
    """Populate the Index sheet."""
    idx_headers = [
        "Input File Sheet",
        "Table",
        "Column Name",
        "Source Sheet",
        "Source Range",
        "Description",
        "Vector Length",
    ]
    col_widths = [22, 22, 30, 22, 24, 45, 14]

    for ci, (h, w) in enumerate(zip(idx_headers, col_widths), 1):
        _apply_header(ws_idx, 1, ci, h)
        ws_idx.column_dimensions[get_column_letter(ci)].width = w

    ws_idx.freeze_panes = "A2"

    for ri, row in enumerate(index_rows, start=2):
        for ci, h in enumerate(idx_headers, 1):
            ws_idx.cell(ri, ci, row.get(h, ""))


# ──────────────────────────────────────────────────────────────────
# Public entry point
# ──────────────────────────────────────────────────────────────────

def generate_structured_input(mapping_path, excel_path=None, output_path=None):
    """Generate a user-friendly structured input Excel workbook.

    Parameters
    ----------
    mapping_path : str
        Path to the mapping report generated by
        :func:`mapper.generate_mapping_report`.
    excel_path : str or None
        Path to the original source Excel workbook.  When provided, the
        generator uses it to look up row labels (metric names) and column
        headers (year / period labels).
    output_path : str or None
        Where to write the structured input file.  Defaults to
        ``<mapping_dir>/structured_input.xlsx``.

    Returns
    -------
    str
        Path to the generated file.
    """
    # ── 1. Load files ──────────────────────────────────────────────
    wb_map = load_workbook(mapping_path)
    wb_src = (load_workbook(excel_path, data_only=True)
              if excel_path else None)

    # ── 2. Resolve output path ────────────────────────────────────
    if output_path is None:
        out_dir = os.path.join(
            os.path.dirname(mapping_path) or ".", "output"
        )
        os.makedirs(out_dir, exist_ok=True)
        output_path = os.path.join(
            os.path.dirname(mapping_path) or ".", "structured_input.xlsx"
        )
    else:
        parent = os.path.dirname(output_path)
        if parent:
            os.makedirs(parent, exist_ok=True)

    # ── 3. Read Input cells from mapping report ────────────────────
    sheet_inputs = _read_mapping_inputs(wb_map)

    # ── 4. Split into vectors and scalars per sheet ────────────────
    sheet_vectors = {}
    sheet_scalars = {}
    for sn, inputs in sheet_inputs.items():
        vecs, scals = _group_into_vectors_and_scalars(inputs)
        # Post-process: vectors where only 1 data cell remains after label
        # extraction are effectively scalar key-value pairs → move to Config.
        true_vecs = []
        for vec in vecs:
            _lbl, data_cells = _split_label_from_vector(vec)
            if len(data_cells) <= 1:
                scals.extend(vec)   # all cells (label + data) → scalars
            else:
                true_vecs.append(vec)
        sheet_vectors[sn] = true_vecs
        sheet_scalars[sn] = scals

    # ── 5. Build output workbook ───────────────────────────────────
    wb_out      = Workbook()
    has_default = "Sheet" in wb_out.sheetnames

    all_index_rows = []

    # Config sheet
    ws_config  = wb_out.create_sheet("Config")
    config_idx = _build_config_sheet(ws_config, sheet_scalars, wb_src)
    all_index_rows.extend(config_idx)

    # One vector sheet per source sheet (skip if no vectors)
    for sn, vecs in sheet_vectors.items():
        if not vecs:
            continue
        # Extract year-header vector if present
        forced_hdrs, vecs = _extract_header_vector(vecs)
        if not vecs:
            continue  # only had a header vector, no data
        vec_idx = _build_vector_sheet(wb_out, sn, vecs, wb_src,
                                      forced_col_headers=forced_hdrs)
        all_index_rows.extend(vec_idx)

    # Index sheet
    ws_idx = wb_out.create_sheet("Index")
    _build_index_sheet(ws_idx, all_index_rows)

    # ── 6. Sheet order: Index → Config → source sheets ────────────
    if has_default and "Sheet" in wb_out.sheetnames and len(wb_out.sheetnames) > 1:
        del wb_out["Sheet"]

    # Move Index to position 0
    wb_out.move_sheet("Index", offset=-len(wb_out.sheetnames) + 1)
    # Move Config to position 1
    if "Config" in wb_out.sheetnames:
        config_pos = wb_out.sheetnames.index("Config")
        wb_out.move_sheet("Config", offset=1 - config_pos)

    # ── 7. Save ────────────────────────────────────────────────────
    wb_out.save(output_path)

    if wb_src:
        wb_src.close()
    wb_map.close()

    print(f"Generated structured input: {output_path}")
    return output_path
