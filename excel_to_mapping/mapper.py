"""
Core mapping logic: classify cells into Inputs / Calculations / Outputs
and generate a **tabular** Excel mapping report.

Reuses the workbook parser and cell classifier from ``excel_to_python``
and the vectorised grouping logic from ``excel_to_python_vectorized``.
"""

import json
import os
import re
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from excel_to_python import (
    load_config,
    parse_workbook,
    classify_cells,
    find_all_references,
    col_letter_to_index,
    index_to_col_letter,
)

from excel_to_python_vectorized.vectorizer import (
    group_formulas,
    extract_references,
)


# ------------------------------------------------------------------
# Determine which formula cells are "outputs" (not referenced by others)
# ------------------------------------------------------------------

def _build_all_referenced_cells(formula_cells):
    """Return the set of (sheet, col, row) cells referenced by any formula."""
    referenced = set()
    for sheet, col, row, formula, _ci in formula_cells:
        refs = extract_references(formula, sheet)
        for ref in refs:
            if ref.kind == "cell":
                referenced.add((ref.sheet, ref.col, ref.row))
            elif ref.kind == "range":
                ci1 = col_letter_to_index(ref.col)
                ci2 = col_letter_to_index(ref.end_col)
                for r in range(ref.row, ref.end_row + 1):
                    for c in range(ci1, ci2 + 1):
                        referenced.add((ref.sheet, index_to_col_letter(c), r))
    return referenced


def _classify_formula_cells(formula_cells):
    """Split formula cells into *calculations* and *outputs*.

    A formula cell is an **output** if no other formula references it.
    Otherwise it is a **calculation** (intermediate result).
    """
    referenced = _build_all_referenced_cells(formula_cells)

    calculations = []
    outputs = []
    for entry in formula_cells:
        sheet, col, row, formula, cell_info = entry
        if (sheet, col, row) in referenced:
            calculations.append(entry)
        else:
            outputs.append(entry)
    return calculations, outputs


# ------------------------------------------------------------------
# Column definitions for the tabular report
# ------------------------------------------------------------------

COLUMNS = [
    "Sheet", "Cell", "Type", "Formula", "Value",
    "GroupID", "GroupDirection", "GroupSize", "PatternFormula",
    "NumberFormat", "FontBold", "FontItalic", "FontSize", "FontColor",
    "FillColor", "HorizAlign", "VertAlign", "WrapText", "IncludeFlag",
]

_TYPE_ORDER = {"Input": 0, "Calculation": 1, "Output": 2}


# ------------------------------------------------------------------
# Formatting extraction helper
# ------------------------------------------------------------------

def _extract_formatting(cell_info):
    """Extract formatting fields from a cell_info dict."""
    font = cell_info.get("font") or {}
    align = cell_info.get("alignment") or {}
    return {
        "NumberFormat": cell_info.get("number_format"),
        "FontBold": bool(font.get("bold")),
        "FontItalic": bool(font.get("italic")),
        "FontSize": font.get("size"),
        "FontColor": font.get("color"),
        "FillColor": cell_info.get("fill_color"),
        "HorizAlign": align.get("horizontal"),
        "VertAlign": align.get("vertical"),
        "WrapText": bool(align.get("wrap_text")),
    }


# ------------------------------------------------------------------
# Build flat tabular rows for one sheet
# ------------------------------------------------------------------

def _sort_key(row_dict):
    """Sort by Type (Input < Calculation < Output), then row, then column."""
    cell = str(row_dict.get("Cell") or "")
    # For ranges like "C2:C6", use the start cell for sorting
    start = cell.split(":")[0]
    m = re.match(r"([A-Z]+)(\d+)", start)
    if not m:
        return (_TYPE_ORDER.get(row_dict["Type"], 9), 0, 0)
    return (_TYPE_ORDER.get(row_dict["Type"], 9),
            int(m.group(2)), col_letter_to_index(m.group(1)))


def _build_sheet_rows(sheet_name, hardcoded_cells, formula_cells):
    """Build tabular rows for *sheet_name*.

    Returns a sorted list of dicts (keys from :data:`COLUMNS`), one per
    input cell, formula group, or single formula cell.
    """
    rows = []

    # --- classify formula cells (using ALL formulas for cross-sheet deps) ---
    calcs_all, outs_all = _classify_formula_cells(formula_cells)
    calc_set = {(s, c, r) for s, c, r, *_ in calcs_all}

    # --- inputs ---
    for s, col, row, val, ci in hardcoded_cells:
        if s != sheet_name:
            continue
        fmt = _extract_formatting(ci)
        rows.append({
            "Sheet": sheet_name,
            "Cell": f"{col}{row}",
            "Type": "Input",
            "Formula": None,
            "Value": val,
            "GroupID": None,
            "GroupDirection": None,
            "GroupSize": None,
            "PatternFormula": None,
            **fmt,
            "IncludeFlag": True,
        })

    # --- formula cells on this sheet ---
    sheet_formulas = [e for e in formula_cells if e[0] == sheet_name]
    if not sheet_formulas:
        rows.sort(key=_sort_key)
        return rows

    groups, singles = group_formulas(sheet_formulas)

    grouped_keys = set()
    group_counter = 0

    for g in groups:
        cells = g["cells"]
        direction = g["direction"]
        s0, col0, row0, formula0, ci0 = cells[0]
        _, col1, row1, _, _ = cells[-1]

        if direction == "vertical":
            rng = f"{col0}{row0}:{col0}{row1}"
        else:
            rng = f"{col0}{row0}:{col1}{row0}"

        cell_type = "Calculation" if (s0, col0, row0) in calc_set else "Output"

        group_counter += 1
        group_id = f"{sheet_name}_G{group_counter}"
        fmt = _extract_formatting(ci0)

        rows.append({
            "Sheet": sheet_name,
            "Cell": rng,
            "Type": cell_type,
            "Formula": formula0,
            "Value": None,
            "GroupID": group_id,
            "GroupDirection": direction,
            "GroupSize": len(cells),
            "PatternFormula": formula0,
            **fmt,
            "IncludeFlag": True,
        })

        for cell in cells:
            grouped_keys.add((cell[0], cell[1], cell[2]))

    for entry in singles:
        s, col, row, formula, ci = entry
        if (s, col, row) in grouped_keys:
            continue
        cell_type = "Calculation" if (s, col, row) in calc_set else "Output"
        fmt = _extract_formatting(ci)
        rows.append({
            "Sheet": sheet_name,
            "Cell": f"{col}{row}",
            "Type": cell_type,
            "Formula": formula,
            "Value": None,
            "GroupID": None,
            "GroupDirection": None,
            "GroupSize": None,
            "PatternFormula": None,
            **fmt,
            "IncludeFlag": True,
        })

    rows.sort(key=_sort_key)
    return rows


# ------------------------------------------------------------------
# Report generation
# ------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4",
                           fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def generate_mapping_report(excel_path, sheet_names=None,
                            config_path=None, output_path=None):
    """Generate the mapping report Excel file.

    Produces one sheet per source sheet with a flat tabular layout
    (columns defined in :data:`COLUMNS`) plus a ``_Metadata`` sheet.

    Parameters
    ----------
    excel_path : str
        Path to the source Excel workbook.
    sheet_names : list[str] or None
        Sheets to include.  ``None`` means *all* sheets.
    config_path : str or None
        Optional path to config YAML.
    output_path : str or None
        Where to write the report.  Defaults to
        ``<excel_dir>/output/mapping_report.xlsx``.

    Returns
    -------
    str
        Path to the generated report file.
    """
    config = load_config(config_path)

    wb_src = load_workbook(excel_path)
    sheets, tables = parse_workbook(wb_src)
    formula_cells, hardcoded_cells = classify_cells(sheets, tables)

    if sheet_names is None:
        sheet_names = list(sheets.keys())
    else:
        sheet_names = [s for s in sheet_names if s in sheets]

    if output_path is None:
        out_dir = os.path.join(os.path.dirname(excel_path) or ".", "output")
        os.makedirs(out_dir, exist_ok=True)
        output_path = os.path.join(out_dir, "mapping_report.xlsx")
    else:
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb_report = Workbook()
    has_default = "Sheet" in wb_report.sheetnames

    for sn in sheet_names:
        ws = wb_report.create_sheet(sn)

        # Write header row
        for ci, h in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL

        # Build and write data rows
        data_rows = _build_sheet_rows(sn, hardcoded_cells, formula_cells)
        for ri, row_dict in enumerate(data_rows, 2):
            for ci, col_name in enumerate(COLUMNS, 1):
                ws.cell(row=ri, column=ci, value=row_dict.get(col_name))

        # Auto-width for key columns
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 14

    # --- _Metadata sheet ---
    ws_meta = wb_report.create_sheet("_Metadata")
    meta_headers = ["SheetName", "MergedCells", "ColWidths", "RowHeights"]
    for ci, h in enumerate(meta_headers, 1):
        cell = ws_meta.cell(row=1, column=ci, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for ri, sn in enumerate(sheet_names, 2):
        sd = sheets[sn]
        ws_meta.cell(row=ri, column=1, value=sn)
        ws_meta.cell(row=ri, column=2,
                     value=";".join(sd.get("merged_cells", [])))
        ws_meta.cell(row=ri, column=3,
                     value=json.dumps(
                         {str(k): v
                          for k, v in sd.get("col_widths", {}).items()}))
        ws_meta.cell(row=ri, column=4,
                     value=json.dumps(
                         {str(k): v
                          for k, v in sd.get("row_heights", {}).items()}))

    # Remove the default "Sheet" if we added real sheets
    if has_default and "Sheet" in wb_report.sheetnames and len(wb_report.sheetnames) > 1:
        del wb_report["Sheet"]

    wb_report.save(output_path)
    wb_src.close()

    print(f"Generated mapping report: {output_path}")
    return output_path
